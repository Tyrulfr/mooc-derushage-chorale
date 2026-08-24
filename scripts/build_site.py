from __future__ import annotations

import csv
import io
import re
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
import unicodedata
import json
from urllib.parse import quote

from lib_derushage import (
    ROOT,
    SITE,
    capsule_bab_duration,
    capsule_duration,
    correct_asr_greffet,
    escape,
    find_overlaps,
    format_seconds,
    html_breadcrumb,
    html_page,
    index_by_id,
    load_affectations,
    load_bab_encode,
    load_bab_encode_index,
    load_capsules,
    load_derushage_edito,
    load_derushage_edito_index,
    load_experts_profils,
    load_programme_table,
    load_segments,
    merge_bab_encode_blocs,
    bab_encode_stats,
    render_bab_encode_export,
    slug,
    total_score,
    write_text,
)

TEST_MAIL_RECIPIENT = "christophe.dubois@universite-paris-saclay.fr"
REVIEW_MAIL_RECIPIENT = "ritanoelle.moussa@agroparistech.fr"


STYLE = """
:root {
  color-scheme: light;
  --ink: #15202b;
  --muted: #5c6b7a;
  --line: #dde4ec;
  --panel: #f0f4f8;
  --surface: #ffffff;
  --accent: #0b6e77;
  --accent-dark: #08545b;
  --accent-soft: #e6f4f5;
  --warn: #b45309;
  --ok: #0f766e;
  --shadow: 0 10px 30px rgba(21, 32, 43, 0.06);
  --radius: 12px;
}
* { box-sizing: border-box; }
html { scroll-behavior: smooth; }
body {
  margin: 0;
  min-height: 100vh;
  display: flex;
  flex-direction: column;
  font-family: "Segoe UI", -apple-system, BlinkMacSystemFont, sans-serif;
  color: var(--ink);
  background: var(--panel);
  line-height: 1.5;
}
a { color: var(--accent); text-decoration: none; font-weight: 600; }
a:hover { color: var(--accent-dark); text-decoration: underline; }
code {
  font-family: ui-monospace, SFMono-Regular, Menlo, monospace;
  font-size: 0.92em;
  background: var(--panel);
  padding: 0.1em 0.35em;
  border-radius: 4px;
}
.site-header {
  position: sticky;
  top: 0;
  z-index: 50;
  background: rgba(255, 255, 255, 0.96);
  border-bottom: 1px solid var(--line);
  backdrop-filter: blur(8px);
  box-shadow: 0 1px 0 rgba(21, 32, 43, 0.04);
}
.site-header__inner {
  width: min(1180px, calc(100vw - 32px));
  margin: 0 auto;
  padding: 14px 0;
  display: flex;
  align-items: center;
  justify-content: space-between;
  gap: 20px;
  flex-wrap: wrap;
}
.site-brand {
  display: flex;
  align-items: center;
  gap: 12px;
  color: inherit;
  text-decoration: none;
  font-weight: 700;
}
.site-brand:hover { text-decoration: none; color: inherit; }
.site-brand__mark {
  width: 40px;
  height: 40px;
  border-radius: 10px;
  display: grid;
  place-items: center;
  background: linear-gradient(145deg, var(--accent), var(--accent-dark));
  color: #fff;
  font-size: 14px;
}
.site-brand__text { display: flex; flex-direction: column; gap: 2px; }
.site-brand__title { font-size: 16px; line-height: 1.2; }
.site-brand__tagline {
  font-size: 12px;
  font-weight: 500;
  color: var(--muted);
}
.site-nav {
  display: flex;
  flex-wrap: wrap;
  gap: 6px;
}
.site-nav__link {
  padding: 8px 12px;
  border-radius: 999px;
  font-size: 14px;
  font-weight: 600;
  color: var(--muted);
  text-decoration: none;
}
.site-nav__link:hover {
  color: var(--accent-dark);
  background: var(--accent-soft);
  text-decoration: none;
}
.site-nav__link--current {
  color: #fff;
  background: var(--accent);
}
.site-nav__link--current:hover {
  color: #fff;
  background: var(--accent-dark);
}
.site-main {
  flex: 1;
  width: min(1180px, calc(100vw - 32px));
  margin: 0 auto;
  padding: 28px 0 48px;
}
.page-home .page-content { padding-top: 0; }
.breadcrumb {
  display: flex;
  flex-wrap: wrap;
  gap: 6px;
  list-style: none;
  margin: 0 0 18px;
  padding: 0;
  font-size: 13px;
  color: var(--muted);
}
.breadcrumb li:not(:last-child)::after {
  content: "›";
  margin-left: 6px;
  color: #9aa8b5;
}
.breadcrumb a { font-weight: 500; }
.page-head { margin-bottom: 22px; }
.page-head h1 {
  margin: 0;
  font-size: clamp(26px, 4vw, 34px);
  line-height: 1.15;
  letter-spacing: -0.02em;
}
.page-head .lead {
  margin: 10px 0 0;
  font-size: 17px;
  color: var(--muted);
  max-width: 62ch;
}
.page-content > h2,
.page-content h2 {
  font-size: 20px;
  margin: 32px 0 14px;
}
.page-content > h2:first-child,
.page-content h2:first-child { margin-top: 0; }
.hero {
  margin: 0 0 36px;
  padding: 40px 32px;
  border-radius: calc(var(--radius) + 4px);
  background:
    radial-gradient(circle at top right, rgba(11, 110, 119, 0.14), transparent 42%),
    linear-gradient(160deg, #ffffff 0%, #eef6f7 100%);
  border: 1px solid var(--line);
  box-shadow: var(--shadow);
}
.hero__eyebrow {
  margin: 0 0 10px;
  font-size: 12px;
  font-weight: 700;
  letter-spacing: 0.08em;
  text-transform: uppercase;
  color: var(--accent);
}
.hero h1 {
  margin: 0 0 12px;
  font-size: clamp(30px, 5vw, 42px);
  line-height: 1.1;
  letter-spacing: -0.03em;
}
.hero__lead {
  margin: 0 0 22px;
  font-size: 18px;
  line-height: 1.55;
  color: var(--muted);
  max-width: 58ch;
}
.hero__stats {
  display: flex;
  flex-wrap: wrap;
  gap: 10px;
}
.stat-pill {
  display: inline-flex;
  align-items: center;
  gap: 8px;
  padding: 8px 14px;
  border-radius: 999px;
  background: #fff;
  border: 1px solid var(--line);
  font-size: 14px;
}
.stat-pill strong { color: var(--accent-dark); }
.section-title {
  margin: 0 0 16px;
  font-size: 13px;
  font-weight: 700;
  letter-spacing: 0.06em;
  text-transform: uppercase;
  color: var(--muted);
}
.sommaire-grid {
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
  gap: 16px;
}
.sommaire-card {
  display: flex;
  flex-direction: column;
  gap: 10px;
  min-height: 150px;
  padding: 22px;
  border: 1px solid var(--line);
  border-radius: var(--radius);
  background: var(--surface);
  color: inherit;
  text-decoration: none;
  box-shadow: var(--shadow);
  transition: transform .15s ease, border-color .15s ease, box-shadow .15s ease;
}
.sommaire-card:hover {
  transform: translateY(-2px);
  border-color: var(--accent);
  box-shadow: 0 14px 34px rgba(11, 110, 119, 0.12);
  text-decoration: none;
  color: inherit;
}
.sommaire-card__icon {
  width: 36px;
  height: 36px;
  border-radius: 9px;
  display: grid;
  place-items: center;
  background: var(--accent-soft);
  color: var(--accent-dark);
  font-size: 18px;
}
.sommaire-card h2 {
  margin: 0;
  font-size: 18px;
  color: var(--ink);
}
.sommaire-card p {
  margin: 0;
  font-size: 14px;
  line-height: 1.5;
  color: var(--muted);
  font-weight: 400;
}
.sommaire-card__cta {
  margin-top: auto;
  font-size: 13px;
  font-weight: 700;
  color: var(--accent);
}
.stats-grid {
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
  gap: 14px;
  margin-bottom: 28px;
}
.stat-card {
  padding: 18px 20px;
  border: 1px solid var(--line);
  border-radius: var(--radius);
  background: var(--surface);
  box-shadow: var(--shadow);
}
.stat-card__label {
  font-size: 13px;
  font-weight: 700;
  text-transform: uppercase;
  letter-spacing: 0.04em;
  color: var(--muted);
  margin-bottom: 8px;
}
.stat-card__value {
  font-size: 28px;
  font-weight: 800;
  line-height: 1.1;
  color: var(--accent-dark);
  margin-bottom: 6px;
}
.stat-card__meta { font-size: 14px; color: var(--muted); }
.panel {
  padding: 20px;
  border: 1px solid var(--line);
  border-radius: var(--radius);
  background: var(--surface);
  box-shadow: var(--shadow);
}
.table-wrap {
  overflow-x: auto;
  border: 1px solid var(--line);
  border-radius: var(--radius);
  background: var(--surface);
  box-shadow: var(--shadow);
}
table { width: 100%; border-collapse: collapse; margin: 0; }
th, td {
  border-bottom: 1px solid var(--line);
  padding: 12px 14px;
  text-align: left;
  vertical-align: top;
}
th {
  font-size: 12px;
  text-transform: uppercase;
  color: var(--muted);
  letter-spacing: 0.04em;
  background: #fafbfc;
}
tbody tr:hover { background: #fbfdfe; }
tbody tr:last-child td { border-bottom: none; }
.grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(260px, 1fr)); gap: 14px; }
.card {
  border: 1px solid var(--line);
  border-radius: var(--radius);
  padding: 16px 18px;
  background: var(--surface);
  box-shadow: var(--shadow);
}
.meta { color: var(--muted); font-size: 14px; }
.tag {
  display: inline-block;
  border: 1px solid var(--line);
  border-radius: 999px;
  padding: 3px 10px;
  margin: 2px 4px 2px 0;
  font-size: 12px;
  font-weight: 600;
  background: #fff;
}
.chip-grid {
  display: flex;
  flex-wrap: wrap;
  gap: 10px;
  margin-top: 8px;
}
.chip {
  display: inline-flex;
  align-items: center;
  padding: 10px 14px;
  border-radius: 999px;
  border: 1px solid var(--line);
  background: #fff;
  font-weight: 600;
  text-decoration: none;
}
.chip:hover {
  border-color: var(--accent);
  background: var(--accent-soft);
  text-decoration: none;
}
.status {
  display: inline-block;
  padding: 3px 10px;
  border-radius: 999px;
  font-size: 12px;
  font-weight: 700;
  letter-spacing: 0.02em;
  background: #eef2f6;
  color: #475569;
}
.status--progress { background: #e0f2fe; color: #0369a1; }
.status--ok { background: #d1fae5; color: #047857; }
.status--warn { background: #ffedd5; color: #c2410c; }
.warn { color: var(--warn); font-weight: 700; }
.script {
  white-space: pre-wrap;
  background: #f8fafc;
  border: 1px solid var(--line);
  padding: 16px;
  border-radius: var(--radius);
  font-size: 14px;
  line-height: 1.55;
}
.script-ref {
  font-size: 0.92em;
  color: #94a3b8;
}
.mail-ready {
  font-family: Aptos, "Segoe UI", Arial, sans-serif;
}
.methodology-panel {
  margin-top: 28px;
  padding: 20px;
  border: 1px solid var(--line);
  border-radius: var(--radius);
  background: var(--surface);
  box-shadow: var(--shadow);
}
.methodology-panel h2 { margin-top: 0; }
.methodology-panel ul { margin: 12px 0 16px; padding-left: 1.25rem; }
.methodology-panel li { margin-bottom: 8px; }
.unites-table th { font-size: 12px; }
.orientation-expert { margin-top: 28px; }
.orientation-expert h3 { margin: 20px 0 8px; font-size: 17px; }
.brief-intervenant-panel { border-left: 4px solid var(--accent); }
.brief-intervenant-panel .meta-lead { font-size: 15px; line-height: 1.5; margin-bottom: 20px; }
.brief-video { margin-top: 24px; padding-top: 20px; border-top: 1px solid var(--line); }
.brief-video:first-of-type { margin-top: 0; padding-top: 0; border-top: none; }
.brief-video h3 { margin: 0 0 8px; font-size: 18px; }
.brief-unites { margin: 16px 0; padding-left: 1.25rem; }
.brief-unites li { margin-bottom: 10px; }
.brief-temoin-phrases { margin: 8px 0 0; padding-left: 1.25rem; }
.brief-temoin-phrases li { margin-bottom: 6px; }
.brief-precaution {
  margin: 20px 0;
  padding: 14px 16px;
  background: #fffbeb;
  border: 1px solid #fde68a;
  border-radius: var(--radius);
  line-height: 1.5;
}
.brief-fascicule {
  margin: 12px 0;
  padding: 12px 14px;
  background: #f0fdf4;
  border: 1px solid #bbf7d0;
  border-radius: var(--radius);
  line-height: 1.5;
}
.script-placeholder {
  margin-top: 16px;
  min-height: 160px;
  padding: 28px 20px;
  border: 2px dashed var(--line);
  border-radius: var(--radius);
  background: #f8fafc;
  color: var(--muted);
  display: grid;
  place-items: center;
  text-align: center;
}
.script-recu-block {
  margin-top: 14px;
  padding: 16px 18px;
  border: 1px solid var(--line);
  border-radius: var(--radius);
  background: #f8fafc;
  line-height: 1.65;
  white-space: normal;
}
.script-attente-panel { border-left: 4px solid var(--warn); }
.script-revue-panel { border-left: 4px solid #0b6e77; }
.script-revue-demandes {
  margin: 12px 0 0;
  padding-left: 1.25rem;
}
.script-revue-demandes li { margin-bottom: 6px; }
.script-revue-mail {
  margin-top: 14px;
  padding: 14px 16px;
  border: 1px dashed var(--line);
  border-radius: var(--radius);
  background: #fff;
  line-height: 1.55;
  white-space: pre-wrap;
  font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace;
  font-size: 13px;
}
.script-revue-legende {
  margin: 10px 0 0;
  padding: 10px 12px;
  background: #f8fafc;
  border: 1px solid var(--line);
  border-radius: var(--radius);
  font-size: 13px;
}
.script-hs {
  display: inline;
  background: #fee2e2;
  border-bottom: 2px solid #ef4444;
  padding: 1px 2px;
}
.script-hs__tag {
  display: inline-block;
  margin: 0 4px 0 0;
  padding: 1px 6px;
  border-radius: 4px;
  background: #ef4444;
  color: #fff;
  font-size: 11px;
  font-weight: 600;
  letter-spacing: 0.02em;
  vertical-align: baseline;
}
.script-rc {
  display: inline;
  background: #ffedd5;
  border-bottom: 2px solid #f97316;
  padding: 1px 2px;
}
.script-rc__tag {
  display: inline-block;
  margin: 0 4px 0 0;
  padding: 1px 6px;
  border-radius: 4px;
  background: #f97316;
  color: #fff;
  font-size: 11px;
  font-weight: 600;
  letter-spacing: 0.02em;
  vertical-align: baseline;
}
.brief-point {
  margin: 14px 0;
  padding: 14px 16px;
  background: #f8fafc;
  border-radius: var(--radius);
  border: 1px solid var(--line);
}
.brief-point h4 { margin: 0 0 8px; font-size: 15px; }
.brief-point p { margin: 0 0 8px; }
.brief-point p:last-child { margin-bottom: 0; }
.brief-sequence { margin: 12px 0 0; padding-left: 1.25rem; }
.brief-sequence li { margin-bottom: 6px; }
.synthese-chorale-panel {
  margin-top: 28px;
  padding: 20px;
  border: 1px solid var(--line);
  border-radius: var(--radius);
  background: #fff;
  box-shadow: var(--shadow);
  border-left: 4px solid #64748b;
}
.synthese-chorale-list { margin: 12px 0 0; padding: 0; list-style: none; }
.synthese-chorale-list li {
  margin-bottom: 12px;
  padding: 12px 14px;
  background: #f8fafc;
  border-radius: var(--radius);
  border: 1px solid var(--line);
  line-height: 1.45;
}
.synthese-chorale-list strong { display: inline-block; min-width: 9rem; }
.synthese-objectifs-expert {
  margin: 16px 0 20px;
  padding: 14px 16px;
  background: #f8fafc;
  border: 1px solid var(--line);
  border-radius: var(--radius);
}
.synthese-objectifs-expert h3 { margin: 0 0 10px; font-size: 15px; }
.synthese-objectifs-expert ul { margin: 0; padding-left: 1.2rem; }
.synthese-objectifs-expert li { margin-bottom: 8px; line-height: 1.4; }
.synthese-chercheur {
  margin: 16px 0 0;
  padding: 14px 16px;
  border: 1px solid var(--line);
  border-radius: var(--radius);
  background: #fff;
}
.synthese-chercheur h3 { margin: 0 0 8px; font-size: 16px; }
.synthese-chercheur h4 { margin: 14px 0 8px; font-size: 14px; color: var(--muted); }
.synthese-chercheur p { margin: 0 0 8px; line-height: 1.45; }
.synthese-appui-list { margin: 0; padding: 0; list-style: none; }
.synthese-appui-list li {
  margin-bottom: 10px;
  padding: 10px 12px;
  background: #f8fafc;
  border-radius: var(--radius);
  border: 1px solid var(--line);
  line-height: 1.45;
}
.synthese-appui-list li:last-child { margin-bottom: 0; }
.script-expertise-projete-panel {
  margin-top: 28px;
  border-left: 4px solid #b45309;
}
.script-expertise-disclaimer {
  margin: 0 0 16px;
  padding: 12px 14px;
  background: #fff7ed;
  border: 1px solid #fdba74;
  border-radius: var(--radius);
  line-height: 1.45;
}
.script-expertise-block {
  margin: 18px 0 0;
  padding: 16px;
  background: #fff;
  border: 1px solid var(--line);
  border-radius: var(--radius);
}
.script-expertise-block h3 { margin: 0 0 8px; font-size: 16px; }
.script-expertise-plan {
  margin: 12px 0 16px;
  padding: 12px 14px;
  background: #f8fafc;
  border: 1px solid var(--line);
  border-radius: var(--radius);
}
.script-expertise-plan h4 {
  margin: 0 0 8px;
  font-size: 14px;
}
.script-expertise-plan ul {
  margin: 0;
  padding-left: 1.25rem;
}
.script-expertise-plan li {
  margin-bottom: 6px;
  line-height: 1.4;
}
.script-expertise-plan li:last-child { margin-bottom: 0; }
.script-expertise-block .script {
  margin-top: 12px;
  white-space: pre-wrap;
}
.cadrage-block { margin-top: 28px; padding-top: 20px; border-top: 1px solid var(--line); }
.cadrage-block h3 { margin: 18px 0 8px; font-size: 16px; }
.cadrage-position { font-size: 13px; color: var(--muted); margin: 0 0 8px; }
.cadrage-quote {
  margin: 8px 0 12px;
  padding: 12px 14px;
  border-left: 3px solid var(--accent);
  background: var(--accent-soft);
  font-style: italic;
}
.cadrage-pancarte {
  margin: 8px 0 12px;
  padding: 12px 14px;
  background: #f8fafc;
  font-family: ui-monospace, monospace;
  font-size: 13px;
  white-space: pre-wrap;
}
.export-panel {
  margin-top: 28px;
  padding: 20px;
  border: 1px solid var(--line);
  border-radius: var(--radius);
  background: var(--surface);
  box-shadow: var(--shadow);
}
.export-panel p { margin: 0 0 14px; }
.btn {
  display: inline-block;
  border: 1px solid var(--accent);
  background: var(--accent);
  color: #fff;
  border-radius: 8px;
  padding: 10px 16px;
  font: inherit;
  font-weight: 650;
  cursor: pointer;
  text-decoration: none;
}
.btn:hover { filter: brightness(1.05); color: #fff; text-decoration: none; }
.btn-secondary { background: #fff; color: var(--accent); }
.modal {
  position: fixed;
  inset: 0;
  background: rgba(15, 23, 42, 0.45);
  display: grid;
  place-items: center;
  padding: 20px;
  z-index: 1000;
}
.modal[hidden] { display: none; }
.modal-card {
  width: min(520px, 100%);
  background: #fff;
  border-radius: 12px;
  border: 1px solid var(--line);
  box-shadow: 0 18px 50px rgba(15, 23, 42, 0.18);
  padding: 22px;
}
.modal-card h2 { margin: 0 0 8px; font-size: 22px; }
.modal-card p { margin: 0 0 16px; }
.export-field { margin-bottom: 14px; }
.export-field label {
  display: block;
  font-size: 13px;
  font-weight: 700;
  margin-bottom: 6px;
  color: var(--muted);
  text-transform: uppercase;
  letter-spacing: .02em;
}
.export-field input {
  width: 100%;
  border: 1px solid var(--line);
  border-radius: 8px;
  padding: 10px 12px;
  font: inherit;
}
.export-folder {
  display: flex;
  align-items: center;
  justify-content: space-between;
  gap: 12px;
  border: 1px dashed var(--line);
  border-radius: 8px;
  padding: 12px;
  background: #fff;
}
.export-folder span { color: var(--muted); font-size: 14px; }
.export-folder--disabled { opacity: 0.72; }
.export-folder--disabled .btn-secondary { cursor: not-allowed; }
.modal-actions { display: flex; justify-content: flex-end; gap: 10px; margin-top: 18px; }
.export-status { margin-top: 14px; font-size: 14px; min-height: 1.2em; }
.export-status.ok { color: var(--ok); }
.export-status.warn { color: var(--warn); }
.export-status.error { color: #b91c1c; }
.page-header {
  display: flex;
  justify-content: space-between;
  align-items: flex-start;
  gap: 16px;
  flex-wrap: wrap;
  margin-bottom: 22px;
}
.page-header h1 { margin: 0 0 6px; }
.export-toolbar { flex-shrink: 0; }
.bab-segment { margin-bottom: 16px; }
.bab-segment .capsules { margin: 8px 0; }
.capsule-tag.utilise { border-color: var(--ok); color: var(--ok); background: #ecfdf5; }
.capsule-tag.reserve { border-color: var(--warn); color: var(--warn); background: #fffbeb; }
.capsule-tag.candidat { border-color: #64748b; color: #475569; }
.capsule-tag.rejete { border-color: #cbd5e1; color: #94a3b8; }
.capsule-tag.non-encode { border-color: #cbd5e1; color: #64748b; background: #f8fafc; }
.bab-segment--non-encode { border-style: dashed; background: #fcfdff; }
.coupe-note { font-size: 13px; color: var(--muted); margin: 6px 0; }
#script-final.sr-export {
  position: absolute;
  width: 1px;
  height: 1px;
  padding: 0;
  margin: -1px;
  overflow: hidden;
  clip: rect(0, 0, 0, 0);
  white-space: pre-wrap;
  border: 0;
}
.site-footer {
  margin-top: auto;
  border-top: 1px solid var(--line);
  background: #fff;
}
.site-footer__inner {
  width: min(1180px, calc(100vw - 32px));
  margin: 0 auto;
  padding: 22px 0 28px;
  font-size: 14px;
  color: var(--muted);
}
.site-footer p { margin: 0 0 6px; }
.site-footer strong { color: var(--ink); }
.heatmap-grid {
  display: grid;
  gap: 20px;
}
.heatmap-card {
  background: #fff;
  border: 1px solid var(--line);
  border-radius: var(--radius);
  padding: 18px;
  box-shadow: var(--shadow);
}
.heatmap-card h2 { margin: 0 0 6px; }
.heatmap-wrap {
  overflow-x: auto;
  border: 1px solid var(--line);
  border-radius: 10px;
  background: #fff;
}
.heatmap-table {
  width: 100%;
  border-collapse: collapse;
  min-width: 860px;
}
.heatmap-table th,
.heatmap-table td {
  border: 1px solid var(--line);
  padding: 8px 10px;
  vertical-align: top;
}
.heatmap-table thead th {
  position: sticky;
  top: 0;
  background: #f8fafc;
  z-index: 1;
  text-align: center;
  font-size: 12px;
}
.heatmap-table .heatmap-row-label {
  background: #f8fafc;
  min-width: 220px;
  font-weight: 650;
}
.heatmap-table td.heatmap-cell {
  text-align: center;
  font-weight: 700;
  min-width: 52px;
}
.heatmap-table--coherence {
  min-width: 1400px;
}
.heatmap-table--coherence thead th {
  min-width: 110px;
  max-width: 140px;
  white-space: normal;
  line-height: 1.25;
  vertical-align: bottom;
}
.heatmap-table--coherence .heatmap-row-label {
  min-width: 260px;
  max-width: 320px;
  font-weight: 650;
  line-height: 1.3;
}
.heatmap-table--coherence td.heatmap-cell--selected {
  outline: 2px solid #0b3d42;
  outline-offset: -2px;
  box-shadow: inset 0 0 0 1px rgba(255,255,255,0.55);
}
.heatmap-legend {
  display: flex;
  align-items: center;
  gap: 10px;
  margin-top: 10px;
  color: var(--muted);
  font-size: 13px;
}
.heatmap-scale {
  width: 180px;
  height: 12px;
  border-radius: 999px;
  border: 1px solid var(--line);
  background: linear-gradient(90deg, rgba(11,110,119,0.15), rgba(11,110,119,0.85));
}
@media (max-width: 720px) {
  .site-header__inner { align-items: flex-start; }
  .site-nav { width: 100%; }
  .hero { padding: 28px 20px; }
}
"""


EXPORT_WORD_MODAL = """
<div class="modal" id="export-word-modal" hidden role="dialog" aria-modal="true" aria-labelledby="export-word-title">
  <div class="modal-card">
    <h2 id="export-word-title">Exporter le dossier capsule</h2>
    <p class="meta">Script final chorale, unités de sens et vidéos expert à produire.</p>
    <div class="export-field">
      <label for="export-word-filename">Nom du fichier</label>
      <input type="text" id="export-word-filename" value="{default_filename}" autocomplete="off">
    </div>
    <div class="export-field" id="export-word-folder-field">
      <label>Dossier de destination</label>
      <div class="export-folder">
        <span id="export-word-folder">Aucun dossier sélectionné</span>
        <button type="button" class="btn btn-secondary" id="export-word-pick-dir">Choisir un dossier</button>
      </div>
    </div>
    <p class="meta" id="export-word-browser-hint">Chrome et Edge permettent de choisir un dossier. Firefox et Safari utilisent le téléchargement du navigateur.</p>
    <div class="modal-actions">
      <button type="button" class="btn btn-secondary" id="export-word-cancel">Annuler</button>
      <button type="button" class="btn" id="export-word-run">Exporter</button>
    </div>
    <p class="export-status" id="export-word-status" aria-live="polite"></p>
  </div>
</div>
"""


def export_word_modal(default_filename: str) -> str:
    return EXPORT_WORD_MODAL.replace("{default_filename}", escape(default_filename))


def selection_methodology_section(capsule: dict, capsule_data: dict) -> str:
    meth = capsule_data.get("methodologie", {})
    fil = meth.get("fil_pedagogique") or capsule.get("message_central", "")
    statut = meth.get("statut_montage", "")
    statut_note = ""
    if statut == "VALIDE_LABORATOIRE":
        statut_note = "<p class='meta'>Montage validé en laboratoire éditorial.</p>"
    elif statut == "A_CARTOGRAPHIER":
        statut_note = "<p class='meta'>Montage à cartographier — méthodologie de référence pour le dérushage à venir.</p>"
    return f"""
<section class="methodology-panel">
  <h2>Méthode de sélection des séquences</h2>
  {statut_note}
  <p>
    Cette capsule chorale est constituée à partir des transcripts BAB d'interviews
    (Jean-Jacques Greffet, Muriel Thomas, Sylvia Cohen-Kaminski, Loïc Rajjou). La sélection
    repose sur une analyse par <strong>unités de sens</strong> (sciences sociales et linguistique),
    pas sur une découpe mécanique du temps de parole.
  </p>
  <p class="meta">Pipeline : analyze_discourse → proposition chorale → montage → sync_unites_de_sens. Voir docs/METHODOLOGIE_ANALYSE.md</p>
  <p><strong>Fil pédagogique de la capsule :</strong> {escape(_normalize_editorial_french(fil))}</p>
  <ul>
    <li><strong>Unités de sens</strong> — identification des passages qui forment une idée complète,
    compréhensible hors du reste de l'entretien.</li>
    <li><strong>Regroupement thématique</strong> — rapprochement des extraits selon le fil pédagogique
    de la capsule.</li>
    <li><strong>Thèmes et sous-thèmes</strong> — chaque extrait est qualifié selon sa contribution
    au message central.</li>
    <li><strong>Redondances</strong> — lorsque plusieurs formulations disent la même chose,
    une seule formulation est retenue pour respecter la durée cible.</li>
    <li><strong>Transitions</strong> — vérification que l'enchaînement entre voix reste lisible.</li>
    <li><strong>Complémentarité entre intervenants</strong> — équilibre des quatre parcours
    sans répétition inutile d'un même angle.</li>
    <li><strong>Autonomie des extraits</strong> — chaque séquence doit pouvoir être entendue
    sans renvoi implicite à un passage non monté.</li>
    <li><strong>Faisabilité du montage</strong> — prise en compte des coupes NON PRONONCE,
    de la durée cible et des réservations d'extraits pour d'autres capsules.</li>
  </ul>
  <p>
    Les extraits retenus, leur ordre et les coupes prévues sont documentés ci-dessus lorsque le montage est établi.
    Les passages écartés ou réservés restent tracés dans les BAB encodés et le registre des extraits.
  </p>
</section>
"""


def _render_orientation_block(orientation: dict, plural: bool = False) -> str:
    concepts = " · ".join(orientation.get("concepts", []))
    consignes = "".join(f"<li>{escape(_normalize_editorial_french(item))}</li>" for item in orientation.get("consignes", []))
    passerelles = []
    for item in orientation.get("passerelles", []):
        passerelles.append(
            "<tr>"
            f"<td>{escape(item.get('extrait', ''))}</td>"
            f"<td>{escape(_normalize_editorial_french(item.get('concept', '')))}</td>"
            f"<td>{escape(_normalize_editorial_french(item.get('orientation', '')))}</td>"
            "</tr>"
        )
    util = orientation.get("utilisation_script_temoin", {})
    util_html = ""
    if util:
        seq_key = next(
            (key for key in util if key.startswith("sequence_recommandee_")),
            "sequence_recommandee_e1",
        )
        seq_items = util.get(seq_key, [])
        seq = "".join(f"<li>{escape(_normalize_editorial_french(s))}</li>" for s in seq_items)
        guide_items = util.get("par_origine") or util.get("par_voix") or []
        guide_label = "origine" if util.get("par_origine") else "voix"
        guide_title = (
            "guide par origine"
            if util.get("par_origine")
            else "guide par extrait témoin"
        )
        guides_html = []
        for item in guide_items:
            titre = item.get("origine") or item.get("angle") or ""
            guides_html.append(
                "<div class='orientation-origine'>"
                f"<h4>{_e_fr(titre)} "
                f"<span class='meta'>— {escape(item.get('extrait_id', ''))} · "
                f"{escape(item.get('timecodes', ''))}</span></h4>"
                f"<p><strong>Verbatim clé :</strong> « {escape(_normalize_editorial_french(item.get('verbatim_cle', '')))} »</p>"
                f"<p><strong>Dans le témoin :</strong> {escape(_normalize_editorial_french(item.get('dans_le_temoin', '')))}</p>"
                f"<p><strong>Travail expert :</strong> {escape(_normalize_editorial_french(item.get('travail_expert', '')))}</p>"
                f"<p class='phrase-amorce'><strong>Phrase d'amorce suggérée :</strong> {escape(_normalize_editorial_french(item.get('phrase_amorce', '')))}</p>"
                f"<p><strong>Question apprenant :</strong> {escape(_normalize_editorial_french(item.get('question_apprenant', '')))}</p>"
                f"<p class='meta'><strong>À éviter :</strong> {escape(_normalize_editorial_french(item.get('erreur_a_eviter', '')))}</p>"
                "</div>"
            )
        util_html = f"""
    <h3>Utilisation du script témoin — {guide_title}</h3>
    <p>{escape(_normalize_editorial_french(util.get('principe', '')))}</p>
    <h4>Séquence recommandée ({escape(orientation.get('code', 'E1'))})</h4>
    <ol>{seq}</ol>
    <div class="orientation-{guide_label}s">
      {''.join(guides_html)}
    </div>
"""
    expert = orientation.get("expert")
    if expert:
        expert_line = escape(expert)
    else:
        expert_line = "Intervenant désigné"
    code = orientation.get("code", "expert")
    heading = "Orientation pour les vidéos expert suivantes" if plural else "Orientation pour la vidéo expert suivante"
    return f"""
  <div class="orientation-expert">
    <h2>{heading}</h2>
    <p>
      <strong>{escape(code)} — {expert_line}</strong><br>
      {escape(_normalize_editorial_french(orientation.get('titre', '')))}
    </p>
    <p class="meta">{escape(_normalize_editorial_french(concepts))}</p>
    <p>{escape(_normalize_editorial_french(orientation.get('introduction', '')))}</p>
    {util_html}
    <h3>Consignes de prise de parole</h3>
    <ul>{consignes}</ul>
    <h3>Passerelles témoin → expert (synthèse)</h3>
    <p class="meta">Tableau récapitulatif extrait / concept / amorce.</p>
    <table>
      <thead><tr><th>Extrait</th><th>Concept {escape(code)}</th><th>Amorce</th></tr></thead>
      <tbody>
        {''.join(passerelles)}
      </tbody>
    </table>
  </div>
"""


BRIEF_PRECAUTION_ORATOIRE = (
    "Les objectifs et unités de sens que nous proposons (ingénierie pédagogique) reflètent "
    "notre niveau de compréhension du sujet à ce stade. C'est sur votre expérience et la "
    "maîtrise de votre discipline que nous nous appuyons : n'hésitez pas à modifier ou "
    "compléter ce travail, en restant aligné avec les objectifs exposés dans le tableau "
    "de conception."
)

EXPORT_BRIEF_SECTION_TITLE = "Proposition de cadrage pour la vidéo expert"
EXPORT_VIDEOS_TABLE_TITLE = "Tableau du programme de conception"
PROGRAMME_TABLE_FIELDS = (
    "module",
    "code",
    "video_temoin",
    "resume_chercheurs",
    "videos_referent",
    "objectif_pedagogique",
    "noms_proposes",
)

BRIEF_CONSIGNES_COMMUNES = [
    "Partir des témoignages vus dans la chorale — pas d'un script à lire mot pour mot.",
    "Nommer les notions ou concepts de votre vidéo en langage clair, avec des exemples concrets entendus.",
    "Tout schéma, liste ou support à incruster dans votre intervention est le bienvenu.",
    "Inviter l'apprenant à faire le lien avec sa pratique de recherche (même sans projet d'innovation).",
    "Ne pas citer les chercheurs phrase pour phrase : résumer dans vos propres mots.",
    "Compléter librement cette trame : ajouter tout élément (exemple, rappel, précision, mise en perspective) que vous jugez complémentaire et nécessaire à ce stade du parcours.",
]

_CONSIGNE_TECHNIQUE_MARKERS = (
    "sequence_recommandee",
    "phrase_amorce",
    "par_origine",
    "par_voix",
    "script_final",
    "id + timecode",
    "cf. par_",
    "transcript monte",
    "transcript monté",
    "ne pas inventer",
    "exemples absents",
    "montage final",
)


def _expert_codes(capsule_data: dict) -> list[str]:
    return [
        item.get("code", "")
        for item in capsule_data.get("videos_expert", [])
        if item.get("code")
    ]


def _grille_values(unite: dict) -> list[str]:
    return [str(value) for key, value in unite.items() if key.startswith("grille_") and value]


def _grille_tags_expert(grille: str, expert_codes: list[str]) -> list[str]:
    tagged = []
    for code in sorted(expert_codes, key=len, reverse=True):
        if re.search(rf"\b{re.escape(code)}\b", grille):
            tagged.append(code)
    return tagged


def _unite_matches_expert(unite: dict, expert_code: str, expert_codes: list[str]) -> bool:
    grilles = _grille_values(unite)
    if not grilles:
        return expert_code == "E1" and unite.get("ordre") == 1
    tagged: list[str] = []
    for grille in grilles:
        tagged.extend(_grille_tags_expert(grille, expert_codes))
    if not tagged:
        return True
    return expert_code in tagged


def _orientation_guides(orientation: dict) -> list[dict]:
    util = orientation.get("utilisation_script_temoin", {})
    guides = util.get("par_origine") or util.get("par_voix") or []
    cleaned: list[dict] = []
    for guide in guides:
        item = dict(guide)
        if item.get("dans_le_temoin"):
            item["dans_le_temoin"] = _normalize_editorial_french(item["dans_le_temoin"])
        cleaned.append(item)
    return cleaned


def _orientation_sequence(orientation: dict) -> list[str]:
    util = orientation.get("utilisation_script_temoin", {})
    for key, value in util.items():
        if key.startswith("sequence_recommandee_") and isinstance(value, list):
            return value
    return []


def _humanize_sequence_step(step: str, guides: list[dict]) -> str:
    by_id = {item.get("extrait_id", ""): item for item in guides if item.get("extrait_id")}
    cleaned = step.strip()
    if "→" in cleaned:
        left, right = [part.strip() for part in cleaned.split("→", 1)]
        if left in by_id:
            left = by_id[left].get("chercheur", left)
        cleaned = f"{left} — {right}"

    def replace_id(match: re.Match[str]) -> str:
        guide = by_id.get(match.group(0))
        return guide.get("chercheur", match.group(0)) if guide else match.group(0)

    cleaned = re.sub(r"\b(?:JJG|MUR|SYL|YAN|LOI)-\d+\b", replace_id, cleaned)
    cleaned = re.sub(r"\bextrait\s+", "", cleaned, flags=re.I)
    cleaned = re.sub(r"\s{2,}", " ", cleaned)
    return cleaned.strip(" —")


def _simplify_consignes(consignes: list[str]) -> list[str]:
    simplified: list[str] = []
    for item in consignes:
        lowered = item.lower()
        if any(marker in lowered for marker in _CONSIGNE_TECHNIQUE_MARKERS):
            continue
        text = re.sub(r"\b(?:JJG|MUR|SYL|YAN|LOI)-\d+\b", "", item)
        text = re.sub(r"\s{2,}", " ", text).strip(" ,;")
        if text:
            simplified.append(text)
    return simplified


CHERCHEUR_LABELS = {
    "Jean-Jacques": "Jean-Jacques Greffet",
    "Muriel": "Muriel Thomas",
    "Sylvia": "Sylvia Cohen-Kaminski",
    "Yann": "Yann Monier",
    "Loic": "Loic Rajjou",
}


def _parse_resume_temoignages(text: str) -> list[tuple[str, str]]:
    text = text.strip()
    if not text:
        return []
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if len(lines) > 1 and all(":" in line for line in lines):
        items: list[tuple[str, str]] = []
        for line in lines:
            name, _, content = line.partition(":")
            items.append((name.strip(), content.strip().rstrip(".")))
        return items
    parts = re.split(r"(?<=\.)\s+(?=[^:]+:)", text)
    items = []
    for part in parts:
        part = part.strip().rstrip(".")
        if not part:
            continue
        if ":" in part:
            name, _, content = part.partition(":")
            items.append((name.strip(), content.strip()))
        else:
            items.append(("", part))
    return items


def _label_chercheur(short_name: str) -> str:
    return CHERCHEUR_LABELS.get(short_name, short_name)


def _render_brief_point(guide: dict) -> str:
    titre = guide.get("origine") or guide.get("angle") or "Point a developper"
    chercheur = guide.get("chercheur", "")
    heading = f"{titre} — {chercheur}" if chercheur else titre
    concepts = guide.get("concepts_e1") or guide.get("concepts") or []
    parts = [
        '<div class="brief-point">',
        f"<h4>{escape(heading)}</h4>",
    ]
    if guide.get("dans_le_temoin"):
        parts.append(
            f"<p><strong>Dans la chorale :</strong> {escape(guide['dans_le_temoin'])}</p>"
        )
    if guide.get("travail_expert"):
        parts.append(f"<p><strong>A apporter :</strong> {escape(guide['travail_expert'])}</p>")
    if concepts:
        parts.append(
            f"<p class='meta'><strong>Concepts :</strong> {escape(' · '.join(concepts))}</p>"
        )
    if guide.get("question_apprenant"):
        parts.append(
            "<p><strong>Question pour l'apprenant :</strong> "
            f"{escape(guide['question_apprenant'])}</p>"
        )
    if guide.get("erreur_a_eviter"):
        parts.append(
            f"<p class='meta'><strong>À éviter :</strong> {escape(guide['erreur_a_eviter'])}</p>"
        )
    parts.append("</div>")
    return "".join(parts)


def _label_video_expert(code: str) -> str:
    match = re.fullmatch(r"E(\d+)(bis)?", code, re.IGNORECASE)
    if match:
        suffix = f" {match.group(2)}" if match.group(2) else ""
        return f"Vidéo Expert {match.group(1)}{suffix}"
    return f"Vidéo Expert ({code})"


def _label_video_temoin(capsule_code: str) -> str:
    if capsule_code == "GEN":
        return "Vidéo témoin 1"
    match = re.fullmatch(r"T(\d+)", capsule_code)
    if match:
        return f"Vidéo témoin {match.group(1)}"
    return f"Vidéo témoin ({capsule_code})"


def _humanize_capsule_labels(text: str) -> str:
    def repl_expert(match: re.Match[str]) -> str:
        suffix = f" {match.group(2)}" if match.group(2) else ""
        return f"Vidéo Expert {match.group(1)}{suffix}"

    text = re.sub(r"\bE(\d+)(bis)?\b", repl_expert, text, flags=re.IGNORECASE)
    text = re.sub(r"\bT(\d+)\b", r"Vidéo témoin \1", text)
    return text


def _strip_chercheur_prefix(text: str, chercheur: str) -> str:
    cleaned = text.strip()
    aliases: list[str] = []
    if chercheur:
        aliases.append(chercheur)
    key = _canonical_name_key(chercheur or "")
    if key == "yann meunier":
        aliases.extend(["Yann Monier", "Yan Monier"])
    for alias in aliases:
        if alias and cleaned.startswith(alias):
            cleaned = cleaned[len(alias) :].lstrip(" :—-\u2014")
            break
    return cleaned.strip()


def _restore_french_accents(text: str) -> str:
    """Retablit les accents sur le francais editorial / genere (hors verbatim BAB)."""
    if not text:
        return text

    # Paires (forme sans accent → forme accentuee). Ordre : plus longues d'abord.
    pairs = [
        ("temoignages", "témoignages"),
        ("temoignage", "témoignage"),
        ("precisement", "précisément"),
        ("declenchee", "déclenchée"),
        ("declenche", "déclenché"),
        ("resultats", "résultats"),
        ("resultat", "résultat"),
        ("matiere", "matière"),
        ("premiere", "première"),
        ("societale", "sociétale"),
        ("societal", "sociétal"),
        ("serendipite", "sérendipité"),
        ("different", "différent"),
        ("differentes", "différentes"),
        ("differents", "différents"),
        ("reconnaitre", "reconnaître"),
        ("connaitre", "connaître"),
        ("apparaitre", "apparaître"),
        ("paraitre", "paraître"),
        ("theorie", "théorie"),
        ("theorique", "théorique"),
        ("pedagogique", "pédagogique"),
        ("pedagogiques", "pédagogiques"),
        ("generale", "générale"),
        ("general", "général"),
        ("generer", "générer"),
        ("geneses", "genèses"),
        ("genese", "genèse"),
        ("interet", "intérêt"),
        ("experience", "expérience"),
        ("experiences", "expériences"),
        ("experimental", "expérimental"),
        ("experimentale", "expérimentale"),
        ("experimentaux", "expérimentaux"),
        ("accumulation", "accumulation"),
        ("accumulee", "accumulée"),
        ("posee", "posée"),
        ("exterieur", "extérieur"),
        ("exterieure", "extérieure"),
        ("orientee", "orientée"),

        ("revele", "révèle"),
        ("revelateur", "révélateur"),
        ("revelatrice", "révélatrice"),
        ("detection", "détection"),
        ("defini", "défini"),
        ("definie", "définie"),
        ("definir", "définir"),
        ("definit", "définit"),
        ("definition", "définition"),
        ("element", "élément"),
        ("elements", "éléments"),
        ("etape", "étape"),
        ("etapes", "étapes"),
        ("etre", "être"),
        ("ete", "été"),
        ("deja", "déjà"),
        ("meme", "même"),
        ("memes", "mêmes"),
        ("idees", "idées"),
        ("idee", "idée"),
        ("apres", "après"),
        ("tres", "très"),
        ("role", "rôle"),
        ("roles", "rôles"),
        ("controle", "contrôle"),
        ("echanges", "échanges"),
        ("echange", "échange"),
        ("ecoutez", "écoutez"),
        ("ecoutons", "écoutons"),
        ("ecoute", "écoute"),
        ("echelle", "échelle"),
        ("evenement", "événement"),
        ("evenements", "événements"),
        ("economique", "économique"),
        ("economiques", "économiques"),
        ("equipe", "équipe"),
        ("equipes", "équipes"),
        ("etudier", "étudier"),
        ("etude", "étude"),
        ("etudes", "études"),
        ("evidemment", "évidemment"),
        ("evidemment", "évidemment"),
        ("eveiller", "éveiller"),
        ("eval", "éval"),  # noop-ish safety — skip short
        ("evaluation", "évaluation"),
        ("evaluer", "évaluer"),
        ("echouer", "échouer"),
        ("echec", "échec"),
        ("echecs", "échecs"),
        ("ecart", "écart"),
        ("ecarts", "écarts"),
        ("ecole", "école"),
        ("ecrire", "écrire"),
        ("ecrit", "écrit"),
        ("ecrite", "écrite"),
        ("numero", "numéro"),
        ("numerique", "numérique"),
        ("numeriques", "numériques"),
        ("strategie", "stratégie"),
        ("strategique", "stratégique"),
        ("creativite", "créativité"),
        ("creer", "créer"),
        ("creation", "création"),
        ("decision", "décision"),
        ("decisions", "décisions"),
        ("decider", "décider"),
        ("decide", "décide"),
        ("declaration", "déclaration"),
        ("declarer", "déclarer"),
        ("depot", "dépôt"),
        ("depots", "dépôts"),
        ("deposer", "déposer"),
        ("developpement", "développement"),
        ("developper", "développer"),
        ("deroulement", "déroulement"),
        ("derouler", "dérouler"),
        ("desormais", "désormais"),
        ("desir", "désir"),
        ("desire", "désire"),
        ("necessaire", "nécessaire"),
        ("necessaires", "nécessaires"),
        ("necessite", "nécessite"),
        ("reponse", "réponse"),
        ("reponses", "réponses"),
        ("repondre", "répondre"),
        ("repond", "répond"),
        ("reseau", "réseau"),
        ("reseaux", "réseaux"),
        ("realite", "réalité"),
        ("realiser", "réaliser"),
        ("referent", "référent"),
        ("referents", "référents"),
        ("reference", "référence"),
        ("references", "références"),
        ("reflexe", "réflexe"),
        ("reflexes", "réflexes"),
        ("reflexion", "réflexion"),
        ("reduire", "réduire"),
        ("resumer", "résumer"),
        ("reutiliser", "réutiliser"),
        ("reutilisation", "réutilisation"),
        ("pretexte", "prétexte"),
        ("pretextes", "prétextes"),
        ("preparer", "préparer"),
        ("preparation", "préparation"),
        ("prematuration", "prématuration"),
        ("precis", "précis"),
        ("precise", "précise"),
        ("precision", "précision"),
        ("precisions", "précisions"),
        ("proteger", "protéger"),
        ("protege", "protège"),
        ("preparer", "préparer"),
        ("prepare", "prépare"),
        ("declarer", "déclarer"),
        ("declare", "déclare"),
        ("resumer", "résumer"),
        ("resume", "résume"),
        ("orienter", "orienter"),
        ("oriente", "oriente"),
        ("creer", "créer"),
        ("cree", "crée"),
        ("developper", "développer"),
        ("developpe", "développe"),
        ("deposer", "déposer"),
        ("depose", "dépose"),
        ("realiser", "réaliser"),
        ("realise", "réalise"),
        ("generer", "générer"),
        ("genere", "génère"),
        ("maniere", "manière"),
        ("manieres", "manières"),
        ("tete", "tête"),
        ("tetes", "têtes"),
        ("derriere", "derrière"),
        ("verifiee", "vérifiée"),
        ("verifie", "vérifié"),
        ("etablie", "établie"),
        ("etabli", "établi"),
        ("editoriale", "éditoriale"),
        ("editorial", "éditorial"),
        ("preciser", "préciser"),
        ("reveler", "révéler"),
        ("verifier", "vérifier"),
        ("reel", "réel"),
        ("reelle", "réelle"),
        ("reels", "réels"),
        ("observee", "observée"),
        ("agees", "âgées"),
        ("agee", "âgée"),
        ("ages", "âgés"),
        ("malgre", "malgré"),
        ("deplacement", "déplacement"),
        ("modifiee", "modifiée"),
        ("presentent", "présentent"),
        ("selectionnees", "sélectionnées"),
        ("selectionnee", "sélectionnée"),
        ("documentee", "documentée"),
        ("documente", "documenté"),
        ("documentes", "documentés"),
        ("basees", "basées"),
        ("basee", "basée"),
        ("interpretation", "interprétation"),
        ("ecran", "écran"),
        ("detnutrition", "dénutrition"),
        ("appauvri", "appauvri"),
        ("detaillee", "détaillée"),
        ("detaille", "détaillé"),
        ("these", "thèse"),
        ("industriel", "industriel"),
        ("syntheses", "synthèses"),
        ("synthese", "synthèse"),
        ("presentations", "présentations"),
        ("presentation", "présentation"),
        ("thematiques", "thématiques"),
        ("thematique", "thématique"),
        ("selections", "sélections"),
        ("selection", "sélection"),
        ("sequences", "séquences"),
        ("sequence", "séquence"),
        ("constituee", "constituée"),
        ("constitue", "constitué"),
        ("mecanique", "mécanique"),
        ("decoupe", "découpe"),
        ("unites", "unités"),
        ("unite", "unité"),
        ("comprehensible", "compréhensible"),
        ("durees", "durées"),
        ("duree", "durée"),
        ("enchainement", "enchaînement"),
        ("complementarite", "complémentarité"),
        ("equilibre", "équilibre"),
        ("repetition", "répétition"),
        ("prevues", "prévues"),
        ("prevue", "prévue"),
        ("ecartes", "écartés"),
        ("ecarte", "écarté"),
        ("suggerée", "suggérée"),
        ("suggeree", "suggérée"),
        ("suggerer", "suggérer"),
        ("suggestee", "suggérée"),
        ("iteratives", "itératives"),
        ("iterative", "itérative"),
        ("iteratif", "itératif"),
        ("concretes", "concrètes"),
        ("concrete", "concrète"),
        ("maitrise", "maîtrise"),
        ("maitriser", "maîtriser"),
        ("hesitez", "hésitez"),
        ("ingenierie", "ingénierie"),
        ("refletent", "reflètent"),
        ("comprehension", "compréhension"),
        ("aligne", "aligné"),
        ("exposes", "exposés"),
        ("expose", "exposé"),
        ("recapitulatif", "récapitulatif"),
        ("recapitulative", "récapitulative"),
        ("integrees", "intégrées"),
        ("integree", "intégrée"),
        ("integre", "intégré"),
        ("marquees", "marquées"),
        ("marquee", "marquée"),
        ("methode", "méthode"),
        ("methodes", "méthodes"),
        ("derushage", "dérushage"),
        ("reference", "référence"),
        ("references", "références"),
        ("generales", "générales"),
        ("proposés", "proposés"),
        ("proposes", "proposés"),
        ("proposee", "proposée"),
        ("proposees", "proposées"),
        ("defini", "défini"),
        ("eviter", "éviter"),
        ("evenement", "événement"),
        ("recherche", "recherche"),
        ("recherches", "recherches"),
        ("maturation", "maturation"),
        ("progressive", "progressive"),
        ("origine", "origine"),
        ("origines", "origines"),
        ("credible", "crédible"),
        ("protection", "protection"),
        ("propriete", "propriété"),
        ("proprietes", "propriétés"),
        ("probleme", "problème"),
        ("problemes", "problèmes"),
        ("problematique", "problématique"),
        ("problematiques", "problématiques"),
        ("methode", "méthode"),
        ("methodes", "méthodes"),
        ("media", "média"),
        ("cloture", "clôture"),
        ("cote", "côté"),
        ("cotes", "côtés"),
        ("controle", "contrôle"),
        ("facon", "façon"),
        ("facons", "façons"),
        ("lecon", "leçon"),
        ("lecons", "leçons"),
        ("anecdote", "anecdote"),
        ("interessant", "intéressant"),
        ("interessante", "intéressante"),
        ("particulierement", "particulièrement"),
        ("regulierement", "régulièrement"),
        ("veritable", "véritable"),
        ("verite", "vérité"),
        ("securite", "sécurité"),
        ("securiser", "sécuriser"),
        ("legitime", "légitime"),
        ("legitimite", "légitimité"),
        ("qualite", "qualité"),
        ("qualites", "qualités"),
        ("activite", "activité"),
        ("activites", "activités"),
        ("universite", "université"),
        ("priorite", "priorité"),
        ("opportunite", "opportunité"),
        ("opportunites", "opportunités"),
        ("possibilite", "possibilité"),
        ("possibilites", "possibilités"),
        ("capacite", "capacité"),
        ("capacites", "capacités"),
        ("difficultes", "difficultés"),
        ("difficulte", "difficulté"),
        ("hesitation", "hésitation"),
        ("hesiter", "hésiter"),
        ("hesite", "hésite"),
        ("arreter", "arrêter"),
        ("arrete", "arrête"),
        ("complet", "complet"),
        ("complete", "complète"),
        ("completer", "compléter"),
        ("complementaire", "complémentaire"),
        ("complementaires", "complémentaires"),
        ("entete", "en-tête"),
        ("entetes", "en-têtes"),
        ("reperes", "repères"),
        ("repere", "repère"),
        ("reperer", "repérer"),
        ("demarche", "démarche"),
        ("demarches", "démarches"),
        ("emerger", "émerger"),
        ("emerge", "émerge"),
        ("videos", "vidéos"),
        ("video", "vidéo"),
        ("numero", "numéro"),
        ("annees", "années"),
        ("annee", "année"),
        ("evenuel", "éventuel"),
        ("eventuel", "éventuel"),
        ("eventuelle", "éventuelle"),
        ("evenutuellement", "éventuellement"),
        ("eventuellement", "éventuellement"),
        ("evidemment", "évidemment"),
        ("la", "là"),  # DANGEROUS - skip
    ]
    # Retirer les paires dangereuses / inutiles
    pairs = [(a, b) for a, b in pairs if a != "la" and a != "eval" and a != b]

    # Appliquer en respectant la casse du mot source.
    def _replace_word(match: re.Match[str], accented: str) -> str:
        src = match.group(0)
        if src.isupper():
            return accented.upper()
        if src[0].isupper():
            return accented[0].upper() + accented[1:]
        return accented

    # Trier par longueur decroissante pour eviter les collisions partielles.
    pairs.sort(key=lambda item: len(item[0]), reverse=True)
    for plain, accented in pairs:
        text = re.sub(
            rf"\b{re.escape(plain)}\b",
            lambda m, acc=accented: _replace_word(m, acc),
            text,
            flags=re.IGNORECASE,
        )

    # Locutions frequentes avec « a » → « à »
    a_patterns = [
        (r"\ba la\b", "à la"),
        (r"\ba l'", "à l'"),
        (r"\ba l’", "à l’"),
        (r"\ba vous\b", "à vous"),
        (r"\ba votre\b", "à votre"),
        (r"\ba vos\b", "à vos"),
        (r"\ba qui\b", "à qui"),
        (r"\ba quoi\b", "à quoi"),
        (r"\ba partir\b", "à partir"),
        (r"\ba travers\b", "à travers"),
        (r"\ba condition\b", "à condition"),
        (r"\ba condition que\b", "à condition que"),
        (r"\ba savoir\b", "à savoir"),
        (r"\ba propos\b", "à propos"),
        (r"\ba garder\b", "à garder"),
        (r"\ba lire\b", "à lire"),
        (r"\ba produire\b", "à produire"),
        (r"\ba atteindre\b", "à atteindre"),
        (r"\ba definir\b", "à définir"),
        (r"\ba confirmer\b", "à confirmer"),
        (r"\ba completer\b", "à compléter"),
        (r"\ba modifier\b", "à modifier"),
        (r"\ba cartographier\b", "à cartographier"),
        (r"\ba venir\b", "à venir"),
        (r"\ba ce stade\b", "à ce stade"),
        (r"\ba faire\b", "à faire"),
        (r"\ba quel\b", "à quel"),
        (r"\ba quelle\b", "à quelle"),
        (r"\ba qui,\b", "à qui,"),
        (r"\bne sont pas la\b", "ne sont pas là"),
        (r"\bpas la pour\b", "pas là pour"),
        (r"\bdeja la\b", "déjà là"),
        (r"\bdéjà la\b", "déjà là"),
        (r"\bOn touche ici a\b", "On touche ici à"),
        (r"\ben tete\b", "en tête"),
        (r"\bou il faut\b", "où il faut"),
        (r"\bou se\b", "où se"),
        (r"\ble maturation\b", "la maturation"),
        (r"\bbesoin exprime\b", "besoin exprimé"),
        (r"\binnovation credible lie\b", "innovation crédible lie"),
        (r"\ben restant aligne\b", "en restant aligné"),
        (r"\bpas de cote\b", "pas de côté"),
        (r"\ble gout\b", "le goût"),
        (r"\bprofessionnel cle\b", "professionnel clé"),
        (r"\bVerbatim cle\b", "Verbatim clé"),
        (r"\bgeste professionnel cle\b", "geste professionnel clé"),
        (r"\bIdee vs\b", "Idée vs"),
        (r"\bidee vs\b", "idée vs"),
        (r"\btire par le marché\b", "tiré par le marché"),
        (r"\b« tire par\b", "« tiré par"),
        (r"«\s*marche\s*»", "« marché »"),
        (r"\bsur le marche\b", "sur le marché"),
        (r"\bdu marche\b", "du marché"),
        (r"\bau marche\b", "au marché"),
        (r"\ble marche\b", "le marché"),
        (r"\bd'un marche\b", "d'un marché"),
        (r"\bun marche\b", "un marché"),
        (r"\bpar le marche\b", "par le marché"),
        (r"\btire par le marche\b", "tiré par le marché"),
        (r"\btire par\b", "tiré par"),
        (r"\bc'est tire\b", "c'est tiré"),
        (r"\bdemande du marche\b", "demande du marché"),
        (r"\bavez rencontres\b", "avez rencontrés"),
        (r"\best nee\b", "est née"),
        (r"\best ne\b", "est né"),
        (r"\bcomment est nee\b", "comment est née"),
        (r"\bmanque observe\b", "manque observé"),
        (r"\bbesoin observe\b", "besoin observé"),
        (r"\bdeja observe\b", "déjà observé"),
        (r"\bdéjà observe\b", "déjà observé"),
        (r"\bCouper apres\b", "Couper après"),
        (r"\bcouper apres\b", "couper après"),
        (r"\bemerger\b", "émerger"),
        (r"\bsocietale\b", "sociétale"),
        (r"\breperer\b", "repérer"),
        (r"\bcoherents\b", "cohérents"),
        (r"\bDuree montage\b", "Durée montage"),
        (r"\bpour reveler\b", "pour révéler"),
        (r"\bou preciser\b", "ou préciser"),
        (r"\ba preciser\b", "à préciser"),
        (r"\bverifier si\b", "vérifier si"),
        (r"\best reel\b", "est réel"),
        (r"\ba l'ecran\b", "à l'écran"),
        (r"\bAnimateur a\b", "Animateur à"),
        (r"\bA l'", "À l'"),
        (r"\bA l’", "À l’"),
        (r"\bA confirmer\b", "À confirmer"),
        (r"\bA eviter\b", "À éviter"),
        (r"\bA definir\b", "À définir"),
        (r"\bA l'issue\b", "À l'issue"),
        (r"\bA l’issue\b", "À l’issue"),
    ]
    for pattern, repl in a_patterns:
        text = re.sub(pattern, repl, text, flags=re.IGNORECASE)

    # Corriger les doubles accidents apres remplacement (déjà déjà, etc.)
    text = text.replace("déjà déjà", "déjà")
    text = text.replace("même même", "même")
    return text


def _normalize_editorial_french(text: str) -> str:
    """Corrige des raccourcis editoriaux fautifs et retablit les accents (hors verbatim BAB)."""
    if not text:
        return text
    # « besoin marche » → « besoin du marche » puis accents
    text = re.sub(
        r"\bbesoin marche\b",
        "besoin du marche",
        text,
        flags=re.IGNORECASE,
    )
    text = re.sub(
        r"\bbesoin marché\b",
        "besoin du marché",
        text,
        flags=re.IGNORECASE,
    )
    text = re.sub(
        r"\ble besoin marche\b",
        "le besoin du marche",
        text,
        flags=re.IGNORECASE,
    )
    text = re.sub(
        r"\ble besoin marché\b",
        "le besoin du marché",
        text,
        flags=re.IGNORECASE,
    )
    return _restore_french_accents(text)



def _e_fr(text: str) -> str:
    """Escape HTML apres normalisation editoriale (accents, formulations)."""
    return escape(_normalize_editorial_french(text or ""))


def _normalize_script_final_editorial(script: str) -> str:
    """Accentue uniquement les blocs cadrage/pancarte du script final (pas les verbatims BAB)."""
    if not script:
        return script
    parts = re.split(r"(\n{2,})", script)
    out: list[str] = []
    for part in parts:
        stripped = part.lstrip()
        if (
            stripped.startswith("[CADRAGE")
            or stripped.startswith("[PAN")
            or stripped.startswith("=== PARTIE")
            or stripped.startswith("[EXPERT]")
        ):
            out.append(_normalize_editorial_french(part))
        else:
            out.append(part)
    return "".join(out)


def _chercheur_prenom(chercheur: str) -> str:
    if " " in chercheur:
        return chercheur.split()[0]
    return chercheur


def _ensure_subject(text: str, chercheur: str) -> str:
    if not text or not chercheur:
        return text
    lower = text.lower()
    verb_starts = (
        "oppose",
        "raconte",
        "part ",
        "illustre",
        "precise",
        "definit",
        "montre",
    )
    if any(lower.startswith(verb) for verb in verb_starts):
        prenom = _chercheur_prenom(chercheur)
        if text[0].isupper():
            return f"{prenom} {text[0].lower()}{text[1:]}"
        return f"{prenom} {text}"
    return text


def _phrases_from_extrait(segment: dict, guide: dict | None) -> list[str]:
    chercheur = segment["chercheur"]
    comment = (segment.get("commentaire") or "").strip()
    detail = ""
    if guide and guide.get("dans_le_temoin"):
        detail = _ensure_subject(
            _strip_chercheur_prefix(guide["dans_le_temoin"], chercheur),
            chercheur,
        )

    if comment:
        return [comment]
    if detail:
        return [detail]
    return []


def _resume_temoignages_map(capsule_data: dict) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for short_name, content in _parse_resume_temoignages(
        capsule_data.get("resume_temoignages", "")
    ):
        if short_name:
            mapping[_label_chercheur(short_name)] = content
    return mapping


def _collect_temoignages_lisibles(
    capsule_data: dict, by_id: dict[str, dict]
) -> list[tuple[str, list[str]]]:
    ordre = capsule_data.get("ordre_montage", [])
    guide_by_id: dict[str, dict] = {}
    for orientation in capsule_data.get("orientations_expert", []):
        util = orientation.get("utilisation_script_temoin", {})
        for guide in util.get("par_origine") or util.get("par_voix") or []:
            extrait_id = guide.get("extrait_id")
            if extrait_id:
                guide_by_id[extrait_id] = guide

    by_chercheur: dict[str, list[str]] = defaultdict(list)
    chercheur_order: list[str] = []

    for extrait_id in ordre:
        segment = by_id.get(extrait_id)
        if not segment:
            continue
        chercheur = segment["chercheur"]
        if chercheur not in chercheur_order:
            chercheur_order.append(chercheur)

        guide = guide_by_id.get(extrait_id)
        for phrase in _phrases_from_extrait(segment, guide):
            if phrase not in by_chercheur[chercheur]:
                by_chercheur[chercheur].append(phrase)

    resume_map = _resume_temoignages_map(capsule_data)
    result: list[tuple[str, list[str]]] = []
    for chercheur in chercheur_order:
        phrases = list(by_chercheur.get(chercheur, []))
        if not phrases and chercheur in resume_map:
            phrases = [resume_map[chercheur]]
        if phrases:
            result.append((chercheur, phrases))

    if result:
        return result

    fallback: list[tuple[str, list[str]]] = []
    for short_name, content in _parse_resume_temoignages(
        capsule_data.get("resume_temoignages", "")
    ):
        if short_name:
            fallback.append((_label_chercheur(short_name), [content]))
        elif content:
            fallback.append(("", [content]))
    return fallback


def _load_cadrage_temoins_narratifs() -> dict:
    path = ROOT / "data" / "cadrage_temoins_narratifs.json"
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def _load_transcripts_videos_finaux() -> dict:
    path = ROOT / "data" / "transcripts_videos_finaux.json"
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def _mounted_transcript_script(capsule_code: str) -> str:
    """Script final = transcript de la vidéo montée, s'il est disponible."""
    data = _load_transcripts_videos_finaux()
    item = (data.get("capsules") or {}).get(capsule_code) or {}
    return (item.get("text") or "").strip()


def _script_final_prefer_mounted_transcript(
    capsule_code: str,
    fallback: str,
) -> str:
    mounted = _mounted_transcript_script(capsule_code)
    return mounted if mounted else fallback


def _narrative_temoin_block(capsule_code: str) -> dict | None:
    data = _load_cadrage_temoins_narratifs()
    block = (data.get("capsules") or {}).get(capsule_code)
    if not block:
        return None
    return block


def _render_temoin_phrases(phrases: list[str]) -> str:
    if len(phrases) == 1:
        return f"<p>{_e_fr(phrases[0])}</p>"
    items = "".join(f"<li>{_e_fr(phrase)}</li>" for phrase in phrases)
    return f'<ul class="brief-temoin-phrases">{items}</ul>'


def _render_brief_temoin_narrative(block: dict) -> str:
    """Rendu fluide « Ce que disent les chercheurs » (version narrative)."""
    title = block.get("titre_section") or "Ce que disent les chercheurs"
    parts = [
        '<article class="brief-video brief-video--temoin">',
        f"<h3>{escape(title)}</h3>",
    ]
    intro = (block.get("intro") or "").strip()
    if intro:
        parts.append(f"<p>{_e_fr(intro)}</p>")
    for voice in block.get("voix") or []:
        chercheur = (voice.get("chercheur") or "").strip()
        if chercheur:
            parts.append(f"<h4>{escape(chercheur)}</h4>")
        for para in voice.get("paragraphes") or []:
            text = (para or "").strip()
            if text:
                parts.append(f"<p>{_e_fr(text)}</p>")
    a_retenir = (block.get("a_retenir") or "").strip()
    if a_retenir:
        parts.append("<h4>À retenir</h4>")
        parts.append(f"<p>{_e_fr(a_retenir)}</p>")
    parts.append("</article>")
    return "\n".join(parts)


def _export_brief_temoin_narrative_plaintext(block: dict) -> list[str]:
    lines: list[str] = []
    title = block.get("titre_section") or "Ce que disent les chercheurs"
    lines.append(f"{title} :")
    lines.append("")
    intro = (block.get("intro") or "").strip()
    if intro:
        lines.append(intro)
        lines.append("")
    for voice in block.get("voix") or []:
        chercheur = (voice.get("chercheur") or "").strip()
        if chercheur:
            lines.append(f"{chercheur} :")
            lines.append("")
        for para in voice.get("paragraphes") or []:
            text = (para or "").strip()
            if text:
                lines.append(text)
                lines.append("")
    a_retenir = (block.get("a_retenir") or "").strip()
    if a_retenir:
        lines.append("À retenir :")
        lines.append("")
        lines.append(a_retenir)
        lines.append("")
    return lines


def _render_brief_temoin(
    capsule_code: str, capsule_data: dict, by_id: dict[str, dict]
) -> str:
    narrative = _narrative_temoin_block(capsule_code)
    if narrative and (narrative.get("voix") or narrative.get("intro")):
        return _render_brief_temoin_narrative(narrative)

    temoins = _collect_temoignages_lisibles(capsule_data, by_id)
    if not temoins:
        return ""

    rows = []
    for chercheur, phrases in temoins:
        if chercheur:
            rows.append(
                "<li>"
                f"<strong>{escape(chercheur)}</strong>"
                f"{_render_temoin_phrases(phrases)}"
                "</li>"
            )
        else:
            rows.append(f"<li>{_render_temoin_phrases(phrases)}</li>")

    return f"""
  <article class="brief-video brief-video--temoin">
    <h3>Ce que disent les chercheurs</h3>
    <ul class="brief-unites">
      {''.join(rows)}
    </ul>
  </article>
"""


def _render_brief_video(video: dict, proposes: list[str] | None = None) -> str:
    label = _label_video_expert(video.get("code", ""))
    intervenant = video.get("intervenant")
    if intervenant:
        who = escape(intervenant)
    else:
        who = "<em>Intervenant désigné</em>"

    titre = video.get("titre", "")
    descriptif = video.get("descriptif", "")
    objectif_html = f"<p><strong>Objectif :</strong> {_e_fr(titre)}</p>"
    if descriptif:
        objectif_html += f"<p>{_e_fr(descriptif)}</p>"

    return f"""
  <article class="brief-video">
    <h3>{escape(label)}</h3>
    <p class='meta'><strong>Intervenant :</strong> {who}</p>
    {objectif_html}
    {_fascicule_refs_html(video.get("code", ""))}
  </article>
"""


def _videos_expert_by_code(capsule_data: dict) -> dict[str, dict]:
    return {
        item.get("code", ""): item
        for item in capsule_data.get("videos_expert", [])
        if item.get("code")
    }


def _names_match_chercheur(left: str, right: str) -> bool:
    if not left or not right:
        return False
    if left == right:
        return True
    left_key = _canonical_name_key(left)
    right_key = _canonical_name_key(right)
    if left_key == right_key:
        return True
    if left_key.split()[0] == right_key.split()[0] and (
        len(left_key.split()) == 1 or len(right_key.split()) == 1
    ):
        return True
    return False


def _is_known_chercheur_name(
    name: str,
    guides_by_chercheur: dict[str, list[tuple[str, dict]]] | None = None,
) -> bool:
    if not name:
        return False
    if name in CHERCHEUR_LABELS or name in CHERCHEUR_LABELS.values():
        return True
    if name in TEMOIN_VOICE_ORDER:
        return True
    key = _canonical_name_key(name)
    if any(_canonical_name_key(item) == key for item in TEMOIN_VOICE_ORDER):
        return True
    if guides_by_chercheur and any(
        _names_match_chercheur(name, chercheur) for chercheur in guides_by_chercheur
    ):
        return True
    return False


def _resolve_chercheur_label(
    name: str,
    guides_by_chercheur: dict[str, list[tuple[str, dict]]],
) -> str:
    if not name:
        return ""
    if name in CHERCHEUR_LABELS:
        full = CHERCHEUR_LABELS[name]
        for key in guides_by_chercheur:
            if _names_match_chercheur(full, key):
                return key
        for voice in TEMOIN_VOICE_ORDER:
            if _names_match_chercheur(full, voice):
                return voice
        return full
    for key in guides_by_chercheur:
        if _names_match_chercheur(name, key):
            return key
    for voice in TEMOIN_VOICE_ORDER:
        if _names_match_chercheur(name, voice):
            return voice
    return name


def _synthese_guides_by_chercheur(
    capsule_data: dict,
) -> dict[str, list[tuple[str, dict]]]:
    """chercheur -> [(code_expert, guide), ...] dans l'ordre des orientations."""
    grouped: dict[str, list[tuple[str, dict]]] = defaultdict(list)
    seen: set[tuple[str, str, str, str]] = set()
    for orientation in capsule_data.get("orientations_expert", []):
        expert_code = orientation.get("code", "")
        for guide in _orientation_guides(orientation):
            chercheur = (guide.get("chercheur") or "").strip()
            if not chercheur:
                continue
            key = (
                chercheur,
                expert_code,
                guide.get("extrait_id", ""),
                guide.get("origine") or guide.get("angle") or "",
            )
            if key in seen:
                continue
            seen.add(key)
            grouped[chercheur].append((expert_code, guide))
    return grouped


def _synthese_chercheur_order(
    capsule_data: dict,
    by_id: dict[str, dict] | None,
    resume_items: list[tuple[str, str]],
    guides_by_chercheur: dict[str, list[tuple[str, dict]]],
) -> list[str]:
    order: list[str] = []

    def _push(name: str) -> None:
        label = _resolve_chercheur_label(name, guides_by_chercheur) if name else ""
        if not label or label in order:
            return
        if not _is_known_chercheur_name(label, guides_by_chercheur):
            return
        order.append(label)

    if by_id:
        for extrait_id in capsule_data.get("ordre_montage", []):
            segment = by_id.get(extrait_id)
            if segment and segment.get("chercheur"):
                _push(segment["chercheur"])

    for name, _content in resume_items:
        if name and _is_known_chercheur_name(name, guides_by_chercheur):
            _push(name)

    for chercheur in guides_by_chercheur:
        _push(chercheur)

    return order


def _synthese_resume_for_chercheur(
    chercheur: str,
    resume_items: list[tuple[str, str]],
    guides: list[tuple[str, dict]],
) -> str:
    for name, content in resume_items:
        if not name or not content:
            continue
        if _names_match_chercheur(name, chercheur) or _names_match_chercheur(
            _label_chercheur(name), chercheur
        ):
            return content
    for _code, guide in guides:
        detail = (guide.get("dans_le_temoin") or "").strip()
        if detail:
            return _strip_chercheur_prefix(detail, chercheur)
    return ""


def _synthese_intro_globale(
    resume_items: list[tuple[str, str]],
    guides_by_chercheur: dict[str, list[tuple[str, dict]]],
) -> str:
    """Resume d'ensemble (ex. GEN) non rattache a une voix."""
    parts = []
    for name, content in resume_items:
        if not content:
            continue
        if name and _is_known_chercheur_name(name, guides_by_chercheur):
            continue
        if name:
            parts.append(f"{name} : {content}")
        else:
            parts.append(content)
    return " ".join(parts).strip()


def _render_synthese_objectifs_expert(capsule_data: dict) -> str:
    videos = capsule_data.get("videos_expert", [])
    if not videos:
        return ""
    rows = []
    for video in videos:
        code = video.get("code", "")
        titre = video.get("titre", "")
        descriptif = video.get("descriptif", "")
        label = _label_video_expert(code)
        line = f"<li><strong>{escape(label)}</strong> — {_e_fr(titre)}"
        if descriptif:
            line += f"<br><span class='meta'>{_e_fr(descriptif)}</span>"
        line += "</li>"
        rows.append(line)
    return f"""
  <div class="synthese-objectifs-expert">
    <h3>Objectifs des vidéos expert à atteindre</h3>
    <ul>
      {''.join(rows)}
    </ul>
  </div>
"""


def _render_synthese_appui_items(
    guides: list[tuple[str, dict]],
    videos_by_code: dict[str, dict],
) -> str:
    if not guides:
        return "<p class='meta'>Pas encore d'orientation expert documentée pour cette voix.</p>"
    rows = []
    for expert_code, guide in guides:
        video = videos_by_code.get(expert_code, {})
        expert_label = _label_video_expert(expert_code)
        angle = guide.get("origine") or guide.get("angle") or ""
        extrait_id = guide.get("extrait_id", "")
        timecodes = guide.get("timecodes", "")
        travail = guide.get("travail_expert", "")
        concepts = guide.get("concepts") or guide.get("concepts_e1") or []
        objectif = video.get("titre", "")
        # Ne pas exposer un id technique « TRANSCRIPT-T* » comme si l'apprenant
        # ou l'expert devait consulter un transcript : c'est un outil interne.
        extrait_display = ""
        if extrait_id:
            m = re.match(r"^TRANSCRIPT[-_]?T?(\d+)$", str(extrait_id).strip(), flags=re.I)
            if m:
                extrait_display = f"vidéo témoin {m.group(1)}"
            elif re.search(r"transcript", str(extrait_id), flags=re.I):
                extrait_display = "vidéo témoin"
            else:
                extrait_display = str(extrait_id)
        meta_parts = [p for p in (extrait_display, timecodes) if p]
        meta = " · ".join(meta_parts)
        header = f"<strong>{escape(expert_label)}</strong>"
        if objectif:
            header += f" — {_e_fr(objectif)}"
        if angle:
            header += f" <span class='meta'>({_e_fr(angle)})</span>"
        body = []
        if meta:
            body.append(f"<p class='meta'>Appui : {escape(meta)}</p>")
            if travail:
                body.append(
                    f"<p><strong>Sur quoi on s'appuie :</strong> "
                    f"{escape(_normalize_editorial_french(travail))}</p>"
                )
        if concepts:
            body.append(
                f"<p class='meta'><strong>Concepts :</strong> "
                f"{_e_fr(' · '.join(concepts))}</p>"
            )
        rows.append(
            "<li>"
            f"<p>{header}</p>"
            f"{''.join(body)}"
            "</li>"
        )
    return f'<ul class="synthese-appui-list">{"".join(rows)}</ul>'


def _guides_for_chercheur(
    chercheur: str,
    guides_by_chercheur: dict[str, list[tuple[str, dict]]],
) -> list[tuple[str, dict]]:
    if chercheur in guides_by_chercheur:
        return guides_by_chercheur[chercheur]
    for key, value in guides_by_chercheur.items():
        if _names_match_chercheur(chercheur, key):
            return value
    return []


def _build_synthese_chercheur_blocks(
    capsule_data: dict,
    by_id: dict[str, dict] | None = None,
) -> tuple[str, list[dict]]:
    resume_items = _parse_resume_temoignages(
        capsule_data.get("resume_temoignages", "")
    )
    guides_by_chercheur = _synthese_guides_by_chercheur(capsule_data)
    order = _synthese_chercheur_order(
        capsule_data, by_id, resume_items, guides_by_chercheur
    )
    videos_by_code = _videos_expert_by_code(capsule_data)
    intro = _synthese_intro_globale(resume_items, guides_by_chercheur)
    blocks: list[dict] = []
    for chercheur in order:
        guides = _guides_for_chercheur(chercheur, guides_by_chercheur)
        dit = _synthese_resume_for_chercheur(chercheur, resume_items, guides)
        if not dit and not guides:
            continue
        blocks.append(
            {
                "chercheur": chercheur,
                "dit": dit,
                "guides": guides,
                "videos_by_code": videos_by_code,
            }
        )
    if not blocks and resume_items:
        for name, content in resume_items:
            if not content:
                continue
            if name and not _is_known_chercheur_name(name, guides_by_chercheur):
                continue
            blocks.append(
                {
                    "chercheur": _resolve_chercheur_label(name, guides_by_chercheur)
                    if name
                    else "",
                    "dit": content,
                    "guides": [],
                    "videos_by_code": videos_by_code,
                }
            )
    return intro, blocks


def synthese_temoignages_section(
    capsule_code: str,
    capsule_data: dict,
    by_id: dict[str, dict] | None = None,
) -> str:
    intro, blocks = _build_synthese_chercheur_blocks(capsule_data, by_id)
    if not blocks and not capsule_data.get("videos_expert") and not intro:
        return ""

    articles = []
    for block in blocks:
        chercheur = block["chercheur"]
        dit = block["dit"]
        heading = escape(chercheur) if chercheur else "Synthèse"
        contexte = ""
        # Contexte factuel tiré de l'orientation "dans le témoin" pour éviter
        # que le lecteur ne doive connaître tout le projet en amont.
        if block.get("guides"):
            for _expert_code, guide in block["guides"]:
                detail = (guide.get("dans_le_temoin") or "").strip()
                if detail:
                    contexte = _strip_chercheur_prefix(detail, chercheur).strip()
                    break

        dit_html = ""
        if contexte:
            dit_html += f"<p class='meta'><strong>Contexte :</strong> {escape(_normalize_editorial_french(contexte))}</p>"
        if dit:
            dit_html += (
                f"<p><strong>Ce que dit le chercheur dans la vidéo témoin :</strong> "
                f"<strong>{escape(_normalize_editorial_french(dit))}</strong></p>"
            )
        else:
            dit_html = "<p class='meta'>Résumé du dit à compléter.</p>"
        appui_html = _render_synthese_appui_items(
            block["guides"], block["videos_by_code"]
        )
        articles.append(
            f"""
  <article class="synthese-chercheur">
    <h3>{heading}</h3>
    {dit_html}
    <h4>Appui vers les objectifs des vidéos expert</h4>
    {appui_html}
  </article>
"""
        )

    if not articles and not capsule_data.get("videos_expert") and not intro:
        return ""

    objectifs_html = _render_synthese_objectifs_expert(capsule_data)
    intro_html = f"<p>{escape(intro)}</p>" if intro else ""
    return f"""
<section class="methodology-panel synthese-chorale-panel">
  <h2>Synthèse des témoignages</h2>
  <p class="meta">Ce que chaque chercheur a dit dans le script, et sur quels éléments on s'appuie pour préparer les objectifs des vidéos expert.</p>
  {intro_html}
  {objectifs_html}
  {''.join(articles)}
</section>
"""


def export_synthese_section_title(capsule_code: str) -> str:
    # Le nom de la vidéo témoin est déjà porté par le chapitre / la page.
    return "Synthèse des témoignages"


def export_synthese_temoignages_plaintext(
    capsule_code: str,
    capsule_data: dict,
    by_id: dict[str, dict] | None = None,
) -> str:
    intro, blocks = _build_synthese_chercheur_blocks(capsule_data, by_id)
    videos = capsule_data.get("videos_expert", [])
    if not blocks and not videos and not intro:
        return ""

    lines = [
        "Ce que chaque chercheur a dit dans le script, et sur quels éléments "
        "on s'appuie pour préparer les objectifs des vidéos expert.",
        "",
    ]
    if intro:
        lines.append(intro)
        lines.append("")
    if videos:
        lines.append("Objectifs des vidéos expert à atteindre")
        for video in videos:
            code = video.get("code", "")
            titre = video.get("titre", "")
            descriptif = video.get("descriptif", "")
            line = f"- {_label_video_expert(code)} — {titre}"
            if descriptif:
                line += f" ({descriptif})"
            lines.append(line)
        lines.append("")

    for block in blocks:
        chercheur = block["chercheur"] or "Synthèse"
        contexte = ""
        if block.get("guides"):
            for _expert_code, guide in block["guides"]:
                detail = (guide.get("dans_le_temoin") or "").strip()
                if detail:
                    contexte = _strip_chercheur_prefix(detail, chercheur).strip()
                    break
        lines.append(chercheur)
        if contexte:
            lines.append(f"Contexte : {contexte}")
        if block["dit"]:
            lines.append(f"Ce que dit le chercheur dans la vidéo témoin : **{block['dit']}**")
        lines.append("Appui vers les objectifs des vidéos expert :")
        if not block["guides"]:
            lines.append("- Pas encore d'orientation expert documentée pour cette voix.")
        else:
            for expert_code, guide in block["guides"]:
                video = block["videos_by_code"].get(expert_code, {})
                angle = guide.get("origine") or guide.get("angle") or ""
                extrait_id = guide.get("extrait_id", "")
                travail = guide.get("travail_expert", "")
                concepts = guide.get("concepts") or guide.get("concepts_e1") or []
                objectif = video.get("titre", "")
                head = f"- {_label_video_expert(expert_code)}"
                if objectif:
                    head += f" — {objectif}"
                if angle:
                    head += f" ({angle})"
                if extrait_id:
                    head += f" [{extrait_id}]"
                lines.append(head)
                if travail:
                    lines.append(f"  Sur quoi on s'appuie : {travail}")
                if concepts:
                    lines.append(f"  Concepts : {' · '.join(concepts)}")
        lines.append("")
    return _normalize_editorial_french("\n".join(lines).strip())


def brief_intervenant_section(
    capsule_code: str, capsule_data: dict, by_id: dict[str, dict]
) -> str:
    videos = capsule_data.get("videos_expert", [])
    if not videos:
        return ""

    temoin_html = _render_brief_temoin(capsule_code, capsule_data, by_id)
    videos_html = "".join(
        _render_brief_video(video)
        for video in videos
    )
    consignes = [_humanize_capsule_labels(item) for item in BRIEF_CONSIGNES_COMMUNES]
    consignes_html = (
        "<h3>Consignes générales</h3>"
        "<ul>"
        + "".join(f"<li>{escape(item)}</li>" for item in consignes)
        + "</ul>"
    )
    precaution_html = (
        f'<p class="brief-precaution"><strong>Précaution :</strong> '
        f"{escape(BRIEF_PRECAUTION_ORATOIRE)}</p>"
    )

    return f"""
<section class="methodology-panel brief-intervenant-panel">
  <h2>{escape(EXPORT_BRIEF_SECTION_TITLE)}</h2>
  <p class="meta">Quelques repères proposés pour préparer la ou les vidéos expertise, en s'appuyant sur les témoignages et les objectifs du programme de conception.</p>
  {precaution_html}
  {temoin_html}
  {videos_html}
  {consignes_html}
  <p class="meta">Version détaillée (extraits, timecodes, passerelles) disponible plus bas sur cette page.</p>
</section>
"""


def export_brief_intervenant_plaintext(
    capsule_code: str, capsule_data: dict, by_id: dict[str, dict]
) -> str:
    videos = capsule_data.get("videos_expert", [])
    if not videos:
        return ""

    lines = [
        "Quelques repères proposés pour préparer la ou les vidéos expertise, "
        "en s'appuyant sur les témoignages et les objectifs du programme de conception.",
        "",
    ]
    lines.append(f"Précaution : {BRIEF_PRECAUTION_ORATOIRE}")
    lines.append("")

    narrative = _narrative_temoin_block(capsule_code)
    if narrative and (narrative.get("voix") or narrative.get("intro")):
        lines.extend(_export_brief_temoin_narrative_plaintext(narrative))
    else:
        temoins = _collect_temoignages_lisibles(capsule_data, by_id)
        if temoins:
            lines.append("Ce que disent les chercheurs")
            lines.append("")
            for chercheur, phrases in temoins:
                if chercheur:
                    lines.append(chercheur)
                    lines.append("")
                for phrase in phrases:
                    lines.append(phrase)
                    lines.append("")

    for video in videos:
        label = _tb_expertise_label(_label_video_expert(video.get("code", "")))
        intervenant = video.get("intervenant")
        if intervenant:
            lines.append(f"{label} — {intervenant}")
        else:
            lines.append(label)
        lines.append(f"Objectif : {video.get('titre', '')}")
        if video.get("descriptif"):
            lines.append(video["descriptif"])
        lines.append("")
        lines.extend(_fascicule_refs_plaintext_lines(video.get("code", "")))

    lines.append("Consignes générales :")
    for item in BRIEF_CONSIGNES_COMMUNES:
        lines.append(f"- {_humanize_capsule_labels(item)}")
    return _normalize_editorial_french("\n".join(lines).strip())


def selection_unites_section(capsule_data: dict) -> str:
    unites = capsule_data.get("unites_de_sens", [])
    orientations = capsule_data.get("orientations_expert") or []
    if not orientations and capsule_data.get("orientation_expert"):
        orientations = [capsule_data["orientation_expert"]]
    if not unites and not orientations:
        return ""

    grille_label = "Grille expert"
    if orientations:
        codes = [o.get("code", "") for o in orientations if o.get("code")]
        if len(codes) == 1:
            grille_label = f"Grille {codes[0]}"

    provisoire = any(u.get("statut") == "PROVISOIRE" for u in unites)
    meta_intro = "Synthèse éditoriale des séquences retenues dans le script témoin."
    if provisoire:
        meta_intro += " Unités provisoires basées sur le programme de conception — à préciser après cartographie BAB."

    rows = []
    for unite in unites:
        extraits = ", ".join(unite.get("extraits", []))
        grille = unite.get("grille_expert") or unite.get("grille_e1") or "—"
        rows.append(
            "<tr>"
            f"<td>{unite.get('ordre', '')}</td>"
            f"<td>{escape(extraits)}</td>"
            f"<td>{_e_fr(unite.get('acte', ''))}</td>"
            f"<td>{_e_fr(unite.get('libelle', ''))}</td>"
            f"<td>{_e_fr(grille)}</td>"
            "</tr>"
        )

    html = f"""
<section class="methodology-panel">
  <h2>Unités de sens sélectionnées</h2>
  <p class="meta">{_e_fr(meta_intro)}</p>
  <table class="unites-table">
    <thead><tr><th>#</th><th>Extraits</th><th>Acte</th><th>Unité de sens</th><th>{escape(grille_label)}</th></tr></thead>
    <tbody>
"""
    html += "\n".join(rows) or "<tr><td colspan='5'>Aucune unité documentée.</td></tr>"
    html += """
    </tbody>
  </table>
"""

    if orientations:
        plural = len(orientations) > 1
        for i, orientation in enumerate(orientations):
            o = dict(orientation)
            if plural and i > 0:
                html += _render_orientation_block(o, plural=False).replace(
                    "<h2>Orientation pour la vidéo expert suivante</h2>",
                    f"<h2>Orientation — {escape(o.get('code', ''))}</h2>",
                )
            else:
                html += _render_orientation_block(o, plural=plural)

    html += "</section>"
    return html


def cadrage_animateur_section(capsule_data: dict) -> str:
    cadrage = capsule_data.get("cadrage_animateur")
    if not cadrage:
        return ""

    def render_bloc(title: str, bloc: dict, kind: str) -> str:
        parts = [
            f'<div class="cadrage-block">',
            f"<h3>{escape(title)}</h3>",
            f'<p class="cadrage-position"><strong>Quand :</strong> {_e_fr(bloc.get("position", ""))}</p>',
        ]
        if bloc.get("duree_cible_secondes"):
            parts.append(
                f'<p class="meta">Durée cible : ~{bloc["duree_cible_secondes"]} s · '
                f'{_e_fr(bloc.get("fonction", ""))}</p>'
            )
        elif bloc.get("fonction"):
            parts.append(f'<p class="meta">{_e_fr(bloc["fonction"])}</p>')
        if kind == "transition":
            parts.append(
                f'<p class="meta">Après <strong>{escape(bloc.get("apres_extrait", ""))}</strong> · '
                f'Avant <strong>{escape(bloc.get("avant_extrait", ""))}</strong></p>'
            )
        if bloc.get("texte_intervenant"):
            parts.append(
                f'<p><strong>Version animateur</strong></p>'
                f'<blockquote class="cadrage-quote">{_e_fr(bloc["texte_intervenant"])}</blockquote>'
            )
        if bloc.get("texte_pancarte"):
            parts.append(
                f'<p><strong>Version pancarte</strong></p>'
                f'<pre class="cadrage-pancarte">{_e_fr(bloc["texte_pancarte"])}</pre>'
            )
        if bloc.get("voix_off_optionnelle"):
            parts.append(
                f'<p class="meta"><strong>Voix off optionnelle :</strong> '
                f'« {_e_fr(bloc["voix_off_optionnelle"])} »</p>'
            )
        if bloc.get("enchainement_expert"):
            parts.append(
                f'<p class="meta"><strong>Enchaînement :</strong> vidéo(s) expert '
                f'{escape(bloc["enchainement_expert"])}</p>'
            )
        parts.append("</div>")
        return "\n".join(parts)

    transitions_html = "".join(
        render_bloc(
            f"Transition — {item.get('id', '').replace('_', ' ')}",
            item,
            "transition",
        )
        for item in cadrage.get("transitions", [])
    )

    return f"""
<section class="methodology-panel cadrage-panel">
  <h2>Cadrage animateur — intro, transitions, outro</h2>
  <p class="meta"><strong>Statut :</strong> {escape(cadrage.get("statut", "NON_PRONONCE"))} — ces propositions sont intégrées au script final ci-dessous (marquées CADRAGE — NON PRONONCE).</p>
  <p>{_e_fr(cadrage.get("dispositif", ""))}</p>
  <p class="meta">{_e_fr(cadrage.get("note", ""))}</p>
  <table>
    <thead><tr><th>Étape</th><th>Position dans le montage</th><th>Fonction</th></tr></thead>
    <tbody>
      <tr><td>Intro</td><td>{_e_fr(cadrage.get("intro", {}).get("position", ""))}</td><td>{_e_fr(cadrage.get("intro", {}).get("fonction", ""))}</td></tr>
      {''.join(f"<tr><td>Transition {escape(item.get('id', ''))}</td><td>{_e_fr(item.get('position', ''))}</td><td>{_e_fr(item.get('fonction', ''))}</td></tr>" for item in cadrage.get("transitions", []))}
      <tr><td>Outro</td><td>{_e_fr(cadrage.get("outro", {}).get("position", ""))}</td><td>{_e_fr(cadrage.get("outro", {}).get("fonction", ""))}</td></tr>
    </tbody>
  </table>
  {render_bloc("Intro", cadrage.get("intro", {}), "intro")}
  {transitions_html}
  {render_bloc("Outro", cadrage.get("outro", {}), "outro")}
</section>
"""


def referents_section(capsule_data: dict) -> str:
    videos = capsule_data.get("videos_expert", [])
    if not videos:
        return ""

    items = []
    for video in videos:
        intervenant = video.get("intervenant")
        if intervenant:
            who = escape(intervenant)
        else:
            who = "<em>Intervenant désigné</em>"
        desc = video.get("descriptif", "")
        desc_html = f" {escape(_normalize_editorial_french(desc))}" if desc else ""
        items.append(
            f"<li><strong>{escape(_label_video_expert(video.get('code', '')))}</strong> — {who} : "
            f"{escape(_normalize_editorial_french(video.get('titre', '')))}.{desc_html}</li>"
        )

    return f"""
<section class="methodology-panel referents-panel">
  <h2>Vidéos expert à produire</h2>
  <p class="meta">Programme de conception mis à jour le 2026-07-10 (source : 20260710_Prev_Vid.xlsx).</p>
  <ul>
    {''.join(items)}
  </ul>
</section>
"""


def _is_script_meta_line(line: str) -> bool:
    """Lignes techniques à griser (refs BAB, projection, volume, etc.)."""
    stripped = (line or "").strip()
    if not stripped:
        return False
    if re.match(r"^\[[A-Z0-9-]+\]\s.+\|\s.+\|\s.+$", stripped):
        return True
    if stripped.startswith("[PROJECTION"):
        return True
    if stripped.startswith("Objectif pédagogique :") or stripped.startswith("Objectif pedagogique :"):
        return True
    if stripped.startswith("Cible :"):
        return True
    if re.match(r"^Volume\s*:\s*\d+\s*mots\b", stripped, flags=re.IGNORECASE):
        return True
    return False


def _script_lines_html(text: str) -> str:
    """Rend un script avec les mentions techniques grisées (style script-ref)."""
    lines = (text or "").splitlines()
    rendered: list[str] = []
    for line in lines:
        safe_line = escape(line)
        if line.strip().startswith("Ce script est une projection"):
            rendered.append(f"<strong>{safe_line}</strong>")
        elif _is_script_meta_line(line):
            rendered.append(f"<span class='script-ref'>{safe_line}</span>")
        else:
            rendered.append(safe_line)
    return f"<div class='script-body'>{'<br>'.join(rendered)}</div>"


SCRIPT_EXPERTISE_DISCLAIMER = (
    "Ce script est une projection construite avec l'IA pour vous orienter, "
    "à partir des consignes et du cadrage éditorial. "
    "La justesse des informations n'est pas vérifiée : elle n'a pas été établie "
    "par une personne qui a l'expertise du sujet. "
    "À reprendre, corriger et valider par l'intervenant expert."
)

SCRIPT_EXPERTISE_WORD_MIN = 450
SCRIPT_EXPERTISE_WORD_MAX = 800


def _strip_guillemets(text: str) -> str:
    cleaned = (text or "").strip()
    for mark in ("«", "»", '"', "'"):
        cleaned = cleaned.strip(mark).strip()
    return cleaned


def _count_words_fr(text: str) -> int:
    return len(re.findall(r"[A-Za-zÀ-ÖØ-öø-ÿ0-9']+", text or ""))


def _sentence_case(text: str) -> str:
    cleaned = (text or "").strip()
    if not cleaned:
        return ""
    if cleaned[0].islower():
        return cleaned[0].upper() + cleaned[1:]
    return cleaned


def _ensure_period(text: str) -> str:
    cleaned = (text or "").strip().rstrip(";")
    if not cleaned:
        return ""
    if cleaned[-1] not in ".!?…":
        return cleaned + "."
    return cleaned


def _chercheur_pronoun(chercheur: str) -> str:
    prenom = _chercheur_prenom(chercheur).lower()
    if prenom in {
        "muriel",
        "sylvia",
        "stephanie",
        "soizic",
        "arielle",
        "eneli",
        "virginia",
        "fatoumata",
    }:
        return "elle"
    return "il"


def _normalize_concept_casing(label: str) -> str:
    raw = (label or "").strip()
    if not raw:
        return raw
    if raw.isupper() and len(raw) <= 4:
        return raw
    return raw[0].lower() + raw[1:] if raw[:1].isupper() else raw


def _fr_label_with_article(label: str) -> str:
    """Ajoute un determinant francais a un concept / angle."""
    raw = _normalize_concept_casing(label)
    if not raw:
        return ""
    lower = raw.lower()
    if lower.startswith(
        (
            "le ",
            "la ",
            "les ",
            "l'",
            "l’",
            "un ",
            "une ",
            "des ",
            "du ",
            "de la ",
            "de l'",
            "de l’",
        )
    ):
        return raw
    feminine_exact = {
        "serendipite",
        "serendipité",
        "rencontre",
        "observation d'usage",
        "observation d usage",
        "maturation progressive",
        "publication",
        "confidentialite",
        "confidentialité",
        "nouveaute",
        "nouveauté",
        "divulgation",
        "prematuration",
        "prématuration",
        "maturation",
        "licence",
        "creation",
        "création",
        "gouvernance",
        "dilution",
        "collaboration",
        "pi",
        "di",
    }
    key = lower.replace("é", "e").replace("è", "e").replace("ê", "e")
    vowel = "aeiouyhàâäéèêëîïôöùûüœ"
    if key in feminine_exact or key.startswith("observation") or key.startswith("idee"):
        if lower[0] in vowel:
            return f"l'{raw}"
        return f"la {raw}"
    if lower[0] in vowel:
        return f"l'{raw}"
    return f"le {raw}"


def _starts_with_verbish(text: str) -> bool:
    first = (text or "").strip().split(" ", 1)[0].lower().rstrip(",;:")
    verbs = {
        "raconte",
        "racontent",
        "part",
        "partent",
        "definit",
        "définit",
        "illustre",
        "illustrent",
        "montre",
        "montrent",
        "precise",
        "précise",
        "explique",
        "expliquent",
        "oppose",
        "ajoute",
        "ajoutent",
        "decrit",
        "décrit",
        "parle",
        "parlent",
        "incarne",
        "croise",
        "croisent",
        "nommer",
        "definir",
        "définir",
        "montrer",
        "expliquer",
        "insister",
        "distinguer",
        "preparer",
        "préparer",
        "installer",
        "lier",
        "traiter",
        "reduire",
        "réduire",
        "confondre",
        "resumer",
        "résumer",
        "poser",
        "rappeler",
        "structurer",
        "clarifier",
        "cartographier",
        "securiser",
        "sécuriser",
        "organiser",
        "concevoir",
        "gerer",
        "gérer",
        "passer",
        "comprendre",
        "adapter",
        "annoncer",
        "suivre",
        "partir",
        "completer",
        "compléter",
        "inviter",
    }
    return first in verbs


_INFINITIVES = {
    "nommer",
    "definir",
    "définir",
    "montrer",
    "expliquer",
    "insister",
    "distinguer",
    "preparer",
    "préparer",
    "installer",
    "lier",
    "traiter",
    "reduire",
    "réduire",
    "confondre",
    "resumer",
    "résumer",
    "poser",
    "rappeler",
    "structurer",
    "clarifier",
    "cartographier",
    "securiser",
    "sécuriser",
    "organiser",
    "concevoir",
    "gerer",
    "gérer",
    "passer",
    "comprendre",
    "adapter",
    "annoncer",
    "suivre",
    "partir",
    "completer",
    "compléter",
    "inviter",
}


def _ensure_oral_subject(text: str, chercheur: str) -> str:
    """Evite « En effet, raconte… » → « En effet, il raconte… »."""
    cleaned = (text or "").strip()
    if not cleaned:
        return ""
    prenom = _chercheur_prenom(chercheur)
    pronoun = _chercheur_pronoun(chercheur)
    lower = cleaned.lower()
    prefixes = (
        prenom.lower() + " ",
        (chercheur or "").lower(),
        f"{pronoun} ",
        "on ",
        "ce ",
        "cet ",
        "cette ",
        "ces ",
        "son ",
        "sa ",
        "ses ",
        "le ",
        "la ",
        "les ",
        "l'",
        "l’",
        "un ",
        "une ",
        "dans ",
        "avec ",
        "pour ",
        "quand ",
        "si ",
        "mais ",
        "et ",
        "ou ",
        "car ",
        "donc ",
        "alors ",
        "ici ",
        "là ",
        "au ",
        "aux ",
    )
    if any(lower.startswith(p) for p in prefixes):
        return _ensure_period(cleaned)
    if _starts_with_verbish(cleaned):
        first = cleaned.split(" ", 1)[0].lower().rstrip(",;:")
        if first in _INFINITIVES:
            return _ensure_period(f"Il s'agit de {cleaned[0].lower() + cleaned[1:]}")
        return _ensure_period(f"{pronoun} {cleaned[0].lower() + cleaned[1:]}")
    return _ensure_period(cleaned)


def _oralize_instruction(text: str) -> str:
    """Transforme une consigne type fiche en tour oral."""
    cleaned = (text or "").strip()
    if not cleaned:
        return ""
    chunks = re.split(r"(?<=[.!?])\s+", cleaned)
    out: list[str] = []
    for chunk in chunks:
        chunk = chunk.strip()
        if not chunk:
            continue
        first = chunk.split(" ", 1)[0].lower().rstrip(",;:")
        if first in _INFINITIVES and not chunk.lower().startswith(("il ", "elle ", "on ")):
            if first == "insister":
                rest = chunk.split(" ", 1)[1] if " " in chunk else ""
                rest = rest.lstrip(" :")
                out.append(_ensure_period(f"J'insiste : {rest}" if rest else "J'insiste."))
            else:
                out.append(_ensure_period(f"Il s'agit de {chunk[0].lower() + chunk[1:]}"))
        else:
            out.append(_ensure_period(chunk))
    return " ".join(out)


def _form_angle_oral(angle: str) -> str:
    return _fr_label_with_article(angle) if angle else ""


def _a_plus_article(phrase: str) -> str:
    """Contracte « a le / a les » → « au / aux »."""
    p = (phrase or "").strip()
    if p.startswith("le "):
        return "au " + p[3:]
    if p.startswith("les "):
        return "aux " + p[4:]
    if p.startswith("la "):
        return "a la " + p[3:]
    if p.startswith("l'"):
        return "a l'" + p[2:]
    if p.startswith("l’"):
        return "a l’" + p[2:]
    return "a " + p


def _list_concepts_oral(concepts: list[str]) -> str:
    labeled = [_fr_label_with_article(c) for c in concepts if c]
    if not labeled:
        return ""
    if len(labeled) == 1:
        return labeled[0]
    if len(labeled) == 2:
        return f"{labeled[0]} et {labeled[1]}"
    return f"{', '.join(labeled[:-1])} et {labeled[-1]}"


def _oralize_editorial_meta(text: str) -> str:
    """
    Retire le jargon de conception d'un texte destiné à l'oral apprenant.
    Le transcript, les codes E et le « côté programme » sont des outils internes.
    L'apprenant a vu une vidéo témoin, pas un transcript.
    """
    if not (text or "").strip():
        return ""
    t = text.strip()
    # Phrases / parenthèses purement techniques.
    t = re.sub(
        r"\([^)]*transcript[^)]*\)",
        "",
        t,
        flags=re.IGNORECASE,
    )
    t = re.sub(
        r"[^.?!]*\btranscripts?\b[^.?!]*[.?!]\s*",
        "",
        t,
        flags=re.IGNORECASE,
    )
    t = re.sub(
        r"\s*\(objectifs?\s+E\d+(?:bis)?(?:\s*/\s*E\d+(?:bis)?)?[^)]*\)",
        "",
        t,
        flags=re.IGNORECASE,
    )
    t = re.sub(r"\s*côté programme\b", "", t, flags=re.IGNORECASE)
    t = re.sub(r"\s*cote programme\b", "", t, flags=re.IGNORECASE)
    # Codes E restants → formulation orale.
    t = re.sub(r"\bE(\d+)(bis)?\b", "cette vidéo", t, flags=re.IGNORECASE)
    t = re.sub(
        r"(^|[.!?]\s+)cette vidéo\b",
        lambda m: f"{m.group(1)}Cette vidéo",
        t,
    )
    # Cadrage conception → cadrage oral.
    t = re.sub(
        r"\bCette vidéo sert les objectifs?\b",
        "Cette vidéo témoin éclaire",
        t,
        flags=re.IGNORECASE,
    )
    t = re.sub(
        r"\bCette vidéo sert\b",
        "Dans la vidéo témoin, on voit surtout",
        t,
        flags=re.IGNORECASE,
    )
    t = re.sub(
        r"\bCette vidéo nourrit les objectifs?\b",
        "La vidéo témoin nourrit",
        t,
        flags=re.IGNORECASE,
    )
    t = re.sub(
        r"\bCette vidéo ouvre le parcours \(objectif\s*:\s*",
        "La vidéo témoin ouvre le parcours — ",
        t,
        flags=re.IGNORECASE,
    )
    # Ne pas présupposer que l'apprenant a déjà un projet d'innovation.
    t = re.sub(
        r"\bdans votre propre projet\b",
        "dans votre activité de recherche",
        t,
        flags=re.IGNORECASE,
    )
    t = re.sub(
        r"\bson propre projet\b",
        "sa pratique de recherche",
        t,
        flags=re.IGNORECASE,
    )
    t = re.sub(
        r"\bdans votre projet\b",
        "dans votre situation",
        t,
        flags=re.IGNORECASE,
    )
    t = re.sub(
        r"\bvers votre projet\b",
        "vers votre situation",
        t,
        flags=re.IGNORECASE,
    )
    t = re.sub(r"\s{2,}", " ", t).strip(" ,;")
    t = re.sub(r"\s+([.!?])", r"\1", t)
    return t.strip()


def _script_variant_seed(code: str, variant_index: int = 0) -> int:
    """Graine stable pour varier le style d'un script expertise."""
    return (sum(ord(c) for c in (code or "E")) * 31 + int(variant_index) * 17) % 997


def _voix_phrase_from_prenoms(prenoms: list[str]) -> str:
    if len(prenoms) >= 2:
        return f"{', '.join(prenoms[:-1])} et {prenoms[-1]}"
    if prenoms:
        return prenoms[0]
    return ""


def _script_expertise_open_paragraph(
    *,
    titre: str,
    titre_oral: str,
    introduction: str,
    descriptif: str,
    guides: list[dict],
    prenoms: list[str],
    voix_phrase: str,
    variant_index: int,
    sibling_count: int,
    seed: int,
) -> str:
    """Ouverture variée : évite le même modèle pour chaque vidéo expertise."""
    q0 = ""
    if guides and guides[0].get("question_apprenant"):
        q0 = _oralize_editorial_meta(
            guides[0]["question_apprenant"].strip().rstrip("?").rstrip()
        )
    first_prenom = prenoms[0] if prenoms else ""
    first_hook = ""
    if guides:
        g0 = guides[0]
        chercheur = (g0.get("chercheur") or "").strip()
        brut = _strip_chercheur_prefix((g0.get("dans_le_temoin") or "").strip(), chercheur)
        if brut:
            first_hook = brut
            if len(first_hook) > 160:
                cut = first_hook[:157].rsplit(" ", 1)[0]
                first_hook = cut.rstrip(" ,;") + "…"

    intro_clean = ""
    if introduction:
        intro = _oralize_editorial_meta(introduction)
        intro = re.sub(r"\bLa chorale T\d+\b", "Ces temoignages", intro, flags=re.I)
        intro = re.sub(r"\bCes temoignages vient\b", "Ces temoignages viennent", intro, flags=re.I)
        intro = re.sub(r"\bCes temoignages montre\b", "Ces temoignages montrent", intro, flags=re.I)
        intro = re.sub(r"\b\(JJG,\s*MUR,\s*SYL,\s*YAN\)", "", intro)
        intro_clean = re.sub(r"\s{2,}", " ", intro).strip()
        if re.search(r"\btranscript\b", intro_clean, flags=re.I):
            intro_clean = ""

    # Quand plusieurs vidéos expertise pour le même intervenant : forcer des familles distinctes.
    if sibling_count > 1:
        family = variant_index % 5
    else:
        family = seed % 5

    bits: list[str] = []

    if family == 0:
        # Accroche directe sur l'objectif.
        bits.append("Bonjour.")
        if titre_oral:
            bits.append(
                f"Je voudrais m'arrêter sur un geste précis : {titre_oral}."
            )
        if first_hook and first_prenom:
            bits.append(
                f"Dans la vidéo témoin, {first_prenom} en donne déjà le goût : "
                f"{first_hook[0].lower() + first_hook[1:] if first_hook[:1].isupper() else first_hook}"
            )
            if not first_hook.endswith((".", "!", "?", "…")):
                bits[-1] = bits[-1].rstrip() + "."
        elif voix_phrase:
            bits.append(
                f"La chorale que vous avez vue — notamment {voix_phrase} — "
                "en montre la portée concrète."
            )
        else:
            bits.append("La chorale que vous avez vue en montre la portée concrète.")
        if intro_clean:
            bits.append(_ensure_period(intro_clean))
        if q0:
            bits.append(f"Une question pour démarrer : {q0} ?")

    elif family == 1:
        # Ouverture par question.
        bits.append("Bonjour.")
        if q0:
            bits.append(f"{q0} ?")
        else:
            bits.append(
                "À quel moment un résultat de recherche cesse-t-il d'être seulement "
                "scientifique pour devenir un enjeu de transfert ?"
            )
        if titre_oral:
            bits.append(
                f"C'est autour de cette question que je voudrais travailler : {titre_oral}."
            )
        if descriptif:
            bits.append(_ensure_period(descriptif))
        elif intro_clean:
            bits.append(_ensure_period(intro_clean))
        if first_prenom:
            bits.append(
                f"Je m'appuierai sur ce que {first_prenom} "
                f"{'et les autres chercheurs ' if len(prenoms) > 1 else ''}"
                "ont déjà partagé — en rappel, sans revenir à la vidéo."
            )

    elif family == 2:
        # Entrée par une voix, puis focale.
        bits.append("Bonjour.")
        if first_prenom and first_hook:
            bits.append(
                f"Souvenez-vous de {first_prenom} : "
                f"{_ensure_oral_subject(first_hook, guides[0].get('chercheur') or first_prenom)}"
            )
        elif first_prenom:
            bits.append(
                f"Souvenez-vous de ce que {first_prenom} a partagé dans la chorale."
            )
        else:
            bits.append("Souvenez-vous de ce que la chorale a déjà mis en lumière.")
        if titre_oral:
            bits.append(
                f"À partir de là, je voudrais éclairer {titre_oral}."
            )
        if intro_clean:
            bits.append(_ensure_period(intro_clean))
        if q0 and seed % 2 == 0:
            bits.append(f"Gardez cette question en tête : {q0} ?")

    elif family == 3:
        # Focale « autre angle » (surtout utile en sibling).
        bits.append("Bonjour.")
        if sibling_count > 1 and variant_index > 0:
            bits.append(
                "Même séquence témoin, autre focale."
            )
            if titre_oral:
                bits.append(f"Cette fois, le fil est : {titre_oral}.")
        else:
            if titre_oral:
                bits.append(
                    f"Le fil de cette intervention, c'est {titre_oral}."
                )
            bits.append(
                "Je ne vais pas ressasser la chorale : je m'en sers comme point d'appui."
            )
        if descriptif:
            bits.append(_ensure_period(f"Concrètement : {descriptif}"))
        elif intro_clean:
            bits.append(_ensure_period(intro_clean))
        if voix_phrase and seed % 2:
            bits.append(
                f"Les parcours de {voix_phrase} restent en arrière-plan, comme repères."
            )
        if q0:
            bits.append(f"Pour vous situer : {q0} ?")

    else:
        # Entrée conceptuelle / pragmatique.
        bits.append("Bonjour.")
        if titre:
            bits.append(
                f"Quand on parle de « {titre} », on parle rarement d'un formalisme. "
                "On parle d'un moment où un choix change la suite."
            )
        if intro_clean:
            bits.append(_ensure_period(intro_clean))
        elif descriptif:
            bits.append(_ensure_period(descriptif))
        if first_prenom and len(prenoms) >= 2:
            bits.append(
                f"La chorale le montre déjà — de {first_prenom} à {prenoms[-1]} — "
                "chacun à sa manière."
            )
        elif voix_phrase:
            bits.append(
                f"La chorale le montre déjà, notamment avec {voix_phrase}."
            )
        if q0:
            bits.append(f"Je vous propose de garder ceci en ligne de mire : {q0} ?")

    return " ".join(b for b in bits if b and str(b).strip())


def _script_expertise_connector_bank(seed: int) -> list[str]:
    banks = [
        [
            "Partons d'abord de ce que {p} a partagé.",
            "Appuyons-nous ensuite sur ce que {p} a dit.",
            "Retenons aussi ce que {p} a mis en lumière.",
            "Autre situation, celle de {p}.",
            "Et aussi ce que {p} a souligné.",
        ],
        [
            "Premier repère : {p}.",
            "Deuxième repère : {p}.",
            "Troisième angle, avec {p}.",
            "Puis {p}.",
            "Enfin {p}.",
        ],
        [
            "Ce que {p} a raconté peut servir de point de départ.",
            "Chez {p}, le même type de moment prend une autre forme.",
            "{p} ajoute une nuance utile.",
            "Le cas de {p} pousse un cran plus loin.",
            "Et {p} ferme la boucle.",
        ],
        [
            "Je m'appuie d'abord sur {p}.",
            "Je croise ensuite avec {p}.",
            "Je retiens aussi {p}.",
            "Je note le cas de {p}.",
            "Je termine ce tour avec {p}.",
        ],
    ]
    return banks[seed % len(banks)]


def _script_expertise_bridge(seed: int, angle: str) -> str:
    bridges = [
        (
            f"Son propos, sur {_form_angle_oral(angle)}, reste éclairant."
            if angle
            else "Son propos reste éclairant."
        ),
        (
            f"Là, sur {_form_angle_oral(angle)}, quelque chose d'utile apparaît."
            if angle
            else "Là, quelque chose d'utile apparaît."
        ),
        (
            f"Ce passage — {_form_angle_oral(angle)} — mérite qu'on s'y arrête."
            if angle
            else "Ce passage mérite qu'on s'y arrête."
        ),
        (
            f"Autrement dit, autour de {_form_angle_oral(angle)} :"
            if angle
            else "Autrement dit :"
        ),
    ]
    return bridges[seed % len(bridges)]


def _build_script_expertise_projete(
    orientation: dict,
    video: dict | None = None,
    *,
    variant_index: int = 0,
    sibling_count: int = 1,
) -> tuple[str, int]:
    """
    Script oral d'expertise (450–800 mots) : l'expert s'adresse aux apprenants.
    Ton naturel ; video temoin deja vue (rappel, pas reecoute).
    variant_index / sibling_count diversifient le modele quand plusieurs
    videos expertise sont proposees au meme intervenant.
    """
    code = orientation.get("code") or (video or {}).get("code", "")
    titre = orientation.get("titre") or (video or {}).get("titre", "")
    concepts = [c for c in orientation.get("concepts", []) if c]
    introduction = (orientation.get("introduction") or "").strip()
    guides = _orientation_guides(orientation)
    descriptif = ((video or {}).get("descriptif") or "").strip()
    label = _label_video_expert(code)
    seed = _script_variant_seed(code, variant_index)

    prenoms = []
    for guide in guides:
        name = (guide.get("chercheur") or "").strip()
        if name:
            p = _chercheur_prenom(name)
            if p and p not in prenoms:
                prenoms.append(p)
    voix_phrase = _voix_phrase_from_prenoms(prenoms)
    titre_oral = ""
    if titre:
        titre_oral = titre[0].lower() + titre[1:] if titre[:1].isupper() else titre

    paragraphs: list[str] = [
        _script_expertise_open_paragraph(
            titre=titre,
            titre_oral=titre_oral,
            introduction=introduction,
            descriptif=descriptif,
            guides=guides,
            prenoms=prenoms,
            voix_phrase=voix_phrase,
            variant_index=variant_index,
            sibling_count=sibling_count,
            seed=seed,
        )
    ]

    # Paragraphe « enjeu » : parfois court, parfois plus développé, parfois omis
    # si l'ouverture a déjà porté l'objectif (surtout siblings).
    teach_mode = (seed + variant_index) % 3
    if not (sibling_count > 1 and variant_index > 0 and teach_mode == 2):
        teach: list[str] = []
        if teach_mode == 0:
            teach.append("Restons un instant sur l'enjeu.")
        elif teach_mode == 1:
            teach.append("Précisons le terrain.")
        else:
            teach.append("Avant d'entrer dans les cas, un mot de cadrage.")
        if titre and teach_mode != 1:
            teach.append(
                f"Quand on dit « {titre} », on ne parle pas d'un formalisme lointain."
            )
        if descriptif and teach_mode != 0:
            parts = [p.strip() for p in re.split(r"[,·;]", descriptif) if p.strip()]
            if len(parts) >= 2 and all(len(p.split()) <= 4 for p in parts):
                teach.append(
                    f"On parle de situations concrètes : {_list_concepts_oral(parts)}."
                )
            else:
                teach.append(
                    f"On parle de situations concrètes : {descriptif.rstrip('.')}."
                )
        if concepts:
            if teach_mode == 2:
                teach.append(
                    f"Les notions qui aident ici : {_list_concepts_oral(concepts)}."
                )
            else:
                teach.append(
                    f"Quelques notions utiles : {_list_concepts_oral(concepts)}. "
                    f"Derrière ces mots, il y a souvent un geste : anticiper avant d'agir, "
                    f"protéger avant de communiquer, préparer avant de rencontrer."
                )
        if teach_mode == 0:
            teach.append(
                "Les chercheurs déjà entendus ne livrent pas une checklist : "
                "ils rendent visible un moment où une question se pose."
            )
        elif teach_mode == 1:
            teach.append(
                "Je m'appuie sur leurs situations, sans les rejouer scène par scène."
            )
        paragraphs.append(" ".join(teach))

    if guides:
        # Ordre des guides : rotation légère pour siblings / seed (stable).
        ordered_guides = list(guides)
        if len(ordered_guides) > 1 and (sibling_count > 1 or seed % 2):
            rot = (seed + variant_index) % len(ordered_guides)
            ordered_guides = ordered_guides[rot:] + ordered_guides[:rot]

        connector_templates = _script_expertise_connector_bank(seed + variant_index)
        for index, guide in enumerate(ordered_guides):
            chercheur = (guide.get("chercheur") or "un chercheur").strip()
            prenom = _chercheur_prenom(chercheur)
            angle = (guide.get("origine") or guide.get("angle") or "").strip()
            brut_temoin = (guide.get("dans_le_temoin") or "").strip()
            dans_temoin = _strip_chercheur_prefix(brut_temoin, chercheur).strip()
            travail = (guide.get("travail_expert") or "").strip()
            travail = re.sub(r"^E\d+(bis)?\s*/\s*E\d+(bis)?\s*:\s*", "", travail)
            travail = re.sub(r"^E\d+(bis)?\s*:\s*", "", travail)
            if any(m in travail.lower() for m in _CONSIGNE_TECHNIQUE_MARKERS):
                travail = ""
            question = _oralize_editorial_meta(
                (guide.get("question_apprenant") or "").strip()
            )
            erreur = (guide.get("erreur_a_eviter") or "").strip()
            if any(m in erreur.lower() for m in _CONSIGNE_TECHNIQUE_MARKERS):
                erreur = ""
            concepts_g = [
                c
                for c in (guide.get("concepts") or guide.get("concepts_e1") or [])
                if c
            ]

            if index < len(connector_templates):
                block = [connector_templates[index].format(p=prenom)]
            else:
                block = [f"Autre repère, avec {prenom}."]
            # Ne pas coller le même pont « Son propos… » à chaque voix.
            if index == 0 or (seed + index) % 3 != 0:
                block.append(_script_expertise_bridge(seed + index, angle))
            elif angle:
                block.append(f"Angle : {_form_angle_oral(angle)}.")

            if dans_temoin:
                block.append("En effet, " + _ensure_oral_subject(dans_temoin, chercheur))
            elif brut_temoin:
                block.append(_ensure_period(brut_temoin))

            if travail:
                if (seed + index) % 2 == 0:
                    block.append(
                        "Ce que cela ouvre, ce n'est pas seulement l'anecdote : "
                        "c'est une lecture possible du parcours. "
                        + _oralize_instruction(travail)
                    )
                else:
                    block.append(
                        "Pour l'expertise, on peut en tirer ceci. "
                        + _oralize_instruction(travail)
                    )
            elif titre:
                block.append(
                    f"Ce rappel aide à éclairer « {titre} »."
                )

            if concepts_g and (seed + index) % 3 != 1:
                labeled = [_a_plus_article(_fr_label_with_article(c)) for c in concepts_g if c]
                if len(labeled) == 1:
                    touch = labeled[0]
                elif len(labeled) == 2:
                    touch = f"{labeled[0]} et {labeled[1]}"
                else:
                    touch = f"{', '.join(labeled[:-1])} et {labeled[-1]}"
                block.append(f"On touche ici {touch}.")

            if erreur:
                err = erreur.strip()
                low = err.lower()
                if low.startswith("ne pas "):
                    err = "On peut éviter de " + err[7:]
                block.append(_ensure_period(err))

            if question and (index == len(ordered_guides) - 1 or (seed + index) % 2 == 0):
                q = question.strip().rstrip("?").rstrip()
                if (seed + index) % 2 == 0:
                    block.append(
                        "Une question utile, pour faire le lien avec sa pratique : "
                        f"{q} ?"
                    )
                else:
                    block.append(f"À vous : {q} ?")
            paragraphs.append(" ".join(block))
    else:
        no_guide = [
            (
                "Même sans reprendre voix par voix, "
                f"ce que la vidéo témoin a déjà montré ouvre la porte à "
                f"{titre or 'un geste professionnel clé'}."
            ),
            (
                f"Sans dérouler chaque voix, le fil reste {titre_oral or 'ce geste'}. "
                "Chacun pourra y reconnaître un moment équivalent dans son parcours."
            ),
            (
                f"Je vais donc aller droit au geste : {titre_oral or 'l’essentiel à retenir'}."
            ),
        ]
        paragraphs.append(no_guide[seed % len(no_guide)])

    practice_openers = [
        "Venons-en à vous.",
        "Passons à votre situation.",
        "Et maintenant, côté projet.",
        "Comment le faire vivre concrètement ?",
    ]
    practice = [practice_openers[(seed + variant_index) % len(practice_openers)]]
    if titre:
        if (seed + variant_index) % 2 == 0:
            practice.append(
                f"Si demain « {titre} » devenait un enjeu dans votre activité, "
                "par où pourrait-on commencer ?"
            )
        else:
            practice.append(
                f"Supposons que « {titre} » devienne demain un enjeu pour vous : "
                "quelle serait la première décision utile ?"
            )
    practice_actions = [
        (
            "Un point d'entrée possible : nommer la situation à risque, "
            "puis se demander à qui parler, avec quoi arriver, "
            "et ce qui ne doit pas sortir trop tôt."
        ),
        (
            "Essayez ceci : décrire en une phrase le moment critique, "
            "nommer l'interlocuteur pertinent, puis ce qu'il ne faut pas dire trop tôt."
        ),
        (
            "Trois questions suffisent souvent : où est le risque, qui doit être dans la boucle, "
            "quelle information garder encore ?"
        ),
    ]
    practice.append(practice_actions[(seed + variant_index) % len(practice_actions)])
    if concepts and (seed % 2 == 0):
        practice.append(
            f"Quelques repères à garder sous les yeux : {_list_concepts_oral(concepts)}."
        )
    practice_closers = [
        "L'idée n'est pas de recopier les chercheurs : "
        "c'est de reconnaître, le cas échéant, le même type de moment critique.",
        "Inutile de calquer leur parcours : repérer le même type de bascule suffit.",
        "Ce qui compte, c'est le transfert vers votre situation, pas la citation.",
    ]
    practice.append(practice_closers[(seed + variant_index) % len(practice_closers)])
    paragraphs.append(" ".join(practice))

    close_openers = [
        "Pour conclure, je voudrais laisser une chose simple.",
        "Je m'arrête sur un réflexe.",
        "En résumé, une chose à emporter.",
        "Avant de terminer, un point d'appui.",
    ]
    close = [close_openers[(seed + variant_index) % len(close_openers)]]
    if titre_oral:
        close.append(
            "Avant une décision irréversible — publication, communication, "
            f"dépôt, rencontre — un réflexe utile reste : {titre_oral}."
        )
    close_tails = [
        "La chorale en donne le goût. À chacun d'en faire, ou non, une pratique.",
        "Vous avez déjà vu ce moment chez d'autres : à vous de le reconnaître chez vous.",
        "Le reste est affaire de jugement, de timing, et de partenaires.",
        "Gardez le geste, pas le récit.",
    ]
    close.append(close_tails[(seed + variant_index) % len(close_tails)])
    if guides and any(g.get("question_apprenant") for g in guides) and seed % 2 == 0:
        last_q = next(g["question_apprenant"] for g in guides if g.get("question_apprenant"))
        last_q = _oralize_editorial_meta(last_q.strip().rstrip("?").rstrip())
        close.append(f"Une dernière question, pour la route : {last_q} ?")
    close.append("Je vous remercie.")
    paragraphs.append(" ".join(close))

    body = "\n\n".join(p for p in paragraphs if p and p.strip())
    word_count = _count_words_fr(body)

    if word_count < SCRIPT_EXPERTISE_WORD_MIN:
        body += (
            "\n\n"
            "Ajoutons un mot de posture. Mon rôle ici n'est pas de remplacer "
            "votre tutelle, votre service de valorisation ou votre accompagnateur. "
            "Il s'agit plutôt d'aider à reconnaître le moment où une question "
            "peut se poser — avant qu'il ne soit trop tard. "
            "Beaucoup de difficultés naissent moins d'un manque de génie "
            "technique que d'un manque d'anticipation : on communique trop tôt, "
            "on arrive sous-préparé, on confond partage scientifique et divulgation. "
            "Les témoins que vous avez déjà entendus rendent ce moment visible. "
            "Chacun pourra le retrouver, éventuellement, dans son laboratoire, "
            "sa thèse, son partenariat ou sa prochaine prise de parole. "
            "Une question simple peut aider : qu'est-ce qui, demain, "
            "pourrait affaiblir la nouveauté, la confidentialité ou la valeur "
            "du résultat si l'on parle trop vite ? "
            "Et à l'inverse : qu'est-ce qui empêcherait d'avancer si l'on reste "
            "trop longtemps seul avec son idée ? "
            "Tenir ces deux exigences ensemble — protéger et avancer — "
            "est souvent le vrai geste d'expertise."
        )
        word_count = _count_words_fr(body)

    if word_count < SCRIPT_EXPERTISE_WORD_MIN:
        body += (
            "\n\n"
            f"Pour ancrer « {titre or label} », une piste possible. "
            "Sur une feuille, trois colonnes : situation à risque, "
            "interlocuteur à contacter, information à préparer. "
            "Les remplir à partir de son projet — pas d'un cas idéal — "
            "puis comparer avec ce que la vidéo témoin a déjà mis en lumière. "
            "Où se situe la différence ? "
            "Souvent, elle révèle une hésitation, un manque d'anticipation, "
            "ou une confusion entre communiquer et divulguer. "
            "Ces écarts sont précieux : ils indiquent où porter l'effort. "
            "Enfin, formuler à voix haute, en une phrase, le réflexe retenu. "
            "S'il est difficile à dire simplement, le transfert n'est peut-être pas encore fait."
        )
        word_count = _count_words_fr(body)

    if word_count > SCRIPT_EXPERTISE_WORD_MAX:
        target = SCRIPT_EXPERTISE_WORD_MAX - 20
        tokens = re.findall(
            r"[A-Za-zÀ-ÖØ-öø-ÿ0-9']+|[^\w\s]+|\s+",
            body,
            flags=re.UNICODE,
        )
        kept = []
        count = 0
        for tok in tokens:
            if re.match(r"[A-Za-zÀ-ÖØ-öø-ÿ0-9']+", tok):
                if count >= target:
                    break
                count += 1
            kept.append(tok)
        body = "".join(kept).rstrip(" ,;:\n")
        if not body.endswith((".", "!", "?", "…")):
            body += "."
        body += "\n\nJe vous remercie."
        word_count = _count_words_fr(body)
        if word_count > SCRIPT_EXPERTISE_WORD_MAX:
            pattern = re.compile(r"[A-Za-zÀ-ÖØ-öø-ÿ0-9']+")
            out = []
            n = 0
            last = 0
            for match in pattern.finditer(body):
                if n >= SCRIPT_EXPERTISE_WORD_MAX:
                    break
                out.append(body[last:match.end()])
                last = match.end()
                n += 1
            body = "".join(out).rstrip(" ,;:") + "."
            word_count = _count_words_fr(body)

    body = _normalize_editorial_french(body)
    word_count = _count_words_fr(body)
    header = (
        f"[PROJECTION — {label} — script oral expert → apprenants]\n"
        f"Objectif pédagogique : {_normalize_editorial_french(titre) or '—'}\n"
        f"Cible : {SCRIPT_EXPERTISE_WORD_MIN}–{SCRIPT_EXPERTISE_WORD_MAX} mots "
        f"(projection actuelle : {word_count} mots)\n"
    )
    return f"{header}\n{body}".strip(), word_count



def _build_script_expertise_plan(
    orientation: dict,
    video: dict | None = None,
    *,
    variant_index: int = 0,
) -> list[str]:
    """Plan en puces du script expertise : etapes + notions traitees."""
    titre = orientation.get("titre") or (video or {}).get("titre", "")
    concepts = [c for c in orientation.get("concepts", []) if c]
    guides = _orientation_guides(orientation)
    descriptif = ((video or {}).get("descriptif") or "").strip()
    code = orientation.get("code") or (video or {}).get("code", "")
    seed = _script_variant_seed(code, variant_index)

    plan: list[str] = []
    open_labels = [
        f"Ouverture — objectif : {titre}" if titre else "Ouverture — ancrage sur la vidéo témoin déjà vue",
        f"Ouverture — question de terrain puis {titre}" if titre else "Ouverture — question de terrain",
        f"Ouverture — rappel d'une voix puis {titre}" if titre else "Ouverture — rappel d'une voix",
        f"Ouverture — autre focale : {titre}" if titre else "Ouverture — autre focale",
        f"Ouverture — cadrage de « {titre} »" if titre else "Ouverture — cadrage",
    ]
    plan.append(open_labels[seed % len(open_labels)])

    if concepts:
        plan.append(f"Notions cadres : {_list_concepts_oral(concepts)}")
    elif descriptif:
        parts = [p.strip() for p in re.split(r"[,·;]", descriptif) if p.strip()]
        if parts:
            plan.append(f"Notions cadres : {_list_concepts_oral(parts)}")

    if guides:
        ordered = list(guides)
        if len(ordered) > 1 and seed % 2:
            rot = seed % len(ordered)
            ordered = ordered[rot:] + ordered[:rot]
        for guide in ordered:
            chercheur = (guide.get("chercheur") or "Temoin").strip()
            prenom = _chercheur_prenom(chercheur) or chercheur
            angle = (guide.get("origine") or guide.get("angle") or "").strip()
            notions = [
                c
                for c in (guide.get("concepts") or guide.get("concepts_e1") or [])
                if c
            ]
            if angle and notions:
                plan.append(
                    f"Rappel — {prenom} ({angle}) — notions : "
                    f"{_list_concepts_oral(notions)}"
                )
            elif angle:
                plan.append(
                    f"Rappel — {prenom} — angle : {_form_angle_oral(angle)}"
                )
            elif notions:
                plan.append(
                    f"Rappel — {prenom} — notions : {_list_concepts_oral(notions)}"
                )
            else:
                plan.append(f"Rappel — {prenom} — illustration du geste")
    else:
        plan.append("Développement — rappel de la vidéo témoin déjà vue")

    plan.append("Mise en pratique — transfert vers la situation de l'apprenant")
    if titre:
        titre_oral = titre[0].lower() + titre[1:] if titre[:1].isupper() else titre
        plan.append(f"Cloture — installer le reflexe : {titre_oral}")
    else:
        plan.append("Clôture — retenir le geste à installer")
    return [_normalize_editorial_french(item) for item in plan]


def scripts_expertise_projetes_section(capsule_data: dict) -> str:
    videos = capsule_data.get("videos_expert", [])
    orientations = capsule_data.get("orientations_expert") or []
    if capsule_data.get("orientation_expert") and not orientations:
        orientations = [capsule_data["orientation_expert"]]
    if not videos and not orientations:
        return ""

    by_code = {item.get("code", ""): item for item in videos if item.get("code")}
    blocks: list[tuple[dict, dict | None]] = []
    seen: set[str] = set()
    for orientation in orientations:
        code = orientation.get("code", "")
        blocks.append((orientation, by_code.get(code)))
        if code:
            seen.add(code)
    for video in videos:
        code = video.get("code", "")
        if code and code not in seen:
            blocks.append(
                (
                    {
                        "code": code,
                        "titre": video.get("titre", ""),
                        "concepts": [],
                        "introduction": video.get("descriptif", ""),
                        "consignes": list(BRIEF_CONSIGNES_COMMUNES),
                        "utilisation_script_temoin": {},
                    },
                    video,
                )
            )

    if not blocks:
        return ""

    articles = []
    sibling_count = len(blocks)
    for variant_index, (orientation, video) in enumerate(blocks):
        code = orientation.get("code") or (video or {}).get("code", "")
        titre = orientation.get("titre") or (video or {}).get("titre", "")
        script, word_count = _build_script_expertise_projete(
            orientation,
            video,
            variant_index=variant_index,
            sibling_count=sibling_count,
        )
        plan = _build_script_expertise_plan(
            orientation, video, variant_index=variant_index
        )
        plan_html = "".join(f"<li>{escape(item)}</li>" for item in plan)
        in_range = SCRIPT_EXPERTISE_WORD_MIN <= word_count <= SCRIPT_EXPERTISE_WORD_MAX
        count_class = "meta" if in_range else "warn"
        articles.append(
            f"""
  <article class="script-expertise-block">
    <h3>{escape(_label_video_expert(code))} — {escape(_normalize_editorial_french(titre or 'Script projeté'))}</h3>
    <p class="meta">Script oral d'expertise adressé aux apprenants. La vidéo témoin sert de <strong>prétexte pédagogique</strong> pour enseigner l'objectif (ex. réflexe de déclaration avant divulgation).</p>
    <p class="{count_class}"><strong>Volume :</strong> {word_count} mots (cible {SCRIPT_EXPERTISE_WORD_MIN}–{SCRIPT_EXPERTISE_WORD_MAX}).</p>
    <div class="script-expertise-plan">
      <h4>Plan du script (notions traitées)</h4>
      <ul>
        {plan_html}
      </ul>
    </div>
    {_script_lines_html(script)}
  </article>
"""
        )

    return f"""
<section class="methodology-panel script-expertise-projete-panel">
  <h2>Proposition de script pour les vidéos expertise</h2>
  <p class="script-expertise-disclaimer"><strong>Attention :</strong> <strong>{escape(SCRIPT_EXPERTISE_DISCLAIMER)}</strong></p>
  <p class="meta">Chaque texte est une prise de parole d'expert (environ 450 à 800 mots) qui s'appuie sur la chorale témoin pour enseigner l'objectif pédagogique de la vidéo expertise — sans se limiter à un résumé des témoignages.</p>
  {''.join(articles)}
</section>
"""


def export_unites_plaintext(capsule_data: dict) -> str:
    unites = capsule_data.get("unites_de_sens", [])
    if not unites:
        return ""
    lines = ["UNITES DE SENS", ""]
    for unite in unites:
        extraits = ", ".join(unite.get("extraits", []))
        grille = unite.get("grille_expert") or unite.get("grille_e1") or unite.get("grille_e20_e21") or ""
        lines.append(f"{unite.get('ordre', '')}. {extraits or '—'}")
        if unite.get("acte"):
            lines.append(f"   Acte : {unite['acte']}")
        if unite.get("libelle"):
            lines.append(f"   Unite de sens : {unite['libelle']}")
        if grille:
            lines.append(f"   Grille expert : {grille}")
        lines.append("")
    return "\n".join(lines).strip()


def export_videos_expert_plaintext(capsule_data: dict) -> str:
    videos = capsule_data.get("videos_expert", [])
    if not videos:
        return ""
    orientations = {
        item.get("code", ""): item
        for item in (capsule_data.get("orientations_expert") or [])
        if item.get("code")
    }
    lines = ["VIDEOS EXPERT A PRODUIRE", ""]
    for video in videos:
        code = video.get("code", "")
        intervenant = video.get("intervenant")
        if intervenant:
            lines.append(f"{_label_video_expert(code)} — {intervenant}")
        else:
            lines.append(f"{_label_video_expert(code)}")
        lines.append(f"   Objectif : {video.get('titre', '')}")
        if video.get("descriptif"):
            lines.append(f"   Descriptif : {video['descriptif']}")
        orientation = orientations.get(code)
        if orientation:
            if orientation.get("introduction"):
                lines.append(f"   Orientation : {orientation['introduction']}")
            consignes = orientation.get("consignes", [])
            if consignes:
                lines.append("   Consignes :")
                for item in consignes:
                    lines.append(f"     - {item}")
        lines.append("")
    return _normalize_editorial_french("\n".join(lines).strip())


def _multiline_cell_html(text: str) -> str:
    if not text:
        return ""
    return escape(text).replace("\n", "<br>")


def _export_highlight_code(capsule_code: str) -> str:
    if capsule_code == "GEN":
        return "T1"
    return capsule_code


def export_programme_complet_table_html(
    current_capsule_code: str, programme_table: dict
) -> str:
    rows_data = programme_table.get("rows", [])
    if not rows_data:
        return ""

    headers = programme_table.get("headers", {})
    highlight_code = _export_highlight_code(current_capsule_code)
    html_rows: list[str] = []
    index = 0

    while index < len(rows_data):
        module = rows_data[index]["module"]
        end = index
        while end < len(rows_data) and rows_data[end]["module"] == module:
            end += 1
        module_count = end - index

        for offset in range(index, end):
            row = rows_data[offset]
            highlight = row["code"] == highlight_code
            row_bg = "background:#dff4f6;" if highlight else ""
            cells: list[str] = []
            if offset == index:
                cells.append(
                    f'<td rowspan="{module_count}" style="{row_bg}vertical-align:top;">'
                    f'{_multiline_cell_html(row["module"])}</td>'
                )
            for field in PROGRAMME_TABLE_FIELDS[1:]:
                cells.append(
                    f'<td style="{row_bg}vertical-align:top;">'
                    f"{_multiline_cell_html(row[field])}</td>"
                )
            html_rows.append("<tr>" + "".join(cells) + "</tr>")
        index = end

    header_cells = "".join(
        f"<th style='background:#f3f3f3;'>{escape(headers.get(field, field))}</th>"
        for field in PROGRAMME_TABLE_FIELDS
    )
    source = programme_table.get("source_document", "tableau de conception")
    date_maj = programme_table.get("date_mise_a_jour", "")
    date_label = f" — mis a jour le {date_maj}" if date_maj else ""
    note = programme_table.get("note", "")
    note_html = (
        f'<p style="font-size:10pt;color:#555;margin:12pt 0 0;">{escape(note)}</p>'
        if note
        else ""
    )
    current_label = _label_video_temoin(current_capsule_code)

    return (
        '<p style="font-size:10pt;color:#555;margin:0 0 10pt;">'
        "Tableau de conception du MOOC (extrait tel quel du fichier source) pour situer "
        "votre intervention dans l'ensemble du parcours et rester aligne sur le grain "
        "prevu. La ligne en surbrillance correspond a "
        f"{escape(current_label)}.</p>"
        f'<p style="font-size:10pt;color:#555;margin:0 0 10pt;">'
        f"Source : {escape(source)}{escape(date_label)}.</p>"
        '<table border="1" cellpadding="6" cellspacing="0" '
        'style="width:100%;border-collapse:collapse;font-size:9pt;">'
        f"<thead><tr>{header_cells}</tr></thead><tbody>"
        + "".join(html_rows)
        + "</tbody></table>"
        + note_html
    )


def _export_hidden_html_block(block_id: str, section_title: str, html: str) -> str:
    return (
        f'<div class="sr-export" id="{block_id}" '
        f'data-section-title="{escape(section_title)}" hidden>{html}</div>'
    )


def _export_hidden_block(block_id: str, section_title: str, body: str) -> str:
    return (
        f'<div class="sr-export" id="{block_id}" '
        f'data-section-title="{escape(section_title)}" hidden>{escape(body)}</div>'
    )


def export_word_section(
    code: str,
    titre: str,
    capsule_data: dict,
    by_id: dict[str, dict],
    programme_table: dict,
) -> str:
    brief_text = export_brief_intervenant_plaintext(code, capsule_data, by_id)
    synthese_text = export_synthese_temoignages_plaintext(code, capsule_data, by_id)
    videos_table_html = export_programme_complet_table_html(code, programme_table)
    hidden_blocks = ""
    if synthese_text:
        hidden_blocks += _export_hidden_block(
            "export-synthese-temoignages",
            export_synthese_section_title(code),
            synthese_text,
        )
    if brief_text:
        hidden_blocks += _export_hidden_block(
            "export-brief-intervenant",
            EXPORT_BRIEF_SECTION_TITLE,
            brief_text,
        )
    if videos_table_html:
        hidden_blocks += _export_hidden_html_block(
            "export-videos-table",
            EXPORT_VIDEOS_TABLE_TITLE,
            videos_table_html,
        )
    return (
        f"""
<section class="export-panel" id="export-word-panel" data-capsule-code="{escape(code)}" data-capsule-title="{escape(titre)}">
  <h2>Export Word</h2>
  <p class="meta">Exporter le script final, la synthèse des témoignages, la proposition de cadrage pour la vidéo expert et le tableau de conception (source Excel).</p>
  <button type="button" class="btn" id="export-word-open" aria-expanded="false" aria-controls="export-word-modal">
    Exporter le dossier capsule
  </button>
  {hidden_blocks}
</section>
"""
        + export_word_modal(f"capsule_{code.lower()}.doc")
    )


def link_segment(segment: dict) -> str:
    return (
        f"<strong>{escape(segment['id'])}</strong> "
        f"<span class='meta'>{escape(segment['debut'])} - {escape(segment['fin'])}</span><br>"
        f"{escape(segment['verbatim'])}"
    )


def status_badge(value: str) -> str:
    css = {
        "EN_CONSTRUCTION": "status--progress",
        "VALIDEE": "status--ok",
        "VERROUILLEE": "status--ok",
        "A_ARBITRER": "status--warn",
        "A_CARTOGRAPHIER": "status--warn",
    }.get(value, "")
    return f'<span class="status {css}">{escape(value)}</span>'


def build_home(capsules: list[dict], segments: list[dict]) -> None:
    used_count = sum(1 for segment in segments if segment.get("statut") == "UTILISE")
    sections = [
        (
            "tableau_de_bord.html",
            "⊞",
            "Tableau de bord",
            "Vue d'ensemble des capsules, durees, chercheurs et acces aux montages.",
        ),
        (
            "suivi_intervenants.html",
            "◎",
            "Suivi Intervenants",
            "Structure principale : modules, intervenants, capsules temoin et videos expert.",
        ),
        (
            "edito.html",
            "✎",
            "Edito",
            "Proposition edito, correspondances et derushage des selections Clarisse.",
        ),
        (
            "fichiers_travail.html",
            "📁",
            "Fichiers de travail",
            "Informations, Prev Vid, videos expert, mails et BAB encodes.",
        ),
    ]
    cards = "".join(
        f"<a class='sommaire-card' href='{escape(href)}'>"
        f"<span class='sommaire-card__icon' aria-hidden='true'>{escape(icon)}</span>"
        f"<h2>{escape(title)}</h2>"
        f"<p>{escape(description)}</p>"
        f"<span class='sommaire-card__cta'>Ouvrir →</span>"
        f"</a>"
        for href, icon, title, description in sections
    )
    body = f"""
<section class="hero">
  <p class="hero__eyebrow">MOOC · L'Esprit d'innover ! Pourquoi pas Vous !</p>
  <h1>Dérushage éditorial chorale</h1>
  <p class="hero__lead">Cartographier, qualifier et assembler les temoignages BAB pour les videos chorales du MOOC.</p>
  <div class="hero__stats">
    <span class="stat-pill"><strong>{len(capsules)}</strong> capsules</span>
    <span class="stat-pill"><strong>{len(segments)}</strong> extraits indexes</span>
    <span class="stat-pill"><strong>{used_count}</strong> utilises dans les montages</span>
  </div>
</section>
<h2 class="section-title">Sommaire</h2>
<nav class="sommaire-grid" aria-label="Sommaire">
{cards}
</nav>
"""
    write_text(
        SITE / "index.html",
        html_page(
            "Accueil",
            body,
            nav_current="index.html",
            page_header="",
            main_class="page-home",
        ),
    )


def _normalize_for_match(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text or "")
    ascii_only = "".join(char for char in normalized if not unicodedata.combining(char))
    return ascii_only.lower()


def _extract_intervenants(raw: str) -> list[str]:
    names = []
    for part in re.split(r"[\n/]+", raw or ""):
        cleaned = " ".join(part.strip().split())
        if cleaned:
            names.append(cleaned)
    return names


def _canonical_name_key(name: str) -> str:
    base = _normalize_for_match(name)
    return re.sub(r"[^a-z0-9]+", " ", base).strip()


EXPERT_NAME_ALIASES = {
    "antoine la treille": "Antoine Latreille",
    "antoine latreille": "Antoine Latreille",
    "soizic lefreuvre": "Soizic Lefeuvre",
    "soizic lefeuvre": "Soizic Lefeuvre",
    "yoan montenot": "Yoann Montenot",
    "yoann montenot": "Yoann Montenot",
    "stephanie oger roussel": "Stephanie Oger-Roussel",
    "stephanie oger-roussel": "Stephanie Oger-Roussel",
    "arielle sante": "Arielle Santé",
    "joel nguen": "Joël Nguen",
    "remi wache": "Rémi Waché",
    "gregoire burge": "Gregoire Burgé",
}


def _build_canonical_labels(names: list[str]) -> tuple[list[str], dict[str, str]]:
    by_key: dict[str, str] = {}
    for name in names:
        key = _canonical_name_key(name)
        if not key:
            continue
        alias = EXPERT_NAME_ALIASES.get(key)
        if alias:
            by_key[key] = alias
            continue
        current = by_key.get(key)
        if current is None or len(name) > len(current):
            by_key[key] = name
    labels = sorted(set(by_key.values()))
    return labels, {key: value for key, value in by_key.items()}


def _extract_temoin_names(text: str) -> list[str]:
    names = []
    for match in re.finditer(r"([A-Za-zÀ-ÿ][A-Za-zÀ-ÿ\-']*)\s*:", text or ""):
        names.append(match.group(1).strip())
    return names


def _heat_cell_style(value: int, max_value: int) -> str:
    if value <= 0 or max_value <= 0:
        return "background:#ffffff;color:#94a3b8;"
    ratio = value / max_value
    alpha = 0.15 + 0.7 * ratio
    text = "#ffffff" if alpha >= 0.55 else "#0f172a"
    return f"background:rgba(11,110,119,{alpha:.2f});color:{text};"


def _render_heatmap_table(
    title: str,
    subtitle: str,
    row_labels: list[str],
    col_labels: list[str],
    matrix: list[list[int]],
) -> str:
    max_value = max((max(row) if row else 0 for row in matrix), default=0)
    header_cells = "".join(f"<th>{escape(label)}</th>" for label in col_labels)
    body_rows = []
    for row_label, values in zip(row_labels, matrix):
        value_cells = "".join(
            f"<td class='heatmap-cell' style='{_heat_cell_style(value, max_value)}'>{value or ''}</td>"
            for value in values
        )
        body_rows.append(
            "<tr>"
            f"<th class='heatmap-row-label'>{escape(row_label)}</th>"
            f"{value_cells}"
            "</tr>"
        )
    return (
        "<article class='heatmap-card'>"
        f"<h2>{escape(title)}</h2>"
        f"<p class='meta'>{escape(subtitle)}</p>"
        "<div class='heatmap-wrap'>"
        "<table class='heatmap-table'>"
        "<thead><tr><th>Thématique</th>"
        f"{header_cells}</tr></thead>"
        f"<tbody>{''.join(body_rows)}</tbody>"
        "</table></div>"
        "<p class='heatmap-legend'>"
        "<span>Faible</span>"
        "<span class='heatmap-scale' aria-hidden='true'></span>"
        "<span>Forte</span>"
        "</p>"
        "</article>"
    )


def build_heatmaps_page(capsules: list[dict], segments: list[dict]) -> None:
    _ = segments
    programme_table = load_programme_table()
    rows = programme_table.get("rows", [])
    rows_by_code = {row.get("code", ""): row for row in rows}
    thematic_rows = [
        capsule
        for capsule in sorted(capsules, key=lambda item: item.get("ordre", 0))
        if capsule.get("code", "").startswith("T")
    ]
    capsule_labels = [f"{capsule['code']} — {capsule['titre']}" for capsule in thematic_rows]

    raw_intervenants = [
        name for row in rows for name in _extract_intervenants(row.get("noms_proposes", ""))
    ]
    intervenants, intervenant_by_key = _build_canonical_labels(raw_intervenants)
    intervenant_col = {label: idx for idx, label in enumerate(intervenants)}
    capsule_matrix = [[0 for _ in intervenants] for _ in thematic_rows]

    for row_idx, capsule in enumerate(thematic_rows):
        row = rows_by_code.get(capsule["code"], {})
        for raw_name in _extract_intervenants(row.get("noms_proposes", "")):
            key = _canonical_name_key(raw_name)
            label = intervenant_by_key.get(key)
            if label is None:
                continue
            capsule_matrix[row_idx][intervenant_col[label]] += 1

    subject_keywords = [
        ("Origines innovation", ("innovation", "origine", "eureka", "serendipite")),
        ("Besoin marche usage", ("besoin", "marche", "utilisateur", "usage", "pivot")),
        ("Preuve et maturation", ("poc", "prototype", "trl", "prematuration", "maturation")),
        ("Protection et PI", ("brevet", "protection", "confidentialite", "propriete intellectuelle", "pi")),
        ("Transfert valorisation", ("licence", "start-up", "valorisation", "transfert", "partenariat")),
        ("Financement", ("financement", "investisseur", "dilution", "bpifrance", "levee")),
        ("Equipe gouvernance", ("equipe", "ceo", "cso", "cto", "gouvernance", "fondateur")),
        ("Pitch et communication", ("pitch", "langage", "valeur", "interlocuteur", "negociation")),
        ("Freins et leviers", ("freins", "legitimite", "incertitude", "echec", "temps")),
        ("Collaboration", ("collaboration", "co-construction", "contrat", "partage de valeur")),
    ]
    subject_labels = [label for label, _ in subject_keywords]
    subject_matrix = [[0 for _ in intervenants] for _ in subject_labels]
    for row in rows:
        text = _normalize_for_match(
            " ".join(
                [
                    row.get("video_temoin", ""),
                    row.get("videos_referent", ""),
                    row.get("objectif_pedagogique", ""),
                ]
            )
        )
        matched = [
            index
            for index, (_, keywords) in enumerate(subject_keywords)
            if any(keyword in text for keyword in keywords)
        ]
        if not matched:
            continue
        for raw_name in _extract_intervenants(row.get("noms_proposes", "")):
            key = _canonical_name_key(raw_name)
            label = intervenant_by_key.get(key)
            if label is None:
                continue
            col_idx = intervenant_col[label]
            for row_index in matched:
                subject_matrix[row_index][col_idx] += 1

    source = programme_table.get("source_document", "")
    date_maj = programme_table.get("date_mise_a_jour", "")
    date_label = f" · mise a jour {date_maj}" if date_maj else ""
    source_line = f"Source : {source}{date_label}." if source else ""

    body = (
        "<div class='page-head'>"
        "<h1>Cartes de chaleur — sujets et intervenants</h1>"
        "<p class='lead'>Vue transversale des liens entre sujets du programme et intervenants proposes.</p>"
        "</div>"
        "<p class='meta'>"
        "Chaque case represente une presence (capsule × intervenant) ou une intensite de recouvrement "
        "(sujet-cle × intervenant). "
        f"{escape(source_line)}"
        "</p>"
        "<section class='heatmap-grid'>"
        + _render_heatmap_table(
            "Sujets-cles × intervenants proposes",
            "Intensite = nombre de capsules ou l'intervenant est associe a un sujet-cle.",
            subject_labels,
            intervenants,
            subject_matrix,
        )
        + _render_heatmap_table(
            "Thématiques du MOOC × intervenants proposes",
            "Presence binaire (1 = intervenant propose sur la capsule).",
            capsule_labels,
            intervenants,
            capsule_matrix,
        )
        + "</section>"
    )
    write_text(
        SITE / "cartes_chaleur.html",
        html_page(
            "Cartes de chaleur",
            body,
            nav_current="cartes_chaleur.html",
            breadcrumb=html_breadcrumb(("Accueil", "index.html"), ("Cartes de chaleur", None)),
        ),
    )


def build_experts_profiles_page(profiles_data: dict) -> None:
    profiles = profiles_data.get("profils", [])

    def status_label(value: str) -> str:
        labels = {
            "confirme": "Profil confirme",
            "probable": "Profil probable",
            "a_verifier": "Profil a verifier",
        }
        return labels.get(value, value)

    cards = []
    for profile in profiles:
        infos = "".join(f"<li>{escape(item)}</li>" for item in profile.get("infos", []))
        mots_cles = "".join(f"<li>{escape(item)}</li>" for item in profile.get("mots_cles", []))
        sources = "".join(
            "<li>"
            f"<a href='{escape(src.get('url', '#'))}' target='_blank' rel='noopener noreferrer'>"
            f"{escape(src.get('label', src.get('url', 'source')))}</a>"
            f" <span class='meta'>({escape(src.get('type', 'source'))})</span>"
            "</li>"
            for src in profile.get("sources", [])
        )
        cards.append(
            "<article class='card'>"
            f"<h2>{escape(profile.get('nom', 'Profil'))}</h2>"
            f"<p class='meta'><strong>Statut :</strong> {escape(status_label(profile.get('statut', 'a_verifier')))}</p>"
            f"<p><strong>Profil cible :</strong> {escape(profile.get('profil_cible', '—'))}</p>"
            "<h3>Mots-cles</h3>"
            f"<ul>{mots_cles or '<li>Aucun mot-cle renseigne.</li>'}</ul>"
            "<h3>Informations recueillies</h3>"
            f"<ul>{infos or '<li>Aucune information validee pour le moment.</li>'}</ul>"
            "<h3>Sources</h3>"
            f"<ul>{sources or '<li>Aucune source.</li>'}</ul>"
            "</article>"
        )

    body = (
        "<div class='page-head'>"
        "<h1>Profils experts — auto-research</h1>"
        "<p class='lead'>Fiches de travail par profil avec traces de collecte et niveau de confiance.</p>"
        "</div>"
        "<p class='meta'>"
        "Les profils marques « a verifier » demandent une validation humaine avant usage editorial."
        "</p>"
        + "".join(cards)
    )
    write_text(
        SITE / "profils_experts.html",
        html_page(
            "Profils experts",
            body,
            nav_current="fichiers_travail.html",
            breadcrumb=html_breadcrumb(("Accueil", "index.html"), ("Fichiers de travail", "fichiers_travail.html"), ("Informations", "informations.html"), ("Profils experts", None)),
        ),
    )


def build_dashboard(capsules: list[dict], segments: list[dict], affectations: dict) -> None:
    by_id = index_by_id(segments)
    rows = []
    for capsule in capsules:
        code = capsule["code"]
        capsule_data = affectations["capsules"].get(code, {})
        used = capsule_data.get("extraits_utilises", [])
        researchers = sorted({by_id[item]["chercheur"] for item in used if item in by_id})
        current = capsule_duration(code, by_id, affectations)
        bab_total = capsule_bab_duration(code, by_id, affectations)
        alert = ""
        if current > capsule["duree_cible_secondes"] * 1.2:
            alert = "<span class='warn'>Duree excessive</span>"
        duration_label = format_seconds(current)
        if bab_total and abs(bab_total - current) > 1:
            duration_label += f" <span class='meta'>(BAB brut {format_seconds(bab_total)})</span>"
        rows.append(
            "<tr>"
            f"<td><a href='capsule_{escape(code)}.html'>{escape(code)}</a></td>"
            f"<td>{escape(capsule['titre'])}</td>"
            f"<td>{status_badge(capsule['statut'])}</td>"
            f"<td>{duration_label} / {format_seconds(capsule['duree_cible_secondes'])}</td>"
            f"<td>{escape(', '.join(researchers) or '-')}</td>"
            f"<td>{len(capsule_data.get('extraits_candidats', []))}</td>"
            f"<td>{alert or '-'}</td>"
            "</tr>"
        )
    researcher_counts = Counter(segment["chercheur"] for segment in segments if segment["statut"] == "UTILISE")
    used_count = sum(researcher_counts.values())
    temoin_capsules = [c for c in capsules if str(c.get("code", "")).startswith("T")]
    en_construction = sum(1 for capsule in temoin_capsules if capsule.get("statut") == "EN_CONSTRUCTION")
    body = f"""
<section class="stats-grid">
  <div class="stat-card">
    <div class="stat-card__label">Vidéos témoins</div>
    <div class="stat-card__value">{len(temoin_capsules)}</div>
    <div class="stat-card__meta">{en_construction} en construction · + GEN laboratoire</div>
  </div>
  <div class="stat-card">
    <div class="stat-card__label">Extraits</div>
    <div class="stat-card__value">{len(segments)}</div>
    <div class="stat-card__meta">segments indexes dans le registre</div>
  </div>
  <div class="stat-card">
    <div class="stat-card__label">Chercheurs</div>
    <div class="stat-card__value">{len(researcher_counts)}</div>
    <div class="stat-card__meta">{used_count} extraits utilises au total</div>
  </div>
</section>
<h2>Capsules</h2>
<div class="table-wrap">
<table>
<thead><tr><th>Code</th><th>Titre</th><th>Statut</th><th>Duree</th><th>Chercheurs utilises</th><th>Candidats</th><th>Alertes</th></tr></thead>
<tbody>
""" + "\n".join(rows) + """
</tbody>
</table>
</div>
<h2>Chercheurs</h2>
<div class="chip-grid">
""" + "".join(
        f"<a class='chip' href='chercheur_{slug(name)}.html'>{escape(name)}</a>"
        for name in sorted({segment["chercheur"] for segment in segments})
    ) + """
</div>
"""
    write_text(
        SITE / "tableau_de_bord.html",
        html_page(
            "Tableau de bord",
            body,
            nav_current="tableau_de_bord.html",
            breadcrumb=html_breadcrumb(("Accueil", "index.html"), ("Tableau de bord", None)),
            page_header='<div class="page-head"><h1>Tableau de bord</h1><p class="lead">Suivi des capsules, montages provisoires et equilibre des voix chercheurs.</p></div>',
        ),
    )


def _tb_edito_parse_videos_expert(raw: str) -> list[dict]:
    videos: list[dict] = []
    text = (raw or "").replace("\r", "").strip()
    if not text:
        return videos

    chunks = re.split(r"(?=(?:^|\n)\s*E\d+(?:\s*bis)?\s*[—–-])", text, flags=re.IGNORECASE)
    for chunk in chunks:
        candidate = chunk.strip()
        if not candidate:
            continue
        match = re.match(r"^\s*(E\d+(?:\s*bis)?)\s*[—–-]\s*(.+)$", candidate, flags=re.IGNORECASE | re.DOTALL)
        if not match:
            continue
        code = re.sub(r"\s+", "", match.group(1))
        desc = " ".join(match.group(2).split())
        videos.append(
            {
                "code": code,
                "intervenant": "",
                "titre": desc,
                "descriptif": "",
            }
        )
    return videos


def _tb_edito_resume_temoignages(sequences: list[dict]) -> str:
    by_voice: dict[str, list[str]] = defaultdict(list)
    for sequence in sequences:
        voice = sequence.get("intervenant", "").strip() or "Temoin"
        text = _strip_chercheur_prefix((sequence.get("texte") or "").strip(), voice)
        if text and text not in by_voice[voice]:
            by_voice[voice].append(text)
    lines: list[str] = []
    for voice in sorted(by_voice):
        snippets = by_voice[voice][:2]
        if snippets:
            lines.append(f"{voice}: {' / '.join(snippets)}")
    return "\n".join(lines)


def _tb_edito_sequences_by_code() -> dict[str, list[dict]]:
    grouped: dict[str, list[dict]] = defaultdict(list)
    for item in load_derushage_edito_index():
        doc = load_derushage_edito(item.get("id", ""))
        if not doc:
            continue
        for sequence in doc.get("sequences", []):
            video_title = (sequence.get("video") or "").strip()
            if not video_title:
                continue
            targets = _target_codes_from_edito_title(video_title)
            for code in targets:
                seq = {
                    **sequence,
                    "intervenant": doc.get("intervenant", item.get("intervenant", "")),
                    "source_doc": doc.get("source", item.get("source", "")),
                    "video": video_title,
                }
                # Correction ASR Greffet dans les couches edito / guides (pas dans data/raw/).
                seq["texte"] = correct_asr_greffet(seq.get("texte", ""))
                grouped[code].append(seq)
    return grouped


TEMOIN_VOICE_ORDER = (
    "Jean-Jacques Greffet",
    "Muriel Thomas",
    "Sylvia Cohen-Kaminski",
    "Loïc Rajjou",
    "Yann Monier",
)


def _tb_edito_voice_sort_key(name: str) -> tuple[int, str]:
    canonical = _canonical_name_key(name)
    for index, preferred in enumerate(TEMOIN_VOICE_ORDER):
        if _canonical_name_key(preferred) == canonical:
            return (index, canonical)
    return (100, canonical)


def _tb_edito_by_chercheur_order(sequences: list[dict]) -> list[dict]:
    """Regroupe les sequences chercheur apres chercheur (ordre documentaire dans chaque voix)."""
    if not sequences:
        return []

    by_voice: dict[str, list[dict]] = defaultdict(list)
    display_name: dict[str, str] = {}
    for sequence in sequences:
        voice = (sequence.get("intervenant") or "Temoin").strip()
        key = _canonical_name_key(voice) or "temoin"
        by_voice[key].append(sequence)
        display_name[key] = voice

    ordered: list[dict] = []
    for key in sorted(by_voice.keys(), key=lambda item: _tb_edito_voice_sort_key(display_name[item])):
        voice_sequences = sorted(
            by_voice[key],
            key=lambda item: (
                int(item.get("ordre", 0) or 0),
                int(item.get("source_paragraphe", 0) or 0),
                item.get("id", ""),
            ),
        )
        ordered.extend(voice_sequences)
    return ordered


def _tb_edito_is_genesis_sequence(sequence: dict) -> bool:
    """Genese explicite (ex. JJGE-0004 : « Comment est née votre innovation ? »)."""
    text = _normalize_for_match(sequence.get("texte", ""))
    question = _normalize_for_match(sequence.get("question", ""))
    genesis_in_question = (
        "comment est nee",
        "comment nait",
        "genese de",
        "d ou vient votre innovation",
        "d ou vient mon innovation",
        "innovation resulte d un declic",
    )
    if any(token in question for token in genesis_in_question):
        return True
    genesis_in_text = (
        "mon innovation est nee",
        "cette innovation est nee",
        "mon innovation decoule",
        "comment est nee mon innovation",
        "au tout debut de la genese",
    )
    return any(token in text[:160] for token in genesis_in_text)


def _tb_edito_t1_split_presentations_geneses(sequences: list[dict]) -> tuple[list[dict], list[dict]]:
    """
    T1 : par voix, tout ce qui precede le premier marqueur de genese reste en presentation.
    Ainsi JJGE-0003 (exemple nano-photonique) reste avec JJGE-0001/0002 ;
    JJGE-0004 (question genese) ouvre la partie geneses.
    """

    def _doc_order(item: dict) -> tuple:
        return (
            int(item.get("ordre", 0) or 0),
            int(item.get("source_paragraphe", 0) or 0),
            item.get("id", ""),
        )

    by_voice: dict[str, list[dict]] = {}
    for seq in sequences:
        voice = (seq.get("intervenant") or "Temoin").strip()
        by_voice.setdefault(voice, []).append(seq)

    presentations: list[dict] = []
    geneses: list[dict] = []
    for voice, items in by_voice.items():
        ordered = sorted(items, key=_doc_order)
        split_at = next(
            (idx for idx, item in enumerate(ordered) if _tb_edito_is_genesis_sequence(item)),
            len(ordered),
        )
        presentations.extend(ordered[:split_at])
        geneses.extend(ordered[split_at:])
    return presentations, geneses


def _tb_edito_t1_presentation_ids(sequences: list[dict]) -> set[str]:
    presentations, _ = _tb_edito_t1_split_presentations_geneses(sequences)
    return {item.get("id", "") for item in presentations if item.get("id")}


def _tb_edito_is_presentation_sequence(sequence: dict, presentation_ids: set[str] | None = None) -> bool:
    """Pour T1, passer presentation_ids issu du split par voix ; sinon heuristique genese."""
    sid = sequence.get("id", "")
    if presentation_ids is not None:
        return sid in presentation_ids
    return not _tb_edito_is_genesis_sequence(sequence)


def _tb_edito_order_for_code(code: str, sequences: list[dict]) -> list[dict]:
    # Priorite au sens : sections chercheur apres chercheur (pas d'alternance chorale).
    if code != "T1":
        return _tb_edito_by_chercheur_order(sequences)

    # T1 : deux ensembles globaux — (1) toutes les presentations, (2) toutes les geneses.
    def _doc_order(item: dict) -> tuple:
        return (
            int(item.get("ordre", 0) or 0),
            int(item.get("source_paragraphe", 0) or 0),
            item.get("id", ""),
        )

    def _voice_then_doc(item: dict) -> tuple:
        voice = (item.get("intervenant") or "Temoin").strip()
        return (*_tb_edito_voice_sort_key(voice), *_doc_order(item))

    presentations, geneses = _tb_edito_t1_split_presentations_geneses(sequences)
    return sorted(presentations, key=_voice_then_doc) + sorted(geneses, key=_voice_then_doc)


def _tb_edito_build_cadrage(code: str, ordered_ids: list[str], by_seq_id: dict[str, dict], videos_expert: list[dict]) -> dict:
    video_label = FIXED_TEMOIN_PLAN.get(code, {}).get("label", _label_video_temoin(code))
    expert_codes = [item.get("code", "") for item in videos_expert if item.get("code")]
    expert_chain = ", ".join(_label_video_expert(code) for code in expert_codes) if expert_codes else ""
    transitions = []
    for idx in range(len(ordered_ids) - 1):
        current = by_seq_id.get(ordered_ids[idx], {})
        nxt = by_seq_id.get(ordered_ids[idx + 1], {})
        current_video = _edito_title_core(current.get("video", ""))
        next_video = _edito_title_core(nxt.get("video", ""))
        if not current_video or not next_video or current_video == next_video:
            continue
        transitions.append(
            {
                "id": f"TR_{len(transitions) + 1:02d}",
                "position": "Entre deux extraits sur changement de sujet",
                "fonction": "Transition de sujet",
                "apres_extrait": ordered_ids[idx],
                "avant_extrait": ordered_ids[idx + 1],
                "texte_intervenant": "Nous changeons maintenant d'angle pour poursuivre la progression pedagogique de cette video.",
                "texte_pancarte": "Transition : nouveau sujet",
                "enchainement_expert": expert_chain,
            }
        )

    intro = {
        "position": "Avant le premier extrait",
        "fonction": "Ouvrir la video temoin et annoncer l'objectif pedagogique",
        "texte_intervenant": f"Dans cette {video_label}, nous allons partager des experiences concretes pour eclairer les points cles du sujet.",
        "texte_pancarte": video_label,
        "enchainement_expert": expert_chain,
    }
    note = "Transitions insérées uniquement lors d'un changement de sujet explicite."

    if code == "T1" and ordered_ids:
        presentation_ids = _tb_edito_t1_presentation_ids(list(by_seq_id.values()))
        presentations = [sid for sid in ordered_ids if sid in presentation_ids]
        geneses = [sid for sid in ordered_ids if sid not in presentation_ids]
        intro = {
            "position": "Avant le premier extrait",
            "fonction": "Annoncer la structure en deux temps : presentations puis geneses.",
            "texte_intervenant": (
                "Comment une innovation commence-t-elle vraiment ? "
                "Quatre chercheurs vont d'abord se presenter — leur domaine, leur laboratoire — "
                "puis chacun racontera comment est nee son innovation."
            ),
            "texte_pancarte": (
                "Comment nait une innovation ?\n"
                "1. Qui sont-ils ?  →  2. Comment est nee leur innovation ?"
            ),
            "enchainement_expert": expert_chain,
        }
        note = (
            "T1 structure en deux ensembles : presentations puis geneses. "
            "JJGE-0004 (et equivalents) sont places en genese, pas en presentation."
        )
        if presentations and geneses:
            transitions.insert(
                0,
                {
                    "id": "TR_PRESENTATIONS_GENESES",
                    "position": "Apres les presentations — avant les geneses",
                    "fonction": "Passer de qui sont-ils a comment est nee leur innovation.",
                    "apres_extrait": presentations[-1],
                    "avant_extrait": geneses[0],
                    "autorise_changement_voix": True,
                    "texte_intervenant": (
                        "Vous les avez rencontres. "
                        "Maintenant, ecoutez comment leur innovation est nee — "
                        "rencontre et besoin de marche, resultat de labo, pas de cote, maturation longue…"
                    ),
                    "texte_pancarte": "Partie 2 — Geneses\nComment est nee leur innovation ?",
                    "enchainement_expert": expert_chain,
                },
            )

    return {
        "statut": "PROPOSITION_AUTO",
        "dispositif": "Proposition de conduite narrative pour la capsule temoin.",
        "note": note,
        "intro": intro,
        "transitions": transitions,
        "outro": {
            "position": "Apres le dernier extrait",
            "fonction": "Clore la video temoin et ouvrir vers les videos expert",
            "texte_intervenant": "Vous avez maintenant les points essentiels. Passons aux videos expert pour approfondir avec des reperes pratiques.",
            "texte_pancarte": "Suite : approfondissements expert",
            "enchainement_expert": expert_chain,
        },
    }


def _tb_edito_script_with_cadrage(ordre: list[str], by_seq_id: dict[str, dict], cadrage: dict) -> str:
    if not ordre:
        return "Aucun extrait edito apparie pour cette video."

    def _seq_line(sequence: dict) -> str:
        verbatim = sequence.get("texte", "")
        return (
            f"[{sequence.get('id', '-')}] {sequence.get('intervenant', '')} | "
            f"{sequence.get('source_doc', '')} | {sequence.get('video', '')}\n"
            f"{verbatim}"
        ).strip()

    def _cadrage_line(kind: str, bloc: dict, label: str = "") -> str:
        kind_label = kind.upper()
        if label:
            kind_label = f"{kind_label} ({label})"
        position = _normalize_editorial_french(bloc.get("position", ""))
        header = f"[CADRAGE — NON PRONONCE — {kind_label}] Animateur | {position}"
        lines = [header]
        if bloc.get("texte_intervenant"):
            lines.append(_normalize_editorial_french(bloc["texte_intervenant"]))
        if bloc.get("texte_pancarte"):
            lines.append(
                f"[PANCARTE]\n{_normalize_editorial_french(bloc['texte_pancarte'])}"
            )
        if bloc.get("enchainement_expert"):
            lines.append(f"[EXPERT] {bloc['enchainement_expert']}")
        return "\n".join(lines)

    parts: list[str] = []
    intro = cadrage.get("intro", {})
    if intro:
        parts.append(_cadrage_line("intro", intro))

    current_voice = None
    current_part = None
    presentation_ids: set[str] | None = None
    if any(t.get("id") == "TR_PRESENTATIONS_GENESES" for t in cadrage.get("transitions", [])):
        presentation_ids = _tb_edito_t1_presentation_ids(list(by_seq_id.values()))
    for idx, seq_id in enumerate(ordre):
        sequence = by_seq_id.get(seq_id)
        if not sequence:
            continue
        # Marqueurs d'ensemble pour T1 (presentations / geneses).
        if presentation_ids is not None:
            part = (
                "presentations"
                if _tb_edito_is_presentation_sequence(sequence, presentation_ids)
                else "geneses"
            )
            if part != current_part:
                label = (
                    "=== PARTIE 1 — PRÉSENTATIONS ==="
                    if part == "presentations"
                    else "=== PARTIE 2 — GENÈSES ==="
                )
                parts.append(label)
                current_part = part
                current_voice = None
        voice = (sequence.get("intervenant") or "Temoin").strip()
        if voice != current_voice:
            parts.append(f"=== {voice} ===")
            current_voice = voice
        parts.append(_seq_line(sequence))
        next_id = ordre[idx + 1] if idx + 1 < len(ordre) else None
        for transition in cadrage.get("transitions", []):
            if transition.get("apres_extrait") != seq_id:
                continue
            before = transition.get("avant_extrait")
            if before is not None and before != next_id:
                continue
            next_seq = by_seq_id.get(next_id or "", {})
            next_voice = (next_seq.get("intervenant") or "").strip()
            # Par defaut, une transition reste dans la meme voix ;
            # sauf si elle marque un changement d'ensemble (presentations → geneses).
            if (
                not transition.get("autorise_changement_voix")
                and next_voice
                and next_voice != voice
            ):
                continue
            parts.append(_cadrage_line("transition", transition, transition.get("id", "")))
            current_voice = None  # forcer un nouveau bandeau voix apres la transition d'ensemble

    outro = cadrage.get("outro", {})
    if outro:
        parts.append(_cadrage_line("outro", outro))

    return "\n\n".join(parts)


def build_tb_edito_capsule_pages(programme_table: dict) -> None:
    rows_by_code = {row.get("code", ""): row for row in programme_table.get("rows", [])}
    grouped = _tb_edito_sequences_by_code()
    all_edito_intervenants = sorted(
        {
            item.get("intervenant", "").strip()
            for item in load_derushage_edito_index()
            if item.get("intervenant", "").strip()
        }
    )
    expected: set[str] = set()
    empty_by_id: dict[str, dict] = {}

    for code, spec in sorted(FIXED_TEMOIN_PLAN.items(), key=lambda item: int(item[0][1:])):
        page_name = f"tb_edito_{code}.html"
        expected.add(page_name)
        row = rows_by_code.get(code, {})
        sequences = grouped.get(code, [])
        sequences_sorted = _tb_edito_order_for_code(code, sequences)
        by_seq_id = {item.get("id", f"{code}-NOID"): item for item in sequences_sorted}
        ordre = [item.get("id", f"{code}-NOID") for item in sequences_sorted]

        videos_expert = _tb_edito_parse_videos_expert(row.get("videos_referent", ""))
        experts_proposes = _extract_intervenants(row.get("noms_proposes", ""))
        resume = ""
        cadrage = _tb_edito_build_cadrage(code, ordre, by_seq_id, videos_expert)

        script_final = _script_final_prefer_mounted_transcript(
            code,
            _tb_edito_script_with_cadrage(ordre, by_seq_id, cadrage),
        )

        capsule_data = {
            "ordre_montage": ordre,
            "script_final": script_final,
            "resume_temoignages": resume,
            "videos_expert": videos_expert,
            "experts_proposes": experts_proposes,
            "cadrage_animateur": cadrage,
        }

        sections = [
            f"<p><strong>Objectif pedagogique :</strong> {escape(row.get('objectif_pedagogique', 'A renseigner.'))}</p>",
            f"<p><strong>Video temoin :</strong> {escape(spec.get('label', _label_video_temoin(code)))}</p>",
            f"<p class='meta'><strong>Sequences edito apparies :</strong> {len(sequences_sorted)}</p>",
            "<h2>Montage édito retenu</h2>",
        ]
        if code == "T1":
            sections.append(
                "<p class='meta'>Organisation en <strong>deux ensembles</strong> : "
                "(1) presentations de tous les temoins, puis (2) geneses de leurs innovations "
                "(chercheur apres chercheur dans chaque ensemble).</p>"
            )
            sections.append(
                "<p class='meta'><strong>Regle T1 :</strong> un passage comme "
                "<code>JJGE-0004</code> (question « Comment est née votre innovation ? ») "
                "est place en genese, jamais en presentation.</p>"
            )
        else:
            sections.append(
                "<p class='meta'>Organisation <strong>chercheur apres chercheur</strong> "
                "(priorite au sens, pas d'alternance chorale).</p>"
            )
        if sequences_sorted:
            current_voice = None
            current_part = None
            presentation_ids = (
                _tb_edito_t1_presentation_ids(sequences_sorted) if code == "T1" else None
            )
            for sequence in sequences_sorted:
                if code == "T1" and presentation_ids is not None:
                    part = (
                        "presentations"
                        if _tb_edito_is_presentation_sequence(sequence, presentation_ids)
                        else "geneses"
                    )
                    if part != current_part:
                        title = (
                            "Partie 1 — Présentations"
                            if part == "presentations"
                            else "Partie 2 — Genèses"
                        )
                        sections.append(f"<h3>{title}</h3>")
                        current_part = part
                        current_voice = None
                voice = (sequence.get("intervenant") or "Temoin").strip()
                if voice != current_voice:
                    sections.append(f"<h4>{escape(voice)}</h4>")
                    current_voice = voice
                verbatim = sequence.get("texte", "")
                sections.append(
                    "<div class='card'>"
                    f"<strong>{escape(sequence.get('id', '-'))}</strong> "
                    f"<span class='meta'>{escape(sequence.get('source_doc', ''))}</span>"
                    f"<p class='meta'><strong>{escape(sequence.get('video', 'Video non renseignee'))}</strong></p>"
                    f"<p>{escape(verbatim)}</p>"
                    "</div>"
                )
        else:
            sections.append("<p class='meta'>Aucun extrait surligne apparié automatiquement a cette video.</p>")
        voices_in_code = {
            item.get("intervenant", "").strip()
            for item in sequences_sorted
            if item.get("intervenant", "").strip()
        }
        missing_intervenants = [name for name in all_edito_intervenants if name not in voices_in_code]
        if missing_intervenants:
            sections.append("<h3>Témoins sans extrait surligné sur cette vidéo</h3>")
            sections.append(
                "<p class='meta'>Presence documentaire uniquement (pas de sequence verbatim retenue pour cette video dans les surlignages).</p>"
            )
            sections.append(
                "<ul>"
                + "".join(f"<li>{escape(name)}</li>" for name in missing_intervenants)
                + "</ul>"
            )
        sections.append(cadrage_animateur_section(capsule_data))
        sections.append("<h2>Script final</h2>")
        sections.append(
            f"<div class='script' id='script-final'>"
            f"{escape(_normalize_script_final_editorial(script_final))}"
            f"</div>"
        )
        sections.append(synthese_temoignages_section(code, capsule_data))
        sections.append(brief_intervenant_section(code, capsule_data, empty_by_id))
        sections.append(referents_section(capsule_data))
        sections.append(scripts_expertise_projetes_section(capsule_data))
        sections.append(export_word_section(code, spec.get("label", code), capsule_data, empty_by_id, programme_table))

        write_text(
            SITE / page_name,
            html_page(
                f"Capsule témoin — {code}",
                "\n".join(part for part in sections if part),
                scripts=["assets/export-word.js"],
                nav_current="suivi_intervenants.html",
                breadcrumb=html_breadcrumb(
                    ("Accueil", "index.html"),
                    ("Capsules témoins", "tb_edito.html"),
                    (code, None),
                ),
                page_header=(
                    "<div class='page-head'>"
                    f"<h1>{escape(code)} — Capsule témoin</h1>"
                    f"<p class='lead'>{escape(spec.get('label', _label_video_temoin(code)))}</p>"
                    "</div>"
                ),
            ),
        )

    for path in SITE.glob("tb_edito_T*.html"):
        if path.name not in expected:
            path.unlink()


def build_tb_edito_page() -> None:
    docs = []
    for item in load_derushage_edito_index():
        doc = load_derushage_edito(item.get("id", ""))
        if doc:
            docs.append(doc)

    total_sequences = sum(len(doc.get("sequences", [])) for doc in docs)
    total_paragraphs = sum(int(doc.get("nb_paragraphes_analyse", 0) or 0) for doc in docs)
    docs_with_sequences = sum(1 for doc in docs if doc.get("sequences"))
    all_sequences = [sequence for doc in docs for sequence in doc.get("sequences", [])]
    unique_video_titles = sorted(
        {
            (sequence.get("video") or "").strip()
            for sequence in all_sequences
            if (sequence.get("video") or "").strip()
        }
    )
    questions_count = sum(1 for sequence in all_sequences if (sequence.get("question") or "").strip())
    unresolved_videos = sum(1 for sequence in all_sequences if not (sequence.get("video") or "").strip())

    coverage_counter = Counter()
    unresolved_titles = Counter()
    for sequence in all_sequences:
        title = (sequence.get("video") or "").strip()
        if not title:
            continue
        targets = _target_codes_from_edito_title(title)
        if targets:
            for code in targets:
                coverage_counter[code] += 1
        else:
            unresolved_titles[title] += 1

    covered_codes = {code for code, count in coverage_counter.items() if count > 0}
    rows = []
    for doc in docs:
        doc_id = doc.get("id", "")
        sequences = doc.get("sequences", [])
        seq_count = len(sequences)
        question_doc_count = sum(1 for sequence in sequences if (sequence.get("question") or "").strip())
        video_titles = {
            (sequence.get("video") or "").strip()
            for sequence in sequences
            if (sequence.get("video") or "").strip()
        }
        targets = set()
        for title in video_titles:
            targets.update(_target_codes_from_edito_title(title))
        targets_label = ", ".join(sorted(targets)) if targets else "-"
        rows.append(
            "<tr>"
            f"<td><a href='derushage_edito_{escape(doc_id)}.html'>{escape(doc.get('intervenant', doc_id))}</a></td>"
            f"<td>{escape(doc.get('source', '-'))}</td>"
            f"<td>{seq_count}</td>"
            f"<td>{escape(doc.get('nb_paragraphes_analyse', 0))}</td>"
            f"<td>{len(video_titles)}</td>"
            f"<td>{escape(targets_label)}</td>"
            f"<td>{question_doc_count}</td>"
            f"<td>{escape(doc.get('date_maj', '-'))}</td>"
            "</tr>"
        )

    coverage_rows = []
    for code, spec in sorted(FIXED_TEMOIN_PLAN.items(), key=lambda item: int(item[0][1:])):
        code_link = f"<a href='tb_edito_{escape(code)}.html'>{escape(code)}</a>"
        coverage_rows.append(
            "<tr>"
            f"<td>{escape(spec.get('module', '-'))}</td>"
            f"<td>{code_link}</td>"
            f"<td>{escape(spec.get('label', '-'))}</td>"
            f"<td>{coverage_counter.get(code, 0)}</td>"
            "</tr>"
        )

    unresolved_rows = "".join(
        f"<li>{escape(title)} <span class='meta'>({count} seq.)</span></li>"
        for title, count in unresolved_titles.most_common(10)
    )
    intervenants = sorted({doc.get("intervenant", "").strip() for doc in docs if doc.get("intervenant")})
    chips = "".join(
        f"<a class='chip' href='derushage_edito_{escape(doc.get('id', ''))}.html'>{escape(doc.get('intervenant', 'Intervenant'))}</a>"
        for doc in sorted(docs, key=lambda item: (item.get("intervenant", ""), item.get("id", "")))
    )
    body = (
        "<section class='stats-grid'>"
        "<div class='stat-card'>"
        "<div class='stat-card__label'>Documents édito</div>"
        f"<div class='stat-card__value'>{len(docs)}</div>"
        f"<div class='stat-card__meta'>{docs_with_sequences} avec sequences retenues</div>"
        "</div>"
        "<div class='stat-card'>"
        "<div class='stat-card__label'>Séquences surlignées</div>"
        f"<div class='stat-card__value'>{total_sequences}</div>"
        f"<div class='stat-card__meta'>{questions_count} avec question renseignee</div>"
        "</div>"
        "<div class='stat-card'>"
        "<div class='stat-card__label'>Couverture vidéos édito</div>"
        f"<div class='stat-card__value'>{len(covered_codes)}/12</div>"
        f"<div class='stat-card__meta'>{len(unique_video_titles)} titres video detectes · {unresolved_videos} sequences sans video</div>"
        "</div>"
        "<div class='stat-card'>"
        "<div class='stat-card__label'>Paragraphes analysés</div>"
        f"<div class='stat-card__value'>{total_paragraphs}</div>"
        "<div class='stat-card__meta'>issus des transcripts corriges fournis</div>"
        "</div>"
        "</section>"
        "<h2>Documents édito</h2>"
        "<div class='table-wrap'><table><thead><tr>"
        "<th>Intervenant</th><th>Source</th><th>Sequences retenues</th><th>Paragraphes analyses</th>"
        "<th>Videos distinctes</th><th>Cibles T1..T12</th><th>Questions</th><th>Mise a jour</th>"
        "</tr></thead><tbody>"
        + ("\n".join(rows) or "<tr><td colspan='8'>Aucun document edito detecte.</td></tr>")
        + "</tbody></table></div>"
        "<h2>Couverture du plan témoin édito</h2>"
        "<div class='table-wrap'><table><thead><tr>"
        "<th>Module</th><th>Code</th><th>Vidéo témoin fixée</th><th>Seq. édito appariées</th>"
        "</tr></thead><tbody>"
        + ("\n".join(coverage_rows) or "<tr><td colspan='4'>Aucune couverture.</td></tr>")
        + "</tbody></table></div>"
        + (
            "<h2>Titres vidéo édito non appariés automatiquement</h2>"
            f"<ul>{unresolved_rows}</ul>"
            if unresolved_rows
            else ""
        )
        + "<h2>Intervenants édito</h2>"
        + (f"<div class='chip-grid'>{chips}</div>" if chips else "<p class='meta'>Aucun intervenant edito.</p>")
    )
    write_text(
        SITE / "tb_edito.html",
        html_page(
            "Capsules témoins",
            body,
            nav_current="suivi_intervenants.html",
            breadcrumb=html_breadcrumb(("Accueil", "index.html"), ("Capsules témoins", None)),
            page_header='<div class="page-head"><h1>Capsules témoins</h1><p class="lead">Suivi des selections surlignees de l\'edito, avec couverture des videos temoins fixees.</p></div>',
        ),
    )


def build_researcher_pages(segments: list[dict]) -> None:
    grouped = defaultdict(list)
    for segment in segments:
        grouped[segment["chercheur"]].append(segment)

    # Nettoie les pages obsoletes (slug change apres correction d'identite).
    expected_slugs = {slug(researcher) for researcher in grouped.keys()}
    for path in SITE.glob("chercheur_*.html"):
        name = path.name
        if not name.startswith("chercheur_") or not name.endswith(".html"):
            continue
        slug_part = name[len("chercheur_") : -len(".html")]
        if slug_part not in expected_slugs:
            path.unlink()

    for researcher, items in grouped.items():
        rows = []
        for segment in sorted(items, key=lambda item: (item["source"], item["debut"])):
            rows.append(
                "<tr>"
                f"<td>{escape(segment['id'])}</td>"
                f"<td>{escape(segment['debut'])} - {escape(segment['fin'])}</td>"
                f"<td>{escape(segment['theme_principal'])}</td>"
                f"<td>{escape(segment['statut'])}</td>"
                f"<td>{total_score(segment)}/12</td>"
                f"<td>{escape(segment['verbatim'])}</td>"
                "</tr>"
            )
        used_duration = sum(item["duree_secondes"] for item in items if item["statut"] == "UTILISE")
        statuses = Counter(item["statut"] for item in items)
        status_tags = "".join(
            f"<span class='tag'>{escape(key)}: {value}</span>" for key, value in statuses.items()
        )
        body = (
            f"<p class='meta'>Duree totale utilisee: {format_seconds(used_duration)}</p>"
            f"<p>{status_tags}</p>"
            '<div class="table-wrap"><table><thead><tr><th>ID</th><th>Timecodes</th><th>Theme</th><th>Statut</th><th>Score</th><th>Verbatim</th></tr></thead><tbody>'
            + "\n".join(rows)
            + "</tbody></table></div>"
        )
        write_text(
            SITE / f"chercheur_{slug(researcher)}.html",
            html_page(
                researcher,
                body,
                nav_current=None,
                breadcrumb=html_breadcrumb(
                    ("Accueil", "index.html"),
                    ("Tableau de bord", "tableau_de_bord.html"),
                    (researcher, None),
                ),
            ),
        )


def build_capsule_pages(
    capsules: list[dict],
    segments: list[dict],
    affectations: dict,
    programme_table: dict,
) -> None:
    by_id = index_by_id(segments)
    overlaps = find_overlaps(segments)
    overlap_ids = {item.first_id for item in overlaps} | {item.second_id for item in overlaps}
    for capsule in capsules:
        code = capsule["code"]
        capsule_data = affectations["capsules"].get(code, {})
        sections = [
            f"<p><strong>Objectif:</strong> {_e_fr(capsule['objectif_pedagogique'])}</p>",
            f"<p><strong>Message central:</strong> {_e_fr(capsule['message_central'])}</p>",
        ]
        if capsule.get("role") == "LABORATOIRE" and capsule.get("equivalent_production"):
            prod = capsule["equivalent_production"]
            sections.append(
                f"<p class='meta'><strong>Laboratoire éditorial :</strong> cette capsule sert à préparer "
                f"<a href='capsule_{escape(prod)}.html'>{escape(prod)}</a> "
                f"(même contenu témoin, montage et cadrage validés ici avant production).</p>"
            )
        else:
            lab = next((c for c in capsules if c.get("equivalent_production") == code), None)
            if lab:
                sections.append(
                    f"<p class='meta'><strong>Production dérivée du laboratoire :</strong> montage et cadrage "
                    f"initialisés depuis <a href='capsule_{escape(lab['code'])}.html'>{escape(lab['code'])}</a>.</p>"
                )
        sections.append("<h2>Extraits candidats</h2>")
        for segment_id in capsule_data.get("extraits_candidats", []):
            segment = by_id.get(segment_id)
            if segment:
                warning = " <span class='warn'>Chevauchement</span>" if segment_id in overlap_ids else ""
                sections.append(f"<div class='card'>{link_segment(segment)}{warning}</div>")
        sections.append("<h2>Montage proposé</h2>")
        plan = capsule_data.get("plan_montage", [])
        if plan:
            montage_total = sum(float(item.get("duree_montage_secondes", 0)) for item in plan)
            sections.append(
                f"<p class='meta'><strong>Durée montage estimée :</strong> {format_seconds(montage_total)} "
                f"(cible 5-7 min). Les timecodes BAB sont des bornes ; les coupes fines sont NON PRONONCE.</p>"
            )
        for index, segment_id in enumerate(capsule_data.get("ordre_montage", []), start=1):
            segment = by_id.get(segment_id)
            if segment:
                plan_item = plan[index - 1] if plan and index - 1 < len(plan) else None
                meta = ""
                if plan_item:
                    role = plan_item.get("role", "")
                    duration = plan_item.get("duree_montage_secondes")
                    coupe = plan_item.get("coupe")
                    meta_parts = [f"#{index}"]
                    if role:
                        meta_parts.append(_normalize_editorial_french(role))
                    if duration is not None:
                        meta_parts.append(f"~{format_seconds(float(duration))}")
                    if coupe:
                        meta_parts.append(
                            f"coupe : {_normalize_editorial_french(coupe)}"
                        )
                    meta = f"<p class='meta'>{escape(' · '.join(meta_parts))}</p>"
                sections.append(f"<div class='card'>{link_segment(segment)}{meta}</div>")
        if capsule_data.get("cadrage_animateur"):
            sections.append(cadrage_animateur_section(capsule_data))
        sections.append("<h2>Script final</h2>")
        sections.append(
            f"<div class='script' id='script-final'>"
            f"{escape(_normalize_script_final_editorial(capsule_data.get('script_final') or 'À construire.'))}"
            f"</div>"
        )
        sections.append(synthese_temoignages_section(code, capsule_data, by_id))
        sections.append("<h2>Manques et décisions</h2>")
        for item in capsule_data.get("manques", []):
            sections.append(f"<p class='warn'>{_e_fr(item)}</p>")
        for item in capsule_data.get("decisions_editoriales", []):
            sections.append(f"<p>{_e_fr(item)}</p>")
        if capsule_data.get("methodologie") or capsule_data.get("unites_de_sens"):
            sections.append(selection_methodology_section(capsule, capsule_data))
        if capsule_data.get("videos_expert"):
            sections.append(brief_intervenant_section(code, capsule_data, by_id))
        if capsule_data.get("methodologie") or capsule_data.get("unites_de_sens"):
            sections.append(selection_unites_section(capsule_data))
        sections.append(referents_section(capsule_data))
        sections.append(scripts_expertise_projetes_section(capsule_data))
        sections.append(
            export_word_section(code, capsule["titre"], capsule_data, by_id, programme_table)
        )
        page_title = f"{code} - {capsule['titre']}"
        write_text(
            SITE / f"capsule_{code}.html",
            html_page(
                page_title,
                "\n".join(sections),
                scripts=["assets/export-word.js"],
                nav_current=None,
                breadcrumb=html_breadcrumb(
                    ("Accueil", "index.html"),
                    ("Tableau de bord", "tableau_de_bord.html"),
                    (code, None),
                ),
            ),
        )


def build_conflicts_page(segments: list[dict]) -> None:
    overlaps = find_overlaps(segments)
    rows = []
    for overlap in overlaps:
        rows.append(
            "<tr>"
            f"<td>{escape(overlap.first_id)}</td>"
            f"<td>{escape(overlap.second_id)}</td>"
            f"<td>{escape(overlap.chercheur)}</td>"
            f"<td>{escape(overlap.source)}</td>"
            f"<td>{overlap.duree}s</td>"
            "</tr>"
        )
    body = '<div class="table-wrap"><table><thead><tr><th>Extrait 1</th><th>Extrait 2</th><th>Chercheur</th><th>Source</th><th>Duree chevauchee</th></tr></thead><tbody>'
    body += "\n".join(rows) or "<tr><td colspan='5'>Aucun conflit detecte.</td></tr>"
    body += "</tbody></table></div>"
    write_text(
        SITE / "conflits.html",
        html_page(
            "Conflits",
            body,
            nav_current="conflits.html",
            breadcrumb=html_breadcrumb(("Accueil", "index.html"), ("Conflits", None)),
            page_header='<div class="page-head"><h1>Conflits</h1><p class="lead">Chevauchements temporels et reutilisations a arbitrer entre extraits.</p></div>',
        ),
    )


def build_registry(segments: list[dict]) -> None:
    rows = []
    for segment in sorted(segments, key=lambda item: item["id"]):
        rows.append(
            "<tr>"
            f"<td>{escape(segment['id'])}</td>"
            f"<td>{escape(segment['chercheur'])}</td>"
            f"<td>{escape(segment['theme_principal'])}</td>"
            f"<td>{escape(', '.join(segment['capsules_candidates']))}</td>"
            f"<td>{escape(segment['statut'])}</td>"
            f"<td>{total_score(segment)}/12</td>"
            f"<td>{escape(segment['qualification'])}</td>"
            f"<td>{'oui' if segment['transcription_a_verifier'] else 'non'}</td>"
            "</tr>"
        )
    body = '<div class="table-wrap"><table><thead><tr><th>ID</th><th>Chercheur</th><th>Theme</th><th>Capsules</th><th>Statut</th><th>Score</th><th>Qualification</th><th>Transcription</th></tr></thead><tbody>'
    body += "\n".join(rows)
    body += "</tbody></table></div>"
    write_text(
        SITE / "registre.html",
        html_page(
            "Registre des extraits",
            body,
            nav_current="registre.html",
            breadcrumb=html_breadcrumb(("Accueil", "index.html"), ("Registre", None)),
            page_header='<div class="page-head"><h1>Registre des extraits</h1><p class="lead">Inventaire complet des segments indexes, leurs statuts et qualifications.</p></div>',
        ),
    )


def render_derushage_edito_export(doc: dict) -> str:
    lines = [
        f"DERUSHAGE EDITO — {doc.get('intervenant', '')}",
        f"Source : {doc.get('source', '')}",
        f"Statut : {doc.get('statut_derushage', '')}",
        f"Derniere mise a jour : {doc.get('date_maj', '')}",
        "",
    ]
    for sequence in doc.get("sequences", []):
        lines.append(
            f"[{sequence.get('id', '-')}] {sequence.get('video') or 'Video non renseignee'}"
        )
        if sequence.get("module"):
            lines.append(f"Module : {sequence['module']}")
        if sequence.get("question"):
            lines.append(f"Question : {sequence['question']}")
        lines.append(
            f"Surlignage : {sequence.get('couleur_surlignage', '-')} · "
            f"Paragraphe source : {sequence.get('source_paragraphe', '-')}"
        )
        lines.append(sequence.get("texte", ""))
        lines.append("")
    return "\n".join(lines).strip() + "\n"


def derushage_edito_sequence_html(sequence: dict) -> str:
    video = sequence.get("video") or "Video non renseignee"
    module = sequence.get("module") or "Module non renseigne"
    question = sequence.get("question") or "Question non renseignee"
    return (
        "<div class='card bab-segment'>"
        f"<strong>{escape(sequence.get('id', '-'))}</strong> "
        f"<span class='meta'>{escape(sequence.get('statut_edito', 'RETENU_PAR_EDITO'))}</span>"
        f"<p class='meta'><strong>{escape(video)}</strong> · {escape(module)}</p>"
        f"<p class='meta'>Question : {escape(question)}</p>"
        f"<p class='meta'>Surlignage : {escape(sequence.get('couleur_surlignage', '-'))} · "
        f"Paragraphe source : {escape(sequence.get('source_paragraphe', '-'))}</p>"
        f"<p>{escape(sequence.get('texte', ''))}</p>"
        "</div>"
    )


def build_derushage_edito_index() -> None:
    rows = []
    for item in load_derushage_edito_index():
        doc = load_derushage_edito(item["id"])
        if not doc:
            continue
        rows.append(
            "<tr>"
            f"<td><a href='derushage_edito_{escape(item['id'])}.html'>{escape(item['intervenant'])}</a></td>"
            f"<td>{escape(item['source'])}</td>"
            f"<td>{escape(item.get('nb_sequences_retenues', 0))}</td>"
            f"<td>{escape(item.get('nb_paragraphes_analyse', 0))}</td>"
            f"<td>{escape(item.get('statut_derushage', '-'))}</td>"
            f"<td>{escape(item.get('date_maj', '-'))}</td>"
            "</tr>"
        )
    body = (
        "<p class='meta'>Synthese des transcripts corriges et surlignes par l'edito (Clarisse). "
        "Les extraits ci-dessous sont auto-extraits depuis les surlignages des fichiers <code>data/raw/*.docx</code>.</p>"
        '<div class="table-wrap"><table><thead><tr><th>Intervenant</th><th>Source</th>'
        "<th>Sequences retenues</th><th>Paragraphes analyses</th><th>Statut</th><th>Mise a jour</th></tr></thead><tbody>"
        + ("\n".join(rows) or "<tr><td colspan='6'>Aucun derushage edito genere.</td></tr>")
        + "</tbody></table></div>"
    )
    write_text(
        SITE / "derushage_edito.html",
        html_page(
            "Dérushage édito",
            body,
            nav_current="edito.html",
            breadcrumb=html_breadcrumb(("Accueil", "index.html"), ("Edito", "edito.html"), ("Dérushage édito", None)),
            page_header="<div class=\"page-head\"><h1>Dérushage édito</h1><p class=\"lead\">Sequences surlignees dans les transcripts corriges, structurees pour l'edition des videos.</p></div>",
        ),
    )


def build_derushage_edito_pages() -> None:
    expected = set()
    for item in load_derushage_edito_index():
        derushage_id = item["id"]
        expected.add(f"derushage_edito_{derushage_id}.html")
        doc = load_derushage_edito(derushage_id)
        if not doc:
            continue
        sequences = doc.get("sequences", [])
        export_text = render_derushage_edito_export(doc)
        sections = [
            f"<p class='meta'><strong>{len(sequences)}</strong> sequences retenues par surlignage "
            f"sur <strong>{escape(doc.get('nb_paragraphes_analyse', 0))}</strong> paragraphes analyses.</p>"
        ]
        for sequence in sequences:
            sections.append(derushage_edito_sequence_html(sequence))
        sections.append(
            f"<div class='script sr-export' id='script-final'>{escape(export_text)}</div>"
        )
        sections.append(export_word_modal(f"derushage_edito_{derushage_id}.doc"))
        page_title = f"Dérushage édito — {doc.get('intervenant', derushage_id)}"
        header = (
            "<div class='page-header'>"
            f"<div><h1>{escape(page_title)}</h1>"
            f"<p class='meta'>Source : {escape(doc.get('source', '-'))} · "
            f"Statut : {escape(doc.get('statut_derushage', '-'))} · "
            f"Mise a jour : {escape(doc.get('date_maj', '-'))}</p></div>"
            "<div class='export-toolbar' id='export-word-panel' "
            f"data-capsule-code='derushage-{escape(derushage_id)}' "
            f"data-capsule-title='{escape(doc.get('intervenant', derushage_id))} — Derushage edito'>"
            "<button type='button' class='btn' id='export-word-open' "
            "aria-expanded='false' aria-controls='export-word-modal'>Exporter en Word</button>"
            "</div></div>"
        )
        write_text(
            SITE / f"derushage_edito_{derushage_id}.html",
            html_page(
                page_title,
                "\n".join(sections),
                scripts=["assets/export-word.js"],
                nav_current="edito.html",
                page_header=header,
                breadcrumb=html_breadcrumb(
                    ("Accueil", "index.html"),
                    ("Edito", "edito.html"),
                    ("Dérushage édito", "derushage_edito.html"),
                    (doc.get("intervenant", derushage_id), None),
                ),
            ),
        )
    for path in SITE.glob("derushage_edito_*.html"):
        if path.name not in expected:
            path.unlink()


def _edito_title_core(title: str) -> str:
    text = _normalize_for_match(title or "")
    text = re.sub(r"\bvideo\s*\d+\b", " ", text)
    text = (
        text.replace(":", " ")
        .replace("/", " ")
        .replace("-", " ")
        .replace("'", " ")
        .replace("’", " ")
        .replace("`", " ")
    )
    return " ".join(text.split())


def _title_similarity(a: str, b: str) -> float:
    tokens_a = set(_edito_title_core(a).split())
    tokens_b = set(_edito_title_core(b).split())
    if not tokens_a or not tokens_b:
        return 0.0
    inter = len(tokens_a & tokens_b)
    union = len(tokens_a | tokens_b)
    jaccard = inter / union if union else 0.0
    overlap = inter / min(len(tokens_a), len(tokens_b))
    return 0.6 * overlap + 0.4 * jaccard


def _collect_edito_video_titles() -> list[dict]:
    by_title: dict[str, dict] = {}
    for item in load_derushage_edito_index():
        doc = load_derushage_edito(item["id"])
        if not doc:
            continue
        witness = doc.get("intervenant", item.get("intervenant", ""))
        for sequence in doc.get("sequences", []):
            title = (sequence.get("video") or "").strip()
            if not title:
                continue
            entry = by_title.setdefault(
                title,
                {
                    "title": title,
                    "witnesses": set(),
                    "nb_sequences": 0,
                    "texts": [],
                },
            )
            entry["witnesses"].add(witness)
            entry["nb_sequences"] += 1
            if sequence.get("texte"):
                entry["texts"].append(sequence["texte"])
    titles = []
    for entry in by_title.values():
        titles.append(
            {
                "title": entry["title"],
                "witnesses": sorted(w for w in entry["witnesses"] if w),
                "nb_sequences": entry["nb_sequences"],
                "texts": entry["texts"],
            }
        )
    return sorted(titles, key=lambda item: item["title"].lower())


FIXED_TEMOIN_PLAN = {
    "T1": {
        "module": "M1",
        "label": "VIDÉO 1 : POURQUOI OSER ?",
        "questions": [
            "Quel est votre domaine de recherche (en une phrase)",
        ],
    },
    "T2": {
        "module": "M1",
        "label": "VIDÉO 2 : DE LA RECHERCHE À L’INNOVATION",
        "questions": [
            "À quel moment vous êtes-vous posé la question de l’usage ou de l’utilité de votre innovation ?",
            "Quel problème concret votre innovation permet-elle de résoudre ?",
        ],
    },
    "T3": {
        "module": "M1",
        "label": "VIDÉO 3 : IDENTIFIER UN BESOIN RÉEL + SORTIR DU LABORATOIRE (fusion)",
        "questions": [
            "Comment avez-vous validé qu’il existait un besoin ou un intérêt",
            "Pourquoi est-il essentiel de sortir du laboratoire pour innover ?",
            "Qu'est-ce qui vous a poussé à aller rencontrer des acteurs extérieurs ?",
            "Y a-t-il eu une rencontre ou un échange déterminant dans votre parcours",
        ],
    },
    "T4": {
        "module": "M1",
        "label": "VIDÉO 4 : UNE IDÉE NE SUFFIT PAS + FREINS ET LEVIERS (fusion)",
        "questions": [],
    },
    "T5": {
        "module": "M2",
        "label": "VIDÉO 5 : PROTECTION ET VALORISATION",
        "questions": [],
    },
    "T6": {
        "module": "M2",
        "label": "VIDÉO 6 : TRANSFERT ET LICENSING",
        "questions": [],
    },
    "T7": {
        "module": "M3",
        "label": "VIDÉO 7 : VOUS N'ÊTES PAS SEUL(E)",
        "questions": [],
    },
    "T8": {
        "module": "M3",
        "label": "VIDÉO 8 : FINANCEMENTS, CONCOURS ET TEMPS POUR ENTREPRENDRE",
        "questions": [],
    },
    "T9": {
        "module": "M4",
        "label": "VIDÉO 9 : PARTENARIATS ET POSTURE : CONSTRUIRE UNE ÉQUIPE",
        "questions": [],
    },
    "T10": {
        "module": "M4",
        "label": "VIDÉO 10 : ENRICHIR SON LANGAGE",
        "questions": [],
    },
    "T11": {
        "module": "M5",
        "label": "VIDÉO 11 : EVOLUTION DANS LE MÉTIER DU CHERCHEUR.EUSE",
        "questions": [
            "Quels sont les principaux enseignements que vous retenez ?",
        ],
    },
    "T12": {
        "module": "M5",
        "label": "VIDÉO 12 : DISPOSITIFS D’ACCOMPAGNEMENT & COLLAB",
        "questions": [
            "Comment les collaborations ou consortiums ont-ils amplifié votre projet ?",
            "À quoi faut-il se préparer pour qu'ils fonctionnent ?",
        ],
    },
    "T13": {
        "module": "M6",
        "label": "VIDÉO 13 : DE CONCLUSION : PASSER À L’ACTION",
        "questions": [
            "Quel conseil donneriez-vous ?",
            "Quelle est l'idée reçue la plus fréquente sur l'innovation qu'il faudrait déconstruire ?",
            "Si vous deviez résumer votre message en un mot ou une phrase",
            "Si vous pouviez vous adresser à la personne que vous étiez avant, que lui diriez-vous aujourd'hui ?",
        ],
    },
}


def _target_codes_from_edito_title(title: str) -> set[str]:
    text = _edito_title_core(title)
    targets: set[str] = set()
    if "pourquoi oser" in text:
        targets.add("T1")
    if "recherche fondamentale" in text:
        targets.add("T2")
    if "besoin reel" in text or "sortir du labo" in text:
        targets.add("T3")
    if "idee ne suffit pas" in text or "freins" in text or "doutes" in text or "legitimite" in text:
        targets.add("T4")
    if "protection intellectuelle" in text:
        targets.add("T5")
    if "transfert" in text or "licensing" in text:
        targets.add("T6")
    if "ne pas avancer seul" in text or "ecosysteme" in text or "vous n etes pas seul" in text:
        targets.add("T7")
    if "financements" in text or "concours" in text:
        targets.add("T8")
    if "partenariat" in text and "equipe" in text:
        targets.add("T9")
    if "changer de langage" in text:
        targets.add("T10")
    # Clarisse: "Video 13 : ce que cela change dans le metier de chercheur"
    if (
        ("evolution" in text and "metier" in text)
        or ("change" in text and "metier" in text and "chercheur" in text)
    ):
        targets.add("T11")
    # Clarisse: "Video 14 : Dispositif accompagnement-collaboration" -> T12 (pas T7)
    if ("dispositif" in text and "accompagnement" in text) or (
        "dispositif" in text and "collaboration" in text
    ):
        targets.add("T12")
    # Clarisse: "Video 15 : Passer a l'action"
    if "passer a l action" in text or (
        "passer" in text and "action" in text and "dispositif" not in text
    ):
        targets.add("T13")
    return targets


def _pedagogical_match_percent(objective_text: str, edito_texts: list[str]) -> int:
    objective_tokens = set(_edito_title_core(objective_text).split())
    if not objective_tokens or not edito_texts:
        return 0
    corpus = " ".join(edito_texts)
    corpus_tokens = set(_edito_title_core(corpus).split())
    if not corpus_tokens:
        return 0
    inter = len(objective_tokens & corpus_tokens)
    coverage = inter / len(objective_tokens)
    jaccard = inter / len(objective_tokens | corpus_tokens)
    score = 0.75 * coverage + 0.25 * jaccard
    return round(score * 100)


def _expert_label(percent: int) -> str:
    if percent >= 70:
        return "Alignement fort"
    if percent >= 45:
        return "Alignement moyen"
    return "Alignement faible"


def _alignment_comment(percent: int, objective: str, top_matches: list[tuple[float, dict]]) -> str:
    if percent >= 70:
        return "Les sequences edito couvrent bien le vocabulaire et l'intention pedagogique de l'objectif."
    if percent >= 45:
        return "Alignement partiel : une partie de l'objectif est couverte, mais certains axes restent peu explicites."

    reasons: list[str] = []
    if not top_matches:
        reasons.append("Aucune correspondance edito suffisamment robuste n'a ete detectee pour cette ligne.")
    else:
        best_score = top_matches[0][0]
        if best_score < 0.2:
            reasons.append("Les titres edito relies sont semantiquement proches mais lexicalement differents du titre cible.")
        total_sequences = sum(item.get("nb_sequences", 0) for _, item in top_matches)
        if total_sequences < 5:
            reasons.append("Le volume de sequences edito exploitees pour cette ligne est faible.")
        objective_tokens = set(_edito_title_core(objective).split())
        match_tokens = set()
        for _, item in top_matches:
            match_tokens.update(_edito_title_core(" ".join(item.get("texts", []))).split())
        missing_ratio = 1.0
        if objective_tokens:
            covered = len(objective_tokens & match_tokens)
            missing_ratio = 1 - (covered / len(objective_tokens))
        if missing_ratio > 0.55:
            reasons.append("L'objectif pedagogique utilise des notions peu presentes explicitement dans les verbatims edito apparies.")
    if not reasons:
        reasons.append("Le recouvrement lexical entre objectif et sequences edito reste limite.")
    return " ".join(reasons)


def _truncate_clean(text: str, limit: int = 220) -> str:
    compact = " ".join((text or "").split())
    if len(compact) <= limit:
        return compact
    return compact[: limit - 1].rstrip() + "…"


def _tb_edito_researcher_summary(code: str, sequences: list[dict]) -> tuple[str, str]:
    if not sequences:
        html = "<span class='meta'>Aucune sequence capsule temoin appariée.</span>"
        return html, "Aucune sequence capsule temoin appariée."

    by_voice: dict[str, list[str]] = defaultdict(list)
    for sequence in sequences:
        voice = (sequence.get("intervenant") or "Temoin").strip()
        text = (sequence.get("texte") or "").strip()
        if text and text not in by_voice[voice]:
            by_voice[voice].append(text)

    corpus_text = _edito_title_core(" ".join(sequence.get("texte", "") for sequence in sequences))
    corpus_tokens = set(corpus_text.split())
    covered_dims, missing_dims = _tb_edito_dimension_coverage(code, corpus_text, corpus_tokens)
    voices = sorted(by_voice)
    voice_label = ", ".join(voices)

    covered_text = "; ".join(covered_dims[:2]) if covered_dims else "le sujet reste encore peu explicite dans les verbatims"
    missing_text = "; ".join(missing_dims[:2]) if missing_dims else "pas de manque majeur au niveau du cadrage sujet"

    html = (
        f"<p class='meta'><strong>{len(sequences)}</strong> sequences capsule temoin appariees, "
        f"<strong>{len(voices)}</strong> temoins mobilises.</p>"
        f"<p><strong>Résumé :</strong> Les chercheurs racontent principalement {escape(covered_text)}.</p>"
        f"<p><strong>Point de vigilance :</strong> {escape(missing_text)}.</p>"
        f"<p class='meta'>Témoins: {escape(voice_label)}.</p>"
    )
    csv_text = (
        f"{len(sequences)} seq., {len(voices)} temoins. "
        f"Resume: {covered_text}. "
        f"Vigilance: {missing_text}. "
        f"Temoins: {voice_label}."
    )
    return html, csv_text


TOPIC_KEYWORDS_BY_CODE = {
    "T1": ["oser", "innovation", "recherche", "utilite", "parcours"],
    "T2": ["besoin", "usage", "probleme", "utilisateur", "valeur"],
    "T3": ["besoin", "terrain", "validation", "preuve", "labo"],
    "T4": ["idee", "freins", "leviers", "doutes", "legitimite"],
    "T5": ["protection", "valorisation", "propriete intellectuelle", "brevet", "secret", "pi"],
    "T6": ["transfert", "licensing", "licence", "startup", "valorisation"],
    "T7": ["ecosysteme", "accompagnement", "maturation", "incubation", "reseau"],
    "T8": ["financement", "concours", "investisseur", "levee", "temps"],
    "T9": ["partenariat", "posture", "equipe", "gouvernance", "competences"],
    "T10": ["langage", "pitch", "communication", "valeur", "interlocuteur"],
    "T11": ["evolution", "metier", "chercheur", "freins", "leviers"],
    "T12": ["collaboration", "accompagnement", "consortium", "partenariat", "contrat"],
    "T13": ["conclusion", "action", "conseil", "engagement", "message"],
}

ALIGNMENT_DIMENSIONS_BY_CODE = {
    "T1": [
        {
            "label": "leurs motivations pour innover et le sens qu'ils donnent a leur demarche",
            "keywords": ["innovation", "innover", "motivation", "envie", "utile", "utilite", "pourquoi"],
        },
        {
            "label": "leur parcours de recherche comme point de depart de l'innovation",
            "keywords": ["recherche", "parcours", "these", "postdoc", "travaux", "laboratoire"],
        },
    ],
    "T2": [
        {
            "label": "la necessite d'aller voir hors du laboratoire pour comprendre le besoin reel",
            "keywords": ["besoin", "usage", "marche", "acteur", "acteurs", "terrain", "exterieur", "utilisateur"],
        },
        {
            "label": "la facon de qualifier le probleme concret a resoudre",
            "keywords": ["probleme", "valeur", "benefice", "client", "situation"],
        },
        {
            "label": "des methodes explicites de validation (tests, hypotheses, indicateurs)",
            "keywords": ["methode", "hypothese", "test", "preuve", "valider", "indicateur", "protocole"],
        },
    ],
    "T3": [
        {
            "label": "la validation terrain du besoin et les echanges avec des acteurs externes",
            "keywords": ["terrain", "besoin", "echange", "acteurs", "marche", "etude"],
        },
        {
            "label": "la sortie du laboratoire pour confronter l'idee aux usages",
            "keywords": ["laboratoire", "labo", "sortir", "usage", "rencontre"],
        },
    ],
    "T4": [
        {
            "label": "les freins concrets rencontres pour transformer une idee en projet viable",
            "keywords": ["freins", "difficile", "obstacle", "pivot", "risque"],
        },
        {
            "label": "les leviers mobilises pour avancer malgre les incertitudes",
            "keywords": ["levier", "accompagnement", "choix", "strategie", "decision"],
        },
    ],
    "T5": [
        {
            "label": "les enjeux de propriete intellectuelle (brevet, secret, divulgation)",
            "keywords": ["propriete intellectuelle", "brevet", "secret", "divulgable", "divulgation", "pi"],
        },
        {
            "label": "la logique de valorisation et de protection en amont des publications",
            "keywords": ["valorisation", "protection", "publier", "publication", "strategie"],
        },
    ],
    "T6": [
        {
            "label": "les options de transfert et de mise en marche de la valorisation (licence, startup, partenariats)",
            "keywords": ["transfert", "licence", "licensing", "startup", "partenariat", "valorisation"],
        },
        {
            "label": "les choix entre plusieurs voies selon le projet et son niveau de maturite",
            "keywords": ["choix", "maturite", "voie", "strategie", "decision"],
        },
    ],
    "T7": [
        {
            "label": "l'importance de l'ecosysteme d'accompagnement et des relais autour du projet",
            "keywords": ["ecosysteme", "accompagnement", "incubateur", "reseau", "satt", "aide"],
        },
        {
            "label": "le besoin de competences complementaires pour ne pas avancer seul",
            "keywords": ["equipe", "competence", "management", "complementaire", "seul"],
        },
    ],
    "T8": [
        {
            "label": "les besoins de financement a differents moments du projet",
            "keywords": ["financement", "aides", "argent", "budget", "concours", "laureat"],
        },
        {
            "label": "l'articulation entre financement et progression concrète du projet",
            "keywords": ["recruter", "etape", "maturation", "jalon", "developpement", "temps"],
        },
    ],
    "T9": [
        {
            "label": "la construction d'une equipe et de partenariats pour porter un projet ambitieux",
            "keywords": ["equipe", "partenariat", "entourage", "collaboration", "competence"],
        },
        {
            "label": "la posture et le cadre relationnel (roles, contrats, gouvernance)",
            "keywords": ["posture", "contrat", "gouvernance", "role", "fondateur"],
        },
    ],
    "T10": [
        {
            "label": "la necessite d'adapter son langage selon les interlocuteurs",
            "keywords": ["langage", "vocabulaire", "interlocuteur", "adapter", "posture"],
        },
        {
            "label": "le passage d'un discours academique a un discours de valeur et d'usage",
            "keywords": ["valeur", "usage", "entreprise", "pitch", "communication"],
        },
    ],
    "T11": [
        {
            "label": "l'evolution du metier de chercheur vers des roles hybrides",
            "keywords": ["evolution", "metier", "chercheur", "posture", "transformation"],
        },
        {
            "label": "les freins et leviers personnels pour passer a l'action",
            "keywords": ["freins", "leviers", "temps", "legitimite", "engagement"],
        },
    ],
    "T12": [
        {
            "label": "le role des collaborations et consortiums pour amplifer le projet",
            "keywords": ["collaboration", "consortium", "partenariat", "collectif", "amplifie"],
        },
        {
            "label": "les conditions pour que ces collaborations fonctionnent (cadre, contrats, preparation)",
            "keywords": ["contrat", "cadre", "preparation", "regles", "gouvernance"],
        },
    ],
    "T13": [
        {
            "label": "la capacite a conclure avec des actions concretes et progressives",
            "keywords": ["conclusion", "action", "premier pas", "engagement", "conseil"],
        },
        {
            "label": "le message cle a transmettre pour oser passer a l'action",
            "keywords": ["message", "conseil", "oser", "innovation", "idee recue"],
        },
    ],
}


def _topic_keyword_covered(keyword: str, corpus_text: str, corpus_tokens: set[str]) -> bool:
    normalized = _edito_title_core(keyword)
    if not normalized:
        return False
    if " " in normalized:
        return normalized in corpus_text
    return normalized in corpus_tokens


def _tb_edito_subject_alignment_percent(code: str, sequences: list[dict]) -> int:
    keywords = TOPIC_KEYWORDS_BY_CODE.get(code, [])
    if not keywords or not sequences:
        return 0
    corpus_text = _edito_title_core(" ".join(sequence.get("texte", "") for sequence in sequences))
    corpus_tokens = set(corpus_text.split())
    covered = sum(1 for keyword in keywords if _topic_keyword_covered(keyword, corpus_text, corpus_tokens))
    return round((covered / len(keywords)) * 100)


def _tb_edito_dimension_coverage(code: str, corpus_text: str, corpus_tokens: set[str]) -> tuple[list[str], list[str]]:
    dims = ALIGNMENT_DIMENSIONS_BY_CODE.get(code, [])
    if not dims:
        return [], []
    covered: list[str] = []
    missing: list[str] = []
    for dim in dims:
        keywords = dim.get("keywords", [])
        is_hit = any(_topic_keyword_covered(str(keyword), corpus_text, corpus_tokens) for keyword in keywords)
        label = str(dim.get("label", ""))
        if is_hit:
            covered.append(label)
        else:
            missing.append(label)
    return covered, missing


def _tb_edito_alignment_comment(code: str, percent: int, seq_count: int, sequences: list[dict]) -> str:
    if seq_count == 0:
        return "Aucune sequence capsule temoin exploitable pour estimer l'alignement sujet."

    corpus_text = _edito_title_core(" ".join(sequence.get("texte", "") for sequence in sequences))
    corpus_tokens = set(corpus_text.split())
    covered_dims, missing_dims = _tb_edito_dimension_coverage(code, corpus_text, corpus_tokens)

    if percent >= 70:
        level_intro = "Alignement sujet fort"
    elif percent >= 45:
        level_intro = "Alignement sujet moyen"
    else:
        level_intro = "Alignement sujet faible"

    if covered_dims:
        concrete_part = f"les temoins abordent clairement {covered_dims[0]}"
    else:
        concrete_part = "les temoins restent generaux sur le sujet"

    if missing_dims:
        gap_part = f"mais ils explicitent peu {missing_dims[0]}"
    elif percent < 45:
        gap_part = "mais la profondeur de traitement reste encore partielle"
    else:
        gap_part = "et les principaux axes attendus apparaissent globalement couverts"

    return f"{level_intro}: {concrete_part}, {gap_part}."


def _tb_edito_coverage_gap(code: str, sequences: list[dict]) -> tuple[str, str]:
    keywords = TOPIC_KEYWORDS_BY_CODE.get(code, [])
    if not keywords:
        html = "<span class='meta'>Referentiel sujet non defini.</span>"
        return html, "Referentiel sujet non defini."

    corpus_text = _edito_title_core(" ".join(sequence.get("texte", "") for sequence in sequences))
    corpus_tokens = set(corpus_text.split())
    covered = [keyword for keyword in keywords if _topic_keyword_covered(keyword, corpus_text, corpus_tokens)]
    missing = [keyword for keyword in keywords if keyword not in covered]
    covered_label = ", ".join(covered) if covered else "Aucun axe sujet explicite detecte"
    missing_label = ", ".join(missing) if missing else "Aucun axe sujet manquant majeur detecte"
    html = (
        f"<p><strong>Abordé :</strong> {escape(covered_label)}</p>"
        f"<p><strong>A développer :</strong> {escape(missing_label)}</p>"
    )
    csv_text = f"Abordé: {covered_label} | A développer: {missing_label}"
    return html, csv_text


def _tb_edito_expertise_preconisation(
    code: str,
    sequences: list[dict],
    videos_expert: list[dict],
    row_match_percent: int,
) -> tuple[str, str]:
    if not videos_expert:
        html = (
            "<p><strong>Préconisation :</strong> Définir au moins une vidéo expertise dédiée à ce sujet.</p>"
            "<p class='meta'>Sans vidéo expert, le passage de l'amorce témoin à l'outillage pédagogique reste incomplet.</p>"
        )
        csv_text = "Definir une video expertise dediee avant arbitrage final."
        return html, csv_text

    corpus_text = _edito_title_core(" ".join(sequence.get("texte", "") for sequence in sequences))
    corpus_tokens = set(corpus_text.split())
    covered_dims, missing_dims = _tb_edito_dimension_coverage(code, corpus_text, corpus_tokens)
    expert_targets = ", ".join(_label_video_expert(item.get("code", "")) for item in videos_expert[:2])

    if not sequences:
        action = (
            "Poser les fondamentaux du sujet, definir le vocabulaire de reference et proposer une methode pas-a-pas."
        )
    elif missing_dims:
        action = (
            f"Prioriser {missing_dims[0]}, puis structurer un cadre operatoire concret (methode, criteres, points de decision)."
        )
    elif row_match_percent < 70:
        action = (
            "Transformer les retours temoins en methode transmissible (etapes, outils, points de vigilance)."
        )
    else:
        action = (
            "Consolider les acquis avec des cas d'application, des erreurs frequentes et des reperes de mise en oeuvre."
        )

    covered_hint = covered_dims[0] if covered_dims else "l'amorce temoin existe mais reste diffuse"
    html = (
        f"<p><strong>Cible expert ({escape(expert_targets)}) :</strong> {escape(action)}</p>"
        f"<p class='meta'>Point d'appui temoins : {escape(covered_hint)}.</p>"
    )
    csv_text = f"Cible expert ({expert_targets}): {action} | Point d'appui temoins: {covered_hint}."
    return html, csv_text


TITLE_RULES_BY_CODE = {
    "T1": [
        ("pourquoi oser", "Pourquoi oser ? De la recherche a l'envie d'innover"),
        ("innovation", "Oser innover : des parcours qui creent de la valeur"),
    ],
    "T2": [
        ("acteurs", "Du laboratoire au terrain : comprendre le besoin reel"),
        ("besoin", "De la recherche a l'innovation : clarifier le besoin"),
    ],
    "T3": [
        ("terrain", "Valider sur le terrain : sortir du laboratoire"),
        ("preuve", "Du laboratoire au reel : eprouver la solution"),
    ],
    "T4": [
        ("frein", "Une idee ne suffit pas : freins, leviers et decisions"),
        ("levier", "Transformer l'idee en projet : freins et leviers"),
    ],
    "T5": [
        ("secret", "Proteger et valoriser : arbitrer entre brevet et secret"),
        ("brevet", "Protection et valorisation : choisir la bonne strategie PI"),
    ],
    "T6": [
        ("licence", "Du transfert a l'impact : choisir la bonne voie"),
        ("startup", "Transfert et licensing : de la recherche au marche"),
    ],
    "T7": [
        ("ecosysteme", "Vous n'etes pas seul : activer l'ecosysteme d'accompagnement"),
        ("accompagnement", "S'entourer pour avancer : maturation, incubation, reseau"),
    ],
    "T8": [
        ("investisseur", "Financer l'innovation : choisir les bons partenaires"),
        ("financement", "Financements et concours : financer au bon rythme"),
    ],
    "T9": [
        ("equipe", "Construire l'equipe : partnerships et posture entrepreneuriale"),
        ("partenariat", "Partenariats et equipe : trouver la bonne complementarite"),
    ],
    "T10": [
        ("langage", "Changer de langage : rendre la science desirable"),
        ("vocabulaire", "Enrichir son langage : de la preuve a la valeur"),
    ],
    "T11": [
        ("frein", "Chercheur et entrepreneur : depasser les freins pour agir"),
        ("metier", "Faire evoluer son metier de chercheur vers l'innovation"),
    ],
    "T12": [
        ("action", "Passer a l'action : de l'intention a la collaboration"),
        ("collaboration", "Conclusion : structurer les premieres collaborations"),
    ],
}


def _proposed_temoin_title(code: str, sequences: list[dict], fallback_label: str) -> tuple[str, str]:
    corpus = _edito_title_core(" ".join(sequence.get("texte", "") for sequence in sequences))
    covered_keywords = [
        keyword
        for keyword in TOPIC_KEYWORDS_BY_CODE.get(code, [])
        if _topic_keyword_covered(keyword, corpus, set(corpus.split()))
    ]
    for trigger, title in TITLE_RULES_BY_CODE.get(code, []):
        if trigger in corpus:
            if covered_keywords:
                justification = f"Indices reperes dans les selections: {', '.join(covered_keywords[:3])}."
            else:
                justification = "Formulation alignee avec les themes explicites dans les selections retenues."
            return title, justification

    dims, _ = _tb_edito_dimension_coverage(code, corpus, set(corpus.split()))
    if dims:
        return (
            fallback_label,
            f"Le titre actuel reste pertinent au regard des selections; point saillant detecte: {dims[0]}.",
        )
    return (
        fallback_label,
        "Le titre actuel est conserve faute de marqueur suffisamment net dans les selections.",
    )


def build_proposition_titres_temoin_page(programme_table: dict) -> None:
    rows_by_code = {row.get("code", ""): row for row in programme_table.get("rows", [])}
    grouped = _tb_edito_sequences_by_code()
    rows_html = []
    title_rows: list[dict] = []

    for code, spec in sorted(FIXED_TEMOIN_PLAN.items(), key=lambda item: int(item[0][1:])):
        current_title = spec.get("label", _label_video_temoin(code))
        sequences = _tb_edito_order_for_code(code, grouped.get(code, []))
        objective = rows_by_code.get(code, {}).get("objectif_pedagogique", "")
        proposed_title, justification = _proposed_temoin_title(code, sequences, current_title)
        corpus = _edito_title_core(" ".join(seq.get("texte", "") for seq in sequences))
        corpus_tokens = set(corpus.split())
        markers = [
            keyword
            for keyword in TOPIC_KEYWORDS_BY_CODE.get(code, [])
            if _topic_keyword_covered(keyword, corpus, corpus_tokens)
        ]
        markers_label = ", ".join(markers[:5]) if markers else "aucun marqueur lexical fort detecte"
        title_rows.append(
            {
                "code": code,
                "titre_actuel": current_title,
                "titre_propose": proposed_title,
                "objectif": objective,
                "nb_extraits": len(sequences),
                "marqueurs": markers_label,
                "justification": justification,
            }
        )
        rows_html.append(
            "<tr>"
            f"<td><a href='tb_edito_{escape(code)}.html'>{escape(code)}</a></td>"
            f"<td>{escape(current_title)}</td>"
            f"<td><strong>{escape(proposed_title)}</strong></td>"
            f"<td>{escape(objective)}</td>"
            f"<td>{len(sequences)}</td>"
            f"<td>{escape(markers_label)}</td>"
            f"<td>{escape(justification)}</td>"
            "</tr>"
        )

    doc_table_rows = []
    doc_title_items = []
    for row in title_rows:
        doc_table_rows.append(
            "<tr>"
            f"<td>{escape(row['code'])}</td>"
            f"<td>{escape(row['titre_actuel'])}</td>"
            f"<td>{escape(row['titre_propose'])}</td>"
            f"<td>{escape(row['objectif'])}</td>"
            f"<td>{escape(str(row['nb_extraits']))}</td>"
            f"<td>{escape(row['marqueurs'])}</td>"
            f"<td>{escape(row['justification'])}</td>"
            "</tr>"
        )
        doc_title_items.append(
            f"<li><strong>{escape(row['code'])}</strong> — {escape(row['titre_propose'])}</li>"
        )
    doc_empty_row = "<tr><td colspan='7'>Aucune capsule disponible.</td></tr>"
    doc_html = (
        "<html><head><meta charset='utf-8'>"
        "<style>"
        "body{font-family:Aptos,Segoe UI,Arial,sans-serif;font-size:11.5pt;line-height:1.5;}"
        "h1{font-size:18pt;margin-bottom:6px;} h2{font-size:14pt;margin:16px 0 8px;}"
        "table{width:100%;border-collapse:collapse;font-size:10.5pt;}"
        "th,td{border:1px solid #cbd5e1;padding:6px 8px;vertical-align:top;}"
        "th{background:#f1f5f9;text-align:left;} ul{margin:8px 0 0 18px;}"
        "</style></head><body>"
        "<h1>Proposition des titres témoins</h1>"
        "<h2>Tableau des titres proposés</h2>"
        "<table><thead><tr>"
        "<th>Capsule</th><th>Titre actuel</th><th>Titre propose</th><th>Objectif video temoin</th><th>Extraits retenus</th><th>Marqueurs reperes</th><th>Justification</th>"
        "</tr></thead><tbody>"
        f"{''.join(doc_table_rows) if doc_table_rows else doc_empty_row}"
        "</tbody></table>"
        "<h2>Liste des titres proposes uniquement</h2>"
        f"<ul>{''.join(doc_title_items) if doc_title_items else '<li>Aucun titre propose.</li>'}</ul>"
        "</body></html>"
    )
    write_text(SITE / "proposition_titres_temoin.doc", doc_html)

    body = (
        "<p class='meta'>Propositions de titres pour chaque video temoin, basees sur le sujet capsule, l'objectif pedagogique et les selections edito retenues.</p>"
        "<p><a class='btn' href='proposition_titres_temoin.doc' download>Exporter les titres témoins (Word)</a></p>"
        "<div class='table-wrap'><table><thead><tr>"
        "<th>Capsule</th><th>Titre actuel</th><th>Titre propose</th><th>Objectif video temoin</th><th>Extraits retenus</th><th>Marqueurs reperes</th><th>Justification</th>"
        "</tr></thead><tbody>"
        + ("".join(rows_html) if rows_html else "<tr><td colspan='7'>Aucune capsule disponible.</td></tr>")
        + "</tbody></table></div>"
    )
    write_text(
        SITE / "proposition_titres_temoin.html",
        html_page(
            "Titres témoins",
            body,
            nav_current="fichiers_travail.html",
            breadcrumb=html_breadcrumb(("Accueil", "index.html"), ("Fichiers de travail", "fichiers_travail.html"), ("Informations", "informations.html"), ("Titres témoins", None)),
            page_header='<div class="page-head"><h1>Titres témoins — Propositions</h1><p class="lead">Un titre proposé par capsule témoin, selon le sujet et le texte effectivement sélectionné.</p></div>',
        ),
    )


def _clean_temoin_intro_text(text: str, intervenant: str) -> str:
    value = " ".join((text or "").split())
    if not value:
        return ""
    value = re.sub(rf"^{re.escape(intervenant)}\s+", "", value, flags=re.IGNORECASE)
    value = re.sub(r"^je m['’]appelle\s+[^,]+,\s*", "", value, flags=re.IGNORECASE)
    value = value.strip(" .")
    if value:
        value = value[0].upper() + value[1:]
    return value


def _extract_temoin_function(doc: dict) -> dict:
    intervenant = (doc.get("intervenant") or "").strip()
    sequences = sorted(
        [
            seq
            for seq in doc.get("sequences", [])
            if seq.get("statut_edito") == "RETENU_PAR_EDITO"
        ],
        key=lambda item: int(item.get("ordre", 0) or 0),
    )
    patterns = (
        r"\bje suis\b",
        r"\bje travaille\b",
        r"\benseignant[-\s]?chercheur\b",
        r"\bprofesseur\b",
    )
    selected = None
    for sequence in sequences:
        text = (sequence.get("texte") or "").strip()
        if not text:
            continue
        normalized = _normalize_for_match(text)
        if any(re.search(pattern, normalized) for pattern in patterns):
            selected = sequence
            break
    if selected is None and sequences:
        selected = sequences[0]

    officiel = (selected or {}).get("texte", "").strip() if selected else ""
    formulation = _clean_temoin_intro_text(officiel, intervenant) if officiel else "Information non explicite dans les selections retenues."
    reference = (
        f"{selected.get('id', '')} (paragraphe {selected.get('source_paragraphe', '-')})"
        if selected
        else "Non disponible"
    )
    fonction_academique = formulation
    fonction_professionnelle = "Non explicite dans les selections retenues."
    officiel_entreprise = ""

    corpus = " ".join((seq.get("texte") or "") for seq in sequences)
    normalized_corpus = _normalize_for_match(corpus)
    if "co-fondatrice" in normalized_corpus or "cofondatrice" in normalized_corpus:
        fonction_professionnelle = "Co-fondatrice d'entreprise."
        officiel_entreprise = "Mention explicite de co-fondatrice dans les verbatims retenus."
    elif "co-fondateur" in normalized_corpus or "cofondateur" in normalized_corpus:
        fonction_professionnelle = "Co-fondateur d'entreprise."
        officiel_entreprise = "Mention explicite de co-fondateur dans les verbatims retenus."
    elif "cso" in normalized_corpus:
        fonction_professionnelle = "CSO d'entreprise."
        officiel_entreprise = "Mention explicite de CSO dans les verbatims retenus."

    return {
        "temoin": intervenant,
        "source": doc.get("source", ""),
        "fonction_academique": fonction_academique,
        "fonction_professionnelle": fonction_professionnelle,
        "fonction_officielle_academique": officiel or "Non disponible dans les selections retenues.",
        "fonction_officielle_professionnelle": officiel_entreprise or "Non disponible dans les selections retenues.",
        "reference_academique": reference,
        "reference_professionnelle": "A preciser",
    }


def build_fonctions_temoins_page() -> None:
    rows = []
    export_rows = []
    for item in load_derushage_edito_index():
        doc = load_derushage_edito(item.get("id", ""))
        if not doc:
            continue
        rec = _extract_temoin_function(doc)
        if rec["temoin"] == "Jean-Jacques Greffet":
            rec["fonction_academique"] = "Professeur émérite."
            rec["fonction_professionnelle"] = "CSO de UNVEIL."
            rec["fonction_officielle_professionnelle"] = "Information fournie pour cadrage éditorial (verification documentaire complémentaire recommandée)."
            rec["reference_professionnelle"] = "A confirmer dans BAB brut complet"
        elif rec["temoin"] == "Muriel Thomas":
            rec["fonction_academique"] = "Chargée de recherche à l'INRAE."
            rec["fonction_professionnelle"] = "Co-fondatrice de Carembouche."
            rec["fonction_officielle_professionnelle"] = "Mentions BAB : creation de start-up et references a Carembouche/Car en bouche."
            rec["reference_professionnelle"] = "MUR-0008 (BAB encode)"
        elif rec["temoin"] == "Loïc Rajjou":
            rec["fonction_academique"] = "Professeur à AgroParisTech."
            rec["fonction_professionnelle"] = "Co-fondateur de son entreprise (nom a confirmer)."
            rec["fonction_officielle_professionnelle"] = "Mentions BAB : co-fondateur et role au comite scientifique de l'entreprise."
            rec["reference_professionnelle"] = "LOI-0011 / LOI-0014 (BAB encode)"

        rows.append(
            "<tr>"
            f"<td>{escape(rec['temoin'])}</td>"
            f"<td>{escape(rec['fonction_academique'])}</td>"
            f"<td>{escape(rec['fonction_professionnelle'])}</td>"
            f"<td>{escape(rec['fonction_officielle_academique'])}</td>"
            f"<td>{escape(rec['fonction_officielle_professionnelle'])}</td>"
            f"<td>{escape(rec['reference_academique'])}</td>"
            f"<td>{escape(rec['reference_professionnelle'])}</td>"
            f"<td>{escape(rec['source'])}</td>"
            "</tr>"
        )
        export_rows.append(rec)

    csv_buffer = io.StringIO()
    writer = csv.DictWriter(
        csv_buffer,
        fieldnames=[
            "temoin",
            "fonction_academique",
            "fonction_professionnelle",
            "fonction_officielle_academique",
            "fonction_officielle_professionnelle",
            "reference_academique",
            "reference_professionnelle",
            "source",
        ],
    )
    writer.writeheader()
    writer.writerows(export_rows)
    write_text(SITE / "fonctions_temoins.csv", csv_buffer.getvalue())

    json_payload = {"temoins": export_rows}
    write_text(
        SITE / "fonctions_temoins.json",
        json.dumps(json_payload, ensure_ascii=False, indent=2),
    )

    body = (
        "<p class='meta'>Fonctions academiques et professionnelles des temoins, avec formulations officielles basees sur les sequences retenues par l'edito.</p>"
        "<p><a class='btn' href='fonctions_temoins.csv' download>Télécharger les fonctions témoins (CSV)</a> "
        "<a class='btn' href='fonctions_temoins.json' download>Télécharger les fonctions témoins (JSON)</a></p>"
        "<div class='table-wrap'><table><thead><tr>"
        "<th>Témoin</th><th>Fonction académique</th><th>Fonction professionnelle</th><th>Verbatim officiel académique</th><th>Verbatim officiel professionnel</th><th>Réf. académique</th><th>Réf. professionnelle</th><th>Source</th>"
        "</tr></thead><tbody>"
        + ("".join(rows) if rows else "<tr><td colspan='8'>Aucune donnée témoin disponible.</td></tr>")
        + "</tbody></table></div>"
    )
    write_text(
        SITE / "fonctions_temoins.html",
        html_page(
            "Fonctions témoins",
            body,
            nav_current="fichiers_travail.html",
            breadcrumb=html_breadcrumb(("Accueil", "index.html"), ("Fichiers de travail", "fichiers_travail.html"), ("Informations", "informations.html"), ("Fonctions témoins", None)),
            page_header='<div class="page-head"><h1>Fonctions témoins</h1><p class="lead">Version academique et professionnelle des fonctions des temoins, avec export.</p></div>',
        ),
    )


PUNCHLINE_HINTS = (
    "il faut",
    "point cle",
    "on s est apercu",
    "ça a change",
    "a change",
    "risque",
    "important",
    "obstacle",
    "levier",
    "pivot",
    "retenez",
    "premier pas",
    "foncez",
    "oser",
    "conseil",
    "impossible",
    "protege",
    "sortir du labo",
)


def _all_clarisse_retained_texts() -> list[str]:
    texts: list[str] = []
    for item in load_derushage_edito_index():
        doc = load_derushage_edito(item.get("id", ""))
        if not doc:
            continue
        for sequence in doc.get("sequences", []):
            if sequence.get("statut_edito") and sequence.get("statut_edito") != "RETENU_PAR_EDITO":
                continue
            text = _edito_title_core(sequence.get("texte", ""))
            if text:
                texts.append(text)
    return list(dict.fromkeys(texts))


def _clarisse_exclusion_corpus(selected_sequences: list[dict], global_texts: list[str] | None = None) -> tuple[str, set[str], list[str]]:
    """Corpus Clarisse pour exclure les passages deja retenus par l'edito."""
    selected_texts: list[str] = []
    for sequence in selected_sequences:
        raw = (sequence.get("texte") or "").strip()
        if not raw:
            continue
        selected_texts.append(_edito_title_core(raw))
    texts = list(dict.fromkeys([*selected_texts, *(global_texts or [])] ))
    texts = [t for t in texts if t]
    corpus = " ".join(texts)
    selected_tokens = set(" ".join(selected_texts).split())
    return corpus, selected_tokens, texts


def _already_used_by_clarisse(verbatim: str, clarisse_corpus: str, clarisse_texts: list[str]) -> bool:
    """True si le BAB non encode recouvre un passage deja selectionne par Clarisse."""
    normalized = _edito_title_core(verbatim)
    tokens = normalized.split()
    if len(tokens) < 8:
        return False
    for size in (12, 10, 8):
        if len(tokens) < size:
            continue
        for index in range(0, len(tokens) - size + 1, max(1, size // 2)):
            window = " ".join(tokens[index : index + size])
            if window and window in clarisse_corpus:
                return True
    for clarisse_text in clarisse_texts:
        clarisse_tokens = clarisse_text.split()
        if len(clarisse_tokens) < 10:
            continue
        compact = " ".join(clarisse_tokens)
        if compact in normalized:
            return True
        if len(tokens) <= 40 and normalized in clarisse_text:
            return True
    return False


def _best_matching_question(
    code: str,
    normalized_verbatim: str,
    selected_sequences: list[dict] | None = None,
) -> str:
    questions: list[str] = []
    for sequence in selected_sequences or []:
        question = (sequence.get("question") or "").strip()
        if question and question not in questions:
            questions.append(question)
    for question in FIXED_TEMOIN_PLAN.get(code, {}).get("questions") or []:
        q = str(question).strip()
        if q and q not in questions:
            questions.append(q)
    if not questions:
        return ""
    best = ""
    best_score = 0.0
    verbatim_tokens = set(normalized_verbatim.split())
    for question in questions:
        q_norm = _edito_title_core(question)
        q_tokens = set(q_norm.split())
        if not q_tokens:
            continue
        score = len(verbatim_tokens & q_tokens) / len(q_tokens)
        if score > best_score:
            best_score = score
            best = question.strip()
    if best_score >= 0.08:
        return best
    return questions[0]


def _editorial_ajout_raison(
    code: str,
    *,
    normalized_verbatim: str,
    covered_dims: list[str],
    keyword_hits: int,
    hint_hits: int,
    missing_dims: list[str] | None = None,
    selected_sequences: list[dict] | None = None,
) -> str:
    """Justification editoriale lisible (sans jargon technique interne)."""
    label = FIXED_TEMOIN_PLAN.get(code, {}).get("label", code)
    title = re.sub(r"^vid[eé]o\s*\d+\s*:\s*", "", label, flags=re.IGNORECASE).strip() or label
    question = _best_matching_question(code, normalized_verbatim, selected_sequences)

    parts: list[str] = [
        "Passage non utilise dans aucun edito Clarisse.",
    ]
    if question:
        parts.append(f"Il ressort autour de la question : « {question} ».")
    else:
        parts.append(f"Il ressort hors selection actuelle, sur le sujet de la video « {title} ».")

    if covered_dims:
        apport = covered_dims[0]
        parts.append(f"Pour la video « {title} », il apporte des elements concrets sur {apport}.")
    elif keyword_hits:
        parts.append(
            f"Pour la video « {title} », il complete le propos temoin avec des elements directement lies au sujet."
        )
    else:
        parts.append(
            f"Pour la video « {title} », il ajoute une formulation utile au recit temoin."
        )

    if missing_dims:
        parts.append(
            f"Pourquoi le proposer : il renforce {missing_dims[0]}, encore peu explicite dans la selection Clarisse."
        )
    elif covered_dims:
        parts.append(
            f"Pourquoi le proposer : il renforce {covered_dims[0]} et donne une prise plus nette a l'apprenant."
        )
    elif hint_hits:
        parts.append(
            "Pourquoi le proposer : formulation dynamique qui peut servir de pivot narratif dans le montage."
        )
    else:
        parts.append(
            "Pourquoi le proposer : il eclaircit un point cle du sujet sans redire ce que Clarisse a deja retenu."
        )
    return " ".join(parts)


def _proposition_edito_candidates(
    code: str,
    selected_sequences: list[dict],
    *,
    clarisse_global_texts: list[str] | None = None,
    missing_dims: list[str] | None = None,
) -> list[dict]:
    """Propose des ajouts depuis les BAB NON_ENCODE, hors passages deja retenus par Clarisse."""
    clarisse_corpus, selected_tokens, clarisse_texts = _clarisse_exclusion_corpus(
        selected_sequences,
        clarisse_global_texts,
    )

    keywords = TOPIC_KEYWORDS_BY_CODE.get(code, [])
    dimensions = ALIGNMENT_DIMENSIONS_BY_CODE.get(code, [])
    candidates: list[dict] = []

    for item in load_bab_encode_index():
        doc = load_bab_encode(item["id"])
        if not doc:
            continue
        chercheur = doc.get("chercheur", item.get("chercheur", ""))
        source = doc.get("source", item.get("source", ""))
        for bloc in merge_bab_encode_blocs(doc):
            if bloc.get("encodage") != "NON_ENCODE":
                continue
            verbatim = (bloc.get("verbatim") or "").strip()
            if not verbatim:
                continue
            word_count = len(verbatim.split())
            if word_count < 12 or word_count > 140:
                continue
            if _already_used_by_clarisse(verbatim, clarisse_corpus, clarisse_texts):
                continue

            normalized = _edito_title_core(verbatim)
            tokens = set(normalized.split())
            if not tokens:
                continue

            keyword_hits = sum(1 for keyword in keywords if _topic_keyword_covered(keyword, normalized, tokens))
            covered_dims = []
            for dim in dimensions:
                dim_keywords = dim.get("keywords", [])
                if any(_topic_keyword_covered(str(keyword), normalized, tokens) for keyword in dim_keywords):
                    covered_dims.append(str(dim.get("label", "")))
            hint_hits = sum(1 for hint in PUNCHLINE_HINTS if hint in normalized)
            overlap_ratio = len(tokens & selected_tokens) / max(1, len(tokens))
            score = (
                keyword_hits * 1.4
                + len(covered_dims) * 1.6
                + hint_hits * 0.8
                + (0.6 if 18 <= word_count <= 90 else 0.0)
                - overlap_ratio * 1.5
            )
            if score < 2.2:
                continue
            if keyword_hits == 0 and not covered_dims and hint_hits < 2:
                continue

            raison = _editorial_ajout_raison(
                code,
                normalized_verbatim=normalized,
                covered_dims=covered_dims,
                keyword_hits=keyword_hits,
                hint_hits=hint_hits,
                missing_dims=missing_dims,
                selected_sequences=selected_sequences,
            )

            first_name = (chercheur.split()[0] if chercheur else "X")
            bloc_id = f"NONENC-{code}-{first_name}-{bloc.get('numero', '?')}"
            candidates.append(
                {
                    "id": bloc_id,
                    "chercheur": chercheur,
                    "source": source,
                    "debut": bloc.get("debut", ""),
                    "fin": bloc.get("fin", ""),
                    "verbatim": verbatim,
                    "score": round(score, 2),
                    "raison": raison,
                    "statut": "NON_ENCODE",
                }
            )

    candidates.sort(key=lambda item: (-item["score"], item.get("debut", ""), item["id"]))
    seen_keys: set[str] = set()
    deduped: list[dict] = []
    for item in candidates:
        key = f"{item.get('chercheur','')}|{item.get('debut','')}|{item.get('fin','')}"
        if key in seen_keys:
            continue
        seen_keys.add(key)
        deduped.append(item)
    return deduped


def build_proposition_edito_pages(programme_table: dict) -> None:
    rows_by_code = {row.get("code", ""): row for row in programme_table.get("rows", [])}
    grouped = _tb_edito_sequences_by_code()
    clarisse_global_texts = _all_clarisse_retained_texts()
    summary_rows = []
    expected: set[str] = set()

    for code, spec in sorted(FIXED_TEMOIN_PLAN.items(), key=lambda item: int(item[0][1:])):
        row = rows_by_code.get(code, {})
        objective = row.get("objectif_pedagogique", "")
        selected_sequences = _tb_edito_order_for_code(code, grouped.get(code, []))
        voices = sorted(
            {
                (sequence.get("intervenant") or "").strip()
                for sequence in selected_sequences
                if (sequence.get("intervenant") or "").strip()
            }
        )
        corpus_text = _edito_title_core(" ".join(sequence.get("texte", "") for sequence in selected_sequences))
        corpus_tokens = set(corpus_text.split())
        covered_dims, missing_dims = _tb_edito_dimension_coverage(code, corpus_text, corpus_tokens)
        align_percent = _tb_edito_subject_alignment_percent(code, selected_sequences)

        candidates = _proposition_edito_candidates(
            code,
            selected_sequences,
            clarisse_global_texts=clarisse_global_texts,
            missing_dims=missing_dims,
        )
        # On ne propose des NON_ENCODE que s'ils apportent vraiment quelque chose.
        if align_percent >= 75 and not missing_dims:
            candidates = []
        else:
            candidates = candidates[:3]

        if not candidates:
            if align_percent >= 70 and not missing_dims:
                verdict = "VALIDER EN L'ETAT"
                validation_msg = "Le script final temoin couvre bien le sujet et l'objectif de la video. Aucun ajout n'est necessaire a ce stade."
            elif align_percent >= 55:
                verdict = "VALIDER EN L'ETAT"
                validation_msg = "La proposition Clarisse est globalement solide et aucun passage complementaire pertinent n'est detecte hors selection edito."
            else:
                verdict = "A CONSOLIDER"
                validation_msg = "Le script final temoin presente encore des zones a clarifier, sans passage complementaire pertinent detecte hors selection Clarisse."
            add_block = "<p class='meta'>Aucun ajout propose : pas de passage inedit pertinent au-dela des selections Clarisse.</p>"
            add_count = 0
        else:
            if align_percent >= 55:
                verdict = "VALIDER AVEC AJOUT OPTIONNEL"
                validation_msg = "La proposition Clarisse est globalement solide. Les passages ci-dessous peuvent renforcer la dynamique."
            else:
                verdict = "A CONSOLIDER AVEC AJOUT"
                validation_msg = "Le script final temoin gagnerait a etre renforce. Les passages ci-dessous sont recommandes car absents des selections Clarisse."
            add_count = len(candidates)
            cards = []
            for candidate in candidates:
                cards.append(
                    "<article class='card'>"
                    f"<p><strong>{escape(candidate.get('chercheur', ''))}</strong>"
                    f" <span class='meta'>{escape(candidate.get('debut', ''))} → {escape(candidate.get('fin', ''))}</span></p>"
                    f"<p class='meta'>{escape(candidate.get('source', ''))}</p>"
                    f"<p><strong>Pourquoi proposer cet ajout :</strong> {escape(candidate.get('raison', ''))}</p>"
                    f"<p><strong>Extrait :</strong> {escape(_truncate_clean(candidate.get('verbatim', ''), 360))}</p>"
                    "</article>"
                )
            add_block = "".join(cards)

        covered_label = covered_dims[0] if covered_dims else "aucun axe clairement explicite"
        missing_label = missing_dims[0] if missing_dims else "pas de manque majeur detecte"
        detail_body = (
            f"<p><strong>Capsule témoin :</strong> {escape(code)} — {escape(spec.get('label', ''))}</p>"
            f"<p><strong>Objectif vidéo témoin :</strong> {escape(objective)}</p>"
            f"<p><strong>Script final Clarisse :</strong> {len(selected_sequences)} extraits retenus · {len(voices)} voix mobilisées.</p>"
            "<h2>Validation de la proposition Clarisse</h2>"
            f"<p><strong>Verdict :</strong> {escape(verdict)}</p>"
            f"<p>{escape(validation_msg)}</p>"
            f"<p><strong>Point couvert :</strong> {escape(covered_label)}</p>"
            f"<p><strong>Point de vigilance :</strong> {escape(missing_label)}</p>"
            "<h2>Ajouts proposés (passages non retenus par Clarisse)</h2>"
            "<p class='meta'>Chaque proposition est absente des éditos Clarisse et justifiée par son apport au sujet de la vidéo témoin.</p>"
            f"{add_block}"
        )

        page_name = f"proposition_edito_{code}.html"
        expected.add(page_name)
        write_text(
            SITE / page_name,
            html_page(
                f"Proposition édito — {code}",
                detail_body,
                nav_current="edito.html",
                breadcrumb=html_breadcrumb(
                    ("Accueil", "index.html"),
                    ("Edito", "edito.html"),
                    ("Proposition édito", "proposition_edito.html"),
                    (code, None),
                ),
                page_header=(
                    "<div class='page-head'>"
                    f"<h1>Proposition édito — {escape(code)}</h1>"
                    f"<p class='lead'>{escape(spec.get('label', ''))}</p>"
                    "</div>"
                ),
            ),
        )

        summary_rows.append(
            "<tr>"
            f"<td><a href='{escape(page_name)}'>{escape(code)}</a></td>"
            f"<td>{escape(spec.get('label', ''))}</td>"
            f"<td>{escape(verdict)}</td>"
            f"<td>{align_percent}%</td>"
            f"<td>{add_count}</td>"
            "</tr>"
        )

    index_body = (
        "<p class='meta'>Validation des scripts finaux temoins (selection Clarisse) et proposition d'ajouts BAB uniquement lorsqu'un gain editorial est detecte.</p>"
        "<div class='table-wrap'><table><thead><tr>"
        "<th>Capsule</th><th>Vidéo témoin</th><th>Validation Clarisse</th><th>Alignement sujet</th><th>Ajouts proposes</th>"
        "</tr></thead><tbody>"
        + ("".join(summary_rows) if summary_rows else "<tr><td colspan='5'>Aucune capsule disponible.</td></tr>")
        + "</tbody></table></div>"
    )
    write_text(
        SITE / "proposition_edito.html",
        html_page(
            "Proposition édito",
            index_body,
            nav_current="edito.html",
            breadcrumb=html_breadcrumb(("Accueil", "index.html"), ("Edito", "edito.html"), ("Proposition édito", None)),
            page_header='<div class="page-head"><h1>Proposition édito</h1><p class="lead">Validation des choix Clarisse + compléments BAB potentiels quand ils apportent une vraie dynamique.</p></div>',
        ),
    )

    for path in SITE.glob("proposition_edito_T*.html"):
        if path.name not in expected:
            path.unlink()


def _expert_org_from_profile(profile: dict | None) -> str:
    if not profile:
        return "Organisme de rattachement à confirmer"
    corpus = _normalize_for_match(
        " ".join(
            [
                profile.get("profil_cible", ""),
                " ".join(profile.get("infos", [])),
                " ".join(profile.get("mots_cles", [])),
            ]
        )
    )
    if "inpi" in corpus:
        return "INPI"
    if "incuballiance" in corpus:
        return "IncubAlliance"
    if "centralesupelec" in corpus:
        return "CentraleSupélec"
    if "agroparistech" in corpus:
        return "AgroParisTech"
    if "cnrs" in corpus:
        return "CNRS"
    if "satt" in corpus:
        return "SATT Paris-Saclay"
    if "ens paris-saclay" in corpus:
        return "ENS Paris-Saclay"
    if "universite paris-saclay" in corpus:
        return "Université Paris-Saclay"
    return "Organisme de rattachement à confirmer"


def _mail_experts_rows(programme_table: dict, experts_profils: dict) -> list[dict]:
    rows = programme_table.get("rows", [])
    profils = experts_profils.get("profils", [])
    profile_by_key = {_canonical_name_key(item.get("nom", "")): item for item in profils}

    assignments: dict[str, dict] = {}
    for row in rows:
        code = row.get("code", "")
        if not code:
            continue
        fixed = FIXED_TEMOIN_PLAN.get(code, {})
        video_temoin_label = fixed.get("label") or row.get("video_temoin", "")
        objective = row.get("objectif_pedagogique", "")
        expert_videos = _tb_edito_parse_videos_expert(row.get("videos_referent", ""))
        expert_codes = [item.get("code", "") for item in expert_videos if item.get("code")]
        expert_labels = [f"{_label_video_expert(item.get('code', ''))} — {item.get('titre', '')}" for item in expert_videos]

        for raw_name in _extract_intervenants(row.get("noms_proposes", "")):
            key = _canonical_name_key(raw_name)
            if not key:
                continue
            profile = profile_by_key.get(key)
            canonical = EXPERT_NAME_ALIASES.get(key) or (profile.get("nom") if profile else raw_name)
            canonical_key = _canonical_name_key(canonical)
            profile = profile_by_key.get(canonical_key) or profile
            bucket = assignments.setdefault(
                canonical_key,
                {
                    "nom": canonical,
                    "profile": profile,
                    "organisme": _expert_org_from_profile(profile),
                    "videos": [],
                },
            )
            bucket["videos"].append(
                {
                    "code": code,
                    "video_temoin_label": video_temoin_label,
                    "tb_edito_href": f"tb_edito_{code}.html",
                    "expert_video_codes": expert_codes,
                    "expert_video_labels": expert_labels,
                    "objectif": objective,
                }
            )

    prepared = []
    for _, item in assignments.items():
        videos = sorted(item["videos"], key=lambda entry: int(entry["code"][1:]) if entry["code"][1:].isdigit() else 999)
        prepared.append(
            {
                "nom": item["nom"],
                "organisme": item["organisme"],
                "slug": slug(item["nom"]),
                "profile": item["profile"],
                "videos": videos,
            }
        )
    return sorted(prepared, key=lambda entry: _normalize_for_match(entry["nom"]))


def _compose_expert_mail(expert: dict) -> tuple[str, str]:
    nom = expert["nom"]
    prenom = " ".join((nom or "").split()).split(" ")[0] if nom else "Madame, Monsieur"
    organisme = expert["organisme"]
    videos = expert["videos"]
    video_codes = []
    for item in videos:
        for code in item.get("expert_video_codes", []):
            if code and code not in video_codes:
                video_codes.append(code)
    video_codes_label = ", ".join(_label_video_expert(code) for code in video_codes) if video_codes else "à définir"
    tb_edito_list = ", ".join(item["code"] for item in videos) if videos else "à définir"

    subject = f"MOOC L'Esprit d'innover — confirmation de vos videos expertise ({nom})"
    mail_text = (
        f"Objet : {subject}\n\n"
        f"Bonjour {prenom},\n\n"
        "Dans le cadre de la conception du MOOC \"L'Esprit d'innover\", nous préparons les vidéos expertise qui "
        "complètent les capsules témoins chorales.\n\n"
        "Nous partageons dans le guide de travail les informations utiles sur les experts et leurs organismes de rattachement ; "
        f"vous y apparaissez comme expert pressenti ({organisme}).\n\n"
        "À ce stade, les documents explicitent les transcripts des cinq chercheurs "
        "(Jean-Jacques Greffet, Muriel Thomas, Loïc Rajjou, Yann Monier et Sylvia Cohen-Kaminski).\n\n"
        f"Selon l'état actuel de la conception, vous êtes pressenti sur : {video_codes_label}.\n\n"
        "Afin d'éviter de produire des scripts inutiles, pourriez-vous nous confirmer les vidéos expertise "
        "sur lesquelles vous souhaitez intervenir selon ce calendrier :\n"
        "- 23 juillet : positionnement de votre part sur les vidéos expertise ;\n"
        "- 27 juillet : retour de notre part sur le positionnement retenu ;\n"
        "- 1er septembre : script pour le prompteur (a minima 15 jours avant la date de tournage).\n\n"
        "Pièces jointes proposées :\n"
        f"- Guide éditorial (propos témoins, objectifs pédagogiques, consignes envisagées et tableau récapitulatif des candidatures) ;\n"
        f"- Capsules témoins concernées : {tb_edito_list}.\n\n"
        "Le travail d'ingénierie pédagogique vise à refléter au mieux votre expertise sans s'y substituer ; "
        "vous êtes bien entendu libre d'aller plus loin, d'ajuster, ou de recadrer selon votre jugement.\n\n"
        "Merci d'avance pour votre retour,\n"
        "Bien cordialement,\n"
        "Equipe Action 2 pilier 1 PUI alliance Paris Saclay."
    )
    return subject, mail_text


def _mailto_href(recipient: str, subject: str, body: str) -> str:
    return f"mailto:{recipient}?subject={quote(subject)}&body={quote(body)}"


def _tb_expertise_label(text: str) -> str:
    value = text or ""
    # Remplacements du plus long au plus court, et garde-fous pour éviter
    # « expertise » → « expertiseise » (préfixe « expert » dans « expertise »).
    replacements = [
        ("Vidéo Expert", "Vidéo expertise"),
        ("Video Expert", "Vidéo expertise"),
        ("videos expertise", "videos expertise"),
        ("vidéos expertise", "vidéos expertise"),
        ("video expertise", "video expertise"),
        ("vidéo expertise", "vidéo expertise"),
        ("videos expert", "videos expertise"),
        ("vidéos expert", "vidéos expertise"),
        ("video expert", "video expertise"),
        ("vidéo expert", "vidéo expertise"),
    ]
    for src, dst in replacements:
        if src == dst:
            continue
        # Ne remplace « … expert » que s'il n'est pas déjà suivi de «ise».
        pattern = re.compile(re.escape(src) + r"(?!ise\b)", flags=re.IGNORECASE)
        value = pattern.sub(dst, value)
    return value


def _display_temoin_title(raw_label: str, capsule_code: str = "") -> str:
    """Libellé compréhensible pour un intervenant externe (sans code T…)."""
    text = _normalize_editorial_french((raw_label or "").strip())
    match = re.match(r"^VID[ÉE]O\s*(\d+)\s*[:\-–]\s*(.+)$", text, flags=re.IGNORECASE)
    if match:
        return f"Vidéo témoin {match.group(1)} — {match.group(2).strip()}"
    base = _label_video_temoin(capsule_code) if capsule_code else "Vidéo témoin"
    if text and text not in {base, capsule_code}:
        # Évite de coller un code brut type T9 derrière le libellé.
        if re.fullmatch(r"T\d+|GEN", text, flags=re.IGNORECASE):
            return base
        return f"{base} — {text}"
    return base


def _display_expertise_title(code: str = "", titre: str = "") -> str:
    """Libellé compréhensible pour un intervenant externe (sans code E…)."""
    title = _normalize_editorial_french((titre or "").strip())
    if title:
        # Si le titre arrive déjà préfixé, ne pas doubler.
        if title.lower().startswith("vidéo expertise"):
            return title
        # Retire un éventuel préfixe interne "Vidéo Expert 17 — …"
        stripped = re.sub(
            r"^Vidéo\s+Expert(?:ise)?\s+\d+(?:\s*bis)?\s*[—\-–:]\s*",
            "",
            title,
            flags=re.IGNORECASE,
        ).strip()
        return f"Vidéo expertise — {stripped or title}"
    labeled = _tb_expertise_label(_label_video_expert(code or ""))
    return labeled if labeled else "Vidéo expertise"


def _expertise_titles_from_item(item: dict) -> list[str]:
    titles: list[str] = []
    codes = item.get("expert_video_codes") or []
    labels = item.get("expert_video_labels") or []
    if codes:
        for idx, code in enumerate(codes):
            raw = labels[idx] if idx < len(labels) else ""
            titre = ""
            if raw:
                parts = re.split(r"\s*[—\-–]\s*", raw, maxsplit=1)
                titre = parts[1].strip() if len(parts) > 1 else raw
            titles.append(_display_expertise_title(code, titre))
        return titles
    for raw in labels:
        titles.append(_display_expertise_title("", raw))
    return titles


def _temoin_with_expertise_heading(item: dict) -> str:
    """Titre de chapitre : vidéo témoin + vidéos expertise associées."""
    temoin = _display_temoin_title(item.get("video_temoin_label", ""), item.get("code", ""))
    expertises = _expertise_titles_from_item(item)
    if not expertises:
        return temoin
    return f"{temoin} → {' · '.join(expertises)}"


def _highlight_tb_edito_syntagmes(text: str, code: str) -> str:
    html = escape(text or "")
    keywords = sorted(TOPIC_KEYWORDS_BY_CODE.get(code, []), key=len, reverse=True)
    for keyword in keywords:
        raw = (keyword or "").strip()
        if len(raw) < 4:
            continue
        pattern = re.compile(re.escape(raw), re.IGNORECASE)
        html = pattern.sub(lambda match: f"<strong>{match.group(0)}</strong>", html, count=1)
    return html


def _guide_editorial_expert_doc_html(expert: dict, grouped_tb: dict[str, list[dict]], rows_by_code: dict[str, dict]) -> str:
    sections = []
    toc_rows = []

    def _ergo_brief_html(text: str) -> str:
        lines = (text or "").splitlines()
        chunks: list[str] = []
        list_items: list[str] = []

        def flush_list() -> None:
            nonlocal list_items
            if list_items:
                chunks.append("<ul>" + "".join(list_items) + "</ul>")
                list_items = []

        for raw in lines:
            stripped = raw.strip()
            if not stripped:
                flush_list()
                continue
            if stripped.startswith("- "):
                list_items.append(f"<li>{escape(stripped[2:].strip())}</li>")
                continue
            flush_list()
            if stripped.startswith("Precaution :"):
                chunks.append(f"<p class='brief-precaution'><strong>Precaution :</strong> {escape(stripped.split(':', 1)[1].strip())}</p>")
            elif stripped.startswith("Repère facultatif"):
                chunks.append(
                    f"<p class='brief-fascicule'><strong>{escape(stripped.split(':', 1)[0].strip())} :</strong> "
                    f"{escape(stripped.split(':', 1)[1].strip()) if ':' in stripped else ''}</p>"
                )
            elif stripped.startswith("Vidéo expertise ") and " — " in stripped:
                left, right = stripped.split(" — ", 1)
                chunks.append(f"<p class='brief-video'><strong>{escape(left)}</strong> — {escape(right)}</p>")
            elif stripped.endswith(":"):
                chunks.append(f"<p class='brief-label'><strong>{escape(stripped)}</strong></p>")
            else:
                chunks.append(f"<p>{escape(stripped)}</p>")
        flush_list()
        return "".join(chunks) if chunks else "<p>Aucun contenu.</p>"

    def _ergo_script_html(text: str) -> str:
        return _script_lines_html(text)

    for item in expert.get("videos", []):
        code = item.get("code", "")
        chapter_anchor = f"chap_{code}"
        row = rows_by_code.get(code, {})
        sequences = grouped_tb.get(code, [])
        ordered = _tb_edito_order_for_code(code, sequences)
        by_seq_id = {seq.get("id", f"{code}-NOID"): seq for seq in ordered}
        ordre = [seq.get("id", f"{code}-NOID") for seq in ordered]
        videos_expert = _tb_edito_parse_videos_expert(row.get("videos_referent", ""))
        cadrage = _tb_edito_build_cadrage(code, ordre, by_seq_id, videos_expert)
        script_final = _tb_expertise_label(
            _script_final_prefer_mounted_transcript(
                code,
                _tb_edito_script_with_cadrage(ordre, by_seq_id, cadrage),
            )
        )
        capsule_data = {
            "videos_expert": videos_expert,
        }
        brief_text = _tb_expertise_label(export_brief_intervenant_plaintext(code, capsule_data, {}))
        toc_rows.append(
            "<tr>"
            f"<td><a href='#{escape(chapter_anchor)}'>{escape(code)} — {escape(item.get('video_temoin_label', ''))}</a></td>"
            "</tr>"
        )
        sections.append(
            "<section style='margin-top:28px;padding-top:10px;border-top:1px solid #cbd5e1;'>"
            f"<a name='{escape(chapter_anchor)}'></a>"
            f"<h2>{escape(code)} — {escape(item.get('video_temoin_label', ''))}</h2>"
            "<h3>Proposition de cadrage pour la video expert</h3>"
            f"<div class='doc-block brief-block'>{_ergo_brief_html(brief_text)}</div>"
            "<h3>Script final</h3>"
            f"<div class='doc-block script-block'>{_ergo_script_html(script_final)}</div>"
            "</section>"
        )

    toc_intro = (
        "<p class='meta'>Liens actifs vers chaque capsule témoin concernée.</p>"
        if len(toc_rows) > 1
        else "<p class='meta'>Lien actif vers la capsule témoin concernée.</p>"
    )
    return (
        "<html><head><meta charset='utf-8'>"
        "<style>"
        "body{font-family:Aptos,Segoe UI,Arial,sans-serif;font-size:12pt;line-height:1.5;}"
        "h1{font-size:18pt;margin-bottom:6px;}"
        "h2{font-size:14pt;margin-bottom:6px;}"
        "h3{font-size:12.5pt;margin-bottom:6px;}"
        ".doc-block{border:1px solid #dbe2ea;border-radius:8px;padding:12px 14px;margin-bottom:14px;}"
        ".brief-block{background:#f8fafc;line-height:1.65;}"
        ".script-block{background:#ffffff;line-height:1.6;}"
        ".doc-block p{margin:0 0 8px 0;}"
        ".doc-block ul{margin:4px 0 10px 24px;padding:0;}"
        ".doc-block li{margin:0 0 6px 0;}"
        ".brief-label{margin-top:10px;}"
        ".brief-video{background:#eef2ff;padding:6px 8px;border-radius:6px;}"
        ".brief-precaution{background:#fff7ed;padding:8px;border-left:3px solid #fdba74;border-radius:4px;}"
        ".brief-fascicule{background:#f0fdf4;padding:8px;border-left:3px solid #86efac;border-radius:4px;}"
        ".script-body{font-size:11pt;line-height:1.62;word-break:break-word;}"
        ".script-ref{font-size:9pt;color:#94a3b8;}"
        ".toc{width:100%;border-collapse:collapse;margin:10px 0 16px;}"
        ".toc td{padding:6px 2px;border-bottom:1px dotted #94a3b8;}"
        "</style>"
        "</head><body>"
        f"<h1>Guide éditorial — {escape(expert.get('nom', 'Expert'))}</h1>"
        "<h2>Sommaire du guide éditorial</h2>"
        f"{toc_intro}"
        f"<table class='toc'><tbody>{''.join(toc_rows) if toc_rows else '<tr><td>Aucune capsule témoin associée à ce stade.</td></tr>'}</tbody></table>"
        f"{''.join(sections) if sections else '<p>Aucune capsule témoin associée à ce stade.</p>'}"
        "</body></html>"
    )


VIDEOS_EXPERT_DATA = ROOT / "data" / "videos_expert"
VIDEOS_EXPERT_SCRIPTS = VIDEOS_EXPERT_DATA / "scripts_recus"
VIDEOS_EXPERT_REVUES = VIDEOS_EXPERT_DATA / "revues"


def _expert_video_sort_key(code: str) -> tuple[int, int, str]:
    match = re.fullmatch(r"E(\d+)(bis)?", code or "", re.IGNORECASE)
    if not match:
        return (9999, 1, code or "")
    return (int(match.group(1)), 1 if match.group(2) else 0, code.upper())


def _expert_video_page_name(code: str) -> str:
    return f"video_expert_{code}.html"


def _load_expert_script_recu(code: str) -> dict:
    """Charge le script renvoye par l'expert s'il est depose dans scripts_recus/."""
    VIDEOS_EXPERT_SCRIPTS.mkdir(parents=True, exist_ok=True)
    for suffix in (".txt", ".md", ".docx.txt"):
        path = VIDEOS_EXPERT_SCRIPTS / f"{code}{suffix}"
        if path.exists() and path.is_file():
            text = path.read_text(encoding="utf-8").strip()
            if text:
                return {
                    "statut": "RECU",
                    "fichier": path.name,
                    "contenu": text,
                }
    return {
        "statut": "EN_ATTENTE",
        "fichier": "",
        "contenu": "",
    }


def _load_expert_script_revues(code: str) -> list[dict]:
    """Charge les revues éditoriales (reprise + mail) pour une vidéo expertise."""
    VIDEOS_EXPERT_REVUES.mkdir(parents=True, exist_ok=True)
    revues: list[dict] = []
    for path in sorted(VIDEOS_EXPERT_REVUES.glob(f"{code}_revue*.json")):
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            continue
        if not isinstance(data, dict):
            continue
        numero = data.get("revue")
        if numero is None:
            match = re.search(r"_revue(\d+)", path.stem, flags=re.IGNORECASE)
            numero = int(match.group(1)) if match else 1
        revues.append(
            {
                "numero": int(numero),
                "date": (data.get("date") or "").strip(),
                "statut": (data.get("statut") or "PROPOSITION").strip(),
                "expert": (data.get("expert") or "").strip(),
                "fichier": path.name,
                "mode_texte": (data.get("mode_texte") or "texte_propose").strip(),
                "legende": data.get("legende") or {},
                "demandes": [str(item).strip() for item in (data.get("demandes") or []) if str(item).strip()],
                "texte_propose": (data.get("texte_propose") or "").strip(),
                "mail": (data.get("mail") or "").strip(),
                "mots_estimes": data.get("mots_estimes")
                or len(re.findall(r"[A-Za-zÀ-ÖØ-öø-ÿ0-9']+", data.get("texte_propose") or "")),
            }
        )
    revues.sort(key=lambda item: item["numero"])
    return revues


def _inventory_videos_expert(programme_table: dict, experts_profils: dict) -> list[dict]:
    """Inventaire des videos expertise depuis programme_table + experts proposes."""
    mail_rows = _mail_experts_rows(programme_table, experts_profils)
    experts_by_capsule: dict[str, list[dict]] = defaultdict(list)
    for expert in mail_rows:
        for video in expert.get("videos", []):
            capsule = video.get("code", "")
            if not capsule:
                continue
            experts_by_capsule[capsule].append(
                {
                    "nom": expert.get("nom", ""),
                    "slug": expert.get("slug", ""),
                    "organisme": expert.get("organisme", ""),
                    "guide_href": f"guide_editorial_{expert.get('slug', '')}.doc",
                    "mail_href": f"mail_expert_{expert.get('slug', '')}.html",
                }
            )

    inventory: list[dict] = []
    for row in programme_table.get("rows", []):
        capsule_code = row.get("code", "")
        if not capsule_code:
            continue
        fixed = FIXED_TEMOIN_PLAN.get(capsule_code, {})
        temoin_label = fixed.get("label") or row.get("video_temoin", "")
        experts = experts_by_capsule.get(capsule_code, [])
        for video in _tb_edito_parse_videos_expert(row.get("videos_referent", "")):
            code = video.get("code", "")
            if not code:
                continue
            script = _load_expert_script_recu(code)
            inventory.append(
                {
                    "code": code,
                    "titre": video.get("titre", ""),
                    "capsule_code": capsule_code,
                    "temoin_label": temoin_label,
                    "module": row.get("module", ""),
                    "objectif_pedagogique": row.get("objectif_pedagogique", ""),
                    "experts": experts,
                    "experts_label": ", ".join(item["nom"] for item in experts if item.get("nom"))
                    or "A confirmer",
                    "script_statut": script["statut"],
                    "script_fichier": script["fichier"],
                    "script_contenu": script["contenu"],
                    "page_href": _expert_video_page_name(code),
                    "tb_edito_href": f"tb_edito_{capsule_code}.html",
                }
            )

    inventory.sort(key=lambda item: _expert_video_sort_key(item["code"]))
    return inventory


def _consignes_envoyees_expert_html(item: dict) -> str:
    """Reprend les consignes transmises dans les guides Word (cadrage + consignes generales)."""
    consignes = [_humanize_capsule_labels(line) for line in BRIEF_CONSIGNES_COMMUNES]
    consignes_html = "".join(f"<li>{escape(line)}</li>" for line in consignes)
    expert_links = []
    for expert in item.get("experts", []):
        nom = expert.get("nom", "")
        guide = expert.get("guide_href", "")
        mail = expert.get("mail_href", "")
        if not nom:
            continue
        bits = [escape(nom)]
        if guide:
            bits.append(f"<a href='{escape(guide)}'>Guide Word</a>")
        if mail:
            bits.append(f"<a href='{escape(mail)}'>Mail</a>")
        expert_links.append("<li>" + " — ".join(bits) + "</li>")

    experts_block = (
        "<ul>" + "".join(expert_links) + "</ul>"
        if expert_links
        else "<p class='meta'>Aucun expert propose pour cette capsule a ce stade.</p>"
    )

    return f"""
<section class="methodology-panel brief-intervenant-panel">
  <h2>Consignes envoyees a l'expert</h2>
  <p class="meta">Contenu repris du guide editorial Word transmis avec le mail de sollicitation
  (proposition de cadrage + consignes generales). Document source : guides
  <code>guide_editorial_*.doc</code>.</p>
  <p class="brief-precaution"><strong>Precaution :</strong> {escape(BRIEF_PRECAUTION_ORATOIRE)}</p>
  <h3>{escape(_label_video_expert(item['code']))}</h3>
  <p><strong>Objectif :</strong> {escape(item.get('titre') or 'A preciser')}</p>
  <p><strong>Capsule temoin associee :</strong>
    <a href="{escape(item['tb_edito_href'])}">{escape(item['capsule_code'])} — {escape(item.get('temoin_label', ''))}</a>
  </p>
  <p><strong>Objectif pedagogique de la capsule :</strong> {escape(item.get('objectif_pedagogique') or '—')}</p>
  <h3>Consignes generales</h3>
  <ul>{consignes_html}</ul>
  <h3>Experts sollicites (guides / mails)</h3>
  {experts_block}
</section>
"""


def _script_expert_recu_html(item: dict) -> str:
    """Bloc d'accueil du script renvoye par l'expert."""
    code = item["code"]
    statut = item.get("script_statut", "EN_ATTENTE")
    if statut == "RECU" and item.get("script_contenu"):
        fichier = item.get("script_fichier", "")
        body = escape(item["script_contenu"]).replace("\n", "<br>")
        return f"""
<section class="methodology-panel">
  <h2>Script renvoye par l'expert</h2>
  <p>{status_badge('VALIDEE')} <span class="meta">Fichier : <code>{escape(fichier)}</code></span></p>
  <div class="script-recu-block">{body}</div>
</section>
"""

    depot = f"data/videos_expert/scripts_recus/{code}.txt"
    return f"""
<section class="methodology-panel script-attente-panel">
  <h2>Script renvoye par l'expert</h2>
  <p>{status_badge('EN_CONSTRUCTION')} <strong>En attente</strong> du script prompteur.</p>
  <p class="meta">Quand l'expert renverra son script, deposer le texte brut dans
  <code>{escape(depot)}</code> puis regenerer le site (<code>python3 scripts/build_site.py</code>).
  Formats acceptes : <code>{escape(code)}.txt</code> ou <code>{escape(code)}.md</code>.</p>
  <div class="script-placeholder" aria-label="Emplacement reserve au script expert">
    <p>Emplacement reserve — script expert a venir.</p>
  </div>
</section>
"""


def _render_annotated_script_html(texte: str) -> str:
    """Rend un script annoté [[HS:...]] / [[RC:...]] avec surlignage HTML."""
    if not texte:
        return ""

    def _replace_hs(match: re.Match[str]) -> str:
        meta = (match.group(1) or "").strip()
        body = match.group(2) or ""
        if "|" in meta:
            codes, note = meta.split("|", 1)
            label = f"HS · {codes.strip()}"
            title = note.strip()
        else:
            label = f"HS · {meta}" if meta else "HS"
            title = meta
        return (
            f'<mark class="script-hs" title="{escape(title)}">'
            f'<span class="script-hs__tag">{escape(label)}</span>'
            f"{escape(body)}"
            f"</mark>"
        )

    def _replace_rc(match: re.Match[str]) -> str:
        meta = (match.group(1) or "").strip()
        body = match.group(2) or ""
        label = "RC"
        title = meta
        if meta:
            short = meta.split("—", 1)[0].strip()
            label = f"RC · {short}" if short else "RC"
        return (
            f'<mark class="script-rc" title="{escape(title)}">'
            f'<span class="script-rc__tag">{escape(label)}</span>'
            f"{escape(body)}"
            f"</mark>"
        )

    # Escape first, then reinject markers from raw via sequential replace on raw
    # Work on raw text, escape only non-tag parts by processing with regex callbacks that escape.
    rendered = texte
    rendered = re.sub(
        r"\[\[HS:([^\]]*)\]\](.*?)\[\[/HS\]\]",
        _replace_hs,
        rendered,
        flags=re.DOTALL,
    )
    rendered = re.sub(
        r"\[\[RC:([^\]]*)\]\](.*?)\[\[/RC\]\]",
        _replace_rc,
        rendered,
        flags=re.DOTALL,
    )
    # Escape remaining plain text while preserving inserted HTML marks.
    parts: list[str] = []
    cursor = 0
    for match in re.finditer(r'<mark class="script-(?:hs|rc)"[^>]*>.*?</mark>', rendered, flags=re.DOTALL):
        parts.append(escape(rendered[cursor : match.start()]).replace("\n", "<br>"))
        parts.append(match.group(0).replace("\n", "<br>"))
        cursor = match.end()
    parts.append(escape(rendered[cursor:]).replace("\n", "<br>"))
    return "".join(parts)


def _script_expert_revues_html(code: str) -> str:
    """Blocs Revue N sous le script reçu : demandes, texte repris/annoté, mail à l'expert."""
    revues = _load_expert_script_revues(code)
    if not revues:
        return ""
    parts: list[str] = []
    for revue in revues:
        numero = revue["numero"]
        demandes = revue.get("demandes") or []
        demandes_html = (
            "<ul class='script-revue-demandes'>"
            + "".join(f"<li>{escape(item)}</li>" for item in demandes)
            + "</ul>"
            if demandes
            else "<p class='meta'>Aucune demande listée.</p>"
        )
        texte = revue.get("texte_propose") or ""
        mode = revue.get("mode_texte") or "texte_propose"
        if texte and mode == "annote_hors_sujet":
            texte_html = f"<div class='script-recu-block'>{_render_annotated_script_html(texte)}</div>"
            texte_title = "Texte de l’expert annoté — passages hors sujet / à recadrer"
        elif texte:
            texte_html = (
                f"<div class='script-recu-block'>{escape(texte).replace(chr(10), '<br>')}</div>"
            )
            texte_title = "Proposition de texte repris"
        else:
            texte_html = "<p class='meta'>Proposition de texte à compléter.</p>"
            texte_title = "Proposition de texte repris"

        legende = revue.get("legende") or {}
        legende_html = ""
        if legende and mode == "annote_hors_sujet":
            items = "".join(
                f"<li><strong>{escape(str(key))}</strong> — {escape(str(value))}</li>"
                for key, value in legende.items()
            )
            legende_html = f"<div class='script-revue-legende'><ul>{items}</ul></div>"

        mail = revue.get("mail") or ""
        mail_html = (
            f"<h3>Mail à envoyer (Revue {numero})</h3>"
            f"<pre class='script-revue-mail'>{escape(mail)}</pre>"
            if mail
            else ""
        )
        meta_bits = [
            status_badge("EN_CONSTRUCTION" if revue.get("statut") == "PROPOSITION" else "VALIDEE"),
            f"<span class='meta'>Fichier : <code>{escape(revue.get('fichier', ''))}</code></span>",
        ]
        if revue.get("date"):
            meta_bits.append(f"<span class='meta'>Date : {escape(revue['date'])}</span>")
        if revue.get("mots_estimes"):
            meta_bits.append(
                f"<span class='meta'>~{escape(str(revue['mots_estimes']))} mots (script source)</span>"
            )
        parts.append(
            f"""
<section class="methodology-panel script-revue-panel">
  <h2>Revue {numero} — reprise et propositions</h2>
  <p>{' · '.join(meta_bits)}</p>
  <h3>Demandes de correction / modification</h3>
  {demandes_html}
  <h3>{escape(texte_title)}</h3>
  {legende_html}
  {texte_html}
  {mail_html}
</section>
"""
        )
    return "\n".join(parts)


def _write_videos_expert_xlsx(inventory: list[dict], path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Videos expert"
    headers = [
        "Code",
        "Titre / objectif",
        "Capsule temoin",
        "Titre temoin",
        "Module",
        "Experts proposes",
        "Statut script",
        "Fichier script",
        "Page site",
        "Guide Word (1er expert)",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="0B6E77")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for item in inventory:
        first_guide = ""
        experts = item.get("experts") or []
        if experts:
            first_guide = experts[0].get("guide_href", "")
        ws.append(
            [
                item["code"],
                item.get("titre", ""),
                item.get("capsule_code", ""),
                item.get("temoin_label", ""),
                item.get("module", ""),
                item.get("experts_label", ""),
                item.get("script_statut", "EN_ATTENTE"),
                item.get("script_fichier", ""),
                item.get("page_href", ""),
                first_guide,
            ]
        )

    widths = [10, 48, 12, 36, 18, 36, 14, 18, 24, 36]
    from openpyxl.utils import get_column_letter

    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _sommaire_cards(sections: list[tuple[str, str, str, str]]) -> str:
    cards = "".join(
        f"<a class='sommaire-card' href='{escape(href)}'>"
        f"<span class='sommaire-card__icon' aria-hidden='true'>{escape(icon)}</span>"
        f"<h2>{escape(title)}</h2>"
        f"<p>{escape(description)}</p>"
        f"<span class='sommaire-card__cta'>Ouvrir →</span>"
        f"</a>"
        for href, icon, title, description in sections
    )
    return f'<nav class="sommaire-grid" aria-label="Sommaire">{cards}</nav>'


def _module_sort_key(module: str) -> tuple[int, str]:
    match = re.fullmatch(r"M(\d+)", module or "", re.IGNORECASE)
    if match:
        return (int(match.group(1)), module.upper())
    return (999, module or "")


def _module_label(module: str) -> str:
    labels = {
        "M1": "Module 1 — Origines, besoin et preuve",
        "M2": "Module 2 — Protection et transfert",
        "M3": "Module 3 — Accompagnement et financements",
        "M4": "Module 4 — Équipe et langage",
        "M5": "Module 5 — Métier et collaborations",
        "M6": "Module 6 — Conclusion",
    }
    return labels.get((module or "").upper(), f"Module {module}")


def _suivi_rows_by_module(programme_table: dict) -> dict[str, list[dict]]:
    grouped: dict[str, list[dict]] = defaultdict(list)
    for row in programme_table.get("rows", []):
        module = (row.get("module") or "").strip() or "M?"
        grouped[module].append(row)
    for module in grouped:
        grouped[module].sort(
            key=lambda item: int(item["code"][1:]) if str(item.get("code", ""))[1:].isdigit() else 999
        )
    return dict(sorted(grouped.items(), key=lambda item: _module_sort_key(item[0])))


def _suivi_canonical_intervenants(
    programme_table: dict, experts_profils: dict
) -> list[dict]:
    """Intervenants dedoublonnes, avec capsules et videos expert rattachees."""
    profils = experts_profils.get("profils", [])
    profile_by_key = {_canonical_name_key(item.get("nom", "")): item for item in profils}
    buckets: dict[str, dict] = {}

    for row in programme_table.get("rows", []):
        capsule = row.get("code", "")
        if not capsule:
            continue
        temoin_label = (
            FIXED_TEMOIN_PLAN.get(capsule, {}).get("label")
            or row.get("video_temoin", "")
        )
        expert_videos = _tb_edito_parse_videos_expert(row.get("videos_referent", ""))
        for raw_name in _extract_intervenants(row.get("noms_proposes", "")):
            key = _canonical_name_key(raw_name)
            if not key:
                continue
            profile = profile_by_key.get(key)
            canonical = EXPERT_NAME_ALIASES.get(key) or (profile.get("nom") if profile else raw_name)
            canonical_key = _canonical_name_key(canonical)
            profile = profile_by_key.get(canonical_key) or profile
            bucket = buckets.setdefault(
                canonical_key,
                {
                    "nom": canonical,
                    "slug": slug(canonical),
                    "organisme": _expert_org_from_profile(profile),
                    "capsules": [],
                    "videos_expert": [],
                },
            )
            if not any(item["code"] == capsule for item in bucket["capsules"]):
                bucket["capsules"].append(
                    {
                        "code": capsule,
                        "label": temoin_label,
                        "module": row.get("module", ""),
                    }
                )
            for video in expert_videos:
                code = video.get("code", "")
                if not code:
                    continue
                if not any(item["code"] == code for item in bucket["videos_expert"]):
                    bucket["videos_expert"].append(
                        {
                            "code": code,
                            "titre": video.get("titre", ""),
                            "capsule_code": capsule,
                        }
                    )

    prepared = list(buckets.values())
    for item in prepared:
        item["capsules"].sort(
            key=lambda entry: int(entry["code"][1:]) if entry["code"][1:].isdigit() else 999
        )
        item["videos_expert"].sort(key=lambda entry: _expert_video_sort_key(entry["code"]))
    return sorted(prepared, key=lambda entry: _normalize_for_match(entry["nom"]))


def _load_suivi_positionnements() -> dict:
    path = ROOT / "data" / "suivi_positionnements.json"
    if not path.exists():
        return {"date_mise_a_jour": "", "note": "", "intervenants": []}
    return json.loads(path.read_text(encoding="utf-8"))


def _save_suivi_positionnements(payload: dict) -> None:
    path = ROOT / "data" / "suivi_positionnements.json"
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _format_proposition_notre_part(intervenant: dict) -> str:
    parts = []
    videos_by_capsule: dict[str, list[str]] = defaultdict(list)
    for video in intervenant.get("videos_expert", []):
        videos_by_capsule[video.get("capsule_code", "")].append(video.get("code", ""))
    for capsule in intervenant.get("capsules", []):
        code = capsule.get("code", "")
        e_codes = [item for item in videos_by_capsule.get(code, []) if item]
        if e_codes:
            parts.append(f"{code} ({', '.join(e_codes)})")
        else:
            parts.append(code)
    return " ; ".join(parts)


def _sync_suivi_positionnements(intervenants: list[dict]) -> list[dict]:
    """Aligne le fichier de suivi sur les intervenants actuels ; conserve les reponses saisies."""
    stored = _load_suivi_positionnements()
    by_slug = {
        item.get("slug") or slug(item.get("nom", "")): item
        for item in stored.get("intervenants", [])
    }
    synced = []
    for intervenant in intervenants:
        key = intervenant["slug"]
        previous = by_slug.get(key, {})
        proposition = _format_proposition_notre_part(intervenant)
        synced.append(
            {
                "nom": intervenant["nom"],
                "slug": key,
                "organisme": previous.get("organisme", "") or intervenant.get("organisme", ""),
                "proposition_notre_part": proposition,
                "capsules": [item.get("code", "") for item in intervenant.get("capsules", [])],
                "videos_expert": [item.get("code", "") for item in intervenant.get("videos_expert", [])],
                "reponse": previous.get("reponse", "") or "",
                "besoin_exprime": previous.get("besoin_exprime", "") or "",
                "positionnement_preferences": previous.get("positionnement_preferences", "") or "",
                "proposition_finale": previous.get("proposition_finale", "") or "",
            }
        )
    payload = {
        "date_mise_a_jour": "2026-07-24",
        "note": stored.get("note")
        or (
            "Suivi des positionnements intervenants. La proposition de notre part est derivee "
            "du programme de conception. Reponse, besoin exprime, preferences et proposition "
            "finale se remplissent au fil des retours."
        ),
        "intervenants": synced,
    }
    _save_suivi_positionnements(payload)
    return synced


def _cell_or_attente(value: str) -> str:
    text = (value or "").strip()
    if text:
        return escape(text).replace("\n", "<br>")
    return '<span class="meta">—</span>'


def _positionnements_finaux_par_video(rows: list[dict]) -> list[dict]:
    grouped: dict[str, list[dict]] = defaultdict(list)
    seen: set[tuple[str, str]] = set()
    for item in rows:
        final = (item.get("proposition_finale") or "").strip()
        if not final:
            continue
        codes = re.findall(r"\bE\d+(?:bis)?\b", final, flags=re.IGNORECASE)
        for raw_code in codes:
            code = raw_code.upper()
            key = (code, item.get("slug", ""))
            if key in seen:
                continue
            seen.add(key)
            grouped[code].append(
                {
                    "nom": item.get("nom", ""),
                    "slug": item.get("slug", ""),
                    "organisme": item.get("organisme", ""),
                }
            )

    def _video_key(code: str) -> tuple[int, int]:
        match = re.fullmatch(r"E(\d+)(bis)?", code or "", re.IGNORECASE)
        if not match:
            return (9999, 1)
        return (int(match.group(1)), 1 if match.group(2) else 0)

    return [
        {"code": code, "intervenants": grouped[code]}
        for code in sorted(grouped.keys(), key=_video_key)
    ]


# Objectifs pédagogiques des vidéos expertise → tags de cohérence.
VIDEO_OBJECTIVE_TAGS: dict[str, set[str]] = {
    "E1": {"origines", "innovation", "usage", "design", "tech_push"},
    "E2": {"probleme", "proposition_valeur", "usage", "design", "innovation"},
    "E3": {"probleme", "validation", "experimentation", "pivot", "apprentissage"},
    "E4": {"trl", "poc", "prototype", "maturation", "preuve"},
    "E5": {"derisking", "prematuration", "maturation", "jalons", "decision"},
    "E6": {"pi", "confidentialite", "declaration", "brevet", "publication"},
    "E7": {"valorisation", "transfert", "premier_contact", "examen_resultat"},
    "E8": {"pi", "protection", "brevet", "secret", "savoir_faire"},
    "E9": {"pi", "strategie", "valorisation", "negociation", "actif"},
    "E10": {"valorisation", "licence", "creation", "partenariat", "criteres"},
    "E11": {"juridique", "contrats", "licence", "pi", "transfert", "negociation"},
    "E12": {"prematuration", "maturation", "accompagnement", "investissement", "jalons"},
    "E13": {"incubateur", "accompagnement", "creation", "equipe", "coaching"},
    "E13BIS": {"accompagnement", "ecosysteme", "design", "structures", "ott"},
    "E14": {"financement", "aides", "investissement", "prematuration", "maturation"},
    "E15": {"investisseurs", "gouvernance", "dilution", "croissance", "compatibilite"},
    "E16": {"competences", "equipe", "recrutement", "complementarite", "entrepreneuriat"},
    "E17": {"juridique", "gouvernance", "pacte", "fondateurs", "parts", "conflits"},
    "E18": {"pitch", "communication", "proposition_valeur", "interlocuteurs", "valorisation"},
    "E19": {"posture", "entrepreneuriat", "mentorat", "apprentissage", "identite"},
    "E20": {"freins", "parcours", "legitimite", "apprentissage", "accompagnement"},
    "E21": {"apprentissage", "incertitude", "pivot", "innovation", "progression"},
    "E22": {"collaboration", "partenariat", "gouvernance", "complementarite", "valeur"},
    "E23": {"juridique", "contrats", "pi", "collaboration", "confidentialite", "partenariat"},
}

# Fonctions des intervenants → tags de cohérence (profil métier).
INTERVENANT_FUNCTION_TAGS: dict[str, set[str]] = {
    "antoine latreille": {"juridique", "pi", "droit", "valorisation", "transfert", "pacte", "contrats"},
    "arielle sante": {
        "incubateur",
        "accompagnement",
        "entrepreneuriat",
        "creation",
        "equipe",
        "pitch",
        "mentorat",
        "posture",
        "communication",
        "freins",
        "competences",
        "interlocuteurs",
        "proposition_valeur",
        "ecosysteme",
        "structures",
        "parcours",
        "investisseurs",
        "gouvernance",
        "dilution",
        "croissance",
        "financement",
        "investissement",
    },
    "bernard yannou": {
        "innovation",
        "design",
        "deeptech",
        "origines",
        "valorisation",
        "freins",
        "apprentissage",
        "probleme",
        "proposition_valeur",
        "usage",
    },
    "eneli vino": {
        "juridique",
        "pi",
        "contrats",
        "partenariat",
        "collaboration",
        "valorisation",
        "confidentialite",
        "negociation",
    },
    "fatoumata aonon": {
        "valorisation",
        "prematuration",
        "maturation",
        "financement",
        "partenariat",
        "trl",
        "investissement",
        "jalons",
        "derisking",
    },
    "gregoire burge": {
        "innovation",
        "transfert",
        "probleme",
        "validation",
        "incubateur",
        "accompagnement",
        "experimentation",
    },
    "pascal corbel": {
        "management",
        "entrepreneuriat",
        "competences",
        "complementarite",
        "posture",
        "innovation",
        "apprentissage",
        "equipe",
        "proposition_valeur",
    },
    "remi wache": {
        "partenariat",
        "collaboration",
        "transfert",
        "business_dev",
        "complementarite",
        "valeur",
        "gouvernance",
    },
    "soizic lefeuvre": {
        "partenariat",
        "valorisation",
        "juridique",
        "contrats",
        "pi",
        "collaboration",
        "confidentialite",
        "transfert",
    },
    "stanislas de lapasse": {"pi", "protection", "brevet", "confidentialite", "actif", "strategie"},
    "stephanie oger roussel": {
        "valorisation",
        "strategie",
        "deeptech",
        "trl",
        "poc",
        "maturation",
        "communication",
        "transfert",
        "preuve",
    },
    "stephanie sano": {"valorisation", "pi", "declaration", "premier_contact", "transfert"},
    "virginia branco": {
        "transfert",
        "valorisation",
        "pi",
        "contrats",
        "appui",
        "partenariat",
        "juridique",
    },
    "yoann montenot": {
        "design",
        "accompagnement",
        "structures",
        "ecosysteme",
        "coaching",
        "usage",
        "innovation",
    },
}

TAG_AFFINITY: dict[str, set[str]] = {
    "juridique": {"contrats", "pi", "pacte", "gouvernance", "confidentialite"},
    "pi": {"juridique", "protection", "brevet", "valorisation", "actif"},
    "incubateur": {"accompagnement", "creation", "equipe", "mentorat", "coaching"},
    "accompagnement": {"incubateur", "mentorat", "structures", "ecosysteme", "parcours"},
    "entrepreneuriat": {"posture", "creation", "competences", "equipe", "apprentissage"},
    "partenariat": {"collaboration", "contrats", "transfert", "valorisation", "negociation"},
    "collaboration": {"partenariat", "gouvernance", "complementarite", "valeur"},
    "valorisation": {"transfert", "licence", "strategie", "pi", "preuve"},
    "financement": {"investissement", "investisseurs", "maturation", "prematuration"},
    "design": {"usage", "probleme", "proposition_valeur", "innovation"},
    "apprentissage": {"pivot", "incertitude", "posture", "freins", "progression"},
}


def _video_sort_key_code(code: str) -> tuple[int, int]:
    match = re.fullmatch(r"E(\d+)(bis)?", code or "", re.IGNORECASE)
    if not match:
        return (9999, 1)
    return (int(match.group(1)), 1 if match.group(2) else 0)


def _load_expert_videos_catalogue() -> list[dict]:
    path = ROOT / "data" / "programme_videos.json"
    if not path.exists():
        return []
    payload = json.loads(path.read_text(encoding="utf-8"))
    videos: dict[str, dict] = {}
    for capsule_code, block in (payload.get("capsules") or {}).items():
        if not isinstance(block, dict):
            continue
        for video in block.get("videos_expert") or []:
            code = (video.get("code") or "").upper().replace(" ", "")
            if not code or code in videos:
                continue
            videos[code] = {
                "code": code,
                "titre": _normalize_editorial_french(video.get("titre") or ""),
                "descriptif": _normalize_editorial_french(video.get("descriptif") or ""),
                "capsule": capsule_code,
            }
    return sorted(videos.values(), key=lambda item: _video_sort_key_code(item["code"]))


def _intervenant_function_label(nom: str, organisme: str = "") -> str:
    path = ROOT / "data" / "experts_profils.json"
    profils = []
    if path.exists():
        profils = json.loads(path.read_text(encoding="utf-8")).get("profils", [])
    by_key = {_canonical_name_key(item.get("nom", "")): item for item in profils}
    key = _canonical_name_key(nom)
    profile = by_key.get(key)
    if profile and (profile.get("profil_cible") or "").strip():
        return _normalize_editorial_french(profile.get("profil_cible", "").strip())
    if organisme:
        return organisme
    return "Fonction à préciser"


def _coherence_score_for_pair(video_code: str, intervenant_nom: str) -> int:
    """Score 1 (peu cohérent) → 5 (très cohérent) selon objectif vidéo × fonction intervenant."""
    code = (video_code or "").upper().replace(" ", "")
    video_tags = VIDEO_OBJECTIVE_TAGS.get(code, set())
    expert_key = _canonical_name_key(intervenant_nom)
    expert_tags = INTERVENANT_FUNCTION_TAGS.get(expert_key, set())
    if not video_tags or not expert_tags:
        return 1

    direct = len(video_tags & expert_tags)
    related = 0
    for tag in video_tags:
        related_set = TAG_AFFINITY.get(tag, set())
        related += len(related_set & expert_tags)
    for tag in expert_tags:
        related_set = TAG_AFFINITY.get(tag, set())
        related += len(related_set & video_tags)
    related = related // 2  # chaque affinité est comptée deux fois

    raw = direct * 2 + related
    if raw >= 7:
        return 5
    if raw >= 5:
        return 4
    if raw >= 3:
        return 3
    if raw >= 1:
        return 2
    return 1


def _render_coherence_heatmap(rows: list[dict]) -> str:
    videos = _load_expert_videos_catalogue()
    intervenants = [
        item
        for item in rows
        if item.get("nom") and _canonical_name_key(item.get("nom", "")) != "joel nguen"
    ]
    if not videos or not intervenants:
        return (
            "<h2>Matrice de cohérence vidéo × intervenant</h2>"
            "<p class='meta'>Aucune donnée disponible pour construire la matrice.</p>"
        )

    selected: set[tuple[str, str]] = set()
    for item in rows:
        final = (item.get("proposition_finale") or "").strip()
        slug = item.get("slug", "")
        for raw in re.findall(r"\bE\d+(?:bis)?\b", final, flags=re.IGNORECASE):
            selected.add((raw.upper().replace(" ", ""), slug))

    header_cells = []
    for item in intervenants:
        fonction = _intervenant_function_label(item.get("nom", ""), item.get("organisme", ""))
        header_cells.append(
            "<th>"
            f"<strong>{escape(item.get('nom', ''))}</strong>"
            f"<br><span class='meta'>{escape(fonction)}</span>"
            "</th>"
        )

    body_rows = []
    matrix_for_xlsx: list[list] = []
    for video in videos:
        code = video["code"]
        objectif = video.get("descriptif") or video.get("titre") or ""
        row_label = (
            f"<strong>{escape(_label_video_expert(code))}</strong>"
            f"<br><span class='meta'>{escape(objectif)}</span>"
        )
        value_cells = []
        xlsx_row = [_label_video_expert(code), objectif]
        for item in intervenants:
            score = _coherence_score_for_pair(code, item.get("nom", ""))
            is_selected = (code, item.get("slug", "")) in selected
            style = _heat_cell_style(score, 5)
            selected_class = " heatmap-cell--selected" if is_selected else ""
            title = f"Cohérence {score}/5"
            if is_selected:
                title += " — sélection finale"
            value_cells.append(
                f"<td class='heatmap-cell{selected_class}' style='{style}' "
                f"title='{escape(title)}'>{score}</td>"
            )
            xlsx_row.append(score)
        body_rows.append(f"<tr><th class='heatmap-row-label'>{row_label}</th>{''.join(value_cells)}</tr>")
        matrix_for_xlsx.append(xlsx_row)

    # Stash for optional Excel export by caller via attribute on function.
    _render_coherence_heatmap._last_matrix = {  # type: ignore[attr-defined]
        "intervenants": [item.get("nom", "") for item in intervenants],
        "fonctions": [
            _intervenant_function_label(item.get("nom", ""), item.get("organisme", ""))
            for item in intervenants
        ],
        "rows": matrix_for_xlsx,
    }

    return (
        "<h2>Matrice de cohérence vidéo × intervenant</h2>"
        "<p class='meta'>Carte de chaleur de 1 (peu cohérent) à 5 (très cohérent), "
        "croisant l'<strong>objectif pédagogique</strong> de chaque vidéo expertise "
        "avec la <strong>fonction</strong> de chaque intervenant. "
        "Les cellules bordées marquent la sélection finale actuelle.</p>"
        "<article class='heatmap-card'>"
        "<div class='heatmap-wrap'>"
        "<table class='heatmap-table heatmap-table--coherence'>"
        "<thead><tr><th>Vidéo / objectif</th>"
        f"{''.join(header_cells)}</tr></thead>"
        f"<tbody>{''.join(body_rows)}</tbody>"
        "</table></div>"
        "<p class='heatmap-legend'>"
        "<span>1 — peu cohérent</span>"
        "<span class='heatmap-scale' aria-hidden='true'></span>"
        "<span>5 — très cohérent</span>"
        "</p>"
        "</article>"
    )


def _export_selection_finale_xlsx(rows: list[dict]) -> str:
    """Exporte un XLSX dédié à la sélection finale par vidéo expertise."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    videos_by_code = {item["code"]: item for item in _load_expert_videos_catalogue()}
    wb = Workbook()
    ws = wb.active
    ws.title = "Selection finale"
    headers = [
        "Code",
        "Vidéo expertise",
        "Objectif",
        "Intervenant",
        "Organisme",
        "Fonction",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="0B6E77")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for item in _positionnements_finaux_par_video(rows):
        code = item["code"]
        video = videos_by_code.get(code, {})
        objectif = video.get("descriptif") or video.get("titre") or ""
        label = _label_video_expert(code)
        intervenants = item.get("intervenants") or []
        if not intervenants:
            ws.append([code, label, objectif, "", "", ""])
            continue
        for entry in intervenants:
            nom = entry.get("nom", "")
            organisme = entry.get("organisme", "")
            fonction = _intervenant_function_label(nom, organisme)
            ws.append([code, label, objectif, nom, organisme, fonction])

    widths = [10, 22, 52, 28, 36, 42]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    filename = "selection_finale_videos_expertise.xlsx"
    wb.save(SITE / filename)
    return filename


def build_suivi_positionnements_page(rows: list[dict]) -> None:
    table_rows = []
    for item in rows:
        table_rows.append(
            "<tr>"
            f"<td><a href='suivi_intervenant_{escape(item['slug'])}.html'><strong>{escape(item['nom'])}</strong></a>"
            f"<br><span class='meta'>{escape(item.get('organisme', ''))}</span></td>"
            f"<td>{_cell_or_attente(item.get('proposition_notre_part', ''))}</td>"
            f"<td>{_cell_or_attente(item.get('reponse', ''))}</td>"
            f"<td>{_cell_or_attente(item.get('besoin_exprime', ''))}</td>"
            f"<td>{_cell_or_attente(item.get('positionnement_preferences', ''))}</td>"
            f"<td>{_cell_or_attente(item.get('proposition_finale', ''))}</td>"
            "</tr>"
        )

    final_rows = []
    for item in _positionnements_finaux_par_video(rows):
        intervenants = item["intervenants"]
        intervenants_html = (
            "<ul>"
            + "".join(
                "<li>"
                f"<a href='suivi_intervenant_{escape(entry['slug'])}.html'><strong>{escape(entry['nom'])}</strong></a>"
                + (
                    f" <span class='meta'>({escape(entry.get('organisme', ''))})</span>"
                    if entry.get("organisme")
                    else ""
                )
                + "</li>"
                for entry in intervenants
            )
            + "</ul>"
            if intervenants
            else '<span class="meta">—</span>'
        )
        final_rows.append(
            "<tr>"
            f"<td><strong>{escape(_label_video_expert(item['code']))}</strong></td>"
            f"<td>{intervenants_html}</td>"
            "</tr>"
        )

    # Export XLSX
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Suivi positionnements"
    headers = [
        "Intervenant",
        "Organisme",
        "Proposition de positionnement (notre part)",
        "Réponse fournie",
        "Besoin exprimé",
        "Positionnement par préférences",
        "Proposition finale",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="0B6E77")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for item in rows:
        ws.append(
            [
                item.get("nom", ""),
                item.get("organisme", ""),
                item.get("proposition_notre_part", ""),
                item.get("reponse", ""),
                item.get("besoin_exprime", ""),
                item.get("positionnement_preferences", ""),
                item.get("proposition_finale", ""),
            ]
        )
    widths = [28, 28, 42, 36, 36, 36, 28]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws_final = wb.create_sheet("Selection finale")
    final_headers = [
        "Code",
        "Vidéo expertise",
        "Objectif",
        "Intervenant",
        "Organisme",
        "Fonction",
    ]
    ws_final.append(final_headers)
    for cell in ws_final[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    videos_by_code = {item["code"]: item for item in _load_expert_videos_catalogue()}
    for item in _positionnements_finaux_par_video(rows):
        code = item["code"]
        video = videos_by_code.get(code, {})
        objectif = video.get("descriptif") or video.get("titre") or ""
        label = _label_video_expert(code)
        intervenants = item.get("intervenants") or []
        if not intervenants:
            ws_final.append([code, label, objectif, "", "", ""])
            continue
        for entry in intervenants:
            nom = entry.get("nom", "")
            organisme = entry.get("organisme", "")
            ws_final.append(
                [
                    code,
                    label,
                    objectif,
                    nom,
                    organisme,
                    _intervenant_function_label(nom, organisme),
                ]
            )
    for idx, width in enumerate((10, 22, 52, 28, 36, 42), start=1):
        ws_final.column_dimensions[get_column_letter(idx)].width = width
    for row in ws_final.iter_rows(
        min_row=2, max_row=ws_final.max_row, min_col=1, max_col=len(final_headers)
    ):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    coherence_html = _render_coherence_heatmap(rows)
    coherence_payload = getattr(_render_coherence_heatmap, "_last_matrix", None)
    if coherence_payload:
        ws_heat = wb.create_sheet("Matrice coherence")
        heat_headers = ["Vidéo expertise", "Objectif"] + list(coherence_payload["intervenants"])
        ws_heat.append(heat_headers)
        for cell in ws_heat[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for heat_row in coherence_payload["rows"]:
            ws_heat.append(heat_row)
        ws_heat.column_dimensions["A"].width = 22
        ws_heat.column_dimensions["B"].width = 48
        for idx in range(3, len(heat_headers) + 1):
            ws_heat.column_dimensions[get_column_letter(idx)].width = 14
        for row in ws_heat.iter_rows(
            min_row=2, max_row=ws_heat.max_row, min_col=1, max_col=len(heat_headers)
        ):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    xlsx_name = "suivi_positionnements.xlsx"
    wb.save(SITE / xlsx_name)
    selection_finale_xlsx = _export_selection_finale_xlsx(rows)

    body = (
        "<p class='meta'>Suivi du dialogue de positionnement avec chaque intervenant. "
        "La <strong>proposition de notre part</strong> reprend le pool actuel du programme de conception. "
        "La <strong>réponse</strong>, le <strong>besoin exprimé</strong> et le "
        "<strong>positionnement par préférences</strong> se remplissent au fil des retours ; "
        "la <strong>proposition finale</strong> sera arbitrée ensuite.</p>"
        f"<p><a class='btn' href='{escape(xlsx_name)}' download>Télécharger le suivi (XLSX)</a></p>"
        "<div class='table-wrap'><table><thead><tr>"
        "<th>Intervenant</th>"
        "<th>Proposition de positionnement<br><span class='meta'>notre part</span></th>"
        "<th>Réponse fournie</th>"
        "<th>Besoin exprimé</th>"
        "<th>Positionnement par préférences</th>"
        "<th>Proposition finale</th>"
        "</tr></thead><tbody>"
        + (
            "".join(table_rows)
            if table_rows
            else "<tr><td colspan='6'>Aucun intervenant dans le programme.</td></tr>"
        )
        + "</tbody></table></div>"
        "<p class='meta'>Source editable : <code>data/suivi_positionnements.json</code> "
        "(champs <code>reponse</code>, <code>besoin_exprime</code>, "
        "<code>positionnement_preferences</code>, <code>proposition_finale</code>).</p>"
        "<h2>Sélection finale par vidéo expertise</h2>"
        "<p class='meta'>Lecture simplifiée des arbitrages finaux, classés par vidéo expertise à partir du champ "
        "<code>proposition_finale</code>.</p>"
        f"<p><a class='btn' href='{escape(selection_finale_xlsx)}' download>"
        "Télécharger la sélection finale (XLSX)</a></p>"
        "<div class='table-wrap'><table><thead><tr>"
        "<th>Vidéo expertise</th>"
        "<th>Intervenant(s) retenu(s)</th>"
        "</tr></thead><tbody>"
        + (
            "".join(final_rows)
            if final_rows
            else "<tr><td colspan='2'>Aucune sélection finale renseignée pour le moment.</td></tr>"
        )
        + "</tbody></table></div>"
        "<p class='meta'>Ce second tableau est alimenté automatiquement à partir des codes "
        "<code>E…</code> présents dans <code>proposition_finale</code>. "
        f"Export dédié : <code>{escape(selection_finale_xlsx)}</code>.</p>"
        + coherence_html
    )
    write_text(
        SITE / "suivi_positionnements.html",
        html_page(
            "Suivi positionnements",
            body,
            nav_current="suivi_intervenants.html",
            breadcrumb=html_breadcrumb(
                ("Accueil", "index.html"),
                ("Suivi Intervenants", "suivi_intervenants.html"),
                ("Suivi positionnements", None),
            ),
            page_header=(
                '<div class="page-head"><h1>Suivi positionnements</h1>'
                '<p class="lead">Proposition, réponse, préférences et arbitrage final par intervenant.</p></div>'
            ),
        ),
    )


def build_suivi_intervenants_pages(programme_table: dict, experts_profils: dict) -> None:
    """Ensemble Suivi Intervenants : sommaires Modules / Intervenants jusqu'aux videos expert."""
    expected: set[str] = set()
    by_module = _suivi_rows_by_module(programme_table)
    intervenants = _suivi_canonical_intervenants(programme_table, experts_profils)
    positionnements = _sync_suivi_positionnements(intervenants)
    build_suivi_positionnements_page(positionnements)
    expected.add("suivi_positionnements.html")
    expected.add("suivi_positionnements.xlsx")
    expected.add("selection_finale_videos_expertise.xlsx")
    rows_by_code = {row.get("code", ""): row for row in programme_table.get("rows", [])}

    def write_sommaire(
        filename: str,
        title: str,
        lead: str,
        sections: list[tuple[str, str, str, str]],
        crumbs: list[tuple[str, str | None]],
        meta: str = "",
    ) -> None:
        expected.add(filename)
        body = (f"<p class='meta'>{meta}</p>" if meta else "") + _sommaire_cards(sections)
        write_text(
            SITE / filename,
            html_page(
                title,
                body,
                nav_current="suivi_intervenants.html",
                breadcrumb=html_breadcrumb(*crumbs),
                page_header=(
                    f'<div class="page-head"><h1>{escape(title)}</h1>'
                    f'<p class="lead">{escape(lead)}</p></div>'
                ),
                main_class="page-home",
            ),
        )

    # 1) Racine
    write_sommaire(
        "suivi_intervenants.html",
        "Suivi Intervenants",
        "Parcourir le plan par modules ou par intervenants.",
        [
            (
                "suivi_modules.html",
                "▦",
                "Modules",
                f"{len(by_module)} modules — videos temoin puis videos expert.",
            ),
            (
                "suivi_intervenants_liste.html",
                "◎",
                "Intervenants",
                f"{len(intervenants)} intervenants — capsules, videos, mails et guides edito.",
            ),
            (
                "suivi_positionnements.html",
                "📋",
                "Suivi positionnements",
                "Proposition de notre part, réponse, préférences et proposition finale.",
            ),
        ],
        [("Accueil", "index.html"), ("Suivi Intervenants", None)],
        meta="Sommaire de suivi operationnel du plan de conception.",
    )

    # 2) Modules
    module_sections = []
    for module, rows in by_module.items():
        codes = ", ".join(row.get("code", "") for row in rows)
        module_sections.append(
            (
                f"suivi_module_{module}.html",
                module,
                _module_label(module),
                f"{len(rows)} video(s) temoin : {codes}.",
            )
        )
    write_sommaire(
        "suivi_modules.html",
        "Modules",
        "Choisir un module pour ouvrir ses videos temoin.",
        module_sections,
        [
            ("Accueil", "index.html"),
            ("Suivi Intervenants", "suivi_intervenants.html"),
            ("Modules", None),
        ],
    )

    # 3) Pages module → videos temoin
    for module, rows in by_module.items():
        temoin_sections = []
        for row in rows:
            code = row.get("code", "")
            label = FIXED_TEMOIN_PLAN.get(code, {}).get("label") or row.get("video_temoin", code)
            experts = _tb_edito_parse_videos_expert(row.get("videos_referent", ""))
            expert_codes = ", ".join(item.get("code", "") for item in experts) or "aucune video expert"
            temoin_sections.append(
                (
                    f"suivi_temoin_{code}.html",
                    code,
                    label,
                    f"Capsule temoin {code} — videos expert : {expert_codes}.",
                )
            )
        write_sommaire(
            f"suivi_module_{module}.html",
            _module_label(module),
            f"Videos temoin du {module}.",
            temoin_sections,
            [
                ("Accueil", "index.html"),
                ("Suivi Intervenants", "suivi_intervenants.html"),
                ("Modules", "suivi_modules.html"),
                (module, None),
            ],
        )

    # 4) Pages temoin → capsule + videos expert
    for code, row in rows_by_code.items():
        label = FIXED_TEMOIN_PLAN.get(code, {}).get("label") or row.get("video_temoin", code)
        module = row.get("module", "")
        experts = _tb_edito_parse_videos_expert(row.get("videos_referent", ""))
        proposes = ", ".join(_extract_intervenants(row.get("noms_proposes", ""))) or "à confirmer"
        sections: list[tuple[str, str, str, str]] = [
            (
                f"tb_edito_{code}.html",
                "🎙",
                f"Capsule témoin {code}",
                label,
            )
        ]
        for video in experts:
            e_code = video.get("code", "")
            sections.append(
                (
                    f"video_expert_{e_code}.html",
                    e_code,
                    _label_video_expert(e_code),
                    video.get("titre", "") or "Objectif à préciser",
                )
            )
        meta = (
            f"Module {escape(module)} — Intervenants proposes : {escape(proposes)}. "
            "La capsule temoin ouvre le montage edito ; chaque bloc expert ouvre la fiche video expert."
        )
        expected.add(f"suivi_temoin_{code}.html")
        body = f"<p class='meta'>{meta}</p>" + _sommaire_cards(sections)
        write_text(
            SITE / f"suivi_temoin_{code}.html",
            html_page(
                f"{code} — Suivi",
                body,
                nav_current="suivi_intervenants.html",
                breadcrumb=html_breadcrumb(
                    ("Accueil", "index.html"),
                    ("Suivi Intervenants", "suivi_intervenants.html"),
                    ("Modules", "suivi_modules.html"),
                    (module, f"suivi_module_{module}.html"),
                    (code, None),
                ),
                page_header=(
                    f'<div class="page-head"><h1>{escape(label)}</h1>'
                    f'<p class="lead">Capsule temoin et videos expert rattachees.</p></div>'
                ),
                main_class="page-home",
            ),
        )

    # 5) Liste intervenants (+ acces rapide mails / guides)
    intervenant_sections = [
        (
            "mails_experts.html",
            "✉",
            "Mails et guides éditoriaux",
            "Index des mails experts (par date d'envoi) et exports Word des guides éditoriaux.",
        )
    ]
    intervenant_sections.extend(
        (
            f"suivi_intervenant_{item['slug']}.html",
            "◎",
            item["nom"],
            f"{item['organisme']} — {len(item['capsules'])} capsule(s), "
            f"{len(item['videos_expert'])} video(s) · mail + guide.",
        )
        for item in intervenants
    )
    write_sommaire(
        "suivi_intervenants_liste.html",
        "Intervenants",
        "Fiche par intervenant : capsules, videos expertise, mail et guide editorial.",
        intervenant_sections,
        [
            ("Accueil", "index.html"),
            ("Suivi Intervenants", "suivi_intervenants.html"),
            ("Intervenants", None),
        ],
        meta="Chaque fiche intervenant donne aussi acces au mail et a l'export Word du guide.",
    )

    # 6) Pages intervenant → mail/guide + capsules temoin + videos expert
    finale_by_slug = {
        item.get("slug") or slug(item.get("nom", "")): (item.get("proposition_finale") or "").strip()
        for item in _load_suivi_positionnements().get("intervenants", [])
    }
    for item in intervenants:
        filename = f"suivi_intervenant_{item['slug']}.html"
        mail_href = f"mail_expert_{item['slug']}.html"
        mail_attendues_href = f"mail_videos_attendues_{item['slug']}.html"
        guide_href = f"guide_editorial_{item['slug']}.doc"
        guide_attendues_href = f"guide_videos_attendues_{item['slug']}.doc"
        has_finale = bool(finale_by_slug.get(item["slug"]))
        sections = [
            (
                mail_href,
                "✉",
                "Mail de positionnement",
                "Mail envoyé le 20/07/2026 — sollicitation de positionnement sur les vidéos expertise.",
            ),
        ]
        if has_finale:
            guide_simple_href = f"guide_editorial_simplifie_{item['slug']}.doc"
            sections.extend(
                [
                    (
                        mail_attendues_href,
                        "✉",
                        "Mail vidéos attendues",
                        "Mail du 27/07/2026 — sélection finale + guide éditorial simplifié (PJ principale).",
                    ),
                    (
                        guide_simple_href,
                        "📄",
                        "Guide éditorial simplifié (PJ principale)",
                        "1–2 pages : objectif, ce que disent les chercheurs, ce qu’on attend.",
                    ),
                    (
                        guide_attendues_href,
                        "📄",
                        "Guide détaillé (optionnel)",
                        "Version complète : scripts, transcript témoin, vue d’ensemble — sur demande.",
                    ),
                ]
            )
        sections.append(
            (
                guide_href,
                "📄",
                "Guide éditorial positionnement (Word)",
                "Export Word historique associé au mail de positionnement (20/07/2026).",
            )
        )
        for capsule in item["capsules"]:
            sections.append(
                (
                    f"suivi_temoin_{capsule['code']}.html",
                    capsule["code"],
                    capsule["label"] or capsule["code"],
                    f"Capsule temoin {capsule['code']} ({capsule.get('module') or '—'}).",
                )
            )
        for video in item["videos_expert"]:
            sections.append(
                (
                    f"video_expert_{video['code']}.html",
                    video["code"],
                    _label_video_expert(video["code"]),
                    video.get("titre", "") or f"Rattachee a {video.get('capsule_code', '')}",
                )
            )
        expected.add(filename)
        if has_finale:
            guide_simple_href = f"guide_editorial_simplifie_{item['slug']}.doc"
            primary_btns = (
                f"<a class='btn' href='{escape(mail_attendues_href)}'>Ouvrir le mail vidéos attendues (27/07)</a> "
                f"<a class='btn' href='{escape(guide_simple_href)}' download>Exporter le guide éditorial simplifié (Word)</a> "
                f"<a class='btn' href='{escape(guide_simple_href)}' target='_blank' rel='noopener'>Ouvrir le guide éditorial simplifié</a> "
                f"<a class='btn' href='{escape(guide_attendues_href)}' download>Guide détaillé (optionnel)</a>"
            )
        else:
            primary_btns = (
                f"<a class='btn' href='{escape(mail_href)}'>Ouvrir le mail de positionnement</a> "
                f"<a class='btn' href='{escape(guide_href)}' download>Exporter le guide positionnement (Word)</a> "
                f"<a class='btn' href='{escape(guide_href)}' target='_blank' rel='noopener'>Ouvrir le guide positionnement</a>"
            )
        body = (
            f"<p class='meta'>{escape(item['organisme'])} — "
            f"{len(item['capsules'])} capsule(s) temoin · "
            f"{len(item['videos_expert'])} video(s) expert."
            + (
                f" · Sélection finale : <strong>{escape(finale_by_slug.get(item['slug'], ''))}</strong>"
                if has_finale
                else ""
            )
            + "</p>"
            f"<p>{primary_btns}</p>"
            + _sommaire_cards(sections)
        )
        write_text(
            SITE / filename,
            html_page(
                item["nom"],
                body,
                nav_current="suivi_intervenants.html",
                breadcrumb=html_breadcrumb(
                    ("Accueil", "index.html"),
                    ("Suivi Intervenants", "suivi_intervenants.html"),
                    ("Intervenants", "suivi_intervenants_liste.html"),
                    (item["nom"], None),
                ),
                page_header=(
                    f'<div class="page-head"><h1>{escape(item["nom"])}</h1>'
                    f'<p class="lead">Mail, guide editorial, capsules temoin et videos expertise.</p></div>'
                ),
                main_class="page-home",
            ),
        )

    for path in SITE.glob("suivi_*.html"):
        if path.name not in expected:
            path.unlink()


def build_edito_hub_page() -> None:
    sections = [
        (
            "script_propose.html",
            "🎬",
            "Script proposé",
            "Montages realistes et comprehensibles (grain tournage), distincts de la banque Clarisse.",
        ),
        (
            "proposition_edito.html",
            "➕",
            "Proposition édito",
            "Validation des selections Clarisse et propositions d'ajouts BAB.",
        ),
        (
            "tableau_correspondances_edito.html",
            "⌗",
            "Correspondances édito",
            "Tableau de conception avec liens vers les titres video proposes.",
        ),
        (
            "derushage_edito.html",
            "✎",
            "Dérushage édito",
            "Sequences surlignees par l'edito dans les transcripts corriges.",
        ),
        (
            "tb_edito.html",
            "🗂",
            "Capsules témoins (Clarisse)",
            "Banque de surlignages Clarisse — grain fin, pas le script de tournage.",
        ),
    ]
    body = (
        "<p class='meta'>Espace edito : script propose (lisible), selections Clarisse et derushage.</p>"
        + _sommaire_cards(sections)
    )
    write_text(
        SITE / "edito.html",
        html_page(
            "Edito",
            body,
            nav_current="edito.html",
            breadcrumb=html_breadcrumb(("Accueil", "index.html"), ("Edito", None)),
            page_header='<div class="page-head"><h1>Edito</h1><p class="lead">Travaux et livrables de l’éditorialisation.</p></div>',
            main_class="page-home",
        ),
    )


def _merge_consecutive_clarisse_blocks(sequences: list[dict]) -> list[dict]:
    """Regroupe les fragments Clarisse consecutifs d'une meme voix en unites lisibles."""
    merged: list[dict] = []
    for seq in sequences:
        voice = (seq.get("intervenant") or "").strip()
        text = (seq.get("texte") or "").strip()
        sid = seq.get("id", "")
        if not text:
            continue
        if merged and merged[-1]["intervenant"] == voice:
            merged[-1]["texte"] = f"{merged[-1]['texte']} {text}".strip()
            merged[-1]["ids"].append(sid)
            continue
        merged.append(
            {
                "intervenant": voice,
                "texte": text,
                "ids": [sid] if sid else [],
                "video": seq.get("video", ""),
                "source_doc": seq.get("source_doc", ""),
            }
        )
    return merged


def _script_propose_from_clarisse(code: str, sequences: list[dict], max_blocks: int = 10) -> tuple[str, list[dict]]:
    """Proposition provisoire T13 : Clarisse consolidee, chorale, plafond de blocs."""
    ordered = _tb_edito_order_for_code(code, sequences)
    merged = _merge_consecutive_clarisse_blocks(ordered)
    # Garde un enchainement choral sans exploser la densite
    selected: list[dict] = []
    voice_counts: dict[str, int] = defaultdict(int)
    for block in merged:
        voice = block["intervenant"] or "?"
        if voice_counts[voice] >= 3:
            continue
        selected.append(block)
        voice_counts[voice] += 1
        if len(selected) >= max_blocks:
            break

    label = FIXED_TEMOIN_PLAN.get(code, {}).get("label", _label_video_temoin(code))
    parts = [
        f"[CADRAGE — INTRO] Animateur | NON PRONONCE | Avant {selected[0]['ids'][0] if selected and selected[0]['ids'] else code}",
        f"Proposition de montage provisoire pour {label} — unites consolidees depuis les surlignages Clarisse "
        "(a arbitrer avec l'edito et le monteur).",
        "",
    ]
    for block in selected:
        ids_label = " + ".join(block["ids"]) if block["ids"] else "SANS-ID"
        parts.append(
            f"[{ids_label}] {block['intervenant']} | {block.get('source_doc') or 'derushage_edito'} | Clarisse consolide"
        )
        parts.append(block["texte"])
        parts.append("")
    return "\n".join(parts).strip(), selected


def _script_propose_stats_for_code(
    code: str,
    affectations: dict,
    segments_by_id: dict[str, dict],
    clarisse_count: int,
) -> dict:
    capsule = affectations.get("capsules", {}).get(code, {})
    script = (capsule.get("script_final") or "").strip()
    ordre = capsule.get("ordre_montage") or capsule.get("extraits_utilises") or []
    duree = capsule_duration(code, segments_by_id, affectations) if ordre else 0.0
    return {
        "has_bab_script": bool(script),
        "nb_blocs": len(ordre) if ordre else 0,
        "duree": duree,
        "clarisse_count": clarisse_count,
        "script": script,
    }


def build_script_propose_pages(programme_table: dict, affectations: dict, segments: list[dict]) -> None:
    """Ensemble Edito > Script propose : montages realistes (BAB) vs banque Clarisse."""
    segments_by_id = index_by_id(segments)
    grouped_clarisse = _tb_edito_sequences_by_code()
    rows_by_code = {row.get("code", ""): row for row in programme_table.get("rows", [])}
    expected = {"script_propose.html"}
    cards = []

    for code, spec in sorted(FIXED_TEMOIN_PLAN.items(), key=lambda item: int(item[0][1:])):
        page_name = f"script_propose_{code}.html"
        expected.add(page_name)
        clarisse_seqs = grouped_clarisse.get(code, [])
        clarisse_ordered = _tb_edito_order_for_code(code, clarisse_seqs)
        stats = _script_propose_stats_for_code(code, affectations, segments_by_id, len(clarisse_ordered))
        label = spec.get("label", _label_video_temoin(code))
        row = rows_by_code.get(code, {})
        objective = row.get("objectif_pedagogique", "")

        source_note = ""
        script_text = stats["script"]
        blocks_meta = stats["nb_blocs"]
        duree_label = format_seconds(stats["duree"]) if stats["duree"] else "—"

        if not script_text:
            # T13 (ou capsule sans montage BAB) : proposition consolidee Clarisse
            script_text, selected = _script_propose_from_clarisse(code, clarisse_seqs)
            blocks_meta = len(selected)
            duree_label = "a estimer"
            source_note = (
                "Source : proposition provisoire construite par consolidation des surlignages Clarisse "
                "(fragments de la meme voix regroupes). Pas encore de montage BAB valide."
            )
        else:
            source_note = (
                "Source : montage BAB valide (<code>affectations.json</code> / script_final) — "
                "grain realiste pour tournage et preparation experts."
            )

        cards.append(
            (
                page_name,
                code,
                label,
                f"{blocks_meta} blocs · {duree_label} · Clarisse banque : {stats['clarisse_count']} fragments",
            )
        )

        voice_preview = []
        for sid in (affectations.get("capsules", {}).get(code, {}).get("ordre_montage")
                    or affectations.get("capsules", {}).get(code, {}).get("extraits_utilises")
                    or []):
            seg = segments_by_id.get(sid)
            if seg:
                voice_preview.append(
                    f"<li><strong>{escape(sid)}</strong> — {escape(seg.get('chercheur', ''))} "
                    f"<span class='meta'>{escape(seg.get('debut', ''))} → {escape(seg.get('fin', ''))}</span></li>"
                )

        if not voice_preview and script_text:
            # T13 consolidated ids
            for match in re.finditer(r"\[([^\]]+)\]\s+([^|]+)\|", script_text):
                voice_preview.append(
                    f"<li><strong>{escape(match.group(1))}</strong> — {escape(match.group(2).strip())}</li>"
                )

        body = (
            f"<p class='meta'><strong>Objectif :</strong> {escape(objective or '—')}</p>"
            f"<p class='meta'>{source_note}</p>"
            "<div class='methodology-panel'>"
            "<p><strong>Lecture :</strong> ce script est propose comme base <em>comprehensible</em> "
            "pour le monteur / la preparation expert. "
            f"La banque Clarisse (<a href='tb_edito_{escape(code)}.html'>tb_edito_{escape(code)}</a>) "
            f"reste disponible ({stats['clarisse_count']} fragments) mais n'est pas le decoupage de tournage.</p>"
            f"<p class='meta'>Blocs : <strong>{blocks_meta}</strong> · Duree estimee : <strong>{escape(duree_label)}</strong></p>"
            "</div>"
            "<h2>Script proposé</h2>"
            f"<div class='script' id='script-final'>{escape(script_text) if script_text else 'A construire.'}</div>"
            "<h2>Ordre des voix / extraits</h2>"
            + (
                "<ul>" + "".join(voice_preview) + "</ul>"
                if voice_preview
                else "<p class='meta'>Aucun extrait structure pour cette capsule.</p>"
            )
            + "<p>"
            f"<a class='btn btn-secondary' href='tb_edito_{escape(code)}.html'>Voir banque Clarisse</a> "
            f"<a class='btn btn-secondary' href='capsule_{escape(code)}.html'>Fiche capsule</a> "
            "<a class='btn btn-secondary' href='script_propose.html'>← Tous les scripts</a>"
            "</p>"
        )
        write_text(
            SITE / page_name,
            html_page(
                f"Script proposé — {code}",
                body,
                nav_current="edito.html",
                breadcrumb=html_breadcrumb(
                    ("Accueil", "index.html"),
                    ("Edito", "edito.html"),
                    ("Script proposé", "script_propose.html"),
                    (code, None),
                ),
                page_header=(
                    f'<div class="page-head"><h1>{escape(label)}</h1>'
                    f'<p class="lead">Script proposé — montage réaliste et compréhensible.</p></div>'
                ),
            ),
        )

    hub_body = (
        "<p class='meta'>Scripts proposes pour un montage <strong>realiste et comprehensible</strong> "
        "(unites de sens / montage BAB). Distinct de la banque Clarisse fragmentee "
        "(<a href='tb_edito.html'>Capsules temoins</a>), destinee a l'edito et au monteur.</p>"
        + _sommaire_cards(cards)
    )
    write_text(
        SITE / "script_propose.html",
        html_page(
            "Script proposé",
            hub_body,
            nav_current="edito.html",
            breadcrumb=html_breadcrumb(
                ("Accueil", "index.html"),
                ("Edito", "edito.html"),
                ("Script proposé", None),
            ),
            page_header=(
                '<div class="page-head"><h1>Script proposé</h1>'
                '<p class="lead">Montages lisibles pour tournage et préparation des experts.</p></div>'
            ),
            main_class="page-home",
        ),
    )

    for path in SITE.glob("script_propose*.html"):
        if path.name not in expected:
            path.unlink()


def build_fichiers_travail_pages() -> None:
    # Sous-ensemble Informations
    info_sections = [
        (
            "profils_experts.html",
            "👤",
            "Profils experts",
            "Informations collectees (LinkedIn/institutions) et sources par profil.",
        ),
        (
            "fonctions_temoins.html",
            "🪪",
            "Fonctions témoins",
            "Formulations academiques et professionnelles des temoins.",
        ),
        (
            "proposition_titres_temoin.html",
            "🏷",
            "Titres témoins",
            "Propositions de titres par video temoin.",
        ),
    ]
    write_text(
        SITE / "informations.html",
        html_page(
            "Informations",
            "<p class='meta'>Fiches de reference sur les intervenants et les temoins.</p>"
            + _sommaire_cards(info_sections),
            nav_current="fichiers_travail.html",
            breadcrumb=html_breadcrumb(
                ("Accueil", "index.html"),
                ("Fichiers de travail", "fichiers_travail.html"),
                ("Informations", None),
            ),
            page_header='<div class="page-head"><h1>Informations</h1><p class="lead">Profils, fonctions et titres.</p></div>',
            main_class="page-home",
        ),
    )

    # Ensemble Fichiers de travail
    sections = [
        (
            "informations.html",
            "ℹ",
            "Informations",
            "Profils experts, fonctions temoins et titres temoins.",
        ),
        (
            "prev_vid.html",
            "▦",
            "Prev Vid",
            "Tableau de conception (13 videos temoin) au format Excel.",
        ),
        (
            "videos_expert.html",
            "▶",
            "Vidéos expert",
            "Tableau des videos expertise, consignes et scripts recus.",
        ),
        (
            "mails_experts.html",
            "✉",
            "Mails experts",
            "Mails envoyés aux experts, classés par date (mail de positionnement, etc.).",
        ),
        (
            "bab_encodes.html",
            "▣",
            "BAB encodé",
            "Parcours des BAB timecodes par chercheur.",
        ),
        (
            "fascicules_oser_innover.html",
            "📘",
            "Oser pour innover (fascicules)",
            "Document PUI en deux fascicules + corrélation grains MOOC ↔ chapitres.",
        ),
    ]
    write_text(
        SITE / "fichiers_travail.html",
        html_page(
            "Fichiers de travail",
            "<p class='meta'>Documents et exports de travail pour le suivi du MOOC.</p>"
            + _sommaire_cards(sections),
            nav_current="fichiers_travail.html",
            breadcrumb=html_breadcrumb(("Accueil", "index.html"), ("Fichiers de travail", None)),
            page_header='<div class="page-head"><h1>Fichiers de travail</h1><p class="lead">Informations et livrables operationnels.</p></div>',
            main_class="page-home",
        ),
    )


def _load_fascicules_oser_innover() -> dict:
    path = ROOT / "data" / "fascicules_oser_innover" / "fascicules.json"
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def _fascicule_chapitre_index(data: dict) -> dict[str, dict]:
    index: dict[str, dict] = {}
    for fasc in data.get("fascicules", []):
        for chap in fasc.get("chapitres", []):
            index[chap["id"]] = {
                **chap,
                "fascicule_id": fasc["id"],
                "fascicule_numero": fasc["numero"],
                "fascicule_titre": fasc["titre"],
                "fichier": fasc["fichier"],
            }
    return index


def _fascicule_refs_for_grain(grain_code: str) -> list[dict]:
    """Pages du fascicule « Oser pour innover » corrélées à un grain T/E."""
    code = (grain_code or "").strip()
    if not code:
        return []
    data = _load_fascicules_oser_innover()
    if not data:
        return []
    chap_index = _fascicule_chapitre_index(data)
    refs: list[dict] = []
    seen: set[tuple[str, str]] = set()
    for corr in data.get("correlations") or []:
        if corr.get("grain") != code:
            continue
        for chap_id in corr.get("chapitres") or []:
            chap = chap_index.get(chap_id)
            if not chap:
                continue
            key = (str(chap.get("fascicule_numero")), str(chap.get("pages") or ""))
            if key in seen:
                continue
            seen.add(key)
            refs.append(
                {
                    "grain": code,
                    "fascicule_numero": chap.get("fascicule_numero"),
                    "fascicule_titre": chap.get("fascicule_titre") or "",
                    "chapitre_titre": chap.get("titre") or "",
                    "pages": chap.get("pages") or "",
                    "fichier": chap.get("fichier") or "",
                    "justification": corr.get("justification") or "",
                }
            )
    return refs


def _format_fascicule_ref_line(ref: dict) -> str:
    numero = ref.get("fascicule_numero")
    pages = (ref.get("pages") or "").strip()
    titre = (ref.get("chapitre_titre") or "").strip()
    base = f"Fascicule {numero}"
    if pages:
        base += f", p. {pages}"
    if titre:
        base += f" — {titre}"
    return base


def _fascicule_refs_html(grain_code: str) -> str:
    refs = _fascicule_refs_for_grain(grain_code)
    if not refs:
        return ""
    items = "".join(
        f"<li>{escape(_format_fascicule_ref_line(ref))}</li>" for ref in refs
    )
    return (
        "<p class='meta'><strong>Repère facultatif — guide « Oser pour innover » :</strong> "
        "si vous souhaitez un appui conceptuel, ces pages peuvent éclairer votre vidéo expertise.</p>"
        f"<ul>{items}</ul>"
    )


def _fascicule_refs_plaintext_lines(grain_code: str) -> list[str]:
    refs = _fascicule_refs_for_grain(grain_code)
    if not refs:
        return []
    lines = [
        "Repère facultatif — guide « Oser pour innover » : "
        "si vous souhaitez un appui conceptuel, ces pages peuvent éclairer votre vidéo expertise.",
    ]
    for ref in refs:
        lines.append(f"- {_format_fascicule_ref_line(ref)}")
    lines.append("")
    return lines


def build_fascicules_oser_innover_pages() -> None:
    """Ensemble Oser pour innover : document (2 fascicules) + corrélation grains."""
    data = _load_fascicules_oser_innover()
    if not data:
        return

    src_dir = ROOT / "data" / "fascicules_oser_innover"
    for fasc in data.get("fascicules", []):
        src = src_dir / fasc["fichier"]
        if src.exists():
            (SITE / fasc["fichier"]).write_bytes(src.read_bytes())

    chap_index = _fascicule_chapitre_index(data)

    # --- Ensemble ---
    write_text(
        SITE / "fascicules_oser_innover.html",
        html_page(
            "Oser pour innover (fascicules)",
            "<p class='meta'>"
            + escape(data.get("note", ""))
            + "</p>"
            + _sommaire_cards(
                [
                    (
                        "fascicules_document.html",
                        "📄",
                        "Le document",
                        "Les deux fascicules imprimés (chapitres) et leurs sommaires.",
                    ),
                    (
                        "fascicules_grains.html",
                        "🔗",
                        "Corrélation grains ↔ fascicules",
                        "Chaque grain MOOC (vidéo témoin / vidéo expertise) relié aux chapitres.",
                    ),
                ]
            ),
            nav_current="fichiers_travail.html",
            breadcrumb=html_breadcrumb(
                ("Accueil", "index.html"),
                ("Fichiers de travail", "fichiers_travail.html"),
                ("Oser pour innover (fascicules)", None),
            ),
            page_header=(
                f'<div class="page-head"><h1>{escape(data.get("titre", "Oser pour innover"))}</h1>'
                f'<p class="lead">{escape(data.get("sous_titre", ""))}</p></div>'
            ),
            main_class="page-home",
        ),
    )

    # --- Sous-ensemble 1 : le document ---
    fasc_cards = []
    for fasc in data.get("fascicules", []):
        chap_rows = []
        for chap in fasc.get("chapitres", []):
            sous = chap.get("sous_sections") or []
            sous_html = (
                "<ul>" + "".join(f"<li>{escape(item)}</li>" for item in sous) + "</ul>"
                if sous
                else "—"
            )
            chap_rows.append(
                "<tr>"
                f"<td><code>{escape(chap['id'])}</code></td>"
                f"<td>{escape(chap['titre'])}</td>"
                f"<td>{escape(chap.get('pages', ''))}</td>"
                f"<td>{sous_html}</td>"
                "</tr>"
            )
        fasc_cards.append(
            "<article class='card'>"
            f"<h2>Fascicule n°{escape(str(fasc['numero']))} — {escape(fasc['titre'])}</h2>"
            f"<p class='meta'>Version {escape(fasc.get('version', ''))} · "
            f"{escape(str(fasc.get('pages', '')))} pages · "
            f"fichier source : <code>{escape(fasc.get('fichier_source', ''))}</code></p>"
            f"<p><a class='btn' href='{escape(fasc['fichier'])}' target='_blank' rel='noopener'>"
            f"Ouvrir le PDF</a> "
            f"<a class='btn' href='{escape(fasc['fichier'])}' download>Télécharger</a></p>"
            "<div class='table-wrap'><table><thead><tr>"
            "<th>Id</th><th>Chapitre</th><th>Pages</th><th>Sous-sections</th>"
            "</tr></thead><tbody>"
            + "".join(chap_rows)
            + "</tbody></table></div>"
            "</article>"
        )

    write_text(
        SITE / "fascicules_document.html",
        html_page(
            "Le document — fascicules Oser pour innover",
            (
                f"<p class='meta'>{escape(data.get('action', ''))}</p>"
                "<p>Document imprimé en <strong>deux fascicules</strong>, "
                "correspondant aux grands chapitres de référence conceptuelle du PUI.</p>"
                + "".join(fasc_cards)
            ),
            nav_current="fichiers_travail.html",
            breadcrumb=html_breadcrumb(
                ("Accueil", "index.html"),
                ("Fichiers de travail", "fichiers_travail.html"),
                ("Oser pour innover (fascicules)", "fascicules_oser_innover.html"),
                ("Le document", None),
            ),
            page_header=(
                '<div class="page-head"><h1>Le document</h1>'
                '<p class="lead">Fascicule 1 (concepts &amp; écosystème) et fascicule 2 (méthode &amp; dispositifs).</p></div>'
            ),
        ),
    )

    # --- Sous-ensemble 2 : corrélation grains ---
    rows_html = []
    for item in data.get("correlations", []):
        grain = item.get("grain", "")
        typ = item.get("type", "")
        type_label = "Vidéo témoin" if typ == "temoin" else "Vidéo expertise"
        fasc_id = item.get("fascicule", "")
        fasc_meta = next(
            (f for f in data.get("fascicules", []) if f["id"] == fasc_id),
            {},
        )
        chap_bits = []
        for cid in item.get("chapitres", []):
            chap = chap_index.get(cid, {})
            label = chap.get("titre") or cid
            pages = chap.get("pages") or ""
            chap_bits.append(
                f"<li><code>{escape(cid)}</code> — {escape(label)}"
                + (f" <span class='meta'>(p. {escape(pages)})</span>" if pages else "")
                + "</li>"
            )
        pdf_href = fasc_meta.get("fichier", "")
        rows_html.append(
            "<tr>"
            f"<td><strong>{escape(grain)}</strong><br>"
            f"<span class='meta'>{escape(type_label)}</span></td>"
            f"<td>{escape(item.get('titre', ''))}</td>"
            f"<td><strong>Fascicule {escape(str(fasc_meta.get('numero', fasc_id)))}</strong><br>"
            f"<span class='meta'>{escape(fasc_meta.get('titre', ''))}</span>"
            + (
                f"<br><a href='{escape(pdf_href)}' target='_blank' rel='noopener'>PDF</a>"
                if pdf_href
                else ""
            )
            + "</td>"
            f"<td><ul>{''.join(chap_bits)}</ul></td>"
            f"<td>{escape(item.get('justification', ''))}</td>"
            "</tr>"
        )

    # Vue par fascicule
    by_fasc: dict[str, list[dict]] = {"F1": [], "F2": []}
    for item in data.get("correlations", []):
        by_fasc.setdefault(item.get("fascicule", ""), []).append(item)

    fasc_blocks = []
    for fasc in data.get("fascicules", []):
        items = by_fasc.get(fasc["id"], [])
        grains = ", ".join(escape(i["grain"]) for i in items) or "—"
        fasc_blocks.append(
            "<article class='card'>"
            f"<h2>Fascicule n°{escape(str(fasc['numero']))}</h2>"
            f"<p class='meta'>{escape(fasc['titre'])}</p>"
            f"<p><strong>Grains corrélés :</strong> {grains}</p>"
            f"<p><a class='btn' href='{escape(fasc['fichier'])}' target='_blank' rel='noopener'>"
            f"Ouvrir le PDF</a></p>"
            "</article>"
        )

    write_text(
        SITE / "fascicules_grains.html",
        html_page(
            "Corrélation grains ↔ fascicules",
            (
                "<p class='meta'>Lecture utile pour cadrer chaque grain du MOOC "
                "(vidéo témoin / vidéo expertise) sur les chapitres des fascicules imprimés. "
                "Corrélation éditoriale indicative — à affiner si besoin.</p>"
                "<h2>Vue par fascicule</h2>"
                + "".join(fasc_blocks)
                + "<h2>Tableau grain par grain</h2>"
                "<div class='table-wrap'><table><thead><tr>"
                "<th>Grain</th><th>Titre</th><th>Fascicule</th><th>Chapitres</th><th>Justification</th>"
                "</tr></thead><tbody>"
                + "".join(rows_html)
                + "</tbody></table></div>"
            ),
            nav_current="fichiers_travail.html",
            breadcrumb=html_breadcrumb(
                ("Accueil", "index.html"),
                ("Fichiers de travail", "fichiers_travail.html"),
                ("Oser pour innover (fascicules)", "fascicules_oser_innover.html"),
                ("Corrélation grains", None),
            ),
            page_header=(
                '<div class="page-head"><h1>Corrélation grains ↔ fascicules</h1>'
                '<p class="lead">Chaque grain du parcours relié aux chapitres des deux fascicules.</p></div>'
            ),
        ),
    )


def build_prev_vid_page(programme_table: dict) -> None:
    """Ensemble Prev Vid : tableau HTML + export Excel (format 20260710, 13 videos)."""
    from export_prev_vid_xlsx import export_prev_vid_xlsx

    data_xlsx = export_prev_vid_xlsx(programme_table, ROOT / "data" / "20260723_Prev_Vid.xlsx")
    site_xlsx_name = "20260723_Prev_Vid.xlsx"
    site_xlsx = SITE / site_xlsx_name
    site_xlsx.write_bytes(data_xlsx.read_bytes())

    headers = programme_table.get("headers", {})
    col_keys = [
        "module",
        "code",
        "video_temoin",
        "resume_chercheurs",
        "videos_referent",
        "objectif_pedagogique",
        "noms_proposes",
    ]
    thead = "".join(f"<th>{escape(headers.get(key, key))}</th>" for key in col_keys)

    rows_html = []
    for row in programme_table.get("rows", []):
        cells = []
        for key in col_keys:
            raw = row.get(key) or ""
            if key == "code":
                code = row.get("code", "")
                cells.append(
                    f"<td><a href='tb_edito_{escape(code)}.html'><strong>{escape(code)}</strong></a></td>"
                )
                continue
            cells.append(f"<td>{escape(raw).replace(chr(10), '<br>')}</td>")
        rows_html.append("<tr>" + "".join(cells) + "</tr>")

    n_temoin = len(programme_table.get("rows", []))
    n_expert = sum(
        1
        for row in programme_table.get("rows", [])
        for _ in _tb_edito_parse_videos_expert(row.get("videos_referent", ""))
    )
    source = programme_table.get("source_document", site_xlsx_name)
    date_maj = programme_table.get("date_mise_a_jour", "")
    note = programme_table.get("note", "")

    body = (
        f"<p class='meta'>Tableau de conception au format Excel historique "
        f"(<code>20260710_Prev_Vid.xlsx</code>), mis a jour pour le plan a "
        f"<strong>{n_temoin} videos temoin</strong> (T1–T13) et "
        f"<strong>{n_expert} videos expert</strong> associees.</p>"
        f"<p class='meta'>Fichier : <code>{escape(source)}</code>"
        f"{f' — mis a jour le {escape(date_maj)}' if date_maj else ''}.</p>"
        f"<p class='meta'>{escape(note)}</p>"
        f"<p><a class='btn' href='{escape(site_xlsx_name)}' download>"
        f"Télécharger Prev_Vid (XLSX)</a></p>"
        "<div class='table-wrap'><table><thead><tr>"
        + thead
        + "</tr></thead><tbody>"
        + (
            "".join(rows_html)
            if rows_html
            else "<tr><td colspan='7'>Aucune ligne dans programme_table.</td></tr>"
        )
        + "</tbody></table></div>"
        "<p class='meta'>L'original a 12 videos reste intact dans "
        "<code>data/raw/20260710_Prev_Vid.xlsx</code>.</p>"
    )
    write_text(
        SITE / "prev_vid.html",
        html_page(
            "Prev Vid",
            body,
            nav_current="fichiers_travail.html",
            breadcrumb=html_breadcrumb(("Accueil", "index.html"), ("Fichiers de travail", "fichiers_travail.html"), ("Prev Vid", None)),
            page_header=(
                '<div class="page-head"><h1>Prev Vid</h1>'
                '<p class="lead">Tableau de conception — 13 vidéos témoin, export Excel.</p></div>'
            ),
        ),
    )


def build_videos_expert_pages(programme_table: dict, experts_profils: dict) -> None:
    inventory = _inventory_videos_expert(programme_table, experts_profils)
    _write_videos_expert_xlsx(inventory, SITE / "videos_expert.xlsx")

    json_rows = [
        {
            "code": item["code"],
            "titre": item.get("titre", ""),
            "capsule_code": item.get("capsule_code", ""),
            "temoin_label": item.get("temoin_label", ""),
            "module": item.get("module", ""),
            "experts": item.get("experts_label", ""),
            "script_statut": item.get("script_statut", "EN_ATTENTE"),
            "script_fichier": item.get("script_fichier", ""),
            "page": item.get("page_href", ""),
        }
        for item in inventory
    ]
    write_text(
        SITE / "videos_expert.json",
        json.dumps({"videos_expert": json_rows}, ensure_ascii=False, indent=2),
    )

    recus = sum(1 for item in inventory if item.get("script_statut") == "RECU")
    table_rows = []
    for item in inventory:
        statut_label = "Reçu" if item["script_statut"] == "RECU" else "En attente"
        badge = status_badge("VALIDEE" if item["script_statut"] == "RECU" else "EN_CONSTRUCTION")
        table_rows.append(
            "<tr>"
            f"<td><a href='{escape(item['page_href'])}'><strong>{escape(item['code'])}</strong></a></td>"
            f"<td>{escape(item.get('titre', ''))}</td>"
            f"<td><a href='{escape(item['tb_edito_href'])}'>{escape(item['capsule_code'])}</a></td>"
            f"<td>{escape(item.get('temoin_label', ''))}</td>"
            f"<td>{escape(item.get('experts_label', ''))}</td>"
            f"<td>{badge} {escape(statut_label)}</td>"
            f"<td><a class='btn' href='{escape(item['page_href'])}'>Ouvrir</a></td>"
            "</tr>"
        )

    body = (
        "<p class='meta'>Inventaire des videos expertise du programme de conception. "
        "Chaque fiche reprend les consignes envoyees dans le guide Word et accueille le script "
        "quand l'expert le renvoie (<code>data/videos_expert/scripts_recus/</code>).</p>"
        f"<p class='meta'><strong>{len(inventory)}</strong> videos — "
        f"<strong>{recus}</strong> script(s) recu(s), "
        f"<strong>{len(inventory) - recus}</strong> en attente.</p>"
        "<p>"
        "<a class='btn' href='videos_expert.xlsx' download>Télécharger le tableau (XLSX)</a> "
        "<a class='btn btn-secondary' href='videos_expert.json' download>JSON</a>"
        "</p>"
        "<div class='table-wrap'><table><thead><tr>"
        "<th>Code</th><th>Titre / objectif</th><th>Capsule</th><th>Titre temoin</th>"
        "<th>Experts proposes</th><th>Script</th><th></th>"
        "</tr></thead><tbody>"
        + (
            "".join(table_rows)
            if table_rows
            else "<tr><td colspan='7'>Aucune video expert dans le programme_table.</td></tr>"
        )
        + "</tbody></table></div>"
    )
    write_text(
        SITE / "videos_expert.html",
        html_page(
            "Vidéos expert",
            body,
            nav_current="fichiers_travail.html",
            breadcrumb=html_breadcrumb(("Accueil", "index.html"), ("Fichiers de travail", "fichiers_travail.html"), ("Vidéos expert", None)),
            page_header=(
                '<div class="page-head"><h1>Vidéos expert</h1>'
                '<p class="lead">Tableau des videos expertise, consignes transmises et scripts recus.</p></div>'
            ),
        ),
    )

    expected = {"videos_expert.html", "videos_expert.xlsx", "videos_expert.json"}
    for item in inventory:
        page_name = item["page_href"]
        expected.add(page_name)
        detail_body = (
            f"<p class='meta'>Module : {escape(item.get('module') or '—')} — "
            f"Capsule : <a href='{escape(item['tb_edito_href'])}'>{escape(item['capsule_code'])}</a> — "
            f"Experts : {escape(item.get('experts_label', ''))}</p>"
            f"{_consignes_envoyees_expert_html(item)}"
            f"{_script_expert_recu_html(item)}"
            f"{_script_expert_revues_html(item['code'])}"
            "<p><a class='btn btn-secondary' href='videos_expert.html'>← Retour au tableau</a> "
            f"<a class='btn btn-secondary' href='videos_expert.xlsx' download>Export XLSX</a></p>"
        )
        write_text(
            SITE / page_name,
            html_page(
                f"{item['code']} — Vidéo expert",
                detail_body,
                nav_current="fichiers_travail.html",
                breadcrumb=html_breadcrumb(
                    ("Accueil", "index.html"),
                    ("Fichiers de travail", "fichiers_travail.html"),
                    ("Vidéos expert", "videos_expert.html"),
                    (item["code"], None),
                ),
                page_header=(
                    f'<div class="page-head"><h1>{escape(_label_video_expert(item["code"]))}</h1>'
                    f'<p class="lead">{escape(item.get("titre") or "")}</p></div>'
                ),
            ),
        )

    for path in SITE.glob("video_expert_*.html"):
        if path.name not in expected:
            path.unlink()



def _filter_capsule_data_for_expert(
    capsule_data: dict,
    expert_video_codes: list[str],
    programme_row: dict | None = None,
) -> dict:
    """Restreint videos_expert / orientations aux codes attendus pour cet intervenant."""
    data = dict(capsule_data or {})
    wanted = {code for code in expert_video_codes if code}
    videos = list(data.get("videos_expert") or [])
    if not videos and programme_row:
        videos = _tb_edito_parse_videos_expert(programme_row.get("videos_referent", ""))
    if wanted:
        videos = [item for item in videos if item.get("code") in wanted]
    data["videos_expert"] = videos

    orientations = list(data.get("orientations_expert") or [])
    if data.get("orientation_expert") and not orientations:
        orientations = [data["orientation_expert"]]
    if wanted:
        orientations = [item for item in orientations if item.get("code") in wanted]
    data["orientations_expert"] = orientations
    if "orientation_expert" in data:
        data.pop("orientation_expert", None)
    return data


def _e_code_to_temoin_map(programme_table: dict) -> dict[str, dict]:
    """Mappe chaque code E* vers la capsule témoin du programme (préfère T* à GEN)."""
    mapping: dict[str, dict] = {}
    for row in programme_table.get("rows", []):
        capsule = row.get("code", "")
        if not capsule:
            continue
        fixed = FIXED_TEMOIN_PLAN.get(capsule, {})
        video_temoin_label = fixed.get("label") or row.get("video_temoin", "")
        objective = row.get("objectif_pedagogique", "")
        for video in _tb_edito_parse_videos_expert(row.get("videos_referent", "")):
            code = (video.get("code") or "").upper().replace(" ", "")
            if not code:
                continue
            current = mapping.get(code)
            prefer_new = current is None or (
                str(current.get("capsule", "")).startswith("GEN")
                and str(capsule).startswith("T")
            )
            if prefer_new:
                mapping[code] = {
                    "capsule": capsule,
                    "video_temoin_label": video_temoin_label,
                    "objectif_temoin": objective,
                    "titre": "",
                    "descriptif": "",
                }
    # Titres / descriptifs propres depuis programme_videos (évite les anciens noms collés au titre).
    catalogue = {item["code"]: item for item in _load_expert_videos_catalogue()}
    prog_path = ROOT / "data" / "programme_videos.json"
    if prog_path.exists():
        payload = json.loads(prog_path.read_text(encoding="utf-8"))
        for capsule, block in (payload.get("capsules") or {}).items():
            if not isinstance(block, dict):
                continue
            for video in block.get("videos_expert") or []:
                code = (video.get("code") or "").upper().replace(" ", "")
                if not code:
                    continue
                current = mapping.get(code)
                prefer_new = current is None or (
                    str(current.get("capsule", "")).startswith("GEN")
                    and str(capsule).startswith("T")
                )
                if prefer_new or current is not None:
                    fixed = FIXED_TEMOIN_PLAN.get(capsule, {})
                    base = current or {}
                    mapping[code] = {
                        "capsule": capsule if prefer_new else base.get("capsule", capsule),
                        "video_temoin_label": (
                            fixed.get("label")
                            or base.get("video_temoin_label")
                            or capsule
                        )
                        if prefer_new
                        else base.get("video_temoin_label", capsule),
                        "objectif_temoin": base.get("objectif_temoin", ""),
                        "titre": video.get("titre")
                        or catalogue.get(code, {}).get("titre")
                        or base.get("titre", ""),
                        "descriptif": video.get("descriptif")
                        or catalogue.get(code, {}).get("descriptif")
                        or base.get("descriptif", ""),
                    }
    for code, video in catalogue.items():
        if code not in mapping:
            mapping[code] = {
                "capsule": video.get("capsule", ""),
                "video_temoin_label": video.get("capsule", ""),
                "objectif_temoin": "",
                "titre": video.get("titre", ""),
                "descriptif": video.get("descriptif", ""),
            }
        else:
            if not mapping[code].get("titre"):
                mapping[code]["titre"] = video.get("titre", "")
            if not mapping[code].get("descriptif"):
                mapping[code]["descriptif"] = video.get("descriptif", "")
    return mapping


def _selection_finale_overview() -> list[dict]:
    """Vue d'ensemble de la sélection finale (une ligne par vidéo expertise)."""
    stored = _load_suivi_positionnements()
    programme = (
        json.loads((ROOT / "data" / "programme_table.json").read_text(encoding="utf-8"))
        if (ROOT / "data" / "programme_table.json").exists()
        else {"rows": []}
    )
    e_map = _e_code_to_temoin_map(programme)
    temoin_by_capsule = {
        row.get("code", ""): row.get("video_temoin", "")
        for row in programme.get("rows", [])
        if row.get("code")
    }
    catalogue = {item["code"]: item for item in _load_expert_videos_catalogue()}
    rows: list[dict] = []
    for item in _positionnements_finaux_par_video(stored.get("intervenants", [])):
        code = item["code"]
        meta = e_map.get(code, {})
        video = catalogue.get(code, {})
        objectif = (
            video.get("descriptif")
            or video.get("titre")
            or meta.get("descriptif")
            or meta.get("titre")
            or ""
        )
        capsule = meta.get("capsule", "")
        titre = _normalize_editorial_french(
            video.get("titre") or meta.get("titre") or ""
        )
        for entry in item.get("intervenants") or []:
            rows.append(
                {
                    "code": code,
                    "label": _display_expertise_title(code, titre),
                    "titre": titre,
                    "objectif": _normalize_editorial_french(objectif),
                    "capsule": capsule,
                    "temoin_label": _display_temoin_title(
                        temoin_by_capsule.get(capsule, "")
                        or meta.get("video_temoin_label", ""),
                        capsule,
                    ),
                    "nom": entry.get("nom", ""),
                    "organisme": entry.get("organisme", ""),
                    "slug": entry.get("slug", ""),
                    "fonction": _intervenant_function_label(
                        entry.get("nom", ""), entry.get("organisme", "")
                    ),
                }
            )
    return rows


def _experts_for_videos_attendues(
    programme_table: dict,
    experts_profils: dict,
) -> list[dict]:
    """Experts du lot « vidéos attendues » filtrés sur la sélection finale."""
    base = {
        _canonical_name_key(item["nom"]): item
        for item in _mail_experts_rows(programme_table, experts_profils)
    }
    stored = _load_suivi_positionnements()
    e_map = _e_code_to_temoin_map(programme_table)
    prepared: list[dict] = []

    for item in stored.get("intervenants", []):
        final = (item.get("proposition_finale") or "").strip()
        if not final:
            continue
        codes = [
            raw.upper().replace(" ", "")
            for raw in re.findall(r"\bE\d+(?:bis)?\b", final, flags=re.IGNORECASE)
        ]
        if not codes:
            continue
        key = _canonical_name_key(item.get("nom", ""))
        base_expert = base.get(key, {})
        by_capsule: dict[str, dict] = {}
        for code in codes:
            meta = e_map.get(code)
            if not meta:
                continue
            capsule = meta["capsule"]
            bucket = by_capsule.setdefault(
                capsule,
                {
                    "code": capsule,
                    "video_temoin_label": meta.get("video_temoin_label", ""),
                    "tb_edito_href": f"tb_edito_{capsule}.html",
                    "expert_video_codes": [],
                    "expert_video_labels": [],
                    "objectif": meta.get("objectif_temoin", ""),
                },
            )
            if code not in bucket["expert_video_codes"]:
                bucket["expert_video_codes"].append(code)
                titre = meta.get("titre") or ""
                bucket["expert_video_labels"].append(
                    _display_expertise_title(code, titre)
                )
                bucket.setdefault("expert_video_objectifs", {})[code] = (
                    _normalize_editorial_french(
                        meta.get("descriptif") or meta.get("titre") or ""
                    )
                )

        if not by_capsule:
            continue

        def _capsule_sort(code: str) -> tuple[int, int]:
            if code == "GEN":
                return (0, 0)
            match = re.fullmatch(r"T(\d+)", code or "")
            return (1, int(match.group(1))) if match else (9, 999)

        videos = sorted(by_capsule.values(), key=lambda entry: _capsule_sort(entry["code"]))
        for video in videos:
            ordered = sorted(video["expert_video_codes"], key=_expert_video_sort_key)
            rebuilt_labels = []
            rebuilt_objectifs: dict[str, str] = {}
            for code in ordered:
                meta = e_map.get(code, {})
                titre = meta.get("titre") or ""
                rebuilt_labels.append(_display_expertise_title(code, titre))
                rebuilt_objectifs[code] = _normalize_editorial_french(
                    meta.get("descriptif") or meta.get("titre") or ""
                )
            video["expert_video_codes"] = ordered
            video["expert_video_labels"] = rebuilt_labels
            video["expert_video_objectifs"] = rebuilt_objectifs
            video["video_temoin_display"] = _display_temoin_title(
                video.get("video_temoin_label", ""), video.get("code", "")
            )
            video["chapter_heading"] = _temoin_with_expertise_heading(video)

        prepared.append(
            {
                "nom": item.get("nom") or base_expert.get("nom", ""),
                "organisme": item.get("organisme")
                or base_expert.get("organisme")
                or "Organisme de rattachement à confirmer",
                "slug": item.get("slug") or slug(item.get("nom", "")),
                "profile": base_expert.get("profile"),
                "videos": videos,
                "proposition_finale": final,
            }
        )

    return sorted(prepared, key=lambda entry: _normalize_for_match(entry["nom"]))


def _selection_finale_overview_html(current_slug: str = "") -> str:
    rows = _selection_finale_overview()
    if not rows:
        return (
            "<h2>Vue d'ensemble — sélection finale</h2>"
            "<p class='meta'>Aucune sélection finale renseignée pour le moment.</p>"
        )
    body = []
    for row in rows:
        highlight = (
            " style='background:#ecfeff;'"
            if current_slug and row.get("slug") == current_slug
            else ""
        )
        titre = row.get("titre") or ""
        objectif = row.get("objectif") or ""
        expertise = row.get("label") or _display_expertise_title(row.get("code", ""), titre)
        temoin = row.get("temoin_label") or ""
        body.append(
            "<tr"
            + highlight
            + ">"
            f"<td><strong>{escape(expertise)}</strong>"
            + (f"<br><span class='meta'>{escape(temoin)}</span>" if temoin else "")
            + "</td>"
            f"<td>{escape(titre)}"
            + (f"<br><span class='meta'>{escape(objectif)}</span>" if objectif else "")
            + "</td>"
            f"<td><strong>{escape(row['nom'])}</strong><br>"
            f"<span class='meta'>{escape(row.get('organisme', ''))}</span></td>"
            "</tr>"
        )
    return (
        "<h2>Vue d'ensemble — sélection finale par vidéo expertise</h2>"
        "<p class='meta'>Lecture partagée : qui est attendu sur quelle <strong>vidéo expertise</strong>, "
        "et à quelle <strong>vidéo témoin</strong> elle se rattache. "
        "Votre ligne est mise en évidence lorsque cela est possible.</p>"
        "<div class='table-wrap'><table>"
        "<thead><tr>"
        "<th>Vidéo expertise</th><th>Objectif</th><th>Intervenant retenu</th>"
        "</tr></thead><tbody>"
        + "".join(body)
        + "</tbody></table></div>"
    )


def _selection_finale_overview_doc_html(current_slug: str = "") -> str:
    rows = _selection_finale_overview()
    if not rows:
        return (
            "<h2>Vue d'ensemble — sélection finale</h2>"
            "<p>Aucune sélection finale renseignée pour le moment.</p>"
        )
    body = []
    for row in rows:
        highlight = (
            " background:#ecfeff;"
            if current_slug and row.get("slug") == current_slug
            else ""
        )
        titre = row.get("titre") or ""
        objectif = row.get("objectif") or ""
        expertise = row.get("label") or _display_expertise_title(row.get("code", ""), titre)
        temoin = row.get("temoin_label") or ""
        body.append(
            f"<tr style='vertical-align:top;{highlight}'>"
            f"<td style='padding:6px;border:1px solid #cbd5e1;'><strong>{escape(expertise)}</strong><br>"
            f"<span style='color:#64748b;font-size:10pt;'>{escape(temoin)}</span></td>"
            f"<td style='padding:6px;border:1px solid #cbd5e1;'>{escape(titre)}"
            + (
                f"<br><span style='color:#64748b;font-size:10pt;'>{escape(objectif)}</span>"
                if objectif
                else ""
            )
            + "</td>"
            f"<td style='padding:6px;border:1px solid #cbd5e1;'><strong>{escape(row['nom'])}</strong><br>"
            f"<span style='color:#64748b;font-size:10pt;'>{escape(row.get('organisme', ''))}</span></td>"
            "</tr>"
        )
    return (
        "<h2>Vue d'ensemble — sélection finale par vidéo expertise</h2>"
        "<p>Pour situer votre intervention dans l'ensemble du parcours : chaque ligne nomme "
        "une <strong>vidéo expertise</strong> et la <strong>vidéo témoin</strong> à laquelle "
        "elle se rattache. Votre ligne est surlignée.</p>"
        "<table style='width:100%;border-collapse:collapse;font-size:10.5pt;margin:10px 0 18px;'>"
        "<thead><tr>"
        "<th style='text-align:left;padding:6px;border:1px solid #cbd5e1;background:#f8fafc;'>Vidéo expertise</th>"
        "<th style='text-align:left;padding:6px;border:1px solid #cbd5e1;background:#f8fafc;'>Objectif</th>"
        "<th style='text-align:left;padding:6px;border:1px solid #cbd5e1;background:#f8fafc;'>Intervenant retenu</th>"
        "</tr></thead><tbody>"
        + "".join(body)
        + "</tbody></table>"
    )


def _expert_videos_attendues_brief_blocks(expert: dict) -> list[str]:
    """Blocs succincts : vidéo expertise + objectif + rattachement témoin."""
    blocks: list[str] = []
    for item in expert.get("videos", []):
        temoin = item.get("video_temoin_display") or _display_temoin_title(
            item.get("video_temoin_label", ""), item.get("code", "")
        )
        codes = item.get("expert_video_codes") or []
        raw_labels = _expertise_titles_from_item(item)
        objectifs = item.get("expert_video_objectifs") or {}
        if not raw_labels:
            blocks.append(
                f"- Vidéo expertise à préciser\n"
                f"  Liée à : {temoin}"
            )
            continue
        for idx, label in enumerate(raw_labels):
            code = codes[idx] if idx < len(codes) else ""
            if code:
                titre = re.sub(
                    r"^Vidéo\s+expertise\s*[—\-–:]\s*",
                    "",
                    label,
                    flags=re.IGNORECASE,
                ).strip()
                mail_label = f"{_label_video_expert(code)} — {titre}" if titre else _label_video_expert(code)
            else:
                mail_label = label
            objectif = (objectifs.get(code) or "").strip()
            block = f"- {mail_label}\n  Liée à : {temoin}"
            if objectif:
                block += f"\n  Objectif à atteindre : {objectif}"
            blocks.append(block)
    return blocks


def _compose_videos_attendues_mail(
    expert: dict,
    page_href: str,
    *,
    simple_doc_name: str = "",
) -> tuple[str, str]:
    nom = expert["nom"]
    prenom = " ".join((nom or "").split()).split(" ")[0] if nom else "Madame, Monsieur"
    organisme = expert["organisme"]
    brief_blocks = _expert_videos_attendues_brief_blocks(expert)
    brief = "\n\n".join(brief_blocks) if brief_blocks else "- À confirmer"
    n_videos = sum(len(item.get("expert_video_codes") or []) for item in expert.get("videos", []))

    subject = f"MOOC L'Esprit d'innover — votre vidéo expertise ({nom})"
    if n_videos > 1:
        subject = f"MOOC L'Esprit d'innover — vos vidéos expertise ({nom})"

    pj_line = (
        f"- Guide éditorial simplifié joint : {simple_doc_name}\n"
        if simple_doc_name
        else ""
    )

    mail_text = (
        f"Objet : {subject}\n\n"
        f"Bonjour {prenom},\n\n"
        "Dans le cadre du MOOC « L'Esprit d'innover », je vous confirme "
        f"votre intervention ({organisme}).\n\n"
        "En deux mots\n"
        "Les apprenants voient d’abord une vidéo témoin (parole de chercheurs). "
        "Ensuite, votre vidéo expertise éclaire une notion précise. "
        "Registre : sensibilisation (montrer / informer), pas une formation technique.\n\n"
        "Ce que l’on vous demande\n"
        f"{brief}\n\n"
        "Format\n"
        "- Environ 5 minutes (± 2 min)\n"
        "- Script prompteur avant le tournage "
        "(on calera le délai avec vous)\n\n"
        "Pièce jointe\n"
        "Un guide éditorial simplifié (1 à 2 pages) : objectif, ce que disent "
        "les chercheurs, et ce que l’on attend de vous. "
        "Ce n’est qu’un appui : libre à vous d’ajuster.\n"
        f"{pj_line}"
        f"- Page en ligne (si besoin) : {page_href}\n\n"
        "Pour toute question, un entretien Teams est possible "
        "(à partir du 24 août).\n\n"
        "Bien cordialement,\n"
        "Christophe Dubois\n"
        "Action 2 — Pilier 1 — PUI Alliance Paris-Saclay\n"
        f"{TEST_MAIL_RECIPIENT}"
    )
    return subject, mail_text


def _expert_videos_attendues_lines(expert: dict) -> list[str]:
    """Lignes compactes pour listes HTML (pages de suivi)."""
    lines: list[str] = []
    for block in _expert_videos_attendues_brief_blocks(expert):
        # Une entrée de liste = première ligne du bloc + suite en retrait
        parts = [p.strip() for p in block.split("\n") if p.strip()]
        if not parts:
            continue
        head = parts[0][2:].strip() if parts[0].startswith("- ") else parts[0]
        if len(parts) == 1:
            lines.append(f"- {head}")
        else:
            lines.append("- " + head + "\n  " + "\n  ".join(parts[1:]))
    return lines


def _simple_guide_temoin_summary(capsule_code: str) -> str:
    """Résumé court « ce que disent les chercheurs » pour le guide éditorial simplifié."""
    narrative = _narrative_temoin_block(capsule_code)
    if not narrative:
        return ""

    def _clip(text: str, max_len: int = 160) -> str:
        t = " ".join((text or "").split())
        if not t:
            return ""
        if len(t) <= max_len:
            return t
        cut = t[: max_len - 1].rsplit(" ", 1)[0]
        return (cut or t[: max_len - 1]).rstrip(".,;:") + "…"

    chunks: list[str] = []
    intro = (narrative.get("intro") or "").strip()
    if intro:
        chunks.append(f"<p>{escape(_oralize_editorial_meta(_clip(intro, 220)))}</p>")
    bullets: list[str] = []
    for voix in narrative.get("voix") or []:
        chercheur = (voix.get("chercheur") or "").strip()
        paras = [p.strip() for p in (voix.get("paragraphes") or []) if (p or "").strip()]
        if not chercheur and not paras:
            continue
        gist = _clip(" ".join(paras), 140)
        if chercheur and gist:
            bullets.append(
                f"<li><strong>{escape(chercheur)}</strong> — {escape(gist)}</li>"
            )
        elif gist:
            bullets.append(f"<li>{escape(gist)}</li>")
    if bullets:
        chunks.append("<ul>" + "".join(bullets) + "</ul>")
    reteni = (narrative.get("a_retenir") or "").strip()
    if reteni:
        chunks.append(
            f"<p><strong>À retenir :</strong> "
            f"{escape(_normalize_editorial_french(_clip(reteni, 220)))}</p>"
        )
    return "".join(chunks)


def _guide_videos_attendues_simple_doc_html(expert: dict) -> str:
    """
    Guide éditorial simplifié (PJ principale) : direct, lisible en quelques minutes.
    Le guide détaillé reste disponible à part.
    """
    nom = expert.get("nom", "Expert")
    sections: list[str] = [
        "<div class='doc-block brief-block'>"
        "<p><strong>En une phrase :</strong> les apprenants voient une vidéo témoin "
        "(chercheurs), puis votre vidéo expertise éclaire une notion. "
        "Sensibilisation, pas formation technique.</p>"
        "<p><strong>Votre livrable :</strong> une prise de parole d’environ 5 minutes "
        "(± 2 min), avec un script prompteur avant tournage. "
        "Ce guide est un appui : libre à vous d’ajuster.</p>"
        "</div>"
    ]

    for item in expert.get("videos", []):
        temoin = item.get("video_temoin_display") or _display_temoin_title(
            item.get("video_temoin_label", ""), item.get("code", "")
        )
        codes = item.get("expert_video_codes") or []
        labels = _expertise_titles_from_item(item)
        objectifs = item.get("expert_video_objectifs") or {}
        capsule_code = item.get("code", "")

        for idx, label in enumerate(labels or ["Vidéo expertise"]):
            code = codes[idx] if idx < len(codes) else ""
            if code:
                titre = re.sub(
                    r"^Vidéo\s+expertise\s*[—\-–:]\s*",
                    "",
                    label,
                    flags=re.IGNORECASE,
                ).strip()
                head = (
                    f"{_label_video_expert(code)} — {titre}"
                    if titre
                    else _label_video_expert(code)
                )
            else:
                head = label
            objectif = (objectifs.get(code) or "").strip()
            fascicule_html = _fascicule_refs_html(code) if code else ""
            temoin_html = _simple_guide_temoin_summary(capsule_code)

            block = [
                "<section style='margin-top:22px;padding-top:10px;border-top:1px solid #cbd5e1;'>",
                f"<h2>{escape(head)}</h2>",
                f"<p><strong>Liée à :</strong> {escape(temoin)}</p>",
            ]
            if objectif:
                block.append(
                    f"<p><strong>Objectif de votre vidéo :</strong> "
                    f"{escape(_normalize_editorial_french(objectif))}</p>"
                )
            block.append("<h3>Ce que disent les chercheurs (vidéo témoin)</h3>")
            block.append(
                temoin_html
                or "<p class='meta'>Synthèse à compléter.</p>"
            )
            block.append("<h3>Ce que l’on attend de vous</h3>")
            block.append(
                "<ul>"
                "<li>Partir de ces témoignages (sans les citer mot à mot).</li>"
                "<li>Éclairer l’objectif ci-dessus, en langage clair.</li>"
                "<li>Montrer / informer : quelques exemples concrets suffisent.</li>"
                "<li>Inviter l’apprenant à faire le lien avec sa pratique de recherche.</li>"
                "</ul>"
            )
            if fascicule_html:
                block.append(fascicule_html)
            block.append("</section>")
            sections.append("".join(block))

    contact = (
        "<div class='doc-block' style='margin-top:22px;'>"
        "<h3>Contact</h3>"
        "<p>Christophe Dubois — "
        f"<a href='mailto:{escape(TEST_MAIL_RECIPIENT)}'>{escape(TEST_MAIL_RECIPIENT)}</a><br>"
        "Teams jusqu’au 31 juillet ; entretiens à partir du 24 août "
        f"(copie possible : {escape(REVIEW_MAIL_RECIPIENT)}).</p>"
        "<p class='meta'>Un guide détaillé existe aussi (scripts, transcript témoin, "
        "vue d’ensemble). Demandez-le si vous en avez besoin — "
        "ce guide simplifié suffit pour démarrer.</p>"
        "</div>"
    )

    return (
        "<html><head><meta charset='utf-8'>"
        "<style>"
        "body{font-family:Aptos,Segoe UI,Arial,sans-serif;font-size:12pt;line-height:1.5;}"
        "h1{font-size:18pt;margin-bottom:8px;}"
        "h2{font-size:14pt;margin:0 0 8px;}"
        "h3{font-size:12.5pt;margin:14px 0 6px;}"
        ".doc-block{border:1px solid #dbe2ea;border-radius:8px;padding:12px 14px;margin-bottom:14px;}"
        ".brief-block{background:#f8fafc;}"
        "ul{margin:6px 0 10px 22px;}"
        "li{margin:0 0 6px 0;}"
        "p{margin:0 0 8px 0;}"
        ".meta{color:#64748b;font-size:10.5pt;}"
        ".brief-fascicule{background:#f0fdf4;padding:8px;border-left:3px solid #86efac;border-radius:4px;}"
        "</style></head><body>"
        f"<h1>Guide éditorial simplifié — {escape(nom)}</h1>"
        "<p class='meta'>MOOC « L'Esprit d'innover » — ce qu’on vous demande, en direct.</p>"
        + "".join(sections)
        + contact
        + "</body></html>"
    )


def export_scripts_expertise_plaintext(capsule_data: dict) -> str:
    videos = capsule_data.get("videos_expert", [])
    orientations = capsule_data.get("orientations_expert") or []
    if capsule_data.get("orientation_expert") and not orientations:
        orientations = [capsule_data["orientation_expert"]]
    if not videos and not orientations:
        return ""

    by_code = {item.get("code", ""): item for item in videos if item.get("code")}
    blocks: list[tuple[dict, dict | None]] = []
    seen: set[str] = set()
    for orientation in orientations:
        code = orientation.get("code", "")
        blocks.append((orientation, by_code.get(code)))
        if code:
            seen.add(code)
    for video in videos:
        code = video.get("code", "")
        if code and code not in seen:
            blocks.append(
                (
                    {
                        "code": code,
                        "titre": video.get("titre", ""),
                        "concepts": [],
                        "introduction": video.get("descriptif", ""),
                        "consignes": list(BRIEF_CONSIGNES_COMMUNES),
                        "utilisation_script_temoin": {},
                    },
                    video,
                )
            )

    lines = [
        "PROPOSITION DE SCRIPT POUR LES VIDÉOS EXPERTISE",
        "",
        SCRIPT_EXPERTISE_DISCLAIMER,
        "",
    ]
    for variant_index, (orientation, video) in enumerate(blocks):
        code = orientation.get("code") or (video or {}).get("code", "")
        titre = orientation.get("titre") or (video or {}).get("titre", "")
        script, word_count = _build_script_expertise_projete(
            orientation,
            video,
            variant_index=variant_index,
            sibling_count=len(blocks),
        )
        plan = _build_script_expertise_plan(
            orientation, video, variant_index=variant_index
        )
        lines.append(f"{_display_expertise_title(code, titre)} — script projeté")
        lines.append(f"Volume : {word_count} mots")
        if plan:
            lines.append("Plan :")
            for item in plan:
                lines.append(f"  - {item}")
        lines.append("")
        lines.append(script)
        lines.append("")
    return _normalize_editorial_french("\n".join(lines).strip())


def _videos_attendues_contact_preface_html(*, for_doc: bool = True) -> str:
    """Coordonnées de contact + mode d'emploi, avant le sommaire."""
    mail = TEST_MAIL_RECIPIENT
    cc = REVIEW_MAIL_RECIPIENT
    reading = (
        "<p><strong>Comment lire ce guide</strong> — le MOOC articule deux types de vidéos :</p>"
        "<ul>"
        "<li><strong>Vidéo témoin</strong> : parole croisée de chercheurs (chorale), "
        "déjà montée éditorialement ; elle donne le contexte de votre intervention.</li>"
        "<li><strong>Vidéo expertise</strong> : votre intervention, pour éclairer ou prolonger "
        "la vidéo témoin. C’est le contenu pour lequel vous êtes attendu(e).</li>"
        "</ul>"
        "<p>Chaque chapitre nomme d’abord la <strong>vidéo témoin</strong>, puis les "
        "<strong>vidéos expertise</strong> qui s’y rattachent.</p>"
        "<p>Un guide conceptuel imprimé — <strong>« Oser pour innover »</strong> "
        "(deux fascicules) — peut éventuellement compléter votre lecture. "
        "Lorsque c’est possible, les pages utiles sont indiquées auprès de chaque "
        "vidéo expertise, en repère facultatif.</p>"
    )
    contact = (
        "<p><strong>Pour toute question, mise en forme ou échange</strong>, "
        "je reste disponible :</p>"
        "<ul>"
        "<li>sur <strong>Teams jusqu’au 31 juillet</strong> ;</li>"
        "<li>puis, <strong>à partir du 24 août</strong>, pour un entretien sur Teams ;</li>"
        "<li>vous pouvez également laisser un message par mail "
        "<strong>à partir du 15 août</strong>, avec en copie "
        f"<a href='mailto:{escape(cc)}'>{escape(cc)}</a>.</li>"
        "</ul>"
        f"<p>Mon adresse : <a href='mailto:{escape(mail)}'>{escape(mail)}</a>. "
        "Je répondrai à vos questions.</p>"
    )
    if for_doc:
        return (
            "<div class='doc-block brief-block' style='margin:12px 0 18px;'>"
            f"{reading}{contact}"
            "</div>"
        )
    return (
        "<section class='methodology-panel'>"
        "<h2>Introduction</h2>"
        f"{reading}{contact}"
        "</section>"
    )


def _guide_doc_shell(
    title: str,
    toc_rows: list[str],
    sections: list[str],
    toc_intro: str,
    preface_html: str = "",
) -> str:
    return (
        "<html><head><meta charset='utf-8'>"
        "<style>"
        "body{font-family:Aptos,Segoe UI,Arial,sans-serif;font-size:12pt;line-height:1.5;}"
        "h1{font-size:18pt;margin-bottom:6px;}"
        "h2{font-size:14pt;margin-bottom:6px;}"
        "h3{font-size:12.5pt;margin-bottom:6px;}"
        ".doc-block{border:1px solid #dbe2ea;border-radius:8px;padding:12px 14px;margin-bottom:14px;}"
        ".brief-block{background:#f8fafc;line-height:1.65;}"
        ".script-block{background:#ffffff;line-height:1.6;}"
        ".doc-block p{margin:0 0 8px 0;}"
        ".doc-block ul{margin:4px 0 10px 24px;padding:0;}"
        ".doc-block li{margin:0 0 6px 0;}"
        ".brief-label{margin-top:10px;}"
        ".brief-video{background:#eef2ff;padding:6px 8px;border-radius:6px;}"
        ".brief-precaution{background:#fff7ed;padding:8px;border-left:3px solid #fdba74;border-radius:4px;}"
        ".brief-fascicule{background:#f0fdf4;padding:8px;border-left:3px solid #86efac;border-radius:4px;}"
        ".script-body{font-size:11pt;line-height:1.62;word-break:break-word;}"
        ".script-ref{font-size:9pt;color:#94a3b8;}"
        ".toc{width:100%;border-collapse:collapse;margin:10px 0 16px;}"
        ".toc td{padding:6px 2px;border-bottom:1px dotted #94a3b8;}"
        "pre.plain{white-space:pre-wrap;font-family:Aptos,Segoe UI,Arial,sans-serif;font-size:11pt;margin:0;}"
        "</style>"
        "</head><body>"
        f"<h1>{escape(title)}</h1>"
        f"{preface_html}"
        "<h2>Sommaire du guide éditorial</h2>"
        f"{toc_intro}"
        f"<table class='toc'><tbody>{''.join(toc_rows) if toc_rows else '<tr><td>Aucune capsule témoin associée à ce stade.</td></tr>'}</tbody></table>"
        f"{''.join(sections) if sections else '<p>Aucune capsule témoin associée à ce stade.</p>'}"
        "</body></html>"
    )


def _guide_videos_attendues_doc_html(
    expert: dict,
    grouped_tb: dict[str, list[dict]],
    rows_by_code: dict[str, dict],
    affectations: dict,
    by_id: dict[str, dict],
) -> str:
    """Guide enrichi : synthèse, cadrage, scripts expertise, script final."""

    def _ergo_brief_html(text: str) -> str:
        lines = (text or "").splitlines()
        chunks: list[str] = []
        list_items: list[str] = []

        def flush_list() -> None:
            nonlocal list_items
            if list_items:
                chunks.append("<ul>" + "".join(list_items) + "</ul>")
                list_items = []

        for raw in lines:
            stripped = raw.strip()
            if not stripped:
                flush_list()
                continue
            if stripped.startswith("- "):
                list_items.append(f"<li>{escape(stripped[2:].strip())}</li>")
                continue
            flush_list()
            if stripped.startswith("Précaution :") or stripped.startswith("Precaution :"):
                chunks.append(
                    f"<p class='brief-precaution'><strong>Précaution :</strong> "
                    f"{escape(stripped.split(':', 1)[1].strip())}</p>"
                )
            elif stripped.startswith("Repère facultatif"):
                chunks.append(
                    f"<p class='brief-fascicule'><strong>{escape(stripped.split(':', 1)[0].strip())} :</strong> "
                    f"{escape(stripped.split(':', 1)[1].strip()) if ':' in stripped else ''}</p>"
                )
            elif stripped.startswith("Vidéo expertise ") and " — " in stripped:
                left, right = stripped.split(" — ", 1)
                chunks.append(
                    f"<p class='brief-video'><strong>{escape(left)}</strong> — {escape(right)}</p>"
                )
            elif stripped.endswith(":"):
                chunks.append(f"<p class='brief-label'><strong>{escape(stripped)}</strong></p>")
            else:
                chunks.append(f"<p>{escape(stripped)}</p>")
        flush_list()
        return "".join(chunks) if chunks else "<p>Aucun contenu.</p>"

    def _ergo_script_html(text: str) -> str:
        return _script_lines_html(text)

    def _plain_block(text: str) -> str:
        if not (text or "").strip():
            return "<p class='meta'>Contenu à compléter.</p>"
        return f"<pre class='plain'>{escape(text)}</pre>"

    sections: list[str] = []
    toc_rows: list[str] = []
    for item in expert.get("videos", []):
        code = item.get("code", "")
        chapter_anchor = f"chap_{code}"
        row = rows_by_code.get(code, {})
        sequences = grouped_tb.get(code, [])
        ordered = _tb_edito_order_for_code(code, sequences)
        by_seq_id = {seq.get("id", f"{code}-NOID"): seq for seq in ordered}
        ordre = [seq.get("id", f"{code}-NOID") for seq in ordered]
        videos_expert_all = _tb_edito_parse_videos_expert(row.get("videos_referent", ""))
        expert_codes = item.get("expert_video_codes") or [
            video.get("code", "") for video in videos_expert_all if video.get("code")
        ]
        aff_capsule = (affectations.get("capsules") or {}).get(code, {})
        capsule_data = _filter_capsule_data_for_expert(aff_capsule, expert_codes, row)
        if not capsule_data.get("videos_expert"):
            capsule_data["videos_expert"] = [
                video for video in videos_expert_all if video.get("code") in set(expert_codes)
            ] or videos_expert_all
        cadrage = _tb_edito_build_cadrage(
            code, ordre, by_seq_id, capsule_data.get("videos_expert") or videos_expert_all
        )
        script_final = _tb_expertise_label(
            _humanize_capsule_labels(
                _normalize_script_final_editorial(
                    _script_final_prefer_mounted_transcript(
                        code,
                        _tb_edito_script_with_cadrage(ordre, by_seq_id, cadrage),
                    )
                )
            )
        )
        brief_text = _tb_expertise_label(
            _humanize_capsule_labels(
                export_brief_intervenant_plaintext(code, capsule_data, by_id)
            )
        )
        scripts_text = _tb_expertise_label(
            _humanize_capsule_labels(export_scripts_expertise_plaintext(capsule_data))
        )

        chapter_title = item.get("chapter_heading") or _temoin_with_expertise_heading(item)
        temoin_title = item.get("video_temoin_display") or _display_temoin_title(
            item.get("video_temoin_label", ""), code
        )
        expertise_titles = _expertise_titles_from_item(item)
        toc_rows.append(
            "<tr>"
            f"<td><a href='#{escape(chapter_anchor)}'>{escape(chapter_title)}</a></td>"
            "</tr>"
        )
        attendues_html = (
            "<ul>"
            + "".join(f"<li>{escape(label)}</li>" for label in expertise_titles)
            + "</ul>"
            if expertise_titles
            else "<p>À définir</p>"
        )
        sections.append(
            "<section style='margin-top:28px;padding-top:10px;border-top:1px solid #cbd5e1;'>"
            f"<p style='page-break-before:always;margin:0;height:0;'>&nbsp;</p>"
            f"<a name='{escape(chapter_anchor)}'></a>"
            f"<h2>{escape(temoin_title)}</h2>"
            "<p><strong>Vidéos expertise associées (votre intervention) :</strong></p>"
            f"{attendues_html}"
            f"<h3>{escape(_tb_expertise_label(EXPORT_BRIEF_SECTION_TITLE))}</h3>"
            f"<div class='doc-block brief-block'>{_ergo_brief_html(brief_text)}</div>"
            "<p style='page-break-before:always;margin:0;height:0;'>&nbsp;</p>"
            "<h3>Proposition de script pour les vidéos expertise</h3>"
            f"<div class='doc-block script-block'>{_script_lines_html(scripts_text)}</div>"
            "<p style='page-break-before:always;margin:0;height:0;'>&nbsp;</p>"
            "<h3>Script final de la vidéo témoin</h3>"
            f"<div class='doc-block script-block'>{_ergo_script_html(script_final)}</div>"
            "</section>"
        )

    toc_rows.append(
        "<tr><td><a href='#vue_ensemble'>Vue d'ensemble — sélection finale</a></td></tr>"
    )
    toc_intro = ""
    overview = (
        "<section style='margin-top:28px;padding-top:10px;border-top:1px solid #cbd5e1;'>"
        "<p style='page-break-before:always;margin:0;height:0;'>&nbsp;</p>"
        "<a name='vue_ensemble'></a>"
        + _selection_finale_overview_doc_html(expert.get("slug", ""))
        + "</section>"
    )
    return _guide_doc_shell(
        f"Guide éditorial — vidéos attendues — {expert.get('nom', 'Expert')}",
        toc_rows,
        sections + [overview],
        toc_intro,
        preface_html=_videos_attendues_contact_preface_html(for_doc=True),
    )


def _videos_attendues_editorial_web_html(
    expert: dict,
    grouped_tb: dict[str, list[dict]],
    rows_by_code: dict[str, dict],
    affectations: dict,
    by_id: dict[str, dict],
) -> str:
    """Version HTML consultable du guide (même contenu que le Word)."""
    parts: list[str] = [_videos_attendues_contact_preface_html(for_doc=False)]
    for item in expert.get("videos", []):
        code = item.get("code", "")
        row = rows_by_code.get(code, {})
        sequences = grouped_tb.get(code, [])
        ordered = _tb_edito_order_for_code(code, sequences)
        by_seq_id = {seq.get("id", f"{code}-NOID"): seq for seq in ordered}
        ordre = [seq.get("id", f"{code}-NOID") for seq in ordered]
        videos_expert_all = _tb_edito_parse_videos_expert(row.get("videos_referent", ""))
        expert_codes = item.get("expert_video_codes") or [
            video.get("code", "") for video in videos_expert_all if video.get("code")
        ]
        aff_capsule = (affectations.get("capsules") or {}).get(code, {})
        capsule_data = _filter_capsule_data_for_expert(aff_capsule, expert_codes, row)
        if not capsule_data.get("videos_expert"):
            capsule_data["videos_expert"] = [
                video for video in videos_expert_all if video.get("code") in set(expert_codes)
            ] or videos_expert_all
        cadrage = _tb_edito_build_cadrage(
            code, ordre, by_seq_id, capsule_data.get("videos_expert") or videos_expert_all
        )
        script_final = _normalize_script_final_editorial(
            _script_final_prefer_mounted_transcript(
                code,
                _tb_edito_script_with_cadrage(ordre, by_seq_id, cadrage),
            )
        )
        attendues = _expertise_titles_from_item(item)
        temoin_title = item.get("video_temoin_display") or _display_temoin_title(
            item.get("video_temoin_label", ""), code
        )
        parts.append(f"<section class='methodology-panel'>")
        parts.append(f"<h2>{escape(temoin_title)}</h2>")
        parts.append("<p><strong>Vidéos expertise associées (votre intervention) :</strong></p>")
        if attendues:
            parts.append("<ul>" + "".join(f"<li>{escape(label)}</li>" for label in attendues) + "</ul>")
        else:
            parts.append("<p class='meta'>À définir</p>")
        parts.append(synthese_temoignages_section(code, capsule_data, by_id) or "")
        parts.append(brief_intervenant_section(code, capsule_data, by_id) or "")
        parts.append(scripts_expertise_projetes_section(capsule_data) or "")
        parts.append("<h2>Script final de la vidéo témoin</h2>")
        parts.append(f"<div class='script'>{escape(script_final)}</div>")
        parts.append("</section>")
    parts.append(_selection_finale_overview_html(expert.get("slug", "")))
    return "\n".join(part for part in parts if part)


def build_mails_experts_pages(
    programme_table: dict,
    experts_profils: dict,
    affectations: dict | None = None,
    segments: list[dict] | None = None,
) -> None:
    """Mails experts : index par date d'envoi, puis type (positionnement, vidéos attendues)."""
    experts = _mail_experts_rows(programme_table, experts_profils)
    experts_attendues = _experts_for_videos_attendues(programme_table, experts_profils)
    affectations = affectations or {}
    by_id = index_by_id(segments or [])
    grouped_tb = _tb_edito_sequences_by_code()
    rows_by_code = {row.get("code", ""): row for row in programme_table.get("rows", [])}

    # Sous-ensemble daté : mails envoyés le 20/07/2026 = mail de positionnement.
    date_slug = "20260720"
    date_label = "20/07/2026"
    date_href = f"mails_experts_{date_slug}.html"
    positionnement_href = f"mails_positionnement_{date_slug}.html"
    positionnement_title = "Mail de positionnement"

    cards = []
    for expert in experts:
        mail_file = f"mail_expert_{expert['slug']}.html"
        video_refs = ", ".join(item["code"] for item in expert["videos"]) or "Aucune capsule témoin"
        cards.append(
            "<article class='card'>"
            f"<h2><a href='{escape(mail_file)}'>{escape(expert['nom'])}</a></h2>"
            f"<p class='meta'>{escape(expert['organisme'])}</p>"
            f"<p>Capsules témoins concernées : <strong>{escape(video_refs)}</strong></p>"
            f"<p><a class='btn' href='{escape(mail_file)}'>Ouvrir le mail</a></p>"
            "</article>"
        )

    # 1) Index Mails experts → dates d'envoi
    date_attendues_slug = "20260727"
    date_attendues_label = "27/07/2026"
    date_attendues_href = f"mails_experts_{date_attendues_slug}.html"
    attendues_href = f"mails_videos_attendues_{date_attendues_slug}.html"
    attendues_title = "Mail vidéos attendues"

    write_text(
        SITE / "mails_experts.html",
        html_page(
            "Mails experts",
            "<p class='meta'>Archives des mails envoyés aux experts, classées par date d'envoi.</p>"
            + _sommaire_cards(
                [
                    (
                        date_href,
                        "📅",
                        date_label,
                        f"Mails envoyés le {date_label} — {positionnement_title.lower()}.",
                    ),
                    (
                        date_attendues_href,
                        "📅",
                        date_attendues_label,
                        f"Mails envoyés le {date_attendues_label} — {attendues_title.lower()}.",
                    ),
                ]
            ),
            nav_current="fichiers_travail.html",
            breadcrumb=html_breadcrumb(
                ("Accueil", "index.html"),
                ("Fichiers de travail", "fichiers_travail.html"),
                ("Mails experts", None),
            ),
            page_header=(
                '<div class="page-head"><h1>Mails experts</h1>'
                '<p class="lead">Messages envoyés aux experts, par date.</p></div>'
            ),
            main_class="page-home",
        ),
    )

    # 2) Sous-ensemble daté 20/07/2026 → types de mail
    write_text(
        SITE / date_href,
        html_page(
            f"Mails experts — {date_label}",
            f"<p class='meta'>Mails envoyés le <strong>{escape(date_label)}</strong>.</p>"
            + _sommaire_cards(
                [
                    (
                        positionnement_href,
                        "✉",
                        positionnement_title,
                        "Sollicitation de positionnement sur les vidéos expertise proposées.",
                    )
                ]
            ),
            nav_current="fichiers_travail.html",
            breadcrumb=html_breadcrumb(
                ("Accueil", "index.html"),
                ("Fichiers de travail", "fichiers_travail.html"),
                ("Mails experts", "mails_experts.html"),
                (date_label, None),
            ),
            page_header=(
                f'<div class="page-head"><h1>Mails experts — {escape(date_label)}</h1>'
                f'<p class="lead">Sous-ensemble des envois du {escape(date_label)}.</p></div>'
            ),
            main_class="page-home",
        ),
    )

    # 3) Mail de positionnement = contenu actuel (liste des mails individuels)
    positionnement_body = (
        f"<p class='meta'><strong>Envoi :</strong> {escape(date_label)}. "
        "Mails individualisés pour solliciter le positionnement des experts sur les vidéos expertise. "
        "Chaque mail reprend les sujets de capsules témoins concernés, les vidéos expertise proposées, "
        "l'information sur les transcripts des cinq chercheurs disponibles et les jalons : "
        "<strong>23 juillet</strong> (positionnement), <strong>27 juillet</strong> (retour d'arbitrage), "
        "<strong>1er septembre</strong> (script prompteur, a minima 15 jours avant tournage).</p>"
        f"<section class='cards'>{''.join(cards) if cards else '<p>Aucun expert proposé dans le programme_table.</p>'}</section>"
    )
    write_text(
        SITE / positionnement_href,
        html_page(
            positionnement_title,
            positionnement_body,
            nav_current="fichiers_travail.html",
            breadcrumb=html_breadcrumb(
                ("Accueil", "index.html"),
                ("Fichiers de travail", "fichiers_travail.html"),
                ("Mails experts", "mails_experts.html"),
                (date_label, date_href),
                (positionnement_title, None),
            ),
            page_header=(
                f'<div class="page-head"><h1>{escape(positionnement_title)}</h1>'
                f'<p class="lead">Mails de sollicitation de positionnement — envoi du {escape(date_label)}.</p></div>'
            ),
        ),
    )

    expected = {
        positionnement_href,
        date_href,
        "mails_experts.html",
        date_attendues_href,
        attendues_href,
    }
    expected_docs = set()
    expected_mail_txt = set()
    for expert in experts:
        subject, mail_text = _compose_expert_mail(expert)
        send_href = _mailto_href(TEST_MAIL_RECIPIENT, subject, mail_text)
        mail_name = f"mail_expert_{expert['slug']}.html"
        doc_name = f"guide_editorial_{expert['slug']}.doc"
        doc_abs_path = str((SITE / doc_name).resolve())
        doc_file_uri = (SITE / doc_name).resolve().as_uri()
        expected.add(mail_name)
        expected_docs.add(doc_name)
        write_text(SITE / doc_name, _guide_editorial_expert_doc_html(expert, grouped_tb, rows_by_code))
        video_rows = []
        for item in expert["videos"]:
            video_rows.append(
                "<tr>"
                f"<td><a href='{escape(item['tb_edito_href'])}'>{escape(item['code'])}</a></td>"
                f"<td>{escape(item['video_temoin_label'])}</td>"
                f"<td>{escape(' | '.join(item['expert_video_labels']) or 'À définir')}</td>"
                f"<td>{escape(item['objectif'])}</td>"
                "</tr>"
            )
        detail_body = (
            f"<p class='meta'><strong>Expert :</strong> {escape(expert['nom'])} · "
            f"<strong>Organisme :</strong> {escape(expert['organisme'])}</p>"
            f"<p class='meta'><strong>Type :</strong> {escape(positionnement_title)} · "
            f"<strong>Envoi :</strong> {escape(date_label)}</p>"
            f"<p class='meta'><strong>Objet proposé :</strong> {escape(subject)}</p>"
            f"<p class='meta'><strong>Destinataire test actuel :</strong> {escape(TEST_MAIL_RECIPIENT)} "
            f"(validation éditoriale ensuite via {escape(REVIEW_MAIL_RECIPIENT)}).</p>"
            f"<p class='meta'>Fiche suivi : "
            f"<a href='suivi_intervenant_{escape(expert['slug'])}.html'>ouvrir dans Suivi Intervenants</a></p>"
            f"<p><a class='btn' href='{escape(send_href)}'>Mail à envoyer</a></p>"
            f"<p><a class='btn' href='{escape(doc_name)}' download>Exporter le guide éditorial (Word)</a></p>"
            f"<p><a class='btn' href='{escape(doc_name)}' target='_blank' rel='noopener'>Ouvrir le guide éditorial (Word)</a></p>"
            f"<p><a class='btn' href='{escape(doc_file_uri)}'>Ouvrir le fichier Word (lien local absolu)</a></p>"
            f"<p class='meta'><strong>Pièce jointe prête :</strong> <code>{escape(doc_name)}</code></p>"
            f"<p class='meta'><strong>Chemin local :</strong> <code>{escape(doc_abs_path)}</code></p>"
            "<h2>Mail prêt à envoyer</h2>"
            f"<pre class='script mail-ready'>{escape(mail_text)}</pre>"
            "<h2>Capsules et sujets concernés</h2>"
            "<div class='table-wrap'><table><thead><tr>"
            "<th>Capsule témoin</th><th>Vidéo témoin</th><th>Vidéos expertise proposées</th><th>Objectif pédagogique</th>"
            "</tr></thead><tbody>"
            + ("".join(video_rows) or "<tr><td colspan='4'>Aucune affectation.</td></tr>")
            + "</tbody></table></div>"
            "<p class='meta'>Pièces jointes recommandées : <code>tableau_correspondances_edito.html</code> "
            "et les pages de capsules témoins listées ci-dessus.</p>"
        )
        write_text(
            SITE / mail_name,
            html_page(
                f"Mail expert — {expert['nom']}",
                detail_body,
                nav_current="fichiers_travail.html",
                breadcrumb=html_breadcrumb(
                    ("Accueil", "index.html"),
                    ("Fichiers de travail", "fichiers_travail.html"),
                    ("Mails experts", "mails_experts.html"),
                    (date_label, date_href),
                    (positionnement_title, positionnement_href),
                    (expert["nom"], None),
                ),
                page_header=(
                    f'<div class="page-head"><h1>Mail expert — {escape(expert["nom"])}</h1>'
                    f'<p class="lead">{escape(positionnement_title)} — envoi du {escape(date_label)}.</p></div>'
                ),
            ),
        )

    # --- Sous-ensemble 27/07/2026 : mail vidéos attendues ---
    write_text(
        SITE / date_attendues_href,
        html_page(
            f"Mails experts — {date_attendues_label}",
            f"<p class='meta'>Mails envoyés le <strong>{escape(date_attendues_label)}</strong>.</p>"
            + _sommaire_cards(
                [
                    (
                        attendues_href,
                        "✉",
                        attendues_title,
                        "Confirmation des vidéos expertise attendues + guide éditorial simplifié (PJ).",
                    )
                ]
            ),
            nav_current="fichiers_travail.html",
            breadcrumb=html_breadcrumb(
                ("Accueil", "index.html"),
                ("Fichiers de travail", "fichiers_travail.html"),
                ("Mails experts", "mails_experts.html"),
                (date_attendues_label, None),
            ),
            page_header=(
                f'<div class="page-head"><h1>Mails experts — {escape(date_attendues_label)}</h1>'
                f'<p class="lead">Sous-ensemble des envois du {escape(date_attendues_label)}.</p></div>'
            ),
            main_class="page-home",
        ),
    )

    zip_simplifies_name = "guides_editoriaux_simplifies.zip"
    attendues_cards = []
    for expert in experts_attendues:
        mail_file = f"mail_videos_attendues_{expert['slug']}.html"
        simple_doc = f"guide_editorial_simplifie_{expert['slug']}.doc"
        video_refs = " · ".join(
            label
            for item in expert["videos"]
            for label in _expertise_titles_from_item(item)
        ) or "Aucune vidéo expertise"
        attendues_cards.append(
            "<article class='card'>"
            f"<h2><a href='{escape(mail_file)}'>{escape(expert['nom'])}</a></h2>"
            f"<p class='meta'>{escape(expert['organisme'])}</p>"
            f"<p>Vidéos expertise attendues : <strong>{escape(video_refs)}</strong></p>"
            f"<p><a class='btn' href='{escape(mail_file)}'>Mail + page</a> "
            f"<a class='btn' href='{escape(simple_doc)}' download>Guide éditorial simplifié</a></p>"
            "</article>"
        )

    write_text(
        SITE / attendues_href,
        html_page(
            attendues_title,
            (
                f"<p class='meta'><strong>Envoi :</strong> {escape(date_attendues_label)}. "
                "Pour chaque intervenant : mail court + <strong>guide éditorial simplifié</strong> "
                "(PJ principale, 1–2 pages). Le guide détaillé reste disponible en option.</p>"
                f"<p><a class='btn' href='{escape(zip_simplifies_name)}' download>"
                f"Télécharger tous les guides éditoriaux simplifiés ({len(experts_attendues)}) — ZIP</a></p>"
                f"<section class='cards'>"
                f"{''.join(attendues_cards) if attendues_cards else '<p>Aucun expert avec sélection finale.</p>'}"
                f"</section>"
            ),
            nav_current="fichiers_travail.html",
            breadcrumb=html_breadcrumb(
                ("Accueil", "index.html"),
                ("Fichiers de travail", "fichiers_travail.html"),
                ("Mails experts", "mails_experts.html"),
                (date_attendues_label, date_attendues_href),
                (attendues_title, None),
            ),
            page_header=(
                f'<div class="page-head"><h1>{escape(attendues_title)}</h1>'
                f'<p class="lead">Confirmation du périmètre — envoi du {escape(date_attendues_label)}.</p></div>'
            ),
        ),
    )

    for expert in experts_attendues:
        page_name = f"mail_videos_attendues_{expert['slug']}.html"
        doc_name = f"guide_videos_attendues_{expert['slug']}.doc"
        simple_doc_name = f"guide_editorial_simplifie_{expert['slug']}.doc"
        mail_txt_name = f"mail_videos_attendues_{expert['slug']}.txt"
        expected.add(page_name)
        expected_docs.add(doc_name)
        expected_docs.add(simple_doc_name)
        expected_mail_txt.add(mail_txt_name)

        subject, mail_text = _compose_videos_attendues_mail(
            expert, page_name, simple_doc_name=simple_doc_name
        )
        write_text(SITE / mail_txt_name, mail_text)
        write_text(
            SITE / doc_name,
            _guide_videos_attendues_doc_html(
                expert, grouped_tb, rows_by_code, affectations, by_id
            ),
        )
        write_text(
            SITE / simple_doc_name,
            _guide_videos_attendues_simple_doc_html(expert),
        )
        send_href = _mailto_href(TEST_MAIL_RECIPIENT, subject, mail_text)
        editorial_html = _videos_attendues_editorial_web_html(
            expert, grouped_tb, rows_by_code, affectations, by_id
        )
        attendues_lines = _expert_videos_attendues_lines(expert)
        if attendues_lines:
            list_items = []
            for line in attendues_lines:
                text = line[2:] if line.startswith("- ") else line
                parts = [part.strip() for part in text.split("\n") if part.strip()]
                list_items.append("<li>" + "<br>".join(escape(part) for part in parts) + "</li>")
            attendues_list = "<ul>" + "".join(list_items) + "</ul>"
        else:
            attendues_list = "<p>Aucune vidéo expertise associée.</p>"
        detail_body = (
            f"<p class='meta'><strong>Expert :</strong> {escape(expert['nom'])} · "
            f"<strong>Organisme :</strong> {escape(expert['organisme'])}</p>"
            f"<p class='meta'><strong>Type :</strong> {escape(attendues_title)} · "
            f"<strong>Envoi :</strong> {escape(date_attendues_label)}</p>"
            f"<p class='meta'><strong>Objet proposé :</strong> {escape(subject)}</p>"
            f"<p class='meta'><strong>Destinataire test actuel :</strong> {escape(TEST_MAIL_RECIPIENT)} "
            f"(validation éditoriale ensuite via {escape(REVIEW_MAIL_RECIPIENT)}).</p>"
            f"<p class='meta'>Fiche suivi : "
            f"<a href='suivi_intervenant_{escape(expert['slug'])}.html'>ouvrir dans Suivi Intervenants</a></p>"
            "<h2>Vidéos expertise attendues</h2>"
            f"{attendues_list}"
            "<h2>Lien à transmettre à l'intervenant</h2>"
            "<p class='meta'>Lien de consultation du guide sur le site de travail "
            "(à ouvrir depuis le dossier <code>site/</code> ou l'URL de publication) :</p>"
            f"<p><code>{escape(page_name)}</code></p>"
            f"<p><a class='btn' href='{escape(page_name)}'>Ouvrir la page destinée à l'intervenant</a></p>"
            "<h2>Mail</h2>"
            f"<p><a class='btn' href='{escape(send_href)}'>Mail à envoyer</a> "
            f"<a class='btn' href='{escape(mail_txt_name)}' download>Exporter le mail (.txt)</a></p>"
            f"<pre class='script mail-ready'>{escape(mail_text)}</pre>"
            "<h2>Pièce jointe principale — guide éditorial simplifié</h2>"
            f"<p><a class='btn' href='{escape(simple_doc_name)}' download>Exporter le guide éditorial simplifié (Word)</a> "
            f"<a class='btn' href='{escape(simple_doc_name)}' target='_blank' rel='noopener'>Ouvrir le guide éditorial simplifié</a></p>"
            "<p class='meta'>1–2 pages : objectif, ce que disent les chercheurs, ce qu’on attend. "
            "À joindre en priorité au mail.</p>"
            "<h2>Guide détaillé (optionnel)</h2>"
            f"<p><a class='btn' href='{escape(doc_name)}' download>Exporter le guide détaillé (Word)</a> "
            f"<a class='btn' href='{escape(doc_name)}' target='_blank' rel='noopener'>Ouvrir le guide détaillé</a></p>"
            f"<p class='meta'>Version complète (scripts, transcript témoin, vue d’ensemble) — "
            f"sur demande ou en second fichier, pas en PJ unique.</p>"
            "<div class='editorial-preview'>"
            f"{editorial_html}"
            "</div>"
        )
        write_text(
            SITE / page_name,
            html_page(
                f"{attendues_title} — {expert['nom']}",
                detail_body,
                nav_current="fichiers_travail.html",
                breadcrumb=html_breadcrumb(
                    ("Accueil", "index.html"),
                    ("Fichiers de travail", "fichiers_travail.html"),
                    ("Mails experts", "mails_experts.html"),
                    (date_attendues_label, date_attendues_href),
                    (attendues_title, attendues_href),
                    (expert["nom"], None),
                ),
                page_header=(
                    f'<div class="page-head"><h1>{escape(attendues_title)} — {escape(expert["nom"])}</h1>'
                    f'<p class="lead">Périmètre confirmé et guide éditorial — envoi du {escape(date_attendues_label)}.</p></div>'
                ),
            ),
        )

    # ZIP de tous les guides éditoriaux simplifiés (envoi / archivage)
    zip_path = SITE / zip_simplifies_name
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for expert in experts_attendues:
            doc = SITE / f"guide_editorial_simplifie_{expert['slug']}.doc"
            if doc.is_file():
                zf.write(doc, arcname=doc.name)
    expected.add(zip_simplifies_name)

    for path in SITE.glob("mail_expert_*.html"):
        if path.name not in expected:
            path.unlink()
    for path in SITE.glob("mail_videos_attendues_*.html"):
        if path.name not in expected:
            path.unlink()
    for path in SITE.glob("mail_videos_attendues_*.txt"):
        if path.name not in expected_mail_txt:
            path.unlink()
    for path in SITE.glob("mail_expert_*.doc"):
        path.unlink()
    for path in SITE.glob("guide_editorial_*.doc"):
        if path.name not in expected_docs:
            path.unlink()
    for path in SITE.glob("guide_videos_attendues_*.doc"):
        if path.name not in expected_docs:
            path.unlink()
    # Ancien nom « fiche simple »
    for path in SITE.glob("guide_videos_attendues_simple_*.doc"):
        path.unlink()
    for path in SITE.glob("package_mail_expert_*.zip"):
        path.unlink()
    for path in SITE.glob("guides_editoriaux_simplifies*.zip"):
        if path.name not in expected:
            path.unlink()
    # Nettoie d'anciennes pages de lots hors des sous-ensembles courants.
    for path in SITE.glob("mails_positionnement_*.html"):
        if path.name not in expected:
            path.unlink()
    for path in SITE.glob("mails_videos_attendues_*.html"):
        if path.name not in expected and path.name.startswith("mails_videos_attendues_20"):
            path.unlink()
    for path in SITE.glob("mails_experts_*.html"):
        if path.name not in expected:
            path.unlink()
    for path in SITE.glob("tb_edito_T*_*.doc"):
        path.unlink()


def build_correspondances_edito_page(programme_table: dict) -> None:
    rows = programme_table.get("rows", [])
    headers = programme_table.get("headers", {})
    grouped_tb = _tb_edito_sequences_by_code()

    table_rows = []
    csv_rows = []
    for row in rows:
        code = row.get("code", "")
        fixed = FIXED_TEMOIN_PLAN.get(code, {})
        module_label = fixed.get("module") or row.get("module", "")
        title_programme = fixed.get("label") or row.get("video_temoin", "")
        sequences = grouped_tb.get(code, [])
        chercheurs_html, chercheurs_csv = _tb_edito_researcher_summary(code, sequences)
        objective = row.get("objectif_pedagogique", "")
        row_match_percent = _tb_edito_subject_alignment_percent(code, sequences)
        expert_badge = _expert_label(row_match_percent)
        expert_comment = _tb_edito_alignment_comment(code, row_match_percent, len(sequences), sequences)
        coverage_html, coverage_csv = _tb_edito_coverage_gap(code, sequences)

        videos_expert = _tb_edito_parse_videos_expert(row.get("videos_referent", ""))
        if videos_expert:
            expert_html = (
                "<ul>"
                + "".join(
                    f"<li><strong>{escape(_label_video_expert(item.get('code', '')))}</strong> — {escape(item.get('titre', ''))}</li>"
                    for item in videos_expert
                )
                + "</ul>"
            )
            expert_csv = " | ".join(
                f"{_label_video_expert(item.get('code', ''))}: {item.get('titre', '')}"
                for item in videos_expert
            )
        else:
            expert_html = "<span class='meta'>Aucune video expert renseignee.</span>"
            expert_csv = "Aucune video expert renseignee."
        preconisation_html, preconisation_csv = _tb_edito_expertise_preconisation(
            code, sequences, videos_expert, row_match_percent
        )

        intervenants = _extract_intervenants(row.get("noms_proposes", ""))
        intervenants_html = escape(", ".join(intervenants)) if intervenants else "<span class='meta'>A definir</span>"
        intervenants_csv = ", ".join(intervenants) if intervenants else "A definir"

        table_rows.append(
            "<tr>"
            f"<td>{escape(module_label)}</td>"
            f"<td>{escape(code)}</td>"
            f"<td><a href='tb_edito_{escape(code)}.html'>{escape(title_programme)}</a></td>"
            f"<td>{chercheurs_html}</td>"
            f"<td>{expert_html}</td>"
            f"<td>{preconisation_html}</td>"
            f"<td>{escape(objective)}</td>"
            f"<td>{coverage_html}</td>"
            f"<td>{intervenants_html}</td>"
            f"<td><strong>{row_match_percent}%</strong> <span class='meta'>({escape(expert_badge)})</span><br>{escape(expert_comment)}</td>"
            "</tr>"
        )
        csv_rows.append(
            {
                "module": module_label,
                "code": code,
                "video_chorale_tb_edito": title_programme,
                "ce_que_racontent_les_chercheurs_tb_edito": chercheurs_csv,
                "videos_expert_envisagees": expert_csv,
                "vue_preconisation_videos_expertise": preconisation_csv,
                "objectif_pedagogique": objective,
                "aborde_par_temoins_et_points_a_developper": coverage_csv,
                "intervenants_experts_proposes": intervenants_csv,
                "alignement_sujet_temoin_pct": row_match_percent,
                "niveau_alignement_sujet": expert_badge,
                "commentaire_alignement_sujet": expert_comment,
            }
        )

    body = (
        "<p class='meta'>Tableau base sur le programme de conception "
        f"<code>{escape(programme_table.get('source_document', '20260710_Prev_Vid.xlsx'))}</code> "
        "avec remplacement de la colonne video chorale par les capsules temoins T1..T12.</p>"
        "<p><a class='btn' href='tableau_correspondances_edito.csv' download>Télécharger le tableau (CSV)</a></p>"
        "<p><a class='btn' href='tableau_corr.html'>Voir le tableau corrigé (HTML)</a></p>"
        "<p class='meta'><strong>Alignement sujet témoin :</strong> estimation du % de presence du sujet de la video "
        "dans les verbatims des capsules temoins (sans attendre une couverture pedagogique complete, qui est portee par la video expertise).</p>"
        "<div class='table-wrap'><table><thead><tr>"
        f"<th>{escape(headers.get('module', 'Module'))}</th>"
        f"<th>{escape(headers.get('code', 'N°'))}</th>"
        "<th>Vidéo chorale (capsule témoin)</th>"
        "<th>Ce que racontent les chercheurs (capsule témoin)</th>"
        "<th>Vidéo(s) expert envisagée(s)</th>"
        "<th>Vue de préconisation pour les vidéos expertise</th>"
        f"<th>{escape(headers.get('objectif_pedagogique', 'Objectif pédagogique atteint'))}</th>"
        "<th>Abordé par les témoins / à développer</th>"
        f"<th>{escape(headers.get('noms_proposes', 'Intervenants experts proposés'))}</th>"
        "<th>Alignement sujet témoin estimé</th>"
        "</tr></thead><tbody>"
        + ("\n".join(table_rows) or "<tr><td colspan='9'>Aucune ligne programme.</td></tr>")
        + "</tbody></table></div>"
    )
    write_text(
        SITE / "tableau_correspondances_edito.html",
        html_page(
            "Correspondances édito",
            body,
            nav_current="edito.html",
            breadcrumb=html_breadcrumb(("Accueil", "index.html"), ("Edito", "edito.html"), ("Correspondances édito", None)),
            page_header="<div class=\"page-head\"><h1>Correspondances édito</h1><p class=\"lead\">Tableau de conception relu via les capsules témoins : videos chorales, contenus temoins, videos expertise et alignement pedagogique.</p></div>",
        ),
    )

    csv_buffer = io.StringIO()
    writer = csv.DictWriter(
        csv_buffer,
        fieldnames=[
            "module",
            "code",
            "video_chorale_tb_edito",
            "ce_que_racontent_les_chercheurs_tb_edito",
            "videos_expert_envisagees",
            "vue_preconisation_videos_expertise",
            "objectif_pedagogique",
            "aborde_par_temoins_et_points_a_developper",
            "intervenants_experts_proposes",
            "alignement_sujet_temoin_pct",
            "niveau_alignement_sujet",
            "commentaire_alignement_sujet",
        ],
    )
    writer.writeheader()
    writer.writerows(csv_rows)
    write_text(SITE / "tableau_correspondances_edito.csv", csv_buffer.getvalue())


def _tableau_corr_rows_from_programme() -> tuple[list[str], list[dict[str, str]]]:
    fieldnames = [
        "module",
        "code",
        "video_temoin_edito",
        "videos_expert",
        "objectif_pedagogique_atteint",
    ]
    programme = load_programme_table()
    rows = []
    for row in programme.get("rows", []):
        rows.append(
            {
                "module": row.get("module", ""),
                "code": row.get("code", ""),
                "video_temoin_edito": row.get("video_temoin", ""),
                "videos_expert": row.get("videos_referent", ""),
                "objectif_pedagogique_atteint": row.get("objectif_pedagogique", ""),
            }
        )
    return fieldnames, rows


def _load_tableau_corr_rows() -> tuple[list[str], list[dict[str, str]]]:
    path = ROOT / "data" / "tableau.corr"
    if not path.exists():
        return _tableau_corr_rows_from_programme()

    with path.open(encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        fieldnames = reader.fieldnames or []
        raw_rows = list(reader)

    has_malformed_rows = any(None in row for row in raw_rows)
    if not fieldnames or has_malformed_rows:
        return _tableau_corr_rows_from_programme()

    rows = [{key: (value or "") for key, value in row.items()} for row in raw_rows]
    return fieldnames, rows


def build_tableau_corr_page() -> None:
    headers, rows = _load_tableau_corr_rows()
    if not headers:
        body = (
            "<p class='meta'>Aucun fichier <code>data/tableau.corr</code> detecte.</p>"
            "<p class='meta'>Generez d'abord le tableau corrige puis relancez <code>python3 scripts/build_site.py</code>.</p>"
        )
        write_text(
            SITE / "tableau_corr.html",
            html_page(
                "Tableau corrige",
                body,
                nav_current="edito.html",
                breadcrumb=html_breadcrumb(
                    ("Accueil", "index.html"),
                    ("Correspondances édito", "tableau_correspondances_edito.html"),
                    ("Tableau corrige", None),
                ),
                page_header="<div class=\"page-head\"><h1>Tableau corrigé</h1><p class=\"lead\">Version corrigée du tableau d'origine avec vidéos témoin fixées selon l'EDITO.</p></div>",
            ),
        )
        return

    table_head = "".join(f"<th>{escape(name)}</th>" for name in headers)
    table_rows = []
    for row in rows:
        table_rows.append(
            "<tr>" + "".join(f"<td>{escape(row.get(col, ''))}</td>" for col in headers) + "</tr>"
        )

    csv_buffer = io.StringIO()
    writer = csv.DictWriter(csv_buffer, fieldnames=headers)
    writer.writeheader()
    writer.writerows(rows)
    write_text(SITE / "tableau_corr.csv", csv_buffer.getvalue())

    body = (
        "<p class='meta'>Tableau corrige derive de <code>data/tableau.corr</code> : "
        "videos temoins fixees selon l'EDITO, avec conservation des videos experts et des objectifs pedagogiques.</p>"
        "<p><a class='btn' href='tableau_corr.csv' download>Télécharger le tableau corrigé (CSV)</a></p>"
        "<div class='table-wrap'><table><thead><tr>"
        + table_head
        + "</tr></thead><tbody>"
        + ("\n".join(table_rows) or f"<tr><td colspan='{len(headers)}'>Aucune ligne.</td></tr>")
        + "</tbody></table></div>"
    )
    write_text(
        SITE / "tableau_corr.html",
        html_page(
            "Tableau corrigé",
            body,
            nav_current="edito.html",
            breadcrumb=html_breadcrumb(
                ("Accueil", "index.html"),
                ("Correspondances édito", "tableau_correspondances_edito.html"),
                ("Tableau corrigé", None),
            ),
            page_header="<div class=\"page-head\"><h1>Tableau corrigé</h1><p class=\"lead\">Correction du tableau d'origine avec cadre vidéos témoin fixé par l'EDITO.</p></div>",
        ),
    )


def pct(value: float) -> str:
    return f"{round((value or 0.0) * 100, 1)} %"


def build_match_pages(match_data: dict) -> None:
    if not match_data:
        body = "<p class='meta'>Aucune analyse match disponible. Lancez d'abord <code>python3 scripts/analyze_match_derushage_edito.py</code>.</p>"
        write_text(
            SITE / "match.html",
            html_page(
                "Match",
                body,
                nav_current="match.html",
                breadcrumb=html_breadcrumb(("Accueil", "index.html"), ("Match", None)),
                page_header='<div class="page-head"><h1>Match</h1><p class="lead">Comparaison derushage interne vs derushage edito.</p></div>',
            ),
        )
        return

    summary = match_data.get("resume", {})
    temoins = match_data.get("temoins", [])
    modules = match_data.get("modules", [])

    witness_rows = []
    for item in temoins:
        witness_slug = slug(item.get("temoin", "temoin"))
        witness_rows.append(
            "<tr>"
            f"<td><a href='match_temoin_{escape(witness_slug)}.html'>{escape(item.get('temoin', '-'))}</a></td>"
            f"<td>{escape(item.get('nb_edito', 0))}</td>"
            f"<td>{escape(item.get('nb_matchs', 0))}</td>"
            f"<td>{pct(item.get('couverture_edito', 0.0))}</td>"
            f"<td>{escape(item.get('nb_segments_derushage', 0))}</td>"
            f"<td>{pct(item.get('recouvrement_derushage', 0.0))}</td>"
            f"<td>{pct(item.get('metrics_derushage', {}).get('fluidite_score', 0.0))}</td>"
            "</tr>"
        )

    module_rows = []
    for module in modules:
        module_key = module.get("module_id", "M?")
        module_rows.append(
            "<tr>"
            f"<td><a href='match_module_{escape(module_key.lower())}.html'>{escape(module.get('module_label', module_key))}</a></td>"
            f"<td>{escape(module.get('nb_edito', 0))}</td>"
            f"<td>{escape(module.get('nb_edito_matches', 0))}</td>"
            f"<td>{pct(module.get('couverture_edito', 0.0))}</td>"
            f"<td>{escape(module.get('nb_derushage', 0))}</td>"
            f"<td>{pct(module.get('recouvrement_derushage', 0.0))}</td>"
            f"<td>{pct(module.get('metrics_derushage', {}).get('fluidite_score', 0.0))}</td>"
            "</tr>"
        )

    body = (
        "<p class='meta'>"
        f"Objectif : {escape(match_data.get('objectif', ''))} "
        f"· Regle de match : {escape(match_data.get('seuil_match_similarite', '-'))} "
        f"· Mise a jour : {escape(match_data.get('date_maj', '-'))}"
        "</p>"
        "<section class='card'>"
        "<h2>Resume global</h2>"
        f"<p><strong>Temoins analyses :</strong> {escape(summary.get('nb_temoins', 0))} "
        f"(dont {escape(summary.get('nb_temoins_avec_doc_edito', 0))} avec document edito "
        f"et {escape(summary.get('nb_temoins_avec_edito', 0))} avec sequences edito retenues).</p>"
        f"<p><strong>Couverture edito :</strong> {pct(summary.get('couverture_edito_globale', 0.0))} "
        f"({escape(summary.get('nb_sequences_match', 0))}/{escape(summary.get('nb_sequences_edito', 0))} sequences).</p>"
        f"<p><strong>Recouvrement derushage :</strong> {pct(summary.get('recouvrement_derushage_global', 0.0))} "
        f"({escape(summary.get('nb_segments_derushage_matches', 0))}/{escape(summary.get('nb_segments_derushage', 0))} segments).</p>"
        "<p><strong>Lecture pedagogique :</strong> plus la couverture edito et le score de fluidite sont hauts, "
        "plus l'alignement editorial et la lisibilite du parcours sont solides.</p>"
        "</section>"
        "<h2>Detail par temoin</h2>"
        '<div class="table-wrap"><table><thead><tr><th>Temoin</th><th>Seq. edito</th><th>Match</th><th>Couverture edito</th>'
        "<th>Seg. derushage</th><th>Recouvrement derushage</th><th>Fluidite</th></tr></thead><tbody>"
        + ("\n".join(witness_rows) or "<tr><td colspan='7'>Aucune donnee temoin.</td></tr>")
        + "</tbody></table></div>"
        "<h2>Detail par module</h2>"
        '<div class="table-wrap"><table><thead><tr><th>Module</th><th>Seq. edito</th><th>Match</th><th>Couverture edito</th>'
        "<th>Seg. derushage</th><th>Recouvrement derushage</th><th>Fluidite</th></tr></thead><tbody>"
        + ("\n".join(module_rows) or "<tr><td colspan='7'>Aucune donnee module.</td></tr>")
        + "</tbody></table></div>"
    )
    write_text(
        SITE / "match.html",
        html_page(
            "Match",
            body,
            nav_current="match.html",
            breadcrumb=html_breadcrumb(("Accueil", "index.html"), ("Match", None)),
            page_header='<div class="page-head"><h1>Match</h1><p class="lead">Analyse de convergence et divergence entre derushage interne et derushage edito.</p></div>',
        ),
    )

    expected_witness_pages = set()
    for item in temoins:
        witness = item.get("temoin", "Temoin")
        witness_slug = slug(witness)
        filename = f"match_temoin_{witness_slug}.html"
        expected_witness_pages.add(filename)
        strong_matches = "".join(
            "<li>"
            f"{escape(match.get('segment_id', '-'))} "
            f"({escape(match.get('segment_debut', '-'))} → {escape(match.get('segment_fin', '-'))}) "
            f"↔ {escape(match.get('edito_id', '-'))} ({escape(match.get('video', '-'))}) "
            f"· similarite {escape(match.get('similarite', '-'))}"
            "</li>"
            for match in item.get("matchs", [])[:8]
        )
        gaps_edito = "".join(
            "<li>"
            f"{escape(seq.get('edito_id', '-'))} · {escape(seq.get('video', '-'))} "
            f"(paragraphe {escape(seq.get('edito_paragraphe', '-'))})"
            "</li>"
            for seq in item.get("non_match_edito", [])[:10]
        )
        gaps_derushage = "".join(
            "<li>"
            f"{escape(seg.get('segment_id', '-'))} "
            f"({escape(seg.get('debut', '-'))} → {escape(seg.get('fin', '-'))}) "
            f"· {escape(seg.get('theme', '-'))}"
            "</li>"
            for seg in item.get("non_match_derushage", [])[:10]
        )
        metrics = item.get("metrics_derushage", {})
        body = (
            f"<p class='meta'>Source edito : {escape(item.get('source_edito', 'non renseignee'))}</p>"
            "<section class='card'>"
            "<h2>Indicateurs</h2>"
            f"<p><strong>Couverture edito :</strong> {pct(item.get('couverture_edito', 0.0))} "
            f"({escape(item.get('nb_matchs', 0))}/{escape(item.get('nb_edito', 0))})</p>"
            f"<p><strong>Recouvrement derushage :</strong> {pct(item.get('recouvrement_derushage', 0.0))} "
            f"({escape(item.get('nb_segments_derushage', 0) - item.get('nb_non_match_derushage', 0))}/{escape(item.get('nb_segments_derushage', 0))})</p>"
            f"<p><strong>Fluidite :</strong> {pct(metrics.get('fluidite_score', 0.0))} "
            f"(autonomie {pct(metrics.get('score_autonomie_moyen', 0.0))}, "
            f"montabilite {pct(metrics.get('score_montage_moyen', 0.0))}, "
            f"adequation composite {pct(metrics.get('score_composite_moyen', 0.0))}).</p>"
            "</section>"
            "<h2>Ce qui matche</h2>"
            f"<ul>{strong_matches or '<li>Aucun match significatif detecte.</li>'}</ul>"
            "<h2>Ce qui ne matche pas cote edito</h2>"
            f"<ul>{gaps_edito or '<li>Aucune divergence edito restante.</li>'}</ul>"
            "<h2>Ce qui ne matche pas cote derushage</h2>"
            f"<ul>{gaps_derushage or '<li>Aucune divergence derushage restante.</li>'}</ul>"
        )
        write_text(
            SITE / filename,
            html_page(
                f"Match — {witness}",
                body,
                nav_current="match.html",
                breadcrumb=html_breadcrumb(
                    ("Accueil", "index.html"),
                    ("Match", "match.html"),
                    (witness, None),
                ),
                page_header=f"<div class='page-head'><h1>Match — {escape(witness)}</h1><p class='lead'>Analyse detaillee des convergences et ecarts editoriaux.</p></div>",
            ),
        )
    for path in SITE.glob("match_temoin_*.html"):
        if path.name not in expected_witness_pages:
            path.unlink()

    expected_module_pages = set()
    for module in modules:
        module_key = module.get("module_id", "M?").lower()
        filename = f"match_module_{module_key}.html"
        expected_module_pages.add(filename)
        metrics = module.get("metrics_derushage", {})
        body = (
            "<section class='card'>"
            "<h2>Indicateurs module</h2>"
            f"<p><strong>Couverture edito :</strong> {pct(module.get('couverture_edito', 0.0))} "
            f"({escape(module.get('nb_edito_matches', 0))}/{escape(module.get('nb_edito', 0))})</p>"
            f"<p><strong>Recouvrement derushage :</strong> {pct(module.get('recouvrement_derushage', 0.0))} "
            f"({escape(module.get('nb_derushage_matches', 0))}/{escape(module.get('nb_derushage', 0))})</p>"
            f"<p><strong>Fluidite :</strong> {pct(metrics.get('fluidite_score', 0.0))}</p>"
            f"<p><strong>Autonomie moyenne :</strong> {pct(metrics.get('score_autonomie_moyen', 0.0))} · "
            f"<strong>Montabilite moyenne :</strong> {pct(metrics.get('score_montage_moyen', 0.0))} · "
            f"<strong>Score composite moyen :</strong> {pct(metrics.get('score_composite_moyen', 0.0))}</p>"
            "</section>"
            "<p class='meta'>Les details temoins de ce module sont accessibles depuis la page resume <code>match.html</code>.</p>"
        )
        write_text(
            SITE / filename,
            html_page(
                f"Match — {module.get('module_label', module_key)}",
                body,
                nav_current="match.html",
                breadcrumb=html_breadcrumb(
                    ("Accueil", "index.html"),
                    ("Match", "match.html"),
                    (module.get("module_label", module_key), None),
                ),
                page_header=f"<div class='page-head'><h1>Match — {escape(module.get('module_label', module_key))}</h1><p class='lead'>Lecture pedagogique et ecarts editoriaux au niveau module.</p></div>",
            ),
        )
    for path in SITE.glob("match_module_*.html"):
        if path.name not in expected_module_pages:
            path.unlink()


def capsule_tags_html(capsules: dict) -> str:
    if not capsules:
        return "<span class='meta'>Aucune capsule</span>"
    tags = []
    for code, info in sorted(capsules.items()):
        statut = escape(str(info.get("statut", "")).lower())
        label = f"{escape(code)}: {escape(info.get('statut', ''))}"
        if info.get("role"):
            label += f" ({escape(info['role'])})"
        tags.append(f"<span class='tag capsule-tag {statut}'>{label}</span>")
    return " ".join(tags)


def bab_bloc_html(bloc: dict) -> str:
    if bloc["encodage"] == "NON_ENCODE":
        return (
            "<div class='card bab-segment bab-segment--non-encode'>"
            f"<strong>Bloc {bloc['numero']}</strong> "
            f"<span class='meta'>{escape(bloc['debut'])} → {escape(bloc['fin'])} · "
            f"{format_seconds(bloc['duree_secondes'])}</span>"
            f"<p class='capsules'><span class='tag capsule-tag non-encode'>Non encode</span></p>"
            f"<p>{escape(bloc['verbatim'])}</p>"
            "</div>"
        )
    coupe_lines = []
    for code, info in sorted(bloc.get("capsules", {}).items()):
        if info.get("coupe"):
            coupe_lines.append(
                f"<p class='coupe-note'><strong>{escape(code)}</strong> — {escape(info['coupe'])}</p>"
            )
    return (
        "<div class='card bab-segment'>"
        f"<strong>{escape(bloc['id'])}</strong> "
        f"<span class='meta'>{escape(bloc['debut'])} → {escape(bloc['fin'])} · "
        f"{format_seconds(bloc['duree_secondes'])} · {escape(bloc.get('statut', '-'))}</span>"
        f"<p class='capsules'>{capsule_tags_html(bloc.get('capsules', {}))}</p>"
        + "".join(coupe_lines)
        + (f"<p class='meta'>{escape(bloc['commentaire'])}</p>" if bloc.get("commentaire") else "")
        + f"<p>{escape(bloc['verbatim'])}</p>"
        "</div>"
    )


def build_bab_encodes_index() -> None:
    rows = []
    for item in load_bab_encode_index():
        doc = load_bab_encode(item["id"])
        if not doc:
            continue
        stats = bab_encode_stats(doc)
        rows.append(
            "<tr>"
            f"<td><a href='bab_encode_{escape(item['id'])}.html'>{escape(item['chercheur'])}</a></td>"
            f"<td>{escape(item['source'])}</td>"
            f"<td>{escape(item['statut_encodage'])}</td>"
            f"<td>{stats['nb_blocs']}</td>"
            f"<td>{stats['nb_encodes']}</td>"
            f"<td>{stats['nb_utilises']}</td>"
            f"<td>{escape(item['date_maj'])}</td>"
            "</tr>"
        )
    body = (
        "<p class='meta'>BAB complets avec decoupage source. Les blocs deja travailles sont encodes ; "
        "les autres sont marques <strong>Non encode</strong>. Evolution dans <code>data/bab_encodes/</code>.</p>"
        '<div class="table-wrap"><table><thead><tr><th>Chercheur</th><th>Source BAB</th><th>Statut</th>'
        "<th>Blocs BAB</th><th>Encodes</th><th>Utilises</th><th>Mise a jour</th></tr></thead><tbody>"
        + ("\n".join(rows) or "<tr><td colspan='7'>Aucun BAB encode.</td></tr>")
        + "</tbody></table></div>"
    )
    write_text(
        SITE / "bab_encodes.html",
        html_page(
            "BAB encodé",
            body,
            nav_current="fichiers_travail.html",
            breadcrumb=html_breadcrumb(("Accueil", "index.html"), ("Fichiers de travail", "fichiers_travail.html"), ("BAB encodé", None)),
            page_header='<div class="page-head"><h1>BAB encodé</h1><p class="lead">Lecture segment par segment des BAB bruts, avec statuts et capsules associees.</p></div>',
        ),
    )


def build_bab_encode_pages() -> None:
    expected = set()
    for item in load_bab_encode_index():
        encode_id = item["id"]
        expected.add(f"bab_encode_{encode_id}.html")
        doc = load_bab_encode(encode_id)
        if not doc:
            continue
        stats = bab_encode_stats(doc)
        export_text = render_bab_encode_export(doc)
        sections = []
        sections.append(
            f"<p class='meta'><strong>{stats['nb_encodes']}</strong> blocs encodes sur "
            f"<strong>{stats['nb_blocs']}</strong> blocs BAB · "
            f"<strong>{stats['nb_utilises']}</strong> utilises dans les capsules.</p>"
        )
        for bloc in merge_bab_encode_blocs(doc):
            sections.append(bab_bloc_html(bloc))
        sections.append(
            f"<div class='script sr-export' id='script-final'>{escape(export_text)}</div>"
        )
        sections.append(export_word_modal(f"bab_encode_{encode_id}.doc"))
        page_title = f"BAB encode — {doc['chercheur']}"
        header = (
            f"<div class='page-header'>"
            f"<div><h1>{escape(page_title)}</h1>"
            f"<p class='meta'>Source : {escape(doc['source'])} · "
            f"Statut : {escape(doc['statut_encodage'])} · "
            f"Blocs : {stats['nb_encodes']}/{stats['nb_blocs']} encodes · "
            f"Mise a jour : {escape(doc['date_maj'])}</p></div>"
            f"<div class='export-toolbar' id='export-word-panel' "
            f"data-capsule-code='bab-{escape(encode_id)}' "
            f"data-capsule-title='{escape(doc['chercheur'])} — BAB encode'>"
            f"<button type='button' class='btn' id='export-word-open' "
            f"aria-expanded='false' aria-controls='export-word-modal'>Exporter en Word</button>"
            f"</div></div>"
        )
        write_text(
            SITE / f"bab_encode_{encode_id}.html",
            html_page(
                page_title,
                "\n".join(sections),
                scripts=["assets/export-word.js"],
                nav_current="fichiers_travail.html",
                page_header=header,
                breadcrumb=html_breadcrumb(
                    ("Accueil", "index.html"),
                    ("BAB encodé", "bab_encodes.html"),
                    (doc["chercheur"], None),
                ),
            ),
        )
    for path in SITE.glob("bab_encode_*.html"):
        if path.name not in expected:
            path.unlink()


if __name__ == "__main__":
    SITE.mkdir(exist_ok=True)
    write_text(SITE / ".nojekyll", "")
    write_text(SITE / "assets" / "style.css", STYLE)
    export_word_js = (ROOT / "assets" / "export-word.js").read_text(encoding="utf-8")
    write_text(SITE / "assets" / "export-word.js", export_word_js)
    all_capsules = load_capsules()
    all_segments = load_segments()
    all_affectations = load_affectations()
    programme_table = load_programme_table()
    experts_profils = load_experts_profils()
    expected_capsule_pages = {f"capsule_{capsule['code']}.html" for capsule in all_capsules}
    for path in SITE.glob("capsule_*.html"):
        if path.name not in expected_capsule_pages:
            path.unlink()
    if (SITE / "cartes_chaleur.html").exists():
        (SITE / "cartes_chaleur.html").unlink()
    if (SITE / "match.html").exists():
        (SITE / "match.html").unlink()
    for path in SITE.glob("match_temoin_*.html"):
        path.unlink()
    for path in SITE.glob("match_module_*.html"):
        path.unlink()
    build_home(all_capsules, all_segments)
    build_experts_profiles_page(experts_profils)
    build_tb_edito_capsule_pages(programme_table)
    build_tb_edito_page()
    build_proposition_titres_temoin_page(programme_table)
    build_proposition_edito_pages(programme_table)
    build_fonctions_temoins_page()
    build_videos_expert_pages(programme_table, experts_profils)
    build_prev_vid_page(programme_table)
    build_suivi_intervenants_pages(programme_table, experts_profils)
    build_edito_hub_page()
    build_script_propose_pages(programme_table, all_affectations, all_segments)
    build_fichiers_travail_pages()
    build_fascicules_oser_innover_pages()
    build_mails_experts_pages(programme_table, experts_profils, all_affectations, all_segments)
    build_correspondances_edito_page(programme_table)
    build_tableau_corr_page()
    build_dashboard(all_capsules, all_segments, all_affectations)
    build_researcher_pages(all_segments)
    build_capsule_pages(all_capsules, all_segments, all_affectations, programme_table)
    for obsolete in ("conflits.html", "registre.html"):
        path = SITE / obsolete
        if path.exists():
            path.unlink()
    build_derushage_edito_index()
    build_derushage_edito_pages()
    build_bab_encodes_index()
    build_bab_encode_pages()
    print(f"Site genere dans {SITE}")

