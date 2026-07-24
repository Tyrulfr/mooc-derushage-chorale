#!/usr/bin/env python3
"""Exporte le tableau de conception Prev_Vid (13 videos temoin) au format Excel historique."""
from __future__ import annotations

import json
from collections import OrderedDict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
DEFAULT_OUT = DATA / "20260723_Prev_Vid.xlsx"
PROGRAMME_TABLE = DATA / "programme_table.json"


def export_prev_vid_xlsx(
    programme_table: dict | None = None,
    out_path: Path | None = None,
) -> Path:
    if programme_table is None:
        programme_table = json.loads(PROGRAMME_TABLE.read_text(encoding="utf-8"))
    path = out_path or DEFAULT_OUT

    wb = Workbook()
    ws = wb.active
    ws.title = "Feuil1"

    # Largeurs calquees sur data/raw/20260710_Prev_Vid.xlsx
    widths = {
        "A": 13.0,
        "B": 4.33,
        "C": 36.83,
        "D": 66.33,
        "E": 42.33,
        "F": 28.0,
        "G": 30.83,
        "H": 13.0,
        "I": 13.0,
        "J": 13.0,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="Aptos Narrow", size=12, bold=True)
    body_font = Font(name="Aptos Narrow", size=11)
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(wrap_text=True, vertical="center", horizontal="center")

    ws["D1"] = "Vidéo Temoin (A)"
    ws["E1"] = "Vidéo Expert (B)"
    for col in ("D", "E"):
        ws[f"{col}1"].font = header_font
        ws[f"{col}1"].alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        ("A", "Module"),
        ("B", "N°"),
        ("C", "Vidéo chorale témoin"),
        ("D", "Ce que racontent les chercheurs"),
        ("E", "Vidéo(s) de référent à produire"),
        ("F", "Objectif pédagogique atteint"),
        ("G", "Nom proposé"),
        ("H", "Intro "),
        ("I", "Transition"),
        ("J", "Outro"),
    ]
    for col, label in headers:
        cell = ws[f"{col}2"]
        cell.value = label
        cell.font = header_font
        cell.alignment = wrap
        cell.border = border

    groups: OrderedDict[str, list] = OrderedDict()
    for row in programme_table.get("rows", []):
        module = row.get("module") or ""
        groups.setdefault(module, []).append(row)

    current_row = 3
    first_module = True
    subheader_cols = ["B", "C", "D", "E", "F"]

    for module, rows in groups.items():
        if not first_module:
            current_row += 1
            for col in subheader_cols:
                cell = ws[f"{col}{current_row}"]
                cell.value = ws[f"{col}2"].value
                cell.font = header_font
                cell.alignment = wrap
                cell.border = border
            current_row += 1
        first_module = False

        start = current_row
        for item in rows:
            ws[f"A{current_row}"] = module
            ws[f"B{current_row}"] = item.get("code", "")
            ws[f"C{current_row}"] = item.get("video_temoin", "")
            ws[f"D{current_row}"] = item.get("resume_chercheurs", "")
            ws[f"E{current_row}"] = item.get("videos_referent", "")
            ws[f"F{current_row}"] = item.get("objectif_pedagogique", "")
            ws[f"G{current_row}"] = item.get("noms_proposes", "")
            for col in "ABCDEFGHIJ":
                cell = ws[f"{col}{current_row}"]
                cell.font = body_font
                cell.alignment = wrap
                cell.border = border
            ws[f"A{current_row}"].alignment = center
            ws[f"B{current_row}"].alignment = center
            ws.row_dimensions[current_row].height = 110
            current_row += 1

        end = current_row - 1
        if end > start:
            ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
            ws[f"A{start}"].alignment = center
            ws[f"A{start}"].font = header_font

    note = programme_table.get("note") or (
        "Les noms de la colonne « Nom propose » sont des propositions ; "
        "l'intervenant de chaque video expert reste a definir."
    )
    ws[f"D{current_row}"] = (
        f"Mise a jour {programme_table.get('date_mise_a_jour', '')} — "
        f"plan a 13 videos temoin (T1–T13). {note}"
    )
    ws[f"D{current_row}"].font = Font(name="Aptos Narrow", size=10, italic=True)
    ws[f"D{current_row}"].alignment = wrap

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


if __name__ == "__main__":
    target = export_prev_vid_xlsx()
    print(f"Excel Prev_Vid ecrit : {target}")
